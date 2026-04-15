from __future__ import annotations

"""
Shared business logic for numbering, totals, PDF generation, and audit helpers.

When a rule should be reused by multiple views or serializers, it usually
belongs here instead of being duplicated inside ``views.py`` or ``web_views.py``.
"""

import io
import base64
import re
import zipfile
from decimal import Decimal
from pathlib import Path
from typing import Optional

from django.db import transaction
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Frame, Image, KeepTogether, LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas
from django.utils import timezone
from django.utils.html import escape
from pypdf import PdfReader, PdfWriter

from . import models

try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover - optional dependency fallback
    PILImage = None


def next_fund_request_number() -> str:
    numbers = (
        models.FundRequest.objects.exclude(fund_request_number__isnull=True)
        .exclude(fund_request_number="")
        .values_list("fund_request_number", flat=True)
    )
    max_seq = 0
    for number in numbers:
        sequence = models.parse_fund_request_sequence(number)
        if sequence is None:
            continue
        max_seq = max(max_seq, sequence)
    return models.format_fund_request_number(f"FR-{max_seq + 1}")


def next_purchase_order_number() -> str:
    """
    Serial PO number format: MASM/MNPXXXYY
    Example for 2026: MASM/MNP00126
    """
    prefix = "MASM/MNP"
    year_suffix = timezone.localdate().strftime("%y")
    numbers = (
        models.PurchaseOrder.objects.exclude(purchase_order_number__isnull=True)
        .exclude(purchase_order_number="")
        .values_list("purchase_order_number", flat=True)
    )
    max_seq = 0
    for number in numbers:
        raw = str(number or "").strip().upper()
        if not raw.startswith(prefix):
            continue
        suffix = raw[len(prefix):]
        if len(suffix) != 5 or not suffix.isdigit():
            continue
        sequence_part = suffix[:3]
        year_part = suffix[3:]
        if year_part != year_suffix:
            continue
        max_seq = max(max_seq, int(sequence_part))
    return f"{prefix}{max_seq + 1:03d}{year_suffix}"


def ensure_purchase_order_number(purchase_order: models.PurchaseOrder) -> str:
    if purchase_order.purchase_order_number:
        return purchase_order.purchase_order_number
    purchase_order.purchase_order_number = next_purchase_order_number()
    purchase_order.save(update_fields=["purchase_order_number"])
    return purchase_order.purchase_order_number


def next_public_application_number() -> str:
    prefix = "P"
    latest = (
        models.PublicBeneficiaryEntry.objects.filter(application_number__startswith=prefix)
        .order_by("-application_number")
        .values_list("application_number", flat=True)
        .first()
    )
    if not latest:
        return f"{prefix}001"
    try:
        seq = int(str(latest).replace(prefix, "", 1)) + 1
    except (TypeError, ValueError):
        seq = 1
    return f"{prefix}{seq:03d}"


def next_institution_application_number() -> str:
    prefix = "I"
    latest = (
        models.InstitutionsBeneficiaryEntry.objects.filter(application_number__startswith=prefix)
        .order_by("-application_number")
        .values_list("application_number", flat=True)
        .first()
    )
    if not latest:
        return f"{prefix}001"
    try:
        seq = int(str(latest).replace(prefix, "", 1)) + 1
    except (TypeError, ValueError):
        seq = 1
    return f"{prefix}{seq:03d}"


def _find_matching_aid_article(aid_type: str | None):
    aid_label = str(aid_type or "").strip()
    queryset = models.Article.objects.filter(item_type=models.ItemTypeChoices.AID)
    if not aid_label:
        return queryset.order_by("article_name").first()

    exact = queryset.filter(article_name__iexact=aid_label).order_by("article_name").first()
    if exact:
        return exact
    exact = queryset.filter(category__iexact=aid_label).order_by("article_name").first()
    if exact:
        return exact
    partial = queryset.filter(article_name__icontains=aid_label).order_by("article_name").first()
    if partial:
        return partial
    partial = queryset.filter(category__icontains=aid_label).order_by("article_name").first()
    if partial:
        return partial
    return queryset.order_by("article_name").first()




def _resolve_fund_request_recipient_source(recipient: models.FundRequestRecipient):
    source_id = recipient.source_entry_id
    if not source_id or not recipient.beneficiary_type:
        return None
    if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        return models.DistrictBeneficiaryEntry.objects.select_related("article").filter(pk=source_id).first()
    if recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        return models.PublicBeneficiaryEntry.objects.select_related("article").filter(pk=source_id).first()
    if recipient.beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
        return models.InstitutionsBeneficiaryEntry.objects.select_related("article").filter(pk=source_id).first()
    return None


def _aid_pdf_beneficiary_label(recipient: models.FundRequestRecipient) -> str:
    source_entry = _resolve_fund_request_recipient_source(recipient)
    beneficiary_type = recipient.beneficiary_type
    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        if source_entry and getattr(source_entry, "district", None):
            return source_entry.district.district_name or "-"
        return recipient.district_name or recipient.recipient_name or "-"
    if beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        if source_entry and getattr(source_entry, "application_number", None):
            return source_entry.application_number or "-"
        text = (recipient.beneficiary or "").strip()
        if text:
            return text.split(" - ", 1)[0].strip() or "-"
        return "-"
    if beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
        if source_entry and getattr(source_entry, "application_number", None):
            return source_entry.application_number or "-"
        text = (recipient.beneficiary or "").strip()
        if text:
            return text.split(" - ", 1)[0].strip() or "-"
        return "-"
    return recipient.recipient_name or "-"


def _article_pdf_beneficiary_label(item: models.FundRequestArticle) -> str:
    article = getattr(item, "article", None)
    if not article:
        return item.beneficiary or "-"

    labels = []
    if article.district_entries.exists():
        labels.append("District")
    if article.public_entries.exists():
        labels.append("Public")
    if article.institution_entries.exists():
        labels.append("Institution")

    if labels:
        return " & ".join(labels)

    stored = (item.beneficiary or "").strip()
    return stored or "-"

def sync_order_entries_from_fund_request(fund_request: models.FundRequest, actor: Optional[models.AppUser] = None) -> None:
    with transaction.atomic():
        models.OrderEntry.objects.filter(fund_request=fund_request).delete()

        if fund_request.status != models.FundRequestStatusChoices.SUBMITTED:
            return

        created_rows = []
        today = timezone.now().date()

        if fund_request.fund_request_type == models.FundRequestTypeChoices.ARTICLE:
            for line in fund_request.articles.select_related("article").all():
                quantity = max(int(line.quantity or 0), 0)
                if not line.article_id or quantity <= 0:
                    continue
                created_rows.append(
                    models.OrderEntry(
                        article=line.article,
                        quantity_ordered=quantity,
                        order_date=today,
                        status=models.OrderStatusChoices.ORDERED,
                        supplier_name=line.vendor_name or line.cheque_in_favour or fund_request.supplier_name or None,
                        unit_price=line.unit_price or 0,
                        notes=f"Created from Fund Request {fund_request.formatted_fund_request_number or 'Draft'}",
                        created_by=actor or fund_request.created_by,
                        fund_request=fund_request,
                    )
                )

        elif fund_request.fund_request_type == models.FundRequestTypeChoices.AID:
            fallback_article = _find_matching_aid_article(fund_request.aid_type)
            for recipient in fund_request.recipients.all():
                source_entry = _resolve_fund_request_recipient_source(recipient)
                article = getattr(source_entry, "article", None) or fallback_article
                quantity = int(getattr(source_entry, "quantity", 1) or 1)
                if not article or quantity <= 0:
                    continue
                amount = recipient.fund_requested or getattr(source_entry, "total_amount", 0) or 0
                recipient_name = recipient.recipient_name or recipient.name_of_beneficiary or recipient.name_of_institution or "Recipient"
                note_bits = [f"Created from Fund Request {fund_request.formatted_fund_request_number or 'Draft'}", recipient_name]
                if recipient.details:
                    note_bits.append(f"Details: {recipient.details}")
                if recipient.cheque_in_favour:
                    note_bits.append(f"Cheque / RTGS in Favour: {recipient.cheque_in_favour}")
                created_rows.append(
                    models.OrderEntry(
                        article=article,
                        quantity_ordered=quantity,
                        order_date=today,
                        status=models.OrderStatusChoices.ORDERED,
                        supplier_name=recipient.cheque_in_favour or None,
                        unit_price=amount,
                        notes=" - ".join(note_bits),
                        created_by=actor or fund_request.created_by,
                        fund_request=fund_request,
                    )
                )

        if created_rows:
            models.OrderEntry.objects.bulk_create(created_rows)


def _infer_aid_type_from_recipients(fund_request: models.FundRequest) -> str:
    for recipient in fund_request.recipients.all():
        text = str(recipient.beneficiary or '').strip()
        if not text:
            continue
        parts = [part.strip() for part in text.split(' - ') if part.strip()]
        if len(parts) >= 2:
            return parts[1]
    return ''


def _ordinal(value: int) -> str:
    if 10 <= (value % 100) <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(value % 10, 'th')
    return f'{value}{suffix}'


def _event_programme_title(event_year: int) -> str:
    birthday_number = max(event_year - 1940, 1)
    return (
        'Payment Request Details for MASM Makkal Nala Pani Programme on the eve of '
        f'{_ordinal(birthday_number)} Birthday Celebrations of His Holiness AMMA at '
        f'Melmaruvathur on 03-03-{event_year}'
    )


def _fund_request_title(fund_request: models.FundRequest) -> str:
    request_no = fund_request.formatted_fund_request_number or 'Draft'
    date_text = timezone.localtime(fund_request.created_at).strftime('%d-%m-%Y') if fund_request.created_at else timezone.localdate().strftime('%d-%m-%Y')
    if fund_request.fund_request_type == models.FundRequestTypeChoices.ARTICLE:
        suffix = 'Article'
    else:
        suffix = (fund_request.aid_type or _infer_aid_type_from_recipients(fund_request) or 'Aid').strip()
    return f'Fund Request No: {request_no}, Dated {date_text} - {suffix}'


def _fund_request_previous_cumulative(fund_request: models.FundRequest) -> Decimal:
    total = (
        models.FundRequest.objects.filter(
            status=models.FundRequestStatusChoices.SUBMITTED,
            created_at__lt=fund_request.created_at,
        )
        .exclude(pk=fund_request.pk)
        .aggregate(total=Sum('total_amount'))
        .get('total')
        or 0
    )
    return Decimal(str(total))


def _indian_grouping(number_text: str) -> str:
    if len(number_text) <= 3:
        return number_text
    last_three = number_text[-3:]
    remaining = number_text[:-3]
    groups = []
    while len(remaining) > 2:
        groups.insert(0, remaining[-2:])
        remaining = remaining[:-2]
    if remaining:
        groups.insert(0, remaining)
    return ",".join(groups + [last_three])


def _pdf_currency(value) -> str:
    amount = Decimal(str(value or 0))
    sign = "-" if amount < 0 else ""
    amount = abs(amount).quantize(Decimal("0.01"))
    whole, fraction = f"{amount:.2f}".split(".", 1)
    return f"{sign}{_indian_grouping(whole)}.{fraction}"


def _pdf_logo_path() -> str | None:
    logo_path = Path(__file__).resolve().parent / 'static' / 'core' / 'images' / 'pdf-logo.png'
    return str(logo_path) if logo_path.exists() else None


def _pdf_guru_logo_path() -> str | None:
    guru_logo_path = Path(__file__).resolve().parent / 'static' / 'core' / 'images' / 'guru-logo.jpg'
    return str(guru_logo_path) if guru_logo_path.exists() else None


def _pdf_signature_path() -> str | None:
    signature_path = Path(__file__).resolve().parent / 'static' / 'core' / 'images' / 'pdf-sign.jpg'
    return str(signature_path) if signature_path.exists() else None


def _fitted_pdf_image(path: str | None, *, max_width_mm: float, max_height_mm: float):
    if not path:
        return Paragraph('', getSampleStyleSheet()['BodyText'])
    try:
        width_px, height_px = ImageReader(path).getSize()
        if not width_px or not height_px:
            raise ValueError("invalid image size")
        scale = min((max_width_mm * mm) / width_px, (max_height_mm * mm) / height_px)
        return Image(path, width=width_px * scale, height=height_px * scale)
    except Exception:
        return Paragraph('', getSampleStyleSheet()['BodyText'])


def _fitted_pdf_image_source(source, *, max_width_mm: float, max_height_mm: float):
    if not source:
        return Paragraph('', getSampleStyleSheet()['BodyText'])
    stream = None
    image_source = source
    try:
        if isinstance(source, (bytes, bytearray)):
            stream = io.BytesIO(source)
            image_source = stream
        width_px, height_px = ImageReader(image_source).getSize()
        if not width_px or not height_px:
            raise ValueError("invalid image size")
        scale = min((max_width_mm * mm) / width_px, (max_height_mm * mm) / height_px)
        if stream is not None:
            stream.seek(0)
            image_source = stream
        return Image(image_source, width=width_px * scale, height=height_px * scale)
    except Exception:
        return Paragraph('', getSampleStyleSheet()['BodyText'])


def _optimized_report_logo(source, mime_type: str | None = None, *, max_width_px: int = 420, max_height_px: int = 520):
    if not source:
        return source, mime_type or "image/png"
    if PILImage is None:
        return source, mime_type or "image/png"
    try:
        raw = Path(source).read_bytes() if isinstance(source, (str, Path)) else bytes(source)
        image = PILImage.open(io.BytesIO(raw))
        image.load()
        resampling = getattr(getattr(PILImage, "Resampling", PILImage), "LANCZOS", getattr(PILImage, "LANCZOS", 1))
        image.thumbnail((max_width_px, max_height_px), resampling)
        has_alpha = image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info)
        buffer = io.BytesIO()
        if has_alpha:
            image.save(buffer, format="PNG", optimize=True)
            return buffer.getvalue(), "image/png"
        image = image.convert("RGB")
        image.save(buffer, format="JPEG", quality=82, optimize=True, progressive=True)
        return buffer.getvalue(), "image/jpeg"
    except Exception:
        return source, mime_type or "image/png"


def extract_pdf_form_fields(pdf_bytes: bytes) -> list[dict]:
    if not pdf_bytes:
        return []
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return []
    fields: list[dict] = []
    seen = set()
    for page_index, page in enumerate(reader.pages, start=1):
        for annot_ref in page.get("/Annots") or []:
            try:
                annot = annot_ref.get_object()
            except Exception:
                continue
            if annot.get("/Subtype") != "/Widget":
                continue
            field_name = str(annot.get("/T") or "").strip()
            if not field_name or field_name in seen:
                continue
            rect = annot.get("/Rect") or []
            fields.append(
                {
                    "name": field_name,
                    "page": page_index,
                    "rect": [float(coord) for coord in rect[:4]] if rect else [],
                }
            )
            seen.add(field_name)
    if fields:
        return fields
    try:
        field_map = reader.get_fields() or {}
    except Exception:
        field_map = {}
    for field_name in field_map.keys():
        if field_name in seen:
            continue
        fields.append({"name": str(field_name), "page": 1, "rect": []})
        seen.add(field_name)
    return fields


def fill_pdf_form_from_rows(
    template_pdf_bytes: bytes,
    rows: list[dict],
    field_map: dict[str, str],
    *,
    derived_aid_value_key: str = "__derived_aid_value__",
) -> io.BytesIO:
    if not template_pdf_bytes:
        raise ValueError("Upload a PDF template first.")
    if not rows:
        raise ValueError("No rows are available to fill.")
    if not field_map:
        raise ValueError("Map at least one PDF field before downloading.")

    def _row_lookup(row: dict, key: str):
        if key in row:
            return row.get(key)
        normalized = str(key or "").strip().casefold()
        for row_key, value in row.items():
            if str(row_key or "").strip().casefold() == normalized:
                return value
        return ""

    def _resolved_value(row: dict, mapped_key: str) -> str:
        mapping = str(mapped_key or "").strip()
        if not mapping or mapping == "__blank__":
            return ""
        if mapping == derived_aid_value_key:
            item_type = str(_row_lookup(row, "Item Type") or "").strip().casefold()
            if item_type == str(models.ItemTypeChoices.AID).strip().casefold():
                return str(_row_lookup(row, "Total Value") or _row_lookup(row, "total_value") or "").strip()
            return ""
        value = _row_lookup(row, mapping)
        return "" if value is None else str(value)

    merged_writer = PdfWriter()
    for row in rows:
        row_reader = PdfReader(io.BytesIO(template_pdf_bytes))
        row_writer = PdfWriter()
        row_writer.append(row_reader)
        row_writer.set_need_appearances_writer()
        fill_values = {
            field_name: _resolved_value(row, mapped_key)
            for field_name, mapped_key in field_map.items()
            if str(mapped_key or "").strip()
        }
        if fill_values:
            row_writer.update_page_form_field_values(row_writer.pages[0], fill_values)
        page_buffer = io.BytesIO()
        row_writer.write(page_buffer)
        page_buffer.seek(0)
        merged_writer.append(PdfReader(page_buffer))
    merged_writer.set_need_appearances_writer()
    output = io.BytesIO()
    merged_writer.write(output)
    output.seek(0)
    return output


def _normalized_docx_report_logo(source, mime_type: str | None = None, *, canvas_width_px: int = 300, canvas_height_px: int = 300):
    if not source:
        return None, "image/png"
    if PILImage is None:
        return source, mime_type or "image/png"
    try:
        raw = Path(source).read_bytes() if isinstance(source, (str, Path)) else bytes(source)
        image = PILImage.open(io.BytesIO(raw))
        image.load()
        resampling = getattr(getattr(PILImage, "Resampling", PILImage), "LANCZOS", getattr(PILImage, "LANCZOS", 1))
        image = image.convert("RGBA")
        image.thumbnail((canvas_width_px, canvas_height_px), resampling)
        canvas_image = PILImage.new("RGBA", (canvas_width_px, canvas_height_px), (255, 255, 255, 0))
        offset_x = max((canvas_width_px - image.width) // 2, 0)
        offset_y = max((canvas_height_px - image.height) // 2, 0)
        canvas_image.paste(image, (offset_x, offset_y), image)
        buffer = io.BytesIO()
        canvas_image.save(buffer, format="PNG", optimize=True)
        return buffer.getvalue(), "image/png"
    except Exception:
        return source, mime_type or "image/png"


def generate_fund_request_pdf(fund_request: models.FundRequest) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle('mnp-normal', parent=styles['BodyText'], fontSize=8.5, leading=10)
    small = ParagraphStyle('mnp-small', parent=normal, fontSize=8.5, leading=10)
    header_title = ParagraphStyle(
        'mnp-header-title',
        parent=styles['Heading2'],
        alignment=1,
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        spaceAfter=3,
        textColor=colors.red,
    )
    header_sub = ParagraphStyle(
        'mnp-header-sub',
        parent=normal,
        alignment=1,
        fontName='Helvetica',
        fontSize=12,
        leading=14,
    )
    title_style = ParagraphStyle(
        'mnp-title',
        parent=styles['Heading3'],
        alignment=1,
        fontSize=15,
        leading=17,
        spaceBefore=6,
        fontName='Helvetica-Bold',
    )
    right_number = ParagraphStyle('mnp-right-number', parent=small, alignment=0, fontSize=8.5, leading=10)
    center_text = ParagraphStyle('mnp-center-text', parent=small, alignment=0, fontSize=8.5, leading=10)
    right_small = ParagraphStyle('mnp-right-small', parent=small, alignment=2, fontSize=8.5, leading=10)
    for_masm_style = ParagraphStyle('mnp-for-masm', parent=right_small, textColor=colors.HexColor('#166534'))
    totals_value_style = ParagraphStyle('mnp-totals-value', parent=small, alignment=2, fontSize=8.5, leading=10)

    story = []
    logo_path = _pdf_logo_path()
    guru_logo_path = _pdf_guru_logo_path()
    left_logo = _fitted_pdf_image(guru_logo_path, max_width_mm=15, max_height_mm=20)
    right_logo = _fitted_pdf_image(logo_path, max_width_mm=18, max_height_mm=24)
    event_year = timezone.localtime(fund_request.created_at).year if fund_request.created_at else timezone.localdate().year
    center_block = [
        Paragraph('OMSAKTHI', header_title),
        Paragraph(_event_programme_title(event_year), header_sub),
        Paragraph(_fund_request_title(fund_request), title_style),
    ]
    header = Table([[left_logo, center_block, right_logo]], colWidths=[24 * mm, 225 * mm, 24 * mm])
    header.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    story.extend([header, Spacer(1, 10)])

    header_bg = colors.HexColor('#e5e7eb')
    grid_color = colors.HexColor('#e5e7eb')

    if fund_request.fund_request_type == models.FundRequestTypeChoices.AID:
        rows = [[
            Paragraph('<b>SL No</b>', small),
            Paragraph('<b>Beneficiary</b>', small),
            Paragraph('<b>Name of beneficiary</b>', small),
            Paragraph('<b>Name of Institution</b>', small),
            Paragraph('<b>Details</b>', small),
            Paragraph('<b>Fund Requested</b>', small),
            Paragraph('<b>Aadhaar No</b>', small),
            Paragraph('<b>Cheque in Favour</b>', small),
            Paragraph('<b>Cheque No.</b>', small),
        ]]
        total = Decimal('0')
        for index, recipient in enumerate(fund_request.recipients.all(), start=1):
            amount = Decimal(str(recipient.fund_requested or 0))
            total += amount
            rows.append([
                Paragraph(str(index), small),
                Paragraph(_aid_pdf_beneficiary_label(recipient), small),
                Paragraph(recipient.name_of_beneficiary or recipient.recipient_name or '-', small),
                Paragraph(recipient.name_of_institution or '-', small),
                Paragraph(recipient.details or recipient.notes or '-', small),
                Paragraph(_pdf_currency(amount), right_number),
                Paragraph(recipient.aadhar_number or '-', small),
                Paragraph(recipient.cheque_in_favour or '-', small),
                Paragraph(recipient.cheque_no or '-', small),
            ])
        rows.append(['', '', '', '', Paragraph('<b>Total:</b>', small), Paragraph(f'<b>{_pdf_currency(total)}</b>', right_number), '', '', ''])
        col_widths = [12 * mm, 24 * mm, 35 * mm, 32 * mm, 26 * mm, 24 * mm, 28 * mm, 68 * mm, 24 * mm]
    else:
        rows = [[
            Paragraph('<b>SL No</b>', small),
            Paragraph('<b>Beneficiary</b>', small),
            Paragraph('<b>Article Name</b>', small),
            Paragraph('<b>GST No.</b>', small),
            Paragraph('<b>Qty</b>', small),
            Paragraph('<b>Price (Incl GST)</b>', small),
            Paragraph('<b>Value</b>', small),
            Paragraph('<b>Cheque in Favour</b>', small),
            Paragraph('<b>Cheque No.</b>', small),
        ]]
        total = Decimal('0')
        for index, item in enumerate(fund_request.articles.all(), start=1):
            value = Decimal(str(item.value or 0))
            total += value
            rows.append([
                Paragraph(str(item.sl_no or index), small),
                Paragraph(_article_pdf_beneficiary_label(item), small),
                Paragraph(item.article_name or '-', small),
                Paragraph(fund_request.gst_number or item.gst_no or '-', small),
                Paragraph(str(item.quantity or 0), center_text),
                Paragraph(_pdf_currency(item.price_including_gst or item.unit_price or 0), right_number),
                Paragraph(_pdf_currency(value), right_number),
                Paragraph(item.cheque_in_favour or '-', small),
                Paragraph(item.cheque_no or '-', small),
            ])
        rows.append(['', '', '', '', Paragraph('<b>Total:</b>', small), '', Paragraph(f'<b>{_pdf_currency(total)}</b>', right_number), '', ''])
        col_widths = [12 * mm, 24 * mm, 50 * mm, 24 * mm, 14 * mm, 24 * mm, 24 * mm, 77 * mm, 24 * mm]

    table = LongTable(rows, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.hAlign = 'LEFT'
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('LINEBELOW', (0, 0), (-1, -1), 0.4, grid_color),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f8fafc')),
    ]))
    story.extend([table, Spacer(1, 12)])

    prev_total = _fund_request_previous_cumulative(fund_request)
    current_total = Decimal(str(fund_request.total_amount or 0))
    grand_total = prev_total + current_total
    signature_path = _pdf_signature_path()
    approvals_left = Paragraph(
        f"1. MASM PRESIDENT'S {event_year} APPROVAL COPY<br/>2. QUOTATION / BANK / REQUEST COPIES",
        normal,
    )
    approvals_right = [Paragraph('FOR MASM', for_masm_style)]
    if signature_path:
        signature_image = Image(signature_path, width=34 * mm, height=14 * mm)
        signature_image.hAlign = 'RIGHT'
        approvals_right.append(signature_image)
    approvals_right.extend([
        Paragraph('R.Surendranath', right_small),
        Paragraph('JS - Social Welfare Activities', right_small),
    ])
    approval_table = Table([[approvals_left, approvals_right]], colWidths=[165 * mm, 108 * mm])
    approval_table.hAlign = 'LEFT'
    approval_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (0, 0), 10 * mm),
        ('LEFTPADDING', (1, 0), (1, 0), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 8 * mm),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    story.extend([approval_table, Spacer(1, 10)])

    totals = Table([
        [Paragraph('<b>PREVIOUS CUMULATIVE(Rs.)</b>', normal), Paragraph(_pdf_currency(prev_total), totals_value_style)],
        [Paragraph('<b>CURRENT FUND REQUEST(Rs.)</b>', normal), Paragraph(_pdf_currency(current_total), totals_value_style)],
        [Paragraph('<b>TOTAL(Rs.)</b>', normal), Paragraph(_pdf_currency(grand_total), totals_value_style)],
    ], colWidths=[205 * mm, 68 * mm])
    totals.hAlign = 'LEFT'
    totals.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e5e7eb')),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.extend([totals, Spacer(1, 10)])

    def draw_footer(canvas, _doc_obj):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawString(doc.leftMargin, 8 * mm, fund_request.formatted_fund_request_number or 'Draft')
        canvas.drawCentredString(landscape(A4)[0] / 2, 8 * mm, f'Page {canvas.getPageNumber()}')
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    buffer.seek(0)
    return buffer


def generate_purchase_order_pdf(purchase_order: models.PurchaseOrder) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("po-normal", parent=styles["BodyText"], fontSize=9, leading=11)
    small = ParagraphStyle("po-small", parent=normal, fontSize=8.5, leading=10)
    center_title = ParagraphStyle(
        "po-title",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=20,
        textColor=colors.HexColor("#008000"),
        spaceAfter=4,
    )
    meta_right = ParagraphStyle(
        "po-meta-right",
        parent=normal,
        alignment=2,
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
    )
    green_header = colors.HexColor("#008000")
    light_border = colors.HexColor("#e5e7eb")

    logo_path = _pdf_logo_path()
    guru_logo_path = _pdf_guru_logo_path()
    signature_path = _pdf_signature_path()
    left_logo = _fitted_pdf_image(guru_logo_path, max_width_mm=28, max_height_mm=24)
    right_logo = _fitted_pdf_image(logo_path, max_width_mm=28, max_height_mm=24)

    po_number = purchase_order.purchase_order_number or next_purchase_order_number()
    current_date = timezone.localtime(purchase_order.created_at).strftime("%d-%m-%y") if purchase_order.created_at else timezone.localdate().strftime("%d-%m-%y")

    left_block = [
        left_logo,
        Spacer(1, 3),
        Paragraph("Melmaruvathur Adhiparasakthi Spiritual Movement", small),
        Paragraph("GST Road, Melmaruvathur 603319", small),
        Paragraph("Chengalpet District, Tamilnadu", small),
        Paragraph("GST NO: 33AACTM0073D1Z5.", small),
        Paragraph("Email: maruvoorhelp@gmail.com", small),
    ]
    center_block = [Spacer(1, 14), Paragraph("PURCHASE ORDER", center_title)]
    right_block = [
        right_logo,
        Spacer(1, 3),
        Paragraph(f"PO No: {po_number}", meta_right),
        Paragraph(f"DATE: {current_date}", meta_right),
    ]

    header = Table([[left_block, center_block, right_block]], colWidths=[60 * mm, 70 * mm, 58 * mm])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    vendor_lines = [
        Paragraph(purchase_order.vendor_name or "-", normal),
        Paragraph((purchase_order.vendor_address or "-").replace("\n", "<br/>"), small),
        Paragraph(
            ", ".join(
                part for part in [
                    purchase_order.vendor_city or "",
                    purchase_order.vendor_state or "",
                    purchase_order.vendor_pincode or "",
                ] if part
            ) or "-",
            small,
        ),
        Paragraph(f"GST No: {purchase_order.gst_number or '-'}", small),
    ]
    ship_to_lines = [
        Paragraph("Melmaruvathur Adhiparasakthi Spiritual Movement", small),
        Paragraph("GST Road, Melmaruvathur 603319", small),
        Paragraph("Chengalpet District, Tamilnadu", small),
    ]
    vendor_table = Table(
        [
            [
                Paragraph("<b>VENDOR</b>", ParagraphStyle("po-section-head", parent=small, textColor=colors.white)),
                Paragraph("<b>SHIP TO</b>", ParagraphStyle("po-section-head-right", parent=small, textColor=colors.white)),
            ],
            [vendor_lines, ship_to_lines],
        ],
        colWidths=[91 * mm, 91 * mm],
    )
    vendor_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), green_header),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (-1, 1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BOX", (0, 1), (-1, 1), 0.5, light_border),
        ("INNERGRID", (0, 1), (-1, 1), 0.5, light_border),
    ]))

    header_left = ParagraphStyle(
        "po-header-left",
        parent=small,
        alignment=0,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )
    header_center = ParagraphStyle(
        "po-header-center",
        parent=small,
        alignment=1,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )
    item_rows = [[
        Paragraph('ITEM NAME', header_left),
        Paragraph('DESCRIPTION', header_left),
        Paragraph('QTY', header_center),
        Paragraph('UNIT PRICE', header_center),
        Paragraph('TOTAL<br/><font size="8">(Inclusive of Tax)</font>', header_center),
    ]]
    total_amount = Decimal("0")
    for item in purchase_order.items.all():
        line_total = Decimal(str(item.total_value or 0))
        total_amount += line_total
        item_rows.append([
            Paragraph(item.supplier_article_name or item.article_name or "-", ParagraphStyle("po-item-left", parent=small, alignment=0)),
            Paragraph((item.description or "").replace("\n", "<br/>") or "-", ParagraphStyle("po-desc-left", parent=small, alignment=0)),
            str(item.quantity or 0),
            _pdf_currency(item.unit_price or 0),
            _pdf_currency(line_total),
        ])
    item_rows.append([
        Paragraph("", small),
        Paragraph("", small),
        Paragraph("", small),
        Paragraph("<b>TOTAL<br/><font size='8'>(Inclusive of Tax)</font></b>", ParagraphStyle("po-total-label", parent=small, alignment=1)),
        f"{_pdf_currency(total_amount)}",
    ])

    item_table = Table(item_rows, colWidths=[46 * mm, 46 * mm, 18 * mm, 36 * mm, 36 * mm], repeatRows=1)
    item_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), green_header),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, light_border),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
        ("ALIGN", (2, 0), (4, 0), "CENTER"),
        ("ALIGN", (2, 1), (4, -1), "CENTER"),
        ("FONTNAME", (2, 1), (4, -2), "Helvetica"),
        ("FONTSIZE", (2, 1), (4, -2), 8.5),
        ("TEXTCOLOR", (2, 1), (4, -2), colors.black),
        ("FONTNAME", (4, -1), (4, -1), "Helvetica-Bold"),
        ("FONTSIZE", (4, -1), (4, -1), 8.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (2, 0), (4, -1), 0),
        ("RIGHTPADDING", (2, 0), (4, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f5f5f5")),
    ]))

    comments_text = (purchase_order.comments or "").strip() or models.PURCHASE_ORDER_DEFAULT_COMMENTS
    comments_text = escape(comments_text).replace("\n", "<br/>")
    comments_table = Table(
        [
            [Paragraph('<font color="white"><b>Comments or Special Instructions</b></font>', small)],
            [Paragraph(comments_text, small)],
        ],
        colWidths=[182 * mm],
    )
    comments_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), green_header),
        ("BOX", (0, 1), (0, 1), 0.5, light_border),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 24),
    ]))

    signature_bits = [Paragraph("<b>Authorised Signatory</b>", small)]
    if signature_path:
        signature_image = Image(signature_path, width=34 * mm, height=14 * mm)
        signature_image.hAlign = "LEFT"
        signature_bits.append(signature_image)
    signature_bits.extend([
        Paragraph("R.Surendranath", small),
        Paragraph("JS - Social Welfare Activities", small),
    ])

    story = [
        header,
        Spacer(1, 10),
        vendor_table,
        Spacer(1, 12),
        item_table,
        Spacer(1, 12),
        comments_table,
        Spacer(1, 10),
        Table([[signature_bits]], colWidths=[70 * mm], hAlign="LEFT", style=TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ])),
    ]

    def draw_footer(canvas, _doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 9)
        canvas.drawCentredString(A4[0] / 2, 10 * mm, "If you have any questions about this purchase order, please contact")
        canvas.drawCentredString(A4[0] / 2, 6 * mm, "+91 98400 46263, maruvoorhelp@gmail.com")
        canvas.restoreState()

    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    buffer.seek(0)
    return buffer


def sync_fund_request_totals(fund_request: models.FundRequest) -> None:
    with transaction.atomic():
        if fund_request.fund_request_type == models.FundRequestTypeChoices.AID:
            total_value = models.FundRequestRecipient.objects.filter(fund_request=fund_request).aggregate(
                total=Sum("fund_requested")
            ).get("total") or 0
        else:
            total_value = models.FundRequestArticle.objects.filter(fund_request=fund_request).aggregate(
                total=Sum("value")
            ).get("total") or 0
        fund_request.total_amount = total_value
        fund_request.save(update_fields=["total_amount"])


def sync_purchase_order_totals(purchase_order: models.PurchaseOrder) -> None:
    with transaction.atomic():
        total_value = models.PurchaseOrderItem.objects.filter(purchase_order=purchase_order).aggregate(
            total=Sum("total_value")
        ).get("total") or 0
        purchase_order.total_amount = total_value
        purchase_order.save(update_fields=["total_amount"])


def log_audit(
    *,
    user: Optional[models.AppUser],
    action_type: str,
    entity_type: str,
    entity_id: str | None,
    details: dict,
    ip_address: str | None = None,
    user_agent: str | None = None,
) -> None:
    models.AuditLog.objects.create(
        user=user,
        action_type=action_type,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=ip_address,
        user_agent=user_agent,
    )


def mark_fund_request_status(
    fund_request: models.FundRequest,
    status: models.FundRequestStatusChoices,
    actor: Optional[models.AppUser] = None,
) -> models.FundRequest:
    prev = fund_request.status
    fund_request.status = status
    if status in {models.FundRequestStatusChoices.SUBMITTED, models.FundRequestStatusChoices.APPROVED, models.FundRequestStatusChoices.REJECTED}:
        sync_fund_request_totals(fund_request)
    fund_request.save(update_fields=["status"])
    log_audit(
        user=actor,
        action_type=models.ActionTypeChoices.STATUS_CHANGE,
        entity_type="fund_request",
        entity_id=str(fund_request.id),
        details={"from": prev, "to": status},
    )
    return fund_request


LABEL_LAYOUTS = {
    "12L": {
        "page_size": portrait(A4),
        "top_margin": 9 * mm,
        "side_margin": 4 * mm,
        "vertical_pitch": 47 * mm,
        "horizontal_pitch": 102 * mm,
        "label_height": 44 * mm,
        "label_width": 100 * mm,
        "number_across": 2,
        "number_down": 6,
        "token_font": 60,
        "token_font_medium": 50,
        "token_font_small": 47,
        "line_font": 15,
        "line_leading": 16,
        "name_leading": 14,
        "name_space_after": 20,
        "article_y": 6 * mm,
        "name_vertical_factor": 0.85,
    },
    "2L": {
        "page_size": portrait(A4),
        "top_margin": 2.5 * mm,
        "side_margin": 5 * mm,
        "vertical_pitch": 146 * mm,
        "horizontal_pitch": 200 * mm,
        "label_height": 146 * mm,
        "label_width": 200 * mm,
        "number_across": 1,
        "number_down": 2,
        "token_font": 80,
        "token_font_medium": 80,
        "token_font_small": 70,
        "line_font": 35,
        "line_leading": 35,
        "name_leading": 35,
        "name_space_after": 0,
        "article_y": 35 * mm,
        "name_vertical_factor": 0.66,
    },
    "A4": {
        "page_size": landscape(A4),
        "top_margin": 12 * mm,
        "side_margin": 12 * mm,
        "vertical_pitch": 0,
        "horizontal_pitch": 0,
        "label_height": landscape(A4)[1] - 24 * mm,
        "label_width": landscape(A4)[0] - 24 * mm,
        "number_across": 1,
        "number_down": 1,
    },
}


def _label_token_font_size(token_text: str, config: dict) -> int:
    if len(token_text) >= 4:
        return config["token_font_small"]
    if len(token_text) == 3:
        return config["token_font_medium"]
    return config["token_font"]


def _draw_standard_label(pdf_canvas, entry: dict, *, x: float, y: float, config: dict) -> None:
    token_text = str(entry.get("token") or "")
    name_text = str(entry.get("name") or "")
    article_text = str(entry.get("article") or "")

    token_style = ParagraphStyle(
        name=f"token-{config['label_width']}",
        fontSize=_label_token_font_size(token_text, config),
        alignment=0,
    )
    article_style = ParagraphStyle(
        name=f"article-{config['label_width']}",
        fontSize=config["line_font"],
        leading=config["line_leading"],
        alignment=1,
    )
    name_style = ParagraphStyle(
        name=f"name-{config['label_width']}",
        fontSize=config["line_font"],
        leading=config["name_leading"],
        spaceAfter=config.get("name_space_after", 0),
        alignment=1,
    )

    token_para = Paragraph(f"<b>{escape(token_text)}</b>", token_style)
    article_para = Paragraph(f"<b>{escape(article_text)}</b>", article_style)
    name_para = Paragraph(f"<b>{escape(name_text)}</b>", name_style)

    label_width = config["label_width"]
    label_height = config["label_height"]

    token_width, token_height = token_para.wrap(label_width / 2, label_height)
    token_para.drawOn(pdf_canvas, x + 8 * mm, y + 0.66 * (label_height - token_height))

    name_width, name_height = name_para.wrap(label_width / 2, label_height)
    name_para.drawOn(
        pdf_canvas,
        x + 0.8 * (label_width - name_width),
        y + config["name_vertical_factor"] * (label_height - name_height),
    )

    article_width, article_height = article_para.wrap(label_width / 2, label_height)
    article_para.drawOn(pdf_canvas, x + 0.8 * (label_width - article_width), y + config["article_y"])


def generate_mnp_labels_pdf(entries: list[dict], *, layout: str = "12L", mode: str = "continuous") -> io.BytesIO:
    config = LABEL_LAYOUTS[layout]
    buffer = io.BytesIO()
    pdf_canvas = canvas.Canvas(buffer, pagesize=config["page_size"])
    _width, height = config["page_size"]
    per_page = config["number_across"] * config["number_down"]
    slot_index = 0
    current_group = None

    for entry in entries:
        group_key = str(entry.get("group") or "")
        if mode == "separate" and current_group is not None and group_key != current_group:
            pdf_canvas.showPage()
            slot_index = 0
        elif slot_index and slot_index % per_page == 0:
            pdf_canvas.showPage()
            slot_index = 0

        current_group = group_key if mode == "separate" else current_group
        col = slot_index % config["number_across"]
        row = slot_index // config["number_across"] % config["number_down"]
        x = config["side_margin"] + col * config["horizontal_pitch"]
        y = height - config["top_margin"] - row * config["vertical_pitch"] - config["label_height"]
        _draw_standard_label(pdf_canvas, entry, x=x, y=y, config=config)
        slot_index += 1

    pdf_canvas.save()
    buffer.seek(0)
    return buffer


def generate_mnp_custom_labels_pdf(entries, *, layout: str = "12L") -> io.BytesIO:
    config = LABEL_LAYOUTS[layout]
    buffer = io.BytesIO()
    pdf_canvas = canvas.Canvas(buffer, pagesize=config["page_size"])
    _width, height = config["page_size"]
    per_page = config["number_across"] * config["number_down"]
    default_font_size = 72 if layout == "A4" else 30 if layout == "12L" else 54

    normalized_entries = []
    if isinstance(entries, list):
        for entry in entries:
            if isinstance(entry, dict):
                text = str(entry.get("text") or "").strip()
                count = int(entry.get("count") or 0)
                font_size = int(entry.get("font_size") or default_font_size)
                line_spacing = entry.get("line_spacing")
            elif isinstance(entry, (list, tuple)) and len(entry) >= 2:
                text = str(entry[0] or "").strip()
                count = int(entry[1] or 0)
                font_size = int(entry[2] or default_font_size) if len(entry) >= 3 else default_font_size
                line_spacing = int(entry[3] or 0) if len(entry) >= 4 else None
            else:
                text = str(entry or "").strip()
                count = 1
                font_size = default_font_size
                line_spacing = None
            font_size = max(8, min(font_size, 120))
            if line_spacing is not None:
                line_spacing = max(0, min(int(line_spacing), 120))
            if text and count > 0:
                normalized_entries.extend([{"text": text, "font_size": font_size, "line_spacing": line_spacing}] * count)
    else:
        text = str(entries or "").strip()
        if text:
            normalized_entries.append({"text": text, "font_size": default_font_size, "line_spacing": None})

    def _custom_label_markup(raw_text: str) -> str:
        escaped = escape(str(raw_text or "").strip())
        escaped = escaped.replace("\r\n", "\n").replace("\r", "\n")
        escaped = escaped.replace("|", "<br/>").replace("\n", "<br/>")
        return escaped

    for index, entry in enumerate(normalized_entries):
        text = str(entry.get("text") or "").strip()
        font_size = int(entry.get("font_size") or default_font_size)
        line_spacing = entry.get("line_spacing")
        line_leading = font_size + (6 if layout == "A4" else 2 if layout == "12L" else 4)
        if line_spacing is not None:
            line_leading = font_size + max(0, min(int(line_spacing), 120))
        if index and index % per_page == 0:
            pdf_canvas.showPage()
        slot_index = index % per_page
        col = slot_index % config["number_across"]
        row = slot_index // config["number_across"] % config["number_down"]
        x = config["side_margin"] + col * config["horizontal_pitch"]
        y = height - config["top_margin"] - row * config["vertical_pitch"] - config["label_height"]

        style = ParagraphStyle(
            name=f"custom-{layout}-{font_size}",
            fontSize=font_size,
            leading=line_leading,
            alignment=1,
        )
        para = Paragraph(f"<b>{_custom_label_markup(text)}</b>", style)
        wrapped_width, wrapped_height = para.wrap(config["label_width"] - 10 * mm, config["label_height"] - 10 * mm)
        para.drawOn(
            pdf_canvas,
            x + (config["label_width"] - wrapped_width) / 2,
            y + (config["label_height"] - wrapped_height) / 2,
        )

    pdf_canvas.save()
    buffer.seek(0)
    return buffer


def _waiting_hall_programme_lines(event_age_label: str, event_date) -> tuple[str, str]:
    age_label = str(event_age_label or "").strip() or "Event"
    if hasattr(event_date, "strftime"):
        line_one_date = event_date.strftime("%d-%m-%Y")
        line_two_date = event_date.strftime("%d.%m.%Y")
    else:
        line_one_date = str(event_date or "").strip()
        line_two_date = line_one_date
    line_one = f"MASM Social Welfare Programme on the eve of {age_label} Birthday ({line_one_date})"
    line_two = f"His Holiness AMMA at Melmaruvathur on {line_two_date}"
    return line_one, line_two


def _district_signature_programme_lines() -> tuple[str, str]:
    current_year = timezone.localdate().year
    birthday_number = max(current_year - 1940, 1)
    line_one = (
        "MASM Social Welfare Programme on the eve of "
        f"{_ordinal(birthday_number)} Birthday (03-03-{current_year})"
    )
    line_two = f"His Holiness AMMA at Melmaruvathur on 03.03.{current_year}"
    return line_one, line_two


def _waiting_hall_pack_pages(groups: list[dict]) -> list[list[dict]]:
    pages: list[list[dict]] = []
    current_page: list[dict] = []
    current_units = 0
    page_capacity = 18
    for group in groups:
        row_count = len(group.get("items") or [])
        block_units = 6 + row_count
        if not current_page:
            current_page = [group]
            current_units = block_units
            continue
        if len(current_page) < 2 and (current_units + block_units) <= page_capacity:
            current_page.append(group)
            current_units += block_units
        else:
            pages.append(current_page)
            current_page = [group]
            current_units = block_units
    if current_page:
        pages.append(current_page)
    return pages


def _waiting_hall_logo_data_uri(source, mime_type: str | None = None) -> str:
    if not source:
        return ""
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            return ""
        raw = path.read_bytes()
        mime = mime_type or "image/png"
    else:
        raw = bytes(source)
        mime = mime_type or "image/png"
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _waiting_hall_layout_profile(item_count: int, *, paired: bool = False) -> dict:
    if item_count >= 14:
        return {
            "logo_width_mm": 24,
            "logo_height_mm": 28,
            "logo_col_width_mm": 26,
            "header_col_width_mm": 132,
            "title_font": 10,
            "title_leading": 11,
            "sub_font": 8,
            "sub_leading": 10,
            "section_font": 10,
            "section_leading": 11,
            "district_font": 13,
            "district_leading": 15,
            "header_font": 14,
            "header_leading": 16,
            "body_font": 11,
            "body_leading": 13,
            "note_font": 12,
            "note_leading": 14,
            "table_top_bottom_padding": 4,
            "space_after_header_mm": 3,
            "space_after_table_mm": 4,
            "space_before_signature_mm": 8,
        }
    if item_count >= 11:
        return {
            "logo_width_mm": 26,
            "logo_height_mm": 30,
            "logo_col_width_mm": 29,
            "header_col_width_mm": 126,
            "title_font": 10,
            "title_leading": 11,
            "sub_font": 9,
            "sub_leading": 10,
            "section_font": 11,
            "section_leading": 12,
            "district_font": 14,
            "district_leading": 16,
            "header_font": 16,
            "header_leading": 18,
            "body_font": 12,
            "body_leading": 14,
            "note_font": 13,
            "note_leading": 15,
            "table_top_bottom_padding": 5,
            "space_after_header_mm": 3,
            "space_after_table_mm": 5,
            "space_before_signature_mm": 10,
        }
    return {
        "logo_width_mm": 28,
        "logo_height_mm": 34,
        "logo_col_width_mm": 32,
        "header_col_width_mm": 120,
        "title_font": 11,
        "title_leading": 12,
        "sub_font": 10,
        "sub_leading": 12,
        "section_font": 12,
        "section_leading": 13,
        "district_font": 15,
        "district_leading": 18,
        "header_font": 18,
        "header_leading": 20,
        "body_font": 14,
        "body_leading": 16,
        "note_font": 14,
        "note_leading": 16,
        "table_top_bottom_padding": 8,
        "space_after_header_mm": 4,
        "space_after_table_mm": 6,
        "space_before_signature_mm": 12,
    }


def _waiting_hall_section_flowables(
    group: dict,
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    right_logo_source=None,
    paired=False,
):
    styles = getSampleStyleSheet()
    item_count = len(group.get("items") or [])
    profile = _waiting_hall_layout_profile(item_count)
    title_style = ParagraphStyle(
        "waiting-hall-om",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["title_font"],
        leading=profile["title_leading"],
        textColor=colors.red,
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "waiting-hall-sub",
        parent=styles["BodyText"],
        alignment=1,
        fontSize=profile["sub_font"],
        leading=profile["sub_leading"],
        textColor=colors.black,
        spaceAfter=1,
    )
    section_title_style = ParagraphStyle(
        "waiting-hall-section-title",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["section_font"],
        leading=profile["section_leading"],
        textColor=colors.black,
        spaceAfter=6,
    )
    district_style = ParagraphStyle(
        "waiting-hall-district",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["district_font"],
        leading=profile["district_leading"],
        spaceAfter=8,
    )
    table_header_text = ParagraphStyle(
        "waiting-hall-table-header-text",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=profile["header_font"],
        leading=profile["header_leading"],
    )
    table_header_qty_text = ParagraphStyle(
        "waiting-hall-table-header-qty-text",
        parent=table_header_text,
        alignment=1,
    )
    table_body_text = ParagraphStyle(
        "waiting-hall-table-body-text",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=profile["body_font"],
        leading=profile["body_leading"],
    )
    table_body_qty_text = ParagraphStyle(
        "waiting-hall-table-body-qty-text",
        parent=table_body_text,
        alignment=1,
    )
    note_style = ParagraphStyle(
        "waiting-hall-note",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["note_font"],
        leading=profile["note_leading"],
    )
    signature_style = ParagraphStyle(
        "waiting-hall-signature",
        parent=styles["BodyText"],
        alignment=2,
        fontName="Helvetica-Bold",
        fontSize=profile["note_font"],
        leading=profile["note_leading"],
    )
    header_line_one, header_line_two = _waiting_hall_programme_lines(event_age_label, event_date)
    left_logo = _fitted_pdf_image_source(custom_logo, max_width_mm=profile["logo_width_mm"], max_height_mm=profile["logo_height_mm"])
    right_logo = _fitted_pdf_image_source(right_logo_source, max_width_mm=profile["logo_width_mm"], max_height_mm=profile["logo_height_mm"])
    header = Table(
        [[left_logo, [
            Paragraph("OM SAKTHI", title_style),
            Paragraph(header_line_one, sub_style),
            Paragraph(header_line_two, sub_style),
            Paragraph("District Waiting Hall Articles Collection List", section_title_style),
        ], right_logo]],
        colWidths=[profile["logo_col_width_mm"] * mm, profile["header_col_width_mm"] * mm, profile["logo_col_width_mm"] * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    table_rows = [
        [Paragraph("<b>Article list</b>", table_header_text), Paragraph("<b>Qty</b>", table_header_qty_text)],
    ]
    for item in group.get("items") or []:
        table_rows.append([
            Paragraph(escape(str(item.get("requested_item") or "")), table_body_text),
            Paragraph(str(item.get("quantity") or 0), table_body_qty_text),
        ])
    table_rows.append([
        Paragraph("<b>Total</b>", table_body_text),
        Paragraph(f"<b>{group.get('total_quantity') or 0}</b>", table_body_qty_text),
    ])
    items_table = Table(table_rows, colWidths=[144 * mm, 28 * mm], repeatRows=1)
    items_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), profile["table_top_bottom_padding"]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), profile["table_top_bottom_padding"]),
    ]))
    return [
        header,
        Spacer(1, profile["space_after_header_mm"] * mm),
        Paragraph(
            f"{escape(str(group.get('label_prefix') or 'Beneficiary'))}: {escape(str(group.get('entity_name') or '-').upper())}",
            district_style,
        ),
        items_table,
        Spacer(1, profile["space_after_table_mm"] * mm),
        Paragraph(
            "I Acknowledge that I have received the above articles",
            ParagraphStyle(
                "waiting-hall-note-smaller",
                parent=note_style,
                fontSize=max(note_style.fontSize - 2, 8),
                leading=max(note_style.leading - 2, 10),
            ),
        ),
        Spacer(1, (profile["space_before_signature_mm"] + 6) * mm),
        Paragraph("Signature", signature_style),
    ]


def _waiting_hall_section_story(
    group: dict,
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    right_logo_source=None,
    paired=False,
):
    return KeepTogether(
        _waiting_hall_section_flowables(
            group,
            event_age_label=event_age_label,
            event_date=event_date,
            custom_logo=custom_logo,
            right_logo_source=right_logo_source,
            paired=paired,
        )
    )


def generate_waiting_hall_acknowledgment_pdf(
    groups: list[dict],
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=12 * mm,
    )
    story = []
    optimized_custom_logo, _ = _optimized_report_logo(custom_logo)
    optimized_right_logo, _ = _optimized_report_logo(_pdf_logo_path(), "image/png")
    packed_pages = _waiting_hall_pack_pages(groups)
    for page_index, page_groups in enumerate(packed_pages):
        for group_index, group in enumerate(page_groups):
            story.append(
                _waiting_hall_section_story(
                    group,
                    event_age_label=event_age_label,
                    event_date=event_date,
                    custom_logo=optimized_custom_logo,
                    right_logo_source=optimized_right_logo,
                    paired=False,
                )
            )
            if group_index < len(page_groups) - 1:
                story.append(Spacer(1, 10 * mm))
        if page_index < len(packed_pages) - 1:
            story.append(PageBreak())
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_waiting_hall_acknowledgment_doc(
    groups: list[dict],
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    custom_logo_mime_type: str | None = None,
) -> bytes:
    def _twips_from_mm(value_mm: float) -> int:
        return int(round(float(value_mm) * 56.6929))

    def _twips_from_pt(value_pt: float) -> int:
        return int(round(float(value_pt) * 20))

    def _half_points(value_pt: float) -> int:
        return int(round(float(value_pt) * 2))

    def _emu_from_mm(value_mm: float) -> int:
        return int(round(float(value_mm) * 36000))

    def _xml_text(value: str) -> str:
        return escape(str(value or ""))

    def _w_paragraph(text: str = "", *, align: str | None = None, bold: bool = False, size_pt: float | None = None,
                     color: str | None = None,
                     spacing_before_pt: float | None = None, spacing_after_pt: float | None = None) -> str:
        paragraph_props = []
        if align:
            paragraph_props.append(f'<w:jc w:val="{align}"/>')
        spacing_attrs = []
        if spacing_before_pt is not None:
            spacing_attrs.append(f'w:before="{_twips_from_pt(spacing_before_pt)}"')
        if spacing_after_pt is not None:
            spacing_attrs.append(f'w:after="{_twips_from_pt(spacing_after_pt)}"')
        if spacing_attrs:
            paragraph_props.append(f"<w:spacing {' '.join(spacing_attrs)}/>")
        ppr = f"<w:pPr>{''.join(paragraph_props)}</w:pPr>" if paragraph_props else ""
        run_props = []
        if bold:
            run_props.append("<w:b/>")
        if size_pt is not None:
            run_props.append(f'<w:sz w:val="{_half_points(size_pt)}"/>')
        if color:
            run_props.append(f'<w:color w:val="{color}"/>')
        rpr = f"<w:rPr>{''.join(run_props)}</w:rPr>" if run_props else ""
        return f"<w:p>{ppr}<w:r>{rpr}<w:t xml:space=\"preserve\">{_xml_text(text)}</w:t></w:r></w:p>"

    def _w_empty_paragraph(*, spacing_after_pt: float | None = None) -> str:
        return _w_paragraph("", spacing_after_pt=spacing_after_pt)

    def _w_table_cell(inner_xml: str, *, width_twips: int, align: str | None = None, shading: str | None = None,
                      borders: bool = True) -> str:
        tc_pr = [f'<w:tcW w:w="{width_twips}" w:type="dxa"/>', '<w:vAlign w:val="center"/>']
        if not borders:
            tc_pr.append(
                "<w:tcBorders><w:top w:val=\"nil\"/><w:left w:val=\"nil\"/><w:bottom w:val=\"nil\"/><w:right w:val=\"nil\"/></w:tcBorders>"
            )
        if shading:
            tc_pr.append(f'<w:shd w:val="clear" w:color="auto" w:fill="{shading}"/>')
        if align:
            inner_xml = inner_xml.replace("<w:p>", f"<w:p><w:pPr><w:jc w:val=\"{align}\"/></w:pPr>", 1)
        return f"<w:tc><w:tcPr>{''.join(tc_pr)}</w:tcPr>{inner_xml}</w:tc>"

    def _w_image_paragraph(rel_id: str, *, width_mm: float, height_mm: float, docpr_id: int) -> str:
        cx = _emu_from_mm(width_mm)
        cy = _emu_from_mm(height_mm)
        return f"""
        <w:p>
          <w:pPr><w:jc w:val="center"/></w:pPr>
          <w:r>
            <w:drawing>
              <wp:inline distT="0" distB="0" distL="0" distR="0">
                <wp:extent cx="{cx}" cy="{cy}"/>
                <wp:docPr id="{docpr_id}" name="Picture {docpr_id}"/>
                <a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
                  <a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">
                    <pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
                      <pic:nvPicPr>
                        <pic:cNvPr id="{docpr_id}" name="Picture {docpr_id}"/>
                        <pic:cNvPicPr/>
                      </pic:nvPicPr>
                      <pic:blipFill>
                        <a:blip r:embed="{rel_id}"/>
                        <a:stretch><a:fillRect/></a:stretch>
                      </pic:blipFill>
                      <pic:spPr>
                        <a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
                        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
                      </pic:spPr>
                    </pic:pic>
                  </a:graphicData>
                </a:graphic>
              </wp:inline>
            </w:drawing>
          </w:r>
        </w:p>
        """

    def _docx_supported_image(image_bytes: bytes | None, mime_type: str | None):
        if not image_bytes:
            return None, None
        normalized_mime = (mime_type or "image/png").strip().lower()
        if normalized_mime in {"image/png", "image/jpeg", "image/jpg"}:
            return image_bytes, "image/jpeg" if normalized_mime == "image/jpg" else normalized_mime
        if PILImage is not None:
            try:
                source = io.BytesIO(image_bytes)
                with PILImage.open(source) as image:
                    converted = io.BytesIO()
                    image.convert("RGBA").save(converted, format="PNG")
                    return converted.getvalue(), "image/png"
            except Exception:
                return None, None
        return None, None

    header_line_one, header_line_two = _waiting_hall_programme_lines(event_age_label, event_date)
    optimized_custom_logo, optimized_custom_logo_mime = _normalized_docx_report_logo(custom_logo, custom_logo_mime_type)
    optimized_right_logo, optimized_right_logo_mime = _normalized_docx_report_logo(_pdf_logo_path(), "image/png")
    left_logo_bytes, left_logo_mime = _docx_supported_image(optimized_custom_logo, optimized_custom_logo_mime)
    right_logo_bytes, right_logo_mime = _docx_supported_image(optimized_right_logo, optimized_right_logo_mime)

    rel_counter = 2
    image_counter = 1
    docpr_counter = 1
    media_files: list[tuple[str, bytes, str, str]] = []
    image_rel_ids: dict[str, str] = {}

    def _register_image(key: str, image_bytes: bytes | None, mime_type: str | None) -> str | None:
        nonlocal rel_counter, image_counter
        if not image_bytes or not mime_type:
            return None
        if key in image_rel_ids:
            return image_rel_ids[key]
        extension = "png" if mime_type == "image/png" else "jpeg"
        target = f"media/image{image_counter}.{extension}"
        rel_id = f"rId{rel_counter}"
        rel_counter += 1
        image_counter += 1
        media_files.append((target, image_bytes, mime_type, rel_id))
        image_rel_ids[key] = rel_id
        return rel_id

    left_logo_rel = _register_image("left_logo", left_logo_bytes, left_logo_mime)
    right_logo_rel = _register_image("right_logo", right_logo_bytes, right_logo_mime)

    body_parts = []
    content_width_twips = _twips_from_mm(186)
    logo_width_twips = _twips_from_mm(30)
    center_width_twips = content_width_twips - (2 * logo_width_twips)
    item_col_twips = _twips_from_mm(145)
    qty_col_twips = _twips_from_mm(27)

    for page_index, page_groups in enumerate(_waiting_hall_pack_pages(groups)):
        for group_index, group in enumerate(page_groups):
            item_count = len(group.get("items") or [])
            profile = _waiting_hall_layout_profile(item_count)

            left_cell_xml = _w_empty_paragraph()
            if left_logo_rel:
                logo_side_mm = min(profile["logo_width_mm"], profile["logo_height_mm"])
                left_cell_xml = _w_image_paragraph(
                    left_logo_rel,
                    width_mm=logo_side_mm,
                    height_mm=logo_side_mm,
                    docpr_id=docpr_counter,
                )
                docpr_counter += 1

            right_cell_xml = _w_empty_paragraph()
            if right_logo_rel:
                logo_side_mm = min(profile["logo_width_mm"], profile["logo_height_mm"])
                right_cell_xml = _w_image_paragraph(
                    right_logo_rel,
                    width_mm=logo_side_mm,
                    height_mm=logo_side_mm,
                    docpr_id=docpr_counter,
                )
                docpr_counter += 1

            title_cell_xml = "".join([
                _w_paragraph("OM SAKTHI", align="center", bold=True, size_pt=profile["title_font"], color="C1121F", spacing_after_pt=1),
                _w_paragraph(header_line_one, align="center", size_pt=profile["sub_font"], spacing_after_pt=0),
                _w_paragraph(header_line_two, align="center", size_pt=profile["sub_font"], spacing_after_pt=1),
                _w_paragraph(
                    "District Waiting Hall Articles Collection List",
                    align="center",
                    bold=True,
                    size_pt=profile["section_font"],
                    spacing_after_pt=2,
                ),
            ])

            header_table = f"""
            <w:tbl>
              <w:tblPr>
                <w:tblW w:w="{content_width_twips}" w:type="dxa"/>
                <w:tblBorders>
                  <w:top w:val="nil"/><w:left w:val="nil"/><w:bottom w:val="nil"/><w:right w:val="nil"/>
                  <w:insideH w:val="nil"/><w:insideV w:val="nil"/>
                </w:tblBorders>
              </w:tblPr>
              <w:tblGrid>
                <w:gridCol w:w="{logo_width_twips}"/>
                <w:gridCol w:w="{center_width_twips}"/>
                <w:gridCol w:w="{logo_width_twips}"/>
              </w:tblGrid>
              <w:tr>
                {_w_table_cell(left_cell_xml, width_twips=logo_width_twips, borders=False)}
                {_w_table_cell(title_cell_xml, width_twips=center_width_twips, borders=False)}
                {_w_table_cell(right_cell_xml, width_twips=logo_width_twips, borders=False)}
              </w:tr>
            </w:tbl>
            """
            body_parts.append(header_table)
            body_parts.append(
                _w_paragraph(
                    f"{str(group.get('label_prefix') or 'Beneficiary')}: {str(group.get('entity_name') or '-').upper()}",
                    align="center",
                    bold=True,
                    size_pt=profile["district_font"],
                    spacing_after_pt=4,
                )
            )

            row_xml = [
                "<w:tr>"
                + _w_table_cell(_w_paragraph("Article list", bold=True, size_pt=profile["header_font"]), width_twips=item_col_twips, shading="F3F4F6")
                + _w_table_cell(_w_paragraph("Qty", align="center", bold=True, size_pt=profile["header_font"]), width_twips=qty_col_twips, shading="F3F4F6")
                + "</w:tr>"
            ]
            for item in (group.get("items") or []):
                row_xml.append(
                    "<w:tr>"
                    + _w_table_cell(_w_paragraph(str(item.get("requested_item") or ""), bold=True, size_pt=profile["body_font"]), width_twips=item_col_twips)
                    + _w_table_cell(_w_paragraph(str(int(item.get("quantity") or 0)), align="center", bold=True, size_pt=profile["body_font"]), width_twips=qty_col_twips)
                    + "</w:tr>"
                )
            row_xml.append(
                "<w:tr>"
                + _w_table_cell(_w_paragraph("Total", bold=True, size_pt=profile["body_font"]), width_twips=item_col_twips)
                + _w_table_cell(_w_paragraph(str(int(group.get("total_quantity") or 0)), align="center", bold=True, size_pt=profile["body_font"]), width_twips=qty_col_twips)
                + "</w:tr>"
            )
            body_parts.append(
                f"""
                <w:tbl>
                  <w:tblPr>
                    <w:tblW w:w="{item_col_twips + qty_col_twips}" w:type="dxa"/>
                    <w:jc w:val="center"/>
                    <w:tblBorders>
                      <w:top w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:left w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:bottom w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:right w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:insideH w:val="single" w:sz="6" w:space="0" w:color="000000"/>
                      <w:insideV w:val="single" w:sz="6" w:space="0" w:color="000000"/>
                    </w:tblBorders>
                  </w:tblPr>
                  <w:tblGrid>
                    <w:gridCol w:w="{item_col_twips}"/>
                    <w:gridCol w:w="{qty_col_twips}"/>
                  </w:tblGrid>
                  {''.join(row_xml)}
                </w:tbl>
                """
            )
            body_parts.append(
                _w_paragraph(
                    "I Acknowledge that I have received the above articles",
                    align="center",
                    bold=True,
                    size_pt=profile["note_font"],
                    spacing_before_pt=8,
                    spacing_after_pt=12,
                )
            )
            body_parts.append(
                _w_paragraph(
                    "Signature",
                    align="right",
                    bold=True,
                    size_pt=profile["note_font"],
                    spacing_before_pt=profile["note_leading"] * 2,
                    spacing_after_pt=0,
                )
            )

            if group_index < len(page_groups) - 1:
                body_parts.append(_w_empty_paragraph(spacing_after_pt=18))
        if page_index < len(_waiting_hall_pack_pages(groups)) - 1:
            body_parts.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')

    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
      <w:style w:type="paragraph" w:default="1" w:styleId="Normal">
        <w:name w:val="Normal"/>
        <w:qFormat/>
        <w:rPr>
          <w:rFonts w:ascii="Arial" w:hAnsi="Arial"/>
          <w:sz w:val="22"/>
        </w:rPr>
      </w:style>
    </w:styles>
    """

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:document
      xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
      xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
      xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
      xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
      xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
      <w:body>
        {''.join(body_parts)}
        <w:sectPr>
          <w:pgSz w:w="11906" w:h="16838"/>
          <w:pgMar w:top="680" w:right="680" w:bottom="680" w:left="680" w:header="708" w:footer="708" w:gutter="0"/>
        </w:sectPr>
      </w:body>
    </w:document>
    """

    image_defaults: dict[str, str] = {}
    document_rels = [
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
    ]
    for target, _image_bytes, mime_type, rel_id in media_files:
        extension = target.rsplit(".", 1)[-1]
        content_type = "image/png" if mime_type == "image/png" else "image/jpeg"
        image_defaults[extension] = content_type
        document_rels.append(
            f'<Relationship Id="{rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{target}"/>'
        )

    content_types_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
      <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
      <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
      <Default Extension="xml" ContentType="application/xml"/>
      {''.join(f'<Default Extension="{extension}" ContentType="{content_type}"/>' for extension, content_type in image_defaults.items())}
      <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
      <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
    </Types>
    """

    package_rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
      <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
    </Relationships>
    """

    document_rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
      {''.join(document_rels)}
    </Relationships>
    """

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as docx_zip:
        docx_zip.writestr("[Content_Types].xml", content_types_xml)
        docx_zip.writestr("_rels/.rels", package_rels_xml)
        docx_zip.writestr("word/document.xml", document_xml)
        docx_zip.writestr("word/styles.xml", styles_xml)
        docx_zip.writestr("word/_rels/document.xml.rels", document_rels_xml)
        for target, image_bytes, _mime_type, _rel_id in media_files:
            docx_zip.writestr(f"word/{target}", image_bytes)
    return buffer.getvalue()


def _public_acknowledgment_pdf_fields(template_bytes: bytes | None) -> list[dict]:
    if not template_bytes:
        return []
    reader = PdfReader(io.BytesIO(template_bytes))
    seen = set()
    fields: list[dict] = []
    for page in reader.pages:
        for annotation_ref in page.get("/Annots") or []:
            annotation = annotation_ref.get_object()
            if str(annotation.get("/Subtype") or "") != "/Widget":
                continue
            if str(annotation.get("/FT") or "") != "/Tx":
                continue
            field_name = str(annotation.get("/T") or "").strip()
            if not field_name or field_name in seen:
                continue
            seen.add(field_name)
            field_key = re.sub(r"[^a-z0-9]+", "_", field_name.casefold()).strip("_") or f"field_{len(fields) + 1}"
            fields.append(
                {
                    "field_name": field_name,
                    "field_key": field_key,
                }
            )
    return fields


def _public_acknowledgment_resolve_value(row: dict, column_name: str | None) -> str:
    if not column_name:
        return ""
    column = str(column_name or "").strip()
    if not column:
        return ""
    blank_tokens = {"n/a", "na", "none", "null", "-"}

    def _clean(value: object) -> str:
        text = str(value or "").strip()
        if text.casefold() in blank_tokens:
            return ""
        return text

    lookup_keys = {
        column,
        column.replace("_", " "),
        column.replace(" ", "_"),
        column.casefold(),
        column.replace("_", " ").casefold(),
        column.replace(" ", "_").casefold(),
    }
    for key, value in row.items():
        key_text = str(key or "").strip()
        if key_text in lookup_keys or key_text.casefold() in {item.casefold() for item in lookup_keys}:
            return _clean(value)
    return _clean(row.get(column) or row.get(column.replace("_", " ")) or row.get(column.replace(" ", "_")) or "")


def _public_acknowledgment_is_aid_row(row: dict) -> bool:
    item_type = str(row.get("Item Type") or row.get("item_type") or "").strip().casefold()
    return item_type == str(models.ItemTypeChoices.AID).strip().casefold()


def _public_acknowledgment_field_rects(template_bytes: bytes) -> dict[str, list[float]]:
    reader = PdfReader(io.BytesIO(template_bytes))
    rects: dict[str, list[float]] = {}
    if not reader.pages:
        return rects
    for annotation_ref in reader.pages[0].get("/Annots") or []:
        annotation = annotation_ref.get_object()
        if str(annotation.get("/Subtype") or "") != "/Widget":
            continue
        field_name = str(annotation.get("/T") or "").strip()
        rect = annotation.get("/Rect")
        if not field_name or not rect:
            continue
        rects[field_name] = [float(value) for value in rect]
    return rects


def _public_acknowledgment_draw_field(
    overlay_canvas,
    field_name: str,
    value: str,
    rect: list[float],
) -> None:
    x1, y1, x2, y2 = rect
    width = max(x2 - x1, 0)
    height = max(y2 - y1, 0)
    text = str(value or "").strip()
    if not text:
        return

    if field_name == "address":
        style = ParagraphStyle(
            "public_ack_address",
            fontName="Helvetica",
            fontSize=12,
            leading=13,
            textColor=colors.black,
            alignment=0,
        )
        paragraph = Paragraph(text.replace("\n", "<br/>"), style)
        frame = Frame(x1 + 6, y1 + 4, max(width - 10, 1), max(height - 8, 1), showBoundary=0)
        frame.addFromList([paragraph], overlay_canvas)
        return

    font_name = "Helvetica"
    font_size = 16
    if field_name in {"district", "bf name", "article", "cheque_no"}:
        font_name = "Helvetica-Bold"
        font_size = 17 if field_name != "district" else 19
    elif field_name in {"App no", "token"}:
        font_name = "Helvetica-Bold"
        font_size = 16
    elif field_name in {"mobile", "Aadhar", "value_aid"}:
        font_name = "Helvetica-Bold"
        font_size = 16

    overlay_canvas.setFillColor(colors.black)
    overlay_canvas.setFont(font_name, font_size)
    text_width = overlay_canvas.stringWidth(text, font_name, font_size)
    if text_width > width - 8:
        shrink = max((width - 8) / max(text_width, 1), 0.6)
        font_size = max(round(font_size * shrink), 8)
        overlay_canvas.setFont(font_name, font_size)
        text_width = overlay_canvas.stringWidth(text, font_name, font_size)
    x = x1 + max((width - text_width) / 2, 2)
    y = y1 + max((height - font_size) / 2, 2)
    overlay_canvas.drawString(x, y, text)


def generate_public_acknowledgment_pdf(
    template_bytes: bytes,
    rows: list[dict],
    field_map: dict[str, str],
) -> io.BytesIO:
    normalized_map = {str(key or "").strip(): str(value or "").strip() for key, value in (field_map or {}).items()}
    if not rows:
        rows = [{}]
    writer = PdfWriter()
    rect_map = _public_acknowledgment_field_rects(template_bytes)
    for row in rows:
        template_reader = PdfReader(io.BytesIO(template_bytes))
        if not template_reader.pages:
            raise ValueError("Uploaded PDF template has no pages.")
        page = template_reader.pages[0]
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)
        overlay_buffer = io.BytesIO()
        overlay_canvas = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
        for field_name, column_name in normalized_map.items():
            if not field_name:
                continue
            if field_name == "value_aid" and not _public_acknowledgment_is_aid_row(row):
                continue
            value = _public_acknowledgment_resolve_value(row, column_name)
            rect = rect_map.get(field_name)
            if not rect:
                continue
            _public_acknowledgment_draw_field(overlay_canvas, field_name, value, rect)
        overlay_canvas.save()
        overlay_buffer.seek(0)
        overlay_reader = PdfReader(overlay_buffer)
        if overlay_reader.pages:
            page.merge_page(overlay_reader.pages[0])
        if "/Annots" in page:
            del page["/Annots"]
        writer.add_page(page)
    buffer = io.BytesIO()
    writer.write(buffer)
    buffer.seek(0)
    return buffer


def _public_signature_app_sort_key(value: str) -> tuple[str, int, str]:
    text = str(value or "").strip()
    match = re.match(r"^([A-Za-z]+)\s*0*([0-9]+)$", text)
    if match:
        return (match.group(1).casefold(), int(match.group(2)), text.casefold())
    return ("", 0, text.casefold())


def generate_public_signature_pdf(rows: list[dict]) -> io.BytesIO:
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "public_signature_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=12,
        alignment=1,
        textColor=colors.black,
    )
    body_style = ParagraphStyle(
        "public_signature_body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=11,
        textColor=colors.black,
    )
    body_bold_style = ParagraphStyle(
        "public_signature_body_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )

    sorted_rows = list(rows or [])

    table_rows = [[
        Paragraph("Sl No", header_style),
        Paragraph("App No", header_style),
        Paragraph("Name", header_style),
        Paragraph("Requested Item", header_style),
        Paragraph("Token No", header_style),
        Paragraph("Signature", header_style),
    ]]

    for index, row in enumerate(sorted_rows, start=1):
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or token_start or 0)
        token_text = str(token_start) if token_end <= token_start else f"{token_start} - {token_end}"
        table_rows.append(
            [
                Paragraph(str(index), body_bold_style),
                Paragraph(str(row.get("application_number") or ""), body_style),
                Paragraph(str(row.get("beneficiary_name") or ""), body_style),
                Paragraph(str(row.get("item_name") or ""), body_style),
                Paragraph(token_text, body_bold_style),
                Paragraph("", body_style),
            ]
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    table = LongTable(
        table_rows,
        colWidths=[12 * mm, 20 * mm, 46 * mm, 54 * mm, 18 * mm, 48 * mm],
        repeatRows=1,
        splitByRow=1,
    )
    table.hAlign = "CENTER"
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9fbbe7")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (4, 0), (4, -1), "CENTER"),
                ("ALIGN", (5, 0), (5, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4.5),
                ("TOPPADDING", (0, 0), (-1, 0), 7),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ]
        )
    )
    doc.build([table])
    buffer.seek(0)
    return buffer



def generate_public_signature_xlsx(rows: list[dict]) -> io.BytesIO:
    sorted_rows = list(rows or [])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Public Signature"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="9FBBE7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Sl No", "App No", "Name", "Requested Item", "Token No", "Signature"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for index, row in enumerate(sorted_rows, start=1):
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or token_start or 0)
        token_text = str(token_start) if token_end <= token_start else f"{token_start} - {token_end}"
        worksheet.append([
            index,
            str(row.get("application_number") or ""),
            str(row.get("beneficiary_name") or ""),
            str(row.get("item_name") or ""),
            token_text,
            "",
        ])

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=6):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = center if idx in {1, 2, 5} else left
            cell.font = Font(size=10.5, bold=idx in {1, 5})

    widths = {1: 10, 2: 16, 3: 32, 4: 34, 5: 16, 6: 40}
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


class _DistrictSignatureNumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.black)
            self.drawCentredString(A4[0] / 2, 8 * mm, f"Page {self._pageNumber} of {total_pages}")
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)


def generate_district_signature_xlsx(
    districts: list[dict],
    *,
    custom_logo: bytes | None = None,
    custom_logo_mime_type: str | None = None,
) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "District Signature"

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    header_fill = PatternFill("solid", fgColor="9FBBE7")
    district_fill = PatternFill("solid", fgColor="F4F8FD")
    title_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A7"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.3
    worksheet.page_margins.bottom = 0.3
    worksheet.page_margins.header = 0.1
    worksheet.page_margins.footer = 0.1

    # Header block to mirror the PDF layout more closely.
    worksheet.merge_cells("B1:E1")
    worksheet.merge_cells("B2:E2")
    worksheet.merge_cells("B3:E3")
    worksheet.merge_cells("B4:E4")
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[2].height = 18
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[4].height = 22
    worksheet.row_dimensions[5].height = 8
    worksheet.row_dimensions[6].height = 22

    left_logo_bytes = None
    right_logo_bytes = None
    try:
        left_logo_source = custom_logo or _pdf_guru_logo_path()
        left_logo_mime = custom_logo_mime_type or ("image/png" if custom_logo else "image/jpeg")
        left_logo_bytes, _ = _normalized_docx_report_logo(left_logo_source, left_logo_mime, canvas_width_px=120, canvas_height_px=120)
        right_logo_bytes, _ = _normalized_docx_report_logo(_pdf_logo_path(), "image/png", canvas_width_px=120, canvas_height_px=120)
    except Exception:
        left_logo_bytes = None
        right_logo_bytes = None

    if left_logo_bytes:
        left_logo = XLImage(io.BytesIO(left_logo_bytes))
        left_logo.width = 50
        left_logo.height = 50
        worksheet.add_image(left_logo, "A1")
    if right_logo_bytes:
        right_logo = XLImage(io.BytesIO(right_logo_bytes))
        right_logo.width = 50
        right_logo.height = 50
        worksheet.add_image(right_logo, "F1")

    current_year = timezone.localdate().year
    birthday_number = max(current_year - 1940, 1)

    title_cell = worksheet["B1"]
    title_cell.value = "OM SAKTHI"
    title_cell.font = Font(size=9, bold=True, color="C1121F")
    title_cell.alignment = center
    title_cell.fill = title_fill

    subtitle_cell = worksheet["B2"]
    subtitle_cell.value = f"MASM Social Welfare Programme on the eve of {_ordinal(birthday_number)} Birthday (03-03-{current_year})"
    subtitle_cell.font = Font(size=11, color="000000")
    subtitle_cell.alignment = center

    subtitle2_cell = worksheet["B3"]
    subtitle2_cell.value = f"His Holiness AMMA at Melmaruvathur on 03.03.{current_year}"
    subtitle2_cell.font = Font(size=11, color="000000")
    subtitle2_cell.alignment = center

    report_title_cell = worksheet["B4"]
    report_title_cell.value = "District Token Distribution Register"
    report_title_cell.font = Font(size=14, bold=True, color="000000")
    report_title_cell.alignment = center

    headers = ["District", "Total Qty", "Token Qty", "Start Token", "End Token", "Signature"]
    header_row_index = 6
    for col_index, header_value in enumerate(headers, start=1):
        cell = worksheet.cell(header_row_index, col_index, header_value)
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = outer_border
        cell.alignment = center

    row_index = 7
    for district in list(districts or []):
        district_name = str(district.get("district_name") or "").strip()
        if not district_name:
            continue

        block_start = row_index
        worksheet.append([district_name, "", "", "", "", ""])
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        for col in range(1, 7):
            worksheet.cell(row_index, col).border = outer_border
        district_cell = worksheet.cell(row_index, 1)
        district_cell.font = Font(size=10.5, bold=True)
        district_cell.fill = district_fill
        district_cell.alignment = center
        row_index += 1

        for item in district.get("items") or []:
            worksheet.append([
                str(item.get("item_name") or ""),
                int(item.get("total_quantity") or 0),
                int(item.get("token_quantity") or 0) if item.get("token_quantity") not in {None, "", 0} else "",
                _district_signature_display_token(item.get("start_token")),
                _district_signature_display_token(item.get("end_token")),
                "",
            ])
            for col in range(1, 7):
                worksheet.cell(row_index, col).border = border
                worksheet.cell(row_index, col).alignment = center if col in {2, 3, 4, 5} else left
                worksheet.cell(row_index, col).font = Font(size=10)
            row_index += 1

        worksheet.append([
            f"{district_name} Total",
            int(district.get("total_quantity") or 0),
            int(district.get("token_quantity") or 0) if district.get("token_quantity") not in {None, "", 0} else "",
            "",
            "",
            "",
        ])
        for col in range(1, 7):
            worksheet.cell(row_index, col).border = outer_border
            worksheet.cell(row_index, col).alignment = center if col in {2, 3} else left
            worksheet.cell(row_index, col).font = Font(size=10.5, bold=True)
            worksheet.cell(row_index, col).fill = district_fill

        # Merge the signature column across the district block so it reads as a
        # single signing area instead of repeated per-row cells.
        worksheet.merge_cells(start_row=block_start, start_column=6, end_row=row_index, end_column=6)
        signature_cell = worksheet.cell(block_start, 6)
        signature_cell.border = outer_border
        signature_cell.alignment = center

        row_index += 1

        # Insert a clean visual separator between districts without disturbing
        # the district box borders.
        worksheet.append(["", "", "", "", "", ""])
        separator_row = row_index
        worksheet.row_dimensions[separator_row].height = 8
        row_index += 1

    widths = {1: 34, 2: 12, 3: 12, 4: 14, 5: 14, 6: 28}
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _district_signature_display_token(value):
    if value in {None, "", 0}:
        return ""
    return str(int(value))


def generate_district_signature_pdf(
    districts: list[dict],
    *,
    custom_logo: bytes | None = None,
) -> io.BytesIO:
    styles = getSampleStyleSheet()
    district_border = colors.HexColor("#5b6572")
    district_light_border = colors.HexColor("#7c8794")
    left_logo_source = custom_logo or _pdf_guru_logo_path()
    left_logo = _fitted_pdf_image_source(left_logo_source, max_width_mm=18, max_height_mm=18)
    right_logo = _fitted_pdf_image(_pdf_logo_path(), max_width_mm=18, max_height_mm=18)
    header_title = ParagraphStyle(
        "district_signature_title",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        textColor=colors.black,
    )
    header_sub = ParagraphStyle(
        "district_signature_sub",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica",
        fontSize=9.3,
        leading=9.2,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.black,
    )
    header_style = ParagraphStyle(
        "district_signature_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=12,
        alignment=1,
        textColor=colors.black,
    )
    district_style = ParagraphStyle(
        "district_signature_district",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=12,
        alignment=1,
        textColor=colors.black,
    )
    item_style = ParagraphStyle(
        "district_signature_item",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=12,
        textColor=colors.black,
    )
    total_style = ParagraphStyle(
        "district_signature_total",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.black,
    )
    numeric_style = ParagraphStyle(
        "district_signature_numeric",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=12,
        alignment=1,
        textColor=colors.black,
    )
    numeric_bold_style = ParagraphStyle(
        "district_signature_numeric_bold",
        parent=numeric_style,
        fontName="Helvetica-Bold",
    )

    table_rows = [[
        Paragraph("District", header_style),
        Paragraph("Total Qty", header_style),
        Paragraph("Token Qty", header_style),
        Paragraph("Start Token", header_style),
        Paragraph("End Token", header_style),
        Paragraph("Signature", header_style),
    ]]
    style_commands = [
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9fbbe7")),
        ("BOX", (0, 0), (-1, 0), 0.7, district_border),
        ("INNERGRID", (0, 0), (-1, 0), 0.55, district_light_border),
        ("ALIGN", (1, 0), (4, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
    ]

    district_prog_line_one, district_prog_line_two = _district_signature_programme_lines()
    header_rows = [[
        left_logo,
        [
            Paragraph("OM SAKTHI", ParagraphStyle("district-signature-om", parent=header_title, fontSize=8.8, leading=9.0, spaceAfter=0, textColor=colors.red)),
            Paragraph(district_prog_line_one, header_sub),
            Paragraph(district_prog_line_two, header_sub),
            Paragraph("District Token Distribution Register", header_title),
        ],
        right_logo,
    ]]

    row_index = 1
    for district in list(districts or []):
        district_name = str(district.get("district_name") or "").strip()
        if not district_name:
            continue
        block_start = row_index
        table_rows.append([
            Paragraph(district_name, district_style),
            "",
            "",
            "",
            "",
            "",
        ])
        style_commands.extend([
            ("SPAN", (0, row_index), (4, row_index)),
            ("BACKGROUND", (0, row_index), (4, row_index), colors.HexColor("#f4f8fd")),
            ("ALIGN", (0, row_index), (4, row_index), "CENTER"),
            ("LINEBELOW", (0, row_index), (4, row_index), 0.75, district_light_border),
            ("TOPPADDING", (0, row_index), (5, row_index), 8),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 8),
        ])
        row_index += 1

        for item in district.get("items") or []:
            table_rows.append([
                Paragraph(str(item.get("item_name") or ""), item_style),
                Paragraph(str(int(item.get("total_quantity") or 0)), numeric_style),
                Paragraph(_district_signature_display_token(item.get("token_quantity")), numeric_style),
                Paragraph(_district_signature_display_token(item.get("start_token")), numeric_style),
                Paragraph(_district_signature_display_token(item.get("end_token")), numeric_style),
                "",
            ])
            style_commands.extend([
                ("LINEBELOW", (0, row_index), (4, row_index), 0.45, district_light_border),
                ("TOPPADDING", (0, row_index), (5, row_index), 7),
                ("BOTTOMPADDING", (0, row_index), (5, row_index), 7),
            ])
            row_index += 1

        table_rows.append([
            Paragraph(f"{district_name} Total", total_style),
            Paragraph(str(int(district.get("total_quantity") or 0)), numeric_bold_style),
            Paragraph(_district_signature_display_token(district.get("token_quantity")), numeric_bold_style),
            "",
            "",
            "",
        ])
        style_commands.extend([
            ("LINEABOVE", (0, row_index), (4, row_index), 1.0, district_border),
            ("LINEBELOW", (0, row_index), (5, row_index), 1.0, district_border),
            ("TOPPADDING", (0, row_index), (5, row_index), 7),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 10),
            ("BOX", (0, block_start), (5, row_index), 1.2, district_border),
            ("INNERGRID", (0, block_start), (4, row_index), 0.45, district_light_border),
            ("LINEBEFORE", (5, block_start), (5, row_index), 1.2, district_border),
        ])
        row_index += 1
        table_rows.append(["", "", "", "", "", ""])
        style_commands.extend([
            ("SPAN", (0, row_index), (5, row_index)),
            ("TOPPADDING", (0, row_index), (5, row_index), 7),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 7),
        ])
        row_index += 1

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=6 * mm,
        bottomMargin=12 * mm,
    )
    table = LongTable(
        table_rows,
        colWidths=[79 * mm, 16 * mm, 16 * mm, 17 * mm, 17 * mm, 45 * mm],
        repeatRows=1,
        splitByRow=1,
    )
    table.hAlign = "CENTER"
    table.setStyle(TableStyle(style_commands))
    header_table = Table(header_rows, colWidths=[18 * mm, 144 * mm, 18 * mm])
    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story = [
        header_table,
        Spacer(1, 0.8 * mm),
        table,
    ]
    doc.build(story, canvasmaker=_DistrictSignatureNumberedCanvas)
    buffer.seek(0)
    return buffer
