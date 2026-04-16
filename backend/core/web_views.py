from __future__ import annotations

"""
Server-rendered UI views, helper functions, list filtering, and form workflows.

This is the main file to inspect when changing the behavior of the Django
templates under ``templates/dashboard/``.
"""

import csv
import base64
import json
import uuid
import zipfile
import io
import os
import re
import mimetypes
import subprocess
import sys
import logging
from datetime import date
from collections import Counter
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pypdf import PdfReader, PdfWriter

from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import CharField, F, Q
from django.db.models.functions import Cast
from django.http import FileResponse, Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
)
from django.shortcuts import get_object_or_404
from django.forms import inlineformset_factory

from . import models
from . import services
from . import google_drive
from .sequence_defaults import SEQUENCE_DEFAULT_ITEMS
from .forms import (
    AppUserCreateForm,
    AppUserPasswordResetForm,
    AppUserUpdateForm,
    ArticleForm,
    ApplicationAttachmentUploadForm,
    FundRequestDocumentUploadForm,
    FundRequestForm,
    FundRequestArticleForm,
    FundRequestRecipientForm,
    MasterDataUploadForm,
    PurchaseOrderForm,
    PurchaseOrderItemForm,
)

FEMALE_STATUS_DESCRIPTIONS = {
    "Single": "A woman who is unmarried and not currently in a marital relationship.",
    "Married": "A woman who is currently married.",
    "Widowed": "A woman whose spouse has died.",
    "Divorced": "A woman whose marriage has been legally dissolved.",
    "Separated": "A woman living apart from her spouse without a finalized divorce.",
    "Deserted": "A woman abandoned by her spouse without support.",
    "Single Mother": "A woman raising one or more children without a partner in the household.",
    "Destitute Woman (no income/support)": "A woman without stable income, family support, or financial security.",
    "Female Head of Household": "A woman who is the primary decision-maker and provider for the household.",
    "Victim of Domestic Violence": "A woman currently affected by violence or abuse within the home.",
    "Survivor of Abuse": "A woman who has survived physical, emotional, or sexual abuse.",
    "Elderly Woman (60+)": "A woman aged 60 or above, often needing age-related support.",
    "Homeless": "A woman without secure or permanent housing.",
    "Orphan / No Family Support": "A woman with no dependable family support structure.",
    "Migrant Woman": "A woman who has moved for work, safety, marriage, or survival and may lack local support.",
    "Caregiver (children / elderly / disabled)": "A woman responsible for regular care of children, elderly persons, or persons with disabilities.",
    "Employed": "A woman currently working in a salaried or wage-based role.",
    "Self-employed": "A woman earning independently through business, trade, farming, or service work.",
    "Unemployed": "A woman currently without paid work and seeking or needing livelihood support.",
    "Student": "A woman currently pursuing school, college, or vocational education.",
}

logger = logging.getLogger(__name__)


class RoleRequiredMixin(UserPassesTestMixin):
    allowed_roles = {"admin", "editor", "viewer"}
    module_key = None
    permission_action = "view"

    def test_func(self):
        user = self.request.user
        if not (user and user.is_authenticated and user.status == models.StatusChoices.ACTIVE):
            return False
        if self.module_key:
            return user.has_module_permission(self.module_key, self.permission_action)
        return user.role in self.allowed_roles


class WriteRoleMixin(RoleRequiredMixin):
    allowed_roles = {"admin", "editor"}



class DashboardView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    allowed_roles = {"admin", "editor", "viewer"}
    template_name = "dashboard/dashboard.html"

    def post(self, request, *args, **kwargs):
        if request.user.role not in {"admin", "editor"}:
            messages.error(request, "You do not have permission to update the event budget.")
            return HttpResponseRedirect(reverse("ui:dashboard"))
        budget_raw = (request.POST.get("event_budget") or "").strip().replace(",", "")
        try:
            event_budget = Decimal(budget_raw or "0")
            if event_budget < 0:
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            messages.error(request, "Enter a valid event budget amount.")
            return HttpResponseRedirect(reverse("ui:dashboard"))
        setting = models.DashboardSetting.objects.order_by("pk").first()
        if setting is None:
            setting = models.DashboardSetting.objects.create(event_budget=event_budget, updated_by=request.user)
        else:
            setting.event_budget = event_budget
            setting.updated_by = request.user
            setting.save(update_fields=["event_budget", "updated_by", "updated_at"])
        messages.success(request, "Event budget updated.")
        return HttpResponseRedirect(reverse("ui:dashboard"))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        metrics = self._build_metrics()
        setting = models.DashboardSetting.objects.order_by("pk").first()
        overall_event_budget = setting.event_budget if setting else metrics["district"]["total_allotted_fund"]
        context.update(
            {
                "metrics": metrics,
                "overall_event_budget": overall_event_budget,
                "balance_to_allot": overall_event_budget - metrics["overall"]["total_value_accrued"],
                "can_edit_event_budget": self.request.user.role in {"admin", "editor"},
            }
        )
        return context


    def _zero(self):
        return Decimal("0")

    def _signed_currency_text(self, value):
        amount = Decimal(str(value or 0))
        sign = "+" if amount >= 0 else "-"
        return {"sign": sign, "amount": abs(amount)}

    def _build_metrics(self):
        zero = self._zero()
        district_entries = list(
            models.DistrictBeneficiaryEntry.objects.select_related("district", "article").all()
        )
        public_entries = list(
            models.PublicBeneficiaryEntry.objects.select_related("article").all()
        )
        institution_entries = list(
            models.InstitutionsBeneficiaryEntry.objects.select_related("article").all()
        )
        districts = list(models.DistrictMaster.objects.filter(is_active=True).order_by("district_name"))
        fund_requests = list(models.FundRequest.objects.all())

        district_article_ids = set()
        district_ids = set()
        district_articles_qty = 0
        district_value_accrued = zero
        district_spend_map = {}
        for entry in district_entries:
            if entry.article_id:
                district_article_ids.add(entry.article_id)
            if entry.district_id:
                district_ids.add(entry.district_id)
                district_spend_map[entry.district_id] = district_spend_map.get(entry.district_id, zero) + Decimal(str(entry.total_amount or 0))
            district_articles_qty += int(entry.quantity or 0)
            district_value_accrued += Decimal(str(entry.total_amount or 0))

        public_article_ids = set()
        public_articles_qty = 0
        public_value_accrued = zero
        gender_counts = {"Male": 0, "Female": 0, "Transgender": 0}
        female_status_counts = {}
        handicapped = 0
        disability_category_counts = {}
        for entry in public_entries:
            if entry.article_id:
                public_article_ids.add(entry.article_id)
            public_articles_qty += int(entry.quantity or 0)
            public_value_accrued += Decimal(str(entry.total_amount or 0))
            if entry.gender in gender_counts:
                gender_counts[entry.gender] += 1
            if entry.female_status:
                female_status_counts[entry.female_status] = female_status_counts.get(entry.female_status, 0) + 1
            if entry.is_handicapped and entry.is_handicapped != models.HandicappedStatusChoices.NO:
                handicapped += 1
                disability_category_counts[entry.is_handicapped] = disability_category_counts.get(entry.is_handicapped, 0) + 1

        institution_article_ids = set()
        institution_applications = set()
        institution_articles_qty = 0
        institution_value_accrued = zero
        for entry in institution_entries:
            if entry.article_id:
                institution_article_ids.add(entry.article_id)
            if entry.application_number:
                institution_applications.add(entry.application_number)
            institution_articles_qty += int(entry.quantity or 0)
            institution_value_accrued += Decimal(str(entry.total_amount or 0))

        overall_article_ids = district_article_ids | public_article_ids | institution_article_ids
        overall_articles_qty = district_articles_qty + public_articles_qty + institution_articles_qty
        total_allotted_fund = sum((Decimal(str(d.allotted_budget or 0)) for d in districts), zero)
        received_district_allotted_fund = sum(
            (Decimal(str(d.allotted_budget or 0)) for d in districts if d.id in district_ids),
            zero,
        )
        district_variance = district_value_accrued - received_district_allotted_fund
        overall_actual_value_accrued = district_value_accrued + public_value_accrued + institution_value_accrued
        overall_planning_value_accrued = received_district_allotted_fund + public_value_accrued + institution_value_accrued
        overall_beneficiaries = district_articles_qty + public_articles_qty + len(institution_applications)

        fund_request_total_value = sum((Decimal(str(f.total_amount or 0)) for f in fund_requests), zero)

        pending_districts = [
            district.district_name
            for district in districts
            if district.id and district.id not in district_ids and district.district_name
        ]

        under_utilized_district_count = 0
        under_utilized_value = zero
        over_utilized_district_count = 0
        over_utilized_value = zero
        for district in districts:
            if district.id not in district_ids:
                continue
            allotted = Decimal(str(district.allotted_budget or 0))
            used = district_spend_map.get(district.id, zero)
            delta = used - allotted
            if delta > 0:
                over_utilized_district_count += 1
                over_utilized_value += delta
            elif delta < 0:
                under_utilized_district_count += 1
                under_utilized_value += abs(delta)

        preferred_female_order = ["Single", "Married", "Widowed", "Single Mother"]
        female_status_lines = []
        for label in preferred_female_order:
            female_status_lines.append(
                {
                    "label": label,
                    "value": female_status_counts.get(label, 0),
                    "class_name": {
                        "Single": "female-unmarried",
                        "Married": "female-married",
                        "Widowed": "female-widow",
                        "Single Mother": "female-single-mother",
                    }.get(label, ""),
                }
            )
        for label, value in female_status_counts.items():
            if label not in preferred_female_order:
                female_status_lines.append({"label": label, "value": value, "class_name": ""})

        return {
            "district": {
                "districts_received": len(district_ids),
                "districts_pending": max(len(districts) - len(district_ids), 0),
                "total_articles_qty": district_articles_qty,
                "unique_articles": len(district_article_ids),
                "total_beneficiaries": len(district_entries),
                "total_allotted_fund": total_allotted_fund,
                "total_value_accrued": district_value_accrued,
                "under_utilized_district_count": under_utilized_district_count,
                "under_utilized_value": under_utilized_value,
                "over_utilized_district_count": over_utilized_district_count,
                "over_utilized_value": over_utilized_value,
                "net_variance": district_variance,
            },
            "public": {
                "total_beneficiaries": len(public_entries),
                "total_articles_qty": public_articles_qty,
                "unique_articles": len(public_article_ids),
                "total_value_accrued": public_value_accrued,
                "gender_lines": [
                    {"label": "Male", "value": gender_counts["Male"], "class_name": "gender-male"},
                    {"label": "Female", "value": gender_counts["Female"], "class_name": "gender-female"},
                    {"label": "Transgender", "value": gender_counts["Transgender"], "class_name": "gender-transgender"},
                ],
                "female_status_lines": female_status_lines,
                "handicapped": handicapped,
                "handicapped_lines": [
                    {
                        "label": label,
                        "value": disability_category_counts.get(label, 0),
                        "color": color,
                    }
                    for label, color in [
                        ("Blindness / Low Vision", "#2563eb"),
                        ("Deaf / Hard of Hearing", "#9333ea"),
                        ("Locomotor Disability", "#ea580c"),
                        ("Cerebral Palsy", "#0f766e"),
                        ("Leprosy Cured", "#ca8a04"),
                        ("Dwarfism", "#db2777"),
                        ("Acid Attack Victim", "#dc2626"),
                        ("Muscular Dystrophy", "#16a34a"),
                        ("Autism Spectrum Disorder", "#4f46e5"),
                        ("Intellectual Disability", "#0891b2"),
                        ("Specific Learning Disability", "#65a30d"),
                        ("Mental Illness", "#d97706"),
                        ("Multiple Disability", "#7c3aed"),
                        ("Deaf-Blindness", "#334155"),
                        ("Other", "#64748b"),
                    ]
                    if disability_category_counts.get(label, 0)
                ] + [
                    {"label": "No", "value": max(len(public_entries) - handicapped, 0), "color": "#cbd5e1"}
                ],
            },
            "institutions": {
                "total_beneficiaries": len(institution_entries),
                "application_count": len(institution_applications),
                "total_articles_qty": institution_articles_qty,
                "unique_articles": len(institution_article_ids),
                "total_value_accrued": institution_value_accrued,
            },
            "overall": {
                "total_beneficiaries": overall_beneficiaries,
                "total_articles_qty": overall_articles_qty,
                "unique_articles": len(overall_article_ids),
                "total_value_accrued": overall_planning_value_accrued,
                "actual_total_value_accrued": overall_actual_value_accrued,
                "district_variance": district_variance,
                "district_variance_signed": self._signed_currency_text(district_variance),
                "district_contribution_signed": self._signed_currency_text(-district_variance),
            },
            "fund_requests": {
                "count": len(fund_requests),
                "total_value": fund_request_total_value,
            },
            "total_districts": len(districts),
            "pending_districts": pending_districts,
        }


class UserGuideView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard/user_guide.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "page_title": "User Guide",
            }
        )
        return context


REPORTS_WAITING_HALL_STATE_KEY = "reports_waiting_hall_acknowledgment"
REPORTS_SHARED_LOGO_KEY = "reports_shared_logo"
REPORTS_PUBLIC_ACK_STATE_KEY = "reports_public_acknowledgment"
REPORTS_TOKEN_LOOKUP_STATE_KEY = "reports_token_lookup"
REPORTS_PUBLIC_SIGNATURE_STATE_KEY = "reports_public_signature"
REPORTS_DISTRICT_SIGNATURE_STATE_KEY = "reports_district_signature"
REPORTS_SEGREGATION_STATE_KEY = "reports_segregation"
REPORTS_DISTRIBUTION_STATE_KEY = "reports_distribution"


def _reports_active_session():
    return models.EventSession.objects.filter(is_active=True).order_by("-event_year", "session_name").first()


def _reports_parse_date(raw_value: str | None) -> date:
    raw = str(raw_value or "").strip()
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return timezone.localdate()


def _reports_waiting_hall_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "ignored_keys": [],
        "rows": [],
        "headers": [],
        "beneficiary_type_filter": "",
        "item_type_filter": models.ItemTypeChoices.AID,
    }


def _reports_waiting_hall_session_state(request):
    state = request.session.get(REPORTS_WAITING_HALL_STATE_KEY) or {}
    merged = _reports_waiting_hall_default_state()
    merged.update(state)
    return merged


def _reports_shared_logo_state(request):
    state = request.session.get(REPORTS_SHARED_LOGO_KEY) or {}
    return {
        "logo_name": str(state.get("logo_name") or ""),
        "logo_content_type": str(state.get("logo_content_type") or ""),
        "logo_base64": str(state.get("logo_base64") or ""),
    }


def _reports_set_shared_logo_state(request, *, uploaded_logo):
    request.session[REPORTS_SHARED_LOGO_KEY] = {
        "logo_name": str(getattr(uploaded_logo, "name", "") or ""),
        "logo_content_type": str(getattr(uploaded_logo, "content_type", "") or "image/png"),
        "logo_base64": base64.b64encode(uploaded_logo.read()).decode("ascii"),
    }
    request.session.modified = True


def _reports_public_ack_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "template_name": "",
        "template_base64": "",
        "template_content_type": "application/pdf",
        "template_fields": [],
        "field_map": {},
    }


def _reports_public_ack_session_state(request):
    state = request.session.get(REPORTS_PUBLIC_ACK_STATE_KEY) or {}
    merged = _reports_public_ack_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    merged["template_fields"] = [dict(field) for field in list(merged.get("template_fields") or []) if isinstance(field, dict)]
    merged["field_map"] = {
        str(key or "").strip(): str(value or "").strip()
        for key, value in dict(merged.get("field_map") or {}).items()
        if str(key or "").strip()
    }
    return merged


def _reports_set_public_ack_state(request, state):
    request.session[REPORTS_PUBLIC_ACK_STATE_KEY] = state
    request.session.modified = True


def _reports_public_signature_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "selected_items": [],
        "sort_modes": [],
    }


def _reports_public_signature_session_state(request):
    state = request.session.get(REPORTS_PUBLIC_SIGNATURE_STATE_KEY) or {}
    merged = _reports_public_signature_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    merged["selected_items"] = [
        str(item or "").strip()
        for item in list(merged.get("selected_items") or [])
        if str(item or "").strip()
    ]
    sort_modes = [
        str(mode or "").strip()
        for mode in list(merged.get("sort_modes") or [])
        if str(mode or "").strip()
    ]
    valid_modes = []
    for mode in sort_modes:
        if mode in {"application_number", "item_name", "token_number"} and mode not in valid_modes:
            valid_modes.append(mode)
    merged["sort_modes"] = valid_modes
    return merged


def _reports_set_public_signature_state(request, state):
    request.session[REPORTS_PUBLIC_SIGNATURE_STATE_KEY] = state
    request.session.modified = True


def _reports_district_signature_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
    }


def _reports_district_signature_session_state(request):
    state = request.session.get(REPORTS_DISTRICT_SIGNATURE_STATE_KEY) or {}
    merged = _reports_district_signature_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    return merged


def _reports_set_district_signature_state(request, state):
    request.session[REPORTS_DISTRICT_SIGNATURE_STATE_KEY] = state
    request.session.modified = True


def _reports_simple_report_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
    }


def _reports_simple_report_session_state(request, state_key: str):
    state = request.session.get(state_key) or {}
    merged = _reports_simple_report_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    return merged


def _reports_set_simple_report_state(request, state_key: str, state):
    request.session[state_key] = state
    request.session.modified = True


SEGREGATION_BENEFICIARY_FILTER_CHOICES = [
    ("", "All"),
    (models.RecipientTypeChoices.DISTRICT, "District"),
    (models.RecipientTypeChoices.PUBLIC, "Public"),
    (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
]

SEGREGATION_ITEM_FILTER_CHOICES = [
    (models.ItemTypeChoices.ARTICLE, "Article"),
    (models.ItemTypeChoices.AID, "Aid"),
    ("", "All"),
]


def _segregation_pick_value(row: dict, aliases: list[str], default=""):
    item = dict(row or {})
    normalized = {
        _phase2_normalize_text(key): value
        for key, value in item.items()
        if str(key or "").strip()
    }
    for alias in aliases:
        normalized_alias = _phase2_normalize_text(alias)
        if normalized_alias not in normalized:
            continue
        value = normalized.get(normalized_alias)
        if value == 0:
            return value
        if str(value or "").strip():
            return value
    return default


def _segregation_display_text(*values):
    for value in values:
        if value == 0:
            return "0"
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _segregation_resolved_item_type(raw_value: str | None, *, default=models.ItemTypeChoices.ARTICLE) -> str:
    value = str(raw_value or "").strip()
    if not value and raw_value == "":
        return ""
    allowed_item_types = {choice[0] for choice in SEGREGATION_ITEM_FILTER_CHOICES}
    if value in allowed_item_types:
        return value
    normalized_value = _phase2_normalize_text(value)
    if normalized_value == _phase2_normalize_text(models.ItemTypeChoices.ARTICLE):
        return models.ItemTypeChoices.ARTICLE
    if normalized_value == _phase2_normalize_text(models.ItemTypeChoices.AID):
        return models.ItemTypeChoices.AID
    if normalized_value == "all":
        return ""
    return default


def _segregation_type_order(value: str) -> int:
    order = {
        models.RecipientTypeChoices.DISTRICT: 0,
        models.RecipientTypeChoices.PUBLIC: 1,
        models.RecipientTypeChoices.INSTITUTIONS: 2,
        models.RecipientTypeChoices.OTHERS: 3,
    }
    return order.get(str(value or "").strip(), 99)


def _segregation_normalize_row(row: dict) -> dict:
    item = dict(row or {})
    application_number = _segregation_display_text(
        _segregation_pick_value(item, ["Application Number", "App No", "application_number"]),
    )
    beneficiary_type = _segregation_display_text(
        _segregation_pick_value(item, ["Beneficiary Type", "beneficiary_type"]),
    )
    item_type = _segregation_display_text(
        _segregation_pick_value(item, ["Item Type", "item_type"]),
    )
    district_name = _segregation_display_text(
        _segregation_pick_value(item, ["District", "district"]),
    )
    beneficiary_name = _segregation_display_text(
        _segregation_pick_value(item, ["Beneficiary Name", "beneficiary_name", "Name"]),
    )
    names_value = _segregation_display_text(
        _segregation_pick_value(item, ["Names", "Beneficiary Name", "Name"]),
    )
    item_name = _segregation_display_text(
        _segregation_pick_value(
            item,
            ["Token Name", "Requested Item", "Article Name", "Article", "Item"],
        ),
    )
    waiting_hall_quantity = _phase2_parse_number(
        _segregation_pick_value(item, ["Waiting Hall Quantity", "waiting_hall_quantity"], 0)
    )
    token_quantity = _phase2_parse_number(
        _segregation_pick_value(item, ["Token Quantity", "Token Qty", "token_quantity"], 0)
    )
    sequence_no = _phase2_parse_number(
        _segregation_pick_value(item, ["Sequence No", "Sequence List", "sequence_no"], 0)
    )
    start_token_no = _phase2_parse_number(
        _segregation_pick_value(item, ["Start Token No", "Start Token No.", "token_start"], 0)
    )
    end_token_no = _phase2_parse_number(
        _segregation_pick_value(item, ["End Token No", "End Token No.", "token_end"], 0)
    )
    if token_quantity <= 0 and start_token_no > 0 and end_token_no >= start_token_no:
        token_quantity = end_token_no - start_token_no + 1

    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        beneficiary_label = _segregation_display_text(district_name, beneficiary_name, names_value, application_number)
    else:
        if application_number and beneficiary_name:
            beneficiary_label = f"{application_number} - {beneficiary_name}"
        else:
            beneficiary_label = _segregation_display_text(names_value, beneficiary_name, application_number)

    return {
        "application_number": application_number,
        "beneficiary_type": beneficiary_type,
        "item_type": item_type,
        "district_name": district_name,
        "beneficiary_name": beneficiary_name,
        "beneficiary_label": beneficiary_label,
        "item_name": item_name,
        "waiting_hall_quantity": waiting_hall_quantity,
        "token_quantity": token_quantity,
        "sequence_no": sequence_no,
        "start_token_no": start_token_no,
        "end_token_no": end_token_no,
    }


def _segregation_normalize_dataset(dataset: dict) -> dict:
    rows = []
    for row in list(dataset.get("rows") or []):
        normalized = _segregation_normalize_row(row)
        if not (
            normalized["beneficiary_label"]
            or normalized["item_name"]
            or normalized["waiting_hall_quantity"]
            or normalized["token_quantity"]
        ):
            continue
        rows.append(normalized)
    return {
        "rows": rows,
        "headers": list(dataset.get("headers") or []),
    }


def _segregation_filter_rows(rows: list[dict], *, beneficiary_type: str, item_type: str) -> list[dict]:
    filtered_rows = []
    for row in list(rows or []):
        row_beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        row_item_type = str(row.get("item_type") or "").strip()
        if beneficiary_type and row_beneficiary_type != beneficiary_type:
            continue
        if item_type and row_item_type != item_type:
            continue
        filtered_rows.append(dict(row))
    return filtered_rows


def _segregation_build_file1(rows: list[dict]) -> dict:
    grouped_map: dict[tuple[str, str], dict] = {}
    for row in list(rows or []):
        waiting_hall_quantity = int(row.get("waiting_hall_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        beneficiary_label = str(row.get("beneficiary_label") or "").strip()
        if waiting_hall_quantity <= 0 or not item_name or not beneficiary_label:
            continue
        group_key = (
            str(row.get("beneficiary_type") or "").strip(),
            beneficiary_label,
        )
        group = grouped_map.setdefault(
            group_key,
            {
                "beneficiary_type": group_key[0],
                "beneficiary_label": beneficiary_label,
                "sort_sequence": _segregation_type_order(group_key[0]),
                "items": {},
                "total_quantity": 0,
            },
        )
        group["items"][item_name] = int(group["items"].get(item_name) or 0) + waiting_hall_quantity
        group["total_quantity"] += waiting_hall_quantity

    groups = []
    row_count = 0
    total_quantity = 0
    for _, group in sorted(
        grouped_map.items(),
        key=lambda item: (
            item[1]["sort_sequence"],
            str(item[1]["beneficiary_label"]).casefold(),
        ),
    ):
        items = [
            {"article_name": item_name, "quantity": quantity}
            for item_name, quantity in sorted(group["items"].items(), key=lambda entry: entry[0].casefold())
        ]
        row_count += len(items)
        total_quantity += int(group["total_quantity"] or 0)
        groups.append(
            {
                "beneficiary_type": group["beneficiary_type"],
                "beneficiary_label": group["beneficiary_label"],
                "items": items,
                "total_quantity": int(group["total_quantity"] or 0),
            }
        )
    return {
        "groups": groups,
        "beneficiary_count": len(groups),
        "row_count": row_count,
        "total_quantity": total_quantity,
    }


def _segregation_build_file2(rows: list[dict]) -> dict:
    grouped_map: dict[str, dict] = {}
    for row in list(rows or []):
        waiting_hall_quantity = int(row.get("waiting_hall_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        beneficiary_label = str(row.get("beneficiary_label") or "").strip()
        if waiting_hall_quantity <= 0 or not item_name or not beneficiary_label:
            continue
        article_group = grouped_map.setdefault(
            item_name,
            {
                "article_name": item_name,
                "beneficiaries": {},
                "total_quantity": 0,
            },
        )
        article_group["beneficiaries"][beneficiary_label] = int(article_group["beneficiaries"].get(beneficiary_label) or 0) + waiting_hall_quantity
        article_group["total_quantity"] += waiting_hall_quantity

    groups = []
    beneficiary_row_count = 0
    total_quantity = 0
    for article_name, group in sorted(grouped_map.items(), key=lambda item: item[0].casefold()):
        beneficiaries = [
            {"beneficiary_label": beneficiary_label, "quantity": quantity}
            for beneficiary_label, quantity in sorted(group["beneficiaries"].items(), key=lambda entry: entry[0].casefold())
        ]
        beneficiary_row_count += len(beneficiaries)
        total_quantity += int(group["total_quantity"] or 0)
        groups.append(
            {
                "article_name": article_name,
                "beneficiaries": beneficiaries,
                "total_quantity": int(group["total_quantity"] or 0),
            }
        )
    return {
        "groups": groups,
        "article_count": len(groups),
        "row_count": beneficiary_row_count,
        "total_quantity": total_quantity,
    }


def _segregation_build_file3(rows: list[dict]) -> dict:
    stage_map: dict[tuple[int, str], dict] = {}
    for row in list(rows or []):
        token_quantity = int(row.get("token_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        if token_quantity <= 0 or not item_name:
            continue
        sequence_no = int(row.get("sequence_no") or 0)
        key = (sequence_no, item_name)
        stage_row = stage_map.setdefault(
            key,
            {
                "sequence_no": sequence_no,
                "item_name": item_name,
                "token_quantity": 0,
                "start_token_no": 0,
                "end_token_no": 0,
            },
        )
        stage_row["token_quantity"] += token_quantity
        start_token_no = int(row.get("start_token_no") or 0)
        end_token_no = int(row.get("end_token_no") or 0)
        if start_token_no > 0:
            if stage_row["start_token_no"] <= 0:
                stage_row["start_token_no"] = start_token_no
            else:
                stage_row["start_token_no"] = min(stage_row["start_token_no"], start_token_no)
        if end_token_no > 0:
            stage_row["end_token_no"] = max(stage_row["end_token_no"], end_token_no)

    rows_list = [
        dict(value)
        for _, value in sorted(
            stage_map.items(),
            key=lambda item: (
                item[0][0] <= 0,
                item[0][0] if item[0][0] > 0 else 10**9,
                item[0][1].casefold(),
            ),
        )
    ]
    return {
        "rows": rows_list,
        "row_count": len(rows_list),
        "total_token_quantity": sum(int(row.get("token_quantity") or 0) for row in rows_list),
    }


def _segregation_master_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Seq No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
            "Beneficiary Type": str(row.get("beneficiary_type") or ""),
            "Beneficiary": str(row.get("beneficiary_label") or ""),
            "Application No": str(row.get("application_number") or ""),
            "Item Type": str(row.get("item_type") or ""),
            "Item": str(row.get("item_name") or ""),
            "Waiting Hall Qty": int(row.get("waiting_hall_quantity") or 0),
            "Token Qty": int(row.get("token_quantity") or 0),
            "Start Token": int(row.get("start_token_no") or 0) if int(row.get("start_token_no") or 0) > 0 else "",
            "End Token": int(row.get("end_token_no") or 0) if int(row.get("end_token_no") or 0) > 0 else "",
            "District": str(row.get("district_name") or ""),
        }
        for row in list(rows or [])
    ]


def _segregation_file1_sheet_rows(groups: list[dict]) -> list[dict]:
    rows = []
    for group in list(groups or []):
        for item in list(group.get("items") or []):
            rows.append(
                {
                    "Beneficiary": str(group.get("beneficiary_label") or ""),
                    "Article": str(item.get("article_name") or ""),
                    "Quantity": int(item.get("quantity") or 0),
                    "Signature": "",
                }
            )
    return rows


def _segregation_file2_sheet_rows(groups: list[dict]) -> list[dict]:
    rows = []
    for group in list(groups or []):
        for beneficiary in list(group.get("beneficiaries") or []):
            rows.append(
                {
                    "Article": str(group.get("article_name") or ""),
                    "Beneficiary": str(beneficiary.get("beneficiary_label") or ""),
                    "Waiting Hall Quantity": int(beneficiary.get("quantity") or 0),
                }
            )
    return rows


def _segregation_file3_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Seq No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
            "Item": str(row.get("item_name") or ""),
            "Token Qty": int(row.get("token_quantity") or 0),
            "Start Token": int(row.get("start_token_no") or 0) if int(row.get("start_token_no") or 0) > 0 else "",
            "End Token": int(row.get("end_token_no") or 0) if int(row.get("end_token_no") or 0) > 0 else "",
        }
        for row in list(rows or [])
    ]


def _reports_token_lookup_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "filters": {
            "token_number": "",
            "application_number": "",
            "beneficiary_name": "",
            "item_name": "",
            "item_type": "",
        },
    }


def _reports_token_lookup_session_state(request):
    state = request.session.get(REPORTS_TOKEN_LOOKUP_STATE_KEY) or {}
    merged = _reports_token_lookup_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    filters = dict(merged.get("filters") or {})
    merged["filters"] = {
        "token_number": str(filters.get("token_number") or "").strip(),
        "application_number": str(filters.get("application_number") or "").strip(),
        "beneficiary_name": str(filters.get("beneficiary_name") or "").strip(),
        "item_name": str(filters.get("item_name") or "").strip(),
        "item_type": str(filters.get("item_type") or "").strip(),
    }
    return merged


def _reports_set_token_lookup_state(request, state):
    request.session[REPORTS_TOKEN_LOOKUP_STATE_KEY] = state
    request.session.modified = True


def _reports_token_lookup_display_value(*values):
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text or value == 0:
            return text
    return ""


def _reports_token_lookup_normalize_row(row):
    item = dict(row or {})
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    beneficiary_name = _reports_token_lookup_display_value(
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    return {
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
        "application_number": _reports_token_lookup_display_value(
            item.get("Application Number"),
            item.get("application_number"),
            item.get("App No"),
        ),
        "beneficiary_name": beneficiary_name,
        "beneficiary_type": beneficiary_type,
        "district": _reports_token_lookup_display_value(
            item.get("District"),
            item.get("district"),
            beneficiary_name if beneficiary_type == models.RecipientTypeChoices.DISTRICT else "",
        ),
        "item_name": _reports_token_lookup_display_value(
            item.get("Requested Item"),
            item.get("Token Name"),
            item.get("Article"),
        ),
        "item_type": _reports_token_lookup_display_value(
            item.get("Item Type"),
            item.get("item_type"),
        ),
        "total_value": _reports_token_lookup_display_value(
            item.get("Total Value"),
            item.get("total_value"),
            item.get("Total Amount"),
            item.get("Amount"),
        ),
        "token_quantity": int(token_quantity),
        "sequence_no": int(_phase2_parse_number(item.get("Sequence No")) or 0),
    }


def _reports_token_lookup_rows_from_session(session):
    rows = list(
        models.TokenGenerationRow.objects.filter(session=session).order_by(
            "sort_order",
            F("sequence_no").asc(nulls_last=True),
            "requested_item",
            "application_number",
            "id",
        )
    )
    prepared_rows = []
    for row in rows:
        row_data = dict(row.row_data or {})
        row_data["Application Number"] = row.application_number or row_data.get("Application Number") or ""
        row_data["Beneficiary Name"] = row.beneficiary_name or row_data.get("Beneficiary Name") or ""
        row_data["Requested Item"] = row.requested_item or row_data.get("Requested Item") or ""
        row_data["Beneficiary Type"] = row.beneficiary_type or row_data.get("Beneficiary Type") or ""
        row_data["Sequence No"] = row.sequence_no if row.sequence_no is not None else row_data.get("Sequence No") or ""
        row_data["Start Token No"] = row.start_token_no if row.start_token_no is not None else row_data.get("Start Token No") or 0
        row_data["End Token No"] = row.end_token_no if row.end_token_no is not None else row_data.get("End Token No") or 0
        existing_token_quantity = row_data.get("Token Quantity")
        if existing_token_quantity in {None, ""}:
            existing_token_quantity = max(
                (row.end_token_no or 0) - (row.start_token_no or 0) + 1,
                0,
            )
        row_data["Token Quantity"] = existing_token_quantity
        row_data["Total Value"] = (
            row_data.get("Total Value")
            or row_data.get("total_value")
            or row_data.get("Total Amount")
            or row_data.get("Amount")
            or ""
        )
        normalized = _reports_token_lookup_normalize_row(row_data)
        if int(normalized.get("token_quantity") or 0) > 0:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_token_lookup_data_rows(rows):
    return [
        normalized
        for normalized in (_reports_token_lookup_normalize_row(row) for row in list(rows or []))
        if int(normalized.get("token_quantity") or 0) > 0
    ]


def _reports_token_lookup_filters_from_post(post_data):
    return {
        "token_number": str(post_data.get("token_number") or "").strip(),
        "application_number": str(post_data.get("application_number") or "").strip(),
        "beneficiary_name": str(post_data.get("beneficiary_name") or "").strip(),
        "item_name": str(post_data.get("item_name") or "").strip(),
        "item_type": str(post_data.get("item_type") or "").strip(),
    }


def _reports_token_lookup_filter_rows(rows, filters):
    filtered = list(rows or [])
    token_number = _phase2_parse_number((filters or {}).get("token_number"))
    application_number = str((filters or {}).get("application_number") or "").strip().casefold()
    beneficiary_name = str((filters or {}).get("beneficiary_name") or "").strip().casefold()
    item_name = str((filters or {}).get("item_name") or "").strip().casefold()
    item_type = str((filters or {}).get("item_type") or "").strip().casefold()

    if token_number:
        filtered = [
            row for row in filtered
            if int(row.get("token_start") or 0) <= int(token_number) <= int(row.get("token_end") or row.get("token_start") or 0)
        ]
    if application_number:
        filtered = [row for row in filtered if application_number in str(row.get("application_number") or "").casefold()]
    if beneficiary_name:
        filtered = [row for row in filtered if beneficiary_name in str(row.get("beneficiary_name") or "").casefold()]
    if item_name:
        filtered = [row for row in filtered if item_name in str(row.get("item_name") or "").casefold()]
    if item_type:
        filtered = [row for row in filtered if item_type in str(row.get("item_type") or "").casefold()]
    return filtered


def _reports_token_lookup_choice_values(rows, key):
    return sorted(
        {
            str(row.get(key) or "").strip()
            for row in list(rows or [])
            if str(row.get(key) or "").strip()
        },
        key=lambda value: value.casefold(),
    )


def _reports_public_signature_normalize_row(row):
    item = dict(row or {})
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    if beneficiary_type != models.RecipientTypeChoices.PUBLIC:
        return None
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    if token_quantity <= 0:
        return None
    application_number = _reports_token_lookup_display_value(
        item.get("Application Number"),
        item.get("application_number"),
        item.get("App No"),
    )
    beneficiary_name = _reports_token_lookup_display_value(
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    item_name = _reports_token_lookup_display_value(
        item.get("Requested Item"),
        item.get("requested_item"),
        item.get("Token Name"),
        item.get("Article"),
    )
    return {
        "application_number": application_number,
        "beneficiary_name": beneficiary_name,
        "item_name": item_name,
        "item_type": _reports_token_lookup_display_value(
            item.get("Item Type"),
            item.get("item_type"),
        ),
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
        "token_quantity": int(token_quantity),
    }


def _reports_public_signature_rows_from_dataset(rows):
    prepared_rows = []
    for row in list(rows or []):
        normalized = _reports_public_signature_normalize_row(row)
        if normalized:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_public_signature_rows_from_session(session):
    dataset = _token_generation_saved_dataset(session) if session else {"rows": []}
    return _reports_public_signature_rows_from_dataset(dataset.get("rows") or [])


def _reports_public_signature_item_options(rows):
    counts = {}
    for row in list(rows or []):
        item_name = str(row.get("item_name") or "").strip()
        if not item_name:
            continue
        counts[item_name] = counts.get(item_name, 0) + 1
    return [
        {
            "item_name": item_name,
            "row_count": counts[item_name],
        }
        for item_name in sorted(counts.keys(), key=lambda value: value.casefold())
    ]


def _reports_public_signature_sort_rows(rows, sort_modes):
    selected_modes = [
        str(mode or "").strip().casefold()
        for mode in list(sort_modes or [])
        if str(mode or "").strip()
    ]
    normalized_modes = []
    for mode in selected_modes:
        if mode in {"application_number", "item_name", "token_number"} and mode not in normalized_modes:
            normalized_modes.append(mode)
    selected_modes = normalized_modes
    if not selected_modes:
        return list(rows or [])

    def _row_key(row):
        application_number = services._public_signature_app_sort_key(str(row.get("application_number") or ""))
        item_name = str(row.get("item_name") or "").strip().casefold()
        token_number = int(row.get("token_start") or 0)
        values = {
            "application_number": application_number,
            "item_name": item_name,
            "token_number": token_number,
        }
        return tuple(values[mode] for mode in selected_modes)

    return sorted(list(rows or []), key=_row_key)


def _reports_district_signature_normalize_row(row):
    item = dict(row or {})
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    if beneficiary_type != models.RecipientTypeChoices.DISTRICT:
        return None
    district_name = _reports_token_lookup_display_value(
        item.get("District"),
        item.get("district"),
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    item_name = _reports_token_lookup_display_value(
        item.get("Requested Item"),
        item.get("requested_item"),
        item.get("Token Name"),
        item.get("Article"),
    )
    total_quantity = max(_phase2_parse_number(item.get("Quantity")) or 0, 0)
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    if not district_name or not item_name or (total_quantity <= 0 and token_quantity <= 0):
        return None
    return {
        "district_name": district_name,
        "item_name": item_name,
        "total_quantity": int(total_quantity),
        "token_quantity": int(token_quantity),
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
    }


def _reports_district_signature_rows_from_dataset(rows):
    prepared_rows = []
    for row in list(rows or []):
        normalized = _reports_district_signature_normalize_row(row)
        if normalized:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_district_signature_rows_from_session(session):
    dataset = _token_generation_saved_dataset(session) if session else {"rows": []}
    return _reports_district_signature_rows_from_dataset(dataset.get("rows") or [])


def _reports_district_signature_grouped(rows):
    grouped = {}
    for row in list(rows or []):
        district_name = str(row.get("district_name") or "").strip()
        item_name = str(row.get("item_name") or "").strip()
        if not district_name or not item_name:
            continue
        district_bucket = grouped.setdefault(district_name, {})
        item_bucket = district_bucket.setdefault(
            item_name,
            {
                "item_name": item_name,
                "total_quantity": 0,
                "token_quantity": 0,
                "start_token": None,
                "end_token": None,
            },
        )
        item_bucket["total_quantity"] += int(row.get("total_quantity") or 0)
        current_token_qty = int(row.get("token_quantity") or 0)
        item_bucket["token_quantity"] += current_token_qty
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or 0)
        if current_token_qty > 0 and token_start > 0:
            item_bucket["start_token"] = token_start if item_bucket["start_token"] is None else min(item_bucket["start_token"], token_start)
            item_bucket["end_token"] = token_end if item_bucket["end_token"] is None else max(item_bucket["end_token"], token_end)

    districts = []
    total_quantity = 0
    total_token_quantity = 0
    for district_name in sorted(grouped.keys(), key=lambda value: value.casefold()):
        items = [grouped[district_name][name] for name in sorted(grouped[district_name].keys(), key=lambda value: value.casefold())]
        district_total_qty = sum(int(item.get("total_quantity") or 0) for item in items)
        district_token_qty = sum(int(item.get("token_quantity") or 0) for item in items)
        positive_starts = [int(item["start_token"]) for item in items if item.get("start_token")]
        positive_ends = [int(item["end_token"]) for item in items if item.get("end_token")]
        districts.append(
            {
                "district_name": district_name,
                "items": items,
                "total_quantity": district_total_qty,
                "token_quantity": district_token_qty,
                "start_token": min(positive_starts) if positive_starts else None,
                "end_token": max(positive_ends) if positive_ends else None,
            }
        )
        total_quantity += district_total_qty
        total_token_quantity += district_token_qty
    return {
        "districts": districts,
        "district_count": len(districts),
        "item_count": total_quantity,
        "total_quantity": total_quantity,
        "total_token_quantity": total_token_quantity,
    }


def _reports_public_ack_data_rows(rows):
    prepared_rows = []
    has_beneficiary_type = any(
        str(row.get("Beneficiary Type") or row.get("beneficiary_type") or "").strip()
        for row in rows
    )
    for row in rows:
        item = dict(row or {})
        beneficiary_type = str(item.get("Beneficiary Type") or item.get("beneficiary_type") or "").strip()
        quantity = _phase2_parse_number(item.get("Quantity"))
        if quantity is None:
            quantity = _phase2_parse_number(item.get("Token Quantity"))
        if quantity is None:
            quantity = 0
        if has_beneficiary_type and beneficiary_type and _phase2_normalize_text(beneficiary_type) != "public":
            continue
        if quantity <= 0 and has_beneficiary_type:
            continue
        if not has_beneficiary_type and quantity <= 0:
            continue
        item["Quantity"] = quantity
        prepared_rows.append(item)
    return prepared_rows


def _reports_public_ack_template_fields(template_bytes):
    fields = services._public_acknowledgment_pdf_fields(template_bytes)
    return [
        {
            "field_name": field["field_name"],
            "field_key": field["field_key"],
        }
        for field in fields
    ]


def _reports_public_ack_column_options(headers):
    fallback = [
        "District",
        "Address",
        "Application Number",
        "Start Token No",
        "Beneficiary Name",
        "Mobile",
        "Aadhar Number",
        "Requested Item",
        "Token Name",
        "Total Value",
        "Cost Per Unit",
        "Cheque / RTGS in Favour",
        "Name of Institution",
        "Quantity",
    ]
    options = []
    seen = set()
    for header in list(headers or []) + fallback:
        text = str(header or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        options.append(text)
    return options


def _reports_public_ack_default_field_map(headers, template_fields):
    normalized_headers = {_phase2_normalize_text(header): header for header in (headers or []) if str(header or "").strip()}
    candidates = {
        "district": ["District"],
        "address": ["Address"],
        "app_no": ["Application Number"],
        "token": ["Start Token No", "Token No", "Sequence No"],
        "bf_name": ["Beneficiary Name", "Name of Beneficiary", "Name"],
        "mobile": ["Mobile"],
        "aadhar": ["Aadhar Number", "Aadhaar Number"],
        "article": ["Requested Item", "Token Name", "Article"],
        "value_aid": ["Total Value", "Cost Per Unit"],
        "cheque_no": ["Cheque / RTGS in Favour", "Cheque No"],
    }
    mapping = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        selected = ""
        for candidate in candidates.get(field_key, []):
            match = normalized_headers.get(_phase2_normalize_text(candidate))
            if match:
                selected = match
                break
        mapping[field_key] = selected
    return mapping


def _reports_public_ack_field_map_from_post(post_data, template_fields):
    mapping = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        mapping[field_key] = str(post_data.get(f"public_ack_map__{field_key}") or "").strip()
    return mapping


def _reports_public_ack_field_map_with_defaults(headers, template_fields, existing_map=None):
    defaults = _reports_public_ack_default_field_map(headers, template_fields)
    existing_map = {
        str(key or "").strip(): str(value or "").strip()
        for key, value in dict(existing_map or {}).items()
        if str(key or "").strip()
    }
    merged = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        merged[field_key] = existing_map.get(field_key) or defaults.get(field_key, "")
    return merged


def _reports_public_ack_normalize_dataset(rows):
    return _reports_public_ack_data_rows(rows)


def _reports_waiting_hall_grouped_data(rows, ignored_keys=None, beneficiary_type_filter="", item_type_filter=""):
    ignored = set(ignored_keys or [])
    selected_type = str(beneficiary_type_filter or "").strip().lower()
    selected_item_type = str(item_type_filter or "").strip().lower()
    if selected_type == "all":
        selected_type = ""
    if selected_item_type == "all":
        selected_item_type = ""
    grouped = {}
    for row in rows:
        beneficiary_type = str(row.get("Beneficiary Type") or row.get("beneficiary_type") or "").strip()
        if selected_type and beneficiary_type.lower() != selected_type:
            continue
        item_type = str(row.get("Item Type") or row.get("item_type") or "").strip()
        if selected_item_type and item_type.lower() != selected_item_type:
            continue
        waiting_quantity = _phase2_parse_number(row.get("Waiting Hall Quantity"))
        if not waiting_quantity or waiting_quantity <= 0:
            continue
        raw_name = str(
            row.get("District")
            or row.get("Names")
            or row.get("Beneficiary Name")
            or row.get("Name of Beneficiary")
            or row.get("Name of Institution")
            or row.get("district")
            or row.get("beneficiary_name")
            or row.get("application_number")
            or row.get("Application Number")
            or ""
        ).strip()
        item_name = str(row.get("Requested Item") or row.get("requested_item") or "").strip()
        if not raw_name or not item_name:
            continue
        beneficiary_type_label = beneficiary_type or "Unknown"
        key = f"{beneficiary_type_label}||{raw_name}||{item_name}"
        group_key = f"{beneficiary_type_label}||{raw_name}"
        group_bucket = grouped.setdefault(
            group_key,
            {
                "entity_name": raw_name,
                "entity_kind": beneficiary_type_label,
                "label_prefix": "District" if beneficiary_type_label == models.RecipientTypeChoices.DISTRICT else "Beneficiary",
                "items_map": {},
            },
        )
        entry = group_bucket["items_map"].setdefault(
            item_name,
            {
                "key": key,
                "entity_name": raw_name,
                "entity_kind": beneficiary_type_label,
                "requested_item": item_name,
                "item_type": item_type,
                "quantity": 0,
            },
        )
        entry["quantity"] += int(waiting_quantity or 0)
    districts = []
    total_items = 0
    total_quantity = 0
    available_keys = []
    for group_key in sorted(grouped.keys(), key=lambda value: value.lower()):
        group_meta = grouped[group_key]
        raw_items = [group_meta["items_map"][name] for name in sorted(group_meta["items_map"].keys(), key=lambda value: value.lower())]
        available_keys.extend(item["key"] for item in raw_items)
        filtered_items = [item for item in raw_items if item["key"] not in ignored]
        district_total = sum(int(item["quantity"] or 0) for item in filtered_items)
        total_items += len(filtered_items)
        total_quantity += district_total
        districts.append(
            {
                "entity_name": group_meta["entity_name"],
                "entity_kind": group_meta["entity_kind"],
                "label_prefix": group_meta["label_prefix"],
                "items": filtered_items,
                "raw_items": raw_items,
                "total_quantity": district_total,
                "item_count": len(filtered_items),
            }
        )
    return {
        "districts": districts,
        "district_count": len(districts),
        "item_count": total_items,
        "total_quantity": total_quantity,
        "available_keys": available_keys,
    }


class ReportsView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.REPORTS
    permission_action = "view"
    template_name = "dashboard/reports.html"

    REPORT_TABS = (
        {
            "key": "token-lookup",
            "label": "Token Lookup",
            "description": "Search and inspect token-generated rows by token number, application number, name, district, or item.",
        },
        {
            "key": "waiting-hall-acknowledgment",
            "label": "Waiting Hall Acknowledgment",
            "description": "District-wise article acknowledgment sheets for waiting hall collections.",
        },
        {
            "key": "public-acknowledgment-form",
            "label": "Public Acknowledgment Form",
            "description": "Public beneficiary acknowledgment forms filled from an uploaded PDF template.",
        },
        {
            "key": "public-signature",
            "label": "Signatures",
            "description": "Signature reports for public and district print sheets.",
        },
        {
            "key": "reports-home",
            "label": "Reports",
            "description": "Segregation and distribution report staging cards.",
        },
    )

    def post(self, request, *args, **kwargs):
        active_tab = (request.POST.get("tab") or self.REPORT_TABS[0]["key"]).strip()
        shared_logo = _reports_shared_logo_state(request)
        if (request.POST.get("action") or "").strip() == "upload_shared_logo":
            uploaded_logo = request.FILES.get("shared_logo")
            if uploaded_logo:
                _reports_set_shared_logo_state(request, uploaded_logo=uploaded_logo)
                messages.success(request, "Reports logo updated.")
            else:
                messages.error(request, "Choose a logo file to upload.")
            target = active_tab or self.REPORT_TABS[0]["key"]
            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab={target}")
        action = (request.POST.get("action") or "").strip()
        if active_tab == "reports-home":
            segregation_state = _reports_simple_report_session_state(request, REPORTS_SEGREGATION_STATE_KEY)
            distribution_state = _reports_simple_report_session_state(request, REPORTS_DISTRIBUTION_STATE_KEY)
            segregation_beneficiary_type = str(request.POST.get("segregation_beneficiary_type") or "").strip()
            allowed_beneficiary_types = {choice[0] for choice in SEGREGATION_BENEFICIARY_FILTER_CHOICES}
            if segregation_beneficiary_type not in allowed_beneficiary_types:
                segregation_beneficiary_type = ""
            segregation_item_type = _segregation_resolved_item_type(
                request.POST.get("segregation_item_type"),
                default=models.ItemTypeChoices.ARTICLE,
            )

            def _reports_home_redirect():
                params = {
                    "tab": "reports-home",
                    "seg_item_type": segregation_item_type or models.ItemTypeChoices.ARTICLE,
                }
                if segregation_beneficiary_type:
                    params["seg_beneficiary_type"] = segregation_beneficiary_type
                return HttpResponseRedirect(f"{reverse('ui:reports')}?{urlencode(params)}")

            def _simple_report_sync(state, label: str):
                session = _reports_active_session()
                dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
                rows = [dict(row) for row in dataset.get("rows") or []]
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": "Synced from Token Generation",
                        "session_id": getattr(session, "pk", None),
                        "rows": rows,
                        "headers": _phase2_unique_headers(dataset.get("headers") or []),
                    }
                )
                _reports_set_simple_report_state(request, label, state)
                return len(rows)

            def _simple_report_upload(state, label: str, uploaded_file):
                source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
                if not source_headers:
                    return None, None
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": uploaded_file.name,
                        "session_id": None,
                        "rows": [dict(row) for row in source_rows],
                        "headers": _phase2_unique_headers(source_headers),
                    }
                )
                _reports_set_simple_report_state(request, label, state)
                return len(source_rows), source_headers

            if action == "sync_reports_segregation":
                rows_count = _simple_report_sync(segregation_state, REPORTS_SEGREGATION_STATE_KEY)
                messages.success(request, f"Segregation synced. Rows: {rows_count}.")
                return _reports_home_redirect()

            if action == "upload_reports_segregation":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return _reports_home_redirect()
                rows_count, source_headers = _simple_report_upload(segregation_state, REPORTS_SEGREGATION_STATE_KEY, uploaded_file)
                if rows_count is None:
                    messages.error(request, "Uploaded file is empty.")
                    return _reports_home_redirect()
                messages.success(request, f"Uploaded {rows_count} segregation row(s).")
                return _reports_home_redirect()

            if action in {
                "download_reports_segregation_excel",
                "download_reports_segregation_file1_pdf",
                "download_reports_segregation_file2_pdf",
                "download_reports_segregation_file3_pdf",
                "preview_reports_segregation_file1_pdf",
                "preview_reports_segregation_file2_pdf",
                "preview_reports_segregation_file3_pdf",
            }:
                normalized_dataset = _segregation_normalize_dataset(segregation_state)
                filtered_rows = _segregation_filter_rows(
                    normalized_dataset.get("rows") or [],
                    beneficiary_type=segregation_beneficiary_type,
                    item_type=segregation_item_type,
                )
                file1_data = _segregation_build_file1(filtered_rows)
                file2_data = _segregation_build_file2(filtered_rows)
                file3_data = _segregation_build_file3(filtered_rows)
                if action == "download_reports_segregation_excel":
                    workbook_stream = services.generate_segregation_xlsx(
                        master_rows=_segregation_master_sheet_rows(normalized_dataset.get("rows") or []),
                        file1_rows=_segregation_file1_sheet_rows(file1_data["groups"]),
                        file2_rows=_segregation_file2_sheet_rows(file2_data["groups"]),
                        file3_rows=_segregation_file3_sheet_rows(file3_data["rows"]),
                    )
                    return FileResponse(
                        workbook_stream,
                        as_attachment=True,
                        filename=f"segregation_reports_{timezone.localtime().strftime('%d_%b_%Y_%H_%M')}.xlsx",
                    )
                if action == "download_reports_segregation_file1_pdf":
                    pdf_stream = services.generate_segregation_file1_pdf(file1_data["groups"])
                    return FileResponse(
                        pdf_stream,
                        as_attachment=True,
                        filename=f"segregation_file_1_{timezone.localtime().strftime('%d_%b_%Y_%H_%M')}.pdf",
                    )
                if action == "preview_reports_segregation_file1_pdf":
                    pdf_stream = services.generate_segregation_file1_pdf(file1_data["groups"])
                    return FileResponse(
                        pdf_stream,
                        as_attachment=False,
                        filename="segregation_file_1_preview.pdf",
                    )
                if action == "download_reports_segregation_file2_pdf":
                    pdf_stream = services.generate_segregation_file2_pdf(file2_data["groups"])
                    return FileResponse(
                        pdf_stream,
                        as_attachment=True,
                        filename=f"segregation_file_2_{timezone.localtime().strftime('%d_%b_%Y_%H_%M')}.pdf",
                    )
                if action == "preview_reports_segregation_file2_pdf":
                    pdf_stream = services.generate_segregation_file2_pdf(file2_data["groups"])
                    return FileResponse(
                        pdf_stream,
                        as_attachment=False,
                        filename="segregation_file_2_preview.pdf",
                    )
                if action == "preview_reports_segregation_file3_pdf":
                    pdf_stream = services.generate_segregation_file3_pdf(file3_data["rows"])
                    return FileResponse(
                        pdf_stream,
                        as_attachment=False,
                        filename="segregation_file_3_preview.pdf",
                    )
                pdf_stream = services.generate_segregation_file3_pdf(file3_data["rows"])
                return FileResponse(
                    pdf_stream,
                    as_attachment=True,
                    filename=f"segregation_file_3_{timezone.localtime().strftime('%d_%b_%Y_%H_%M')}.pdf",
                )

            if action == "sync_reports_distribution":
                rows_count = _simple_report_sync(distribution_state, REPORTS_DISTRIBUTION_STATE_KEY)
                messages.success(request, f"Distribution synced. Rows: {rows_count}.")
                return _reports_home_redirect()

            if action == "upload_reports_distribution":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return _reports_home_redirect()
                rows_count, source_headers = _simple_report_upload(distribution_state, REPORTS_DISTRIBUTION_STATE_KEY, uploaded_file)
                if rows_count is None:
                    messages.error(request, "Uploaded file is empty.")
                    return _reports_home_redirect()
                messages.success(request, f"Uploaded {rows_count} distribution row(s).")
                return _reports_home_redirect()

            return _reports_home_redirect()

        if active_tab == "token-lookup":
            state = _reports_token_lookup_session_state(request)
            if action == "sync_token_lookup":
                session = _reports_active_session()
                dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
                rows = _reports_token_lookup_rows_from_session(session) if session else []
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": "Synced from Token Generation",
                        "session_id": getattr(session, "pk", None),
                        "rows": rows,
                        "headers": _phase2_unique_headers(dataset.get("headers") or []),
                    }
                )
                _reports_set_token_lookup_state(request, state)
                messages.success(request, f"Token Lookup synced. Rows: {len(rows)}.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")

            if action == "upload_token_lookup":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")
                source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
                if not source_headers:
                    messages.error(request, "Uploaded file is empty.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")
                rows = _reports_token_lookup_data_rows(source_rows)
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": uploaded_file.name,
                        "session_id": None,
                        "rows": rows,
                        "headers": _phase2_unique_headers(source_headers),
                    }
                )
                _reports_set_token_lookup_state(request, state)
                messages.success(request, f"Uploaded {len(rows)} token lookup row(s).")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")

            if action == "clear_token_lookup_filters":
                state["filters"] = _reports_token_lookup_default_state()["filters"]
                _reports_set_token_lookup_state(request, state)
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")

            if action == "search_token_lookup":
                if not state.get("loaded"):
                    messages.warning(request, "Sync Data or Upload a file first before searching tokens.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")
                state["filters"] = _reports_token_lookup_filters_from_post(request.POST)
                _reports_set_token_lookup_state(request, state)
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")

            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=token-lookup")

        if active_tab == "public-acknowledgment-form":
            state = _reports_public_ack_session_state(request)
            if action == "sync_public_ack":
                session = _reports_active_session()
                dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
                rows = _reports_public_ack_normalize_dataset(dataset.get("rows") or [])
                headers = _phase2_unique_headers(dataset.get("headers") or [])
                template_fields = list(state.get("template_fields") or [])
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": "Synced from Token Generation",
                        "session_id": getattr(session, "pk", None),
                        "rows": rows,
                        "headers": headers,
                        "field_map": _reports_public_ack_field_map_with_defaults(
                            headers,
                            template_fields,
                            state.get("field_map"),
                        ),
                    }
                )
                _reports_set_public_ack_state(request, state)
                messages.success(request, f"Public Acknowledgment Form synced. Rows: {len(rows)}.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")

            if action == "upload_public_ack_data":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
                if not source_headers:
                    messages.error(request, "Uploaded file is empty.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                rows = _reports_public_ack_normalize_dataset(source_rows)
                template_fields = list(state.get("template_fields") or [])
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": uploaded_file.name,
                        "session_id": None,
                        "rows": rows,
                        "headers": _phase2_unique_headers(source_headers),
                        "field_map": _reports_public_ack_field_map_with_defaults(
                            source_headers,
                            template_fields,
                            state.get("field_map"),
                        ),
                    }
                )
                _reports_set_public_ack_state(request, state)
                messages.success(request, f"Uploaded {len(rows)} public acknowledgment row(s).")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")

            if action == "upload_public_ack_template":
                uploaded_template = request.FILES.get("template_pdf") or request.FILES.get("template")
                if not uploaded_template:
                    messages.error(request, "Choose a PDF template to upload.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                template_bytes = uploaded_template.read()
                template_fields = _reports_public_ack_template_fields(template_bytes)
                if not template_fields:
                    messages.error(request, "No fillable fields were found in the uploaded PDF template.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                state.update(
                    {
                        "template_name": uploaded_template.name,
                        "template_base64": base64.b64encode(template_bytes).decode("ascii"),
                        "template_content_type": "application/pdf",
                        "template_fields": template_fields,
                        "field_map": _reports_public_ack_field_map_with_defaults(
                            state.get("headers") or [],
                            template_fields,
                            state.get("field_map"),
                        ),
                    }
                )
                _reports_set_public_ack_state(request, state)
                messages.success(request, f"Template loaded with {len(template_fields)} fillable field(s).")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")

            if action in {"download_public_ack", "preview_public_ack"}:
                if not state.get("loaded"):
                    messages.error(request, "Sync Data or Upload a file first before downloading this report.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                if not state.get("template_base64"):
                    messages.error(request, "Upload a PDF template first.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                rows = list(state.get("rows") or [])
                if not rows:
                    messages.error(request, "No public acknowledgment rows are available for the current selection.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                template_fields = list(state.get("template_fields") or [])
                submitted_map = _reports_public_ack_field_map_from_post(request.POST, template_fields)
                field_map = _reports_public_ack_field_map_with_defaults(
                    state.get("headers") or [],
                    template_fields,
                    {**state.get("field_map", {}), **{key: value for key, value in submitted_map.items() if value}},
                )
                state["field_map"] = field_map
                _reports_set_public_ack_state(request, state)
                template_bytes = base64.b64decode(state.get("template_base64") or "")
                if not template_bytes:
                    messages.error(request, "Upload a valid PDF template first.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                field_name_map = {
                    field["field_name"]: field_map.get(field["field_key"], "")
                    for field in template_fields
                    if field_map.get(field["field_key"], "")
                }
                if not field_name_map:
                    messages.error(request, "Map at least one template field before downloading.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")
                buffer = services.generate_public_acknowledgment_pdf(
                    template_bytes,
                    rows,
                    field_name_map,
                )
                filename = f"public_acknowledgment_{timezone.localtime().strftime('%d_%b_%y_%H_%M')}.pdf"
                response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
                disposition = "inline" if action == "preview_public_ack" else "attachment"
                response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
                return response

            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-acknowledgment-form")

        if active_tab == "public-signature":
            state = _reports_public_signature_session_state(request)
            district_state = _reports_district_signature_session_state(request)
            if action == "sync_public_signature":
                session = _reports_active_session()
                dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
                rows = _reports_public_signature_rows_from_session(session) if session else []
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": "Synced from Token Generation",
                        "session_id": getattr(session, "pk", None),
                        "rows": rows,
                        "headers": _phase2_unique_headers(dataset.get("headers") or []),
                        "selected_items": [],
                        "sort_modes": [],
                    }
                )
                _reports_set_public_signature_state(request, state)
                messages.success(request, f"Public Signature synced. Rows: {len(rows)}.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

            if action == "upload_public_signature":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
                if not source_headers:
                    messages.error(request, "Uploaded file is empty.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                rows = _reports_public_signature_rows_from_dataset(source_rows)
                state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": uploaded_file.name,
                        "session_id": None,
                        "rows": rows,
                        "headers": _phase2_unique_headers(source_headers),
                        "selected_items": [],
                        "sort_modes": [],
                    }
                )
                _reports_set_public_signature_state(request, state)
                messages.success(request, f"Uploaded {len(rows)} public signature row(s).")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

            if action == "set_public_signature_sort":
                ordered_modes = [
                    str(mode or "").strip()
                    for mode in (request.POST.get("sort_modes_order") or "").split(",")
                    if str(mode or "").strip() in {"application_number", "item_name", "token_number"}
                ]
                requested_modes = []
                for mode in ordered_modes:
                    if mode not in requested_modes:
                        requested_modes.append(mode)
                if not requested_modes:
                    requested_modes = [
                        str(mode or "").strip()
                        for mode in request.POST.getlist("sort_modes")
                        if str(mode or "").strip() in {"application_number", "item_name", "token_number"} and str(mode or "").strip() not in requested_modes
                    ]
                state["sort_modes"] = requested_modes
                _reports_set_public_signature_state(request, state)
                messages.success(request, "Public Signature sort updated.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

            if action in {"download_public_signature", "preview_public_signature"}:
                if not state.get("loaded"):
                    messages.error(request, "Sync Data or Upload a file first before downloading this report.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                download_format = "pdf" if action == "preview_public_signature" else (request.POST.get("download_format") or "pdf").strip().lower()
                selected_items = [
                    str(item or "").strip()
                    for item in request.POST.getlist("selected_items")
                    if str(item or "").strip()
                ]
                state["selected_items"] = selected_items
                _reports_set_public_signature_state(request, state)
                if not selected_items:
                    messages.error(request, "Select at least one public item before downloading.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                selected_rows = [
                    row for row in list(state.get("rows") or [])
                    if str(row.get("item_name") or "").strip() in set(selected_items)
                ]
                if not selected_rows:
                    messages.error(request, "No public rows are available for the selected items.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                selected_rows = _reports_public_signature_sort_rows(selected_rows, state.get("sort_modes"))
                stamp = timezone.localtime().strftime('%d_%b_%y_%H_%M')
                if download_format == "xlsx":
                    buffer = services.generate_public_signature_xlsx(selected_rows)
                    filename = f"public_signature_{stamp}.xlsx"
                    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    buffer = services.generate_public_signature_pdf(selected_rows)
                    filename = f"public_signature_{stamp}.pdf"
                    content_type = "application/pdf"
                response = HttpResponse(buffer.getvalue(), content_type=content_type)
                disposition = "inline" if action == "preview_public_signature" and content_type == "application/pdf" else "attachment"
                response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
                return response

            if action == "sync_district_signature":
                session = _reports_active_session()
                dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
                rows = _reports_district_signature_rows_from_session(session) if session else []
                district_state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": "Synced from Token Generation",
                        "session_id": getattr(session, "pk", None),
                        "rows": rows,
                        "headers": _phase2_unique_headers(dataset.get("headers") or []),
                    }
                )
                _reports_set_district_signature_state(request, district_state)
                messages.success(request, f"District Signature synced. Rows: {len(rows)}.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

            if action == "upload_district_signature":
                uploaded_file = request.FILES.get("file")
                if not uploaded_file:
                    messages.error(request, "Choose a CSV or Excel file to upload.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
                if not source_headers:
                    messages.error(request, "Uploaded file is empty.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                rows = _reports_district_signature_rows_from_dataset(source_rows)
                district_state.update(
                    {
                        "loaded": True,
                        "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                        "source": uploaded_file.name,
                        "session_id": None,
                        "rows": rows,
                        "headers": _phase2_unique_headers(source_headers),
                    }
                )
                _reports_set_district_signature_state(request, district_state)
                messages.success(request, f"Uploaded {len(rows)} district signature row(s).")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

            if action in {"download_district_signature", "preview_district_signature"}:
                if not district_state.get("loaded"):
                    messages.error(request, "Sync Data or Upload a file first before downloading this report.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                download_format = "pdf" if action == "preview_district_signature" else (request.POST.get("download_format") or "pdf").strip().lower()
                grouped = _reports_district_signature_grouped(district_state.get("rows") or [])
                if not grouped.get("districts"):
                    messages.error(request, "No district rows are available for this report.")
                    return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")
                stamp = timezone.localtime().strftime('%d_%b_%y_%H_%M')
                logo_bytes = base64.b64decode(shared_logo["logo_base64"]) if shared_logo.get("logo_base64") else None
                if logo_bytes:
                    logo_bytes, logo_mime = services._optimized_report_logo(
                        logo_bytes,
                        shared_logo.get("logo_content_type") or "image/png",
                        max_width_px=180,
                        max_height_px=180,
                    )
                else:
                    logo_mime = shared_logo.get("logo_content_type") or "image/png"
                if download_format == "xlsx":
                    buffer = services.generate_district_signature_xlsx(
                        grouped.get("districts") or [],
                        custom_logo=logo_bytes,
                        custom_logo_mime_type=logo_mime,
                    )
                    filename = f"district_signature_{stamp}.xlsx"
                    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    buffer = services.generate_district_signature_pdf(
                        grouped.get("districts") or [],
                        custom_logo=logo_bytes,
                    )
                    filename = f"district_signature_{stamp}.pdf"
                    content_type = "application/pdf"
                response = HttpResponse(buffer.getvalue(), content_type=content_type)
                disposition = "inline" if action == "preview_district_signature" and content_type == "application/pdf" else "attachment"
                response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
                return response

            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=public-signature")

        if active_tab != "waiting-hall-acknowledgment":
            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab={active_tab}")

        state = _reports_waiting_hall_session_state(request)
        shared_logo = _reports_shared_logo_state(request)

        if action == "sync_waiting_hall":
            session = _reports_active_session()
            dataset = _token_generation_saved_dataset(session) if session else {"rows": [], "headers": []}
            rows = [dict(row) for row in dataset.get("rows") or []]
            grouped = _reports_waiting_hall_grouped_data(rows, ignored_keys=[])
            state.update(
                {
                    "loaded": True,
                    "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                    "source": "Synced from Token Generation",
                    "session_id": getattr(session, "pk", None),
                    "ignored_keys": [],
                    "rows": rows,
                    "headers": list(dataset.get("headers") or []),
                    "beneficiary_type_filter": "",
                    "item_type_filter": models.ItemTypeChoices.AID,
                }
            )
            request.session[REPORTS_WAITING_HALL_STATE_KEY] = state
            request.session.modified = True
            messages.success(request, f"Waiting Hall Acknowledgment synced. Districts: {grouped['district_count']}, items: {grouped['item_count']}.")
            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")

        if action == "upload_waiting_hall":
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Choose a CSV or Excel file to upload.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")
            source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
            if not source_headers:
                messages.error(request, "Uploaded file is empty.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")
            grouped = _reports_waiting_hall_grouped_data(source_rows, ignored_keys=[])
            state.update(
                {
                    "loaded": True,
                    "synced_at": timezone.localtime().strftime("%d %b %Y at %H:%M"),
                    "source": uploaded_file.name,
                    "session_id": None,
                    "ignored_keys": [],
                    "rows": source_rows,
                    "headers": source_headers,
                    "beneficiary_type_filter": "",
                    "item_type_filter": models.ItemTypeChoices.AID,
                }
            )
            request.session[REPORTS_WAITING_HALL_STATE_KEY] = state
            request.session.modified = True
            messages.success(request, f"Uploaded {len(source_rows)} row(s) into Waiting Hall Acknowledgment.")
            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")

        selected_ignored = request.POST.getlist("ignored_keys")
        state["ignored_keys"] = selected_ignored
        if "beneficiary_type_filter" in request.POST:
            state["beneficiary_type_filter"] = str(request.POST.get("beneficiary_type_filter") or "").strip()
        else:
            state["beneficiary_type_filter"] = str(state.get("beneficiary_type_filter") or "").strip()
        if "item_type_filter" in request.POST:
            state["item_type_filter"] = str(request.POST.get("item_type_filter") or "").strip()
        else:
            state["item_type_filter"] = str(state.get("item_type_filter") or models.ItemTypeChoices.AID).strip()

        request.session[REPORTS_WAITING_HALL_STATE_KEY] = state
        request.session.modified = True

        if action == "save_waiting_hall":
            messages.success(request, "Waiting Hall Acknowledgment selections saved.")
            return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")

        if action in {"download_waiting_hall", "preview_waiting_hall"}:
            if not state.get("loaded"):
                messages.error(request, "Sync Data or Upload a file first before downloading this report.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")
            grouped = _reports_waiting_hall_grouped_data(
                state.get("rows") or [],
                ignored_keys=state.get("ignored_keys"),
                beneficiary_type_filter=state.get("beneficiary_type_filter"),
                item_type_filter=state.get("item_type_filter"),
            )
            report_groups = [group for group in grouped["districts"] if group["items"]]
            if not report_groups:
                messages.error(request, "No printable waiting hall rows are available for the current filter. Check whether all rows under this filter were ignored.")
                return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")
            event_year = timezone.localdate().year
            event_date = date(event_year, 3, 3)
            age_label = services._ordinal(max(event_date.year - 1940, 1))
            logo_bytes = base64.b64decode(shared_logo["logo_base64"]) if shared_logo.get("logo_base64") else None
            output_format = "pdf" if action == "preview_waiting_hall" else (request.POST.get("download_format") or "pdf").strip().lower()
            if output_format == "pdf":
                buffer = services.generate_waiting_hall_acknowledgment_pdf(
                    report_groups,
                    event_age_label=age_label,
                    event_date=event_date,
                    custom_logo=logo_bytes,
                )
                filename = f"district_waiting_hall_acknowledgment_{timezone.localtime().strftime('%d_%b_%y_%H_%M')}.pdf"
                response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
                disposition = "inline" if action == "preview_waiting_hall" else "attachment"
                response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
                return response
            document_bytes = services.generate_waiting_hall_acknowledgment_doc(
                report_groups,
                event_age_label=age_label,
                event_date=event_date,
                custom_logo=logo_bytes,
                custom_logo_mime_type=shared_logo.get("logo_content_type") or "image/png",
            )
            filename = f"district_waiting_hall_acknowledgment_{timezone.localtime().strftime('%d_%b_%y_%H_%M')}.docx"
            response = HttpResponse(
                document_bytes,
                content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        return HttpResponseRedirect(f"{reverse('ui:reports')}?tab=waiting-hall-acknowledgment")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        active_tab = (self.request.GET.get("tab") or self.REPORT_TABS[0]["key"]).strip()
        tab_keys = {tab["key"] for tab in self.REPORT_TABS}
        if active_tab not in tab_keys:
            active_tab = self.REPORT_TABS[0]["key"]
        active_tab_meta = next(tab for tab in self.REPORT_TABS if tab["key"] == active_tab)
        token_lookup_state = _reports_token_lookup_session_state(self.request)
        token_lookup_rows = list(token_lookup_state.get("rows") or [])
        token_lookup_filtered_rows = _reports_token_lookup_filter_rows(token_lookup_rows, token_lookup_state.get("filters"))
        token_lookup_item_options = _reports_token_lookup_choice_values(token_lookup_rows, "item_name")
        token_lookup_item_type_options = _reports_token_lookup_choice_values(token_lookup_rows, "item_type")
        public_signature_state = _reports_public_signature_session_state(self.request)
        public_signature_rows = list(public_signature_state.get("rows") or [])
        public_signature_item_options = _reports_public_signature_item_options(public_signature_rows)
        district_signature_state = _reports_district_signature_session_state(self.request)
        district_signature_rows = list(district_signature_state.get("rows") or [])
        district_signature_grouped = _reports_district_signature_grouped(district_signature_rows) if district_signature_state.get("loaded") else {
            "districts": [],
            "district_count": 0,
            "item_count": 0,
            "total_quantity": 0,
            "total_token_quantity": 0,
        }
        waiting_hall_state = _reports_waiting_hall_session_state(self.request)
        waiting_hall_grouped = _reports_waiting_hall_grouped_data(
            waiting_hall_state.get("rows") or [],
            ignored_keys=waiting_hall_state.get("ignored_keys"),
        ) if waiting_hall_state.get("loaded") else {
            "districts": [],
            "district_count": 0,
            "item_count": 0,
            "total_quantity": 0,
            "available_keys": [],
        }
        public_ack_state = _reports_public_ack_session_state(self.request)
        public_ack_template_fields = list(public_ack_state.get("template_fields") or [])
        public_ack_rows = list(public_ack_state.get("rows") or [])
        segregation_state = _reports_simple_report_session_state(self.request, REPORTS_SEGREGATION_STATE_KEY)
        segregation_rows = list(segregation_state.get("rows") or [])
        segregation_beneficiary_type = str(self.request.GET.get("seg_beneficiary_type") or "").strip()
        allowed_beneficiary_types = {choice[0] for choice in SEGREGATION_BENEFICIARY_FILTER_CHOICES}
        if segregation_beneficiary_type not in allowed_beneficiary_types:
            segregation_beneficiary_type = ""
        segregation_item_type = _segregation_resolved_item_type(
            self.request.GET.get("seg_item_type"),
            default=models.ItemTypeChoices.ARTICLE,
        )
        segregation_dataset = _segregation_normalize_dataset(segregation_state)
        segregation_filtered_rows = _segregation_filter_rows(
            segregation_dataset.get("rows") or [],
            beneficiary_type=segregation_beneficiary_type,
            item_type=segregation_item_type,
        )
        segregation_file1 = _segregation_build_file1(segregation_filtered_rows)
        segregation_file2 = _segregation_build_file2(segregation_filtered_rows)
        segregation_file3 = _segregation_build_file3(segregation_filtered_rows)
        distribution_state = _reports_simple_report_session_state(self.request, REPORTS_DISTRIBUTION_STATE_KEY)
        distribution_rows = list(distribution_state.get("rows") or [])
        context.update(
            {
                "page_title": "Reports",
                "report_tabs": self.REPORT_TABS,
                "active_report_tab": active_tab,
                "active_report_tab_meta": active_tab_meta,
                "token_lookup_state": token_lookup_state,
                "token_lookup_row_count": len(token_lookup_rows),
                "token_lookup_filtered_rows": token_lookup_filtered_rows,
                "token_lookup_match_count": len(token_lookup_filtered_rows),
                "token_lookup_total_token_quantity": sum(int(row.get("token_quantity") or 0) for row in token_lookup_filtered_rows),
                "token_lookup_item_options": token_lookup_item_options,
                "token_lookup_item_type_options": token_lookup_item_type_options,
                "public_signature_state": public_signature_state,
                "public_signature_row_count": len(public_signature_rows),
                "public_signature_item_options": public_signature_item_options,
                "district_signature_state": district_signature_state,
                "district_signature_row_count": len(district_signature_rows),
                "district_signature_grouped": district_signature_grouped,
                "waiting_hall_state": waiting_hall_state,
                "waiting_hall_grouped": waiting_hall_grouped,
                "reports_shared_logo": _reports_shared_logo_state(self.request),
                "public_ack_state": public_ack_state,
                "public_ack_rows": public_ack_rows,
                "public_ack_row_count": len(public_ack_rows),
                "public_ack_template_fields": public_ack_template_fields,
                "segregation_state": segregation_state,
                "segregation_row_count": len(segregation_rows),
                "segregation_filter_values": {
                    "beneficiary_type": segregation_beneficiary_type,
                    "item_type": segregation_item_type,
                },
                "segregation_beneficiary_type_choices": SEGREGATION_BENEFICIARY_FILTER_CHOICES,
                "segregation_item_type_choices": SEGREGATION_ITEM_FILTER_CHOICES,
                "segregation_filtered_row_count": len(segregation_filtered_rows),
                "segregation_file1": segregation_file1,
                "segregation_file2": segregation_file2,
                "segregation_file3": segregation_file3,
                "distribution_state": distribution_state,
                "distribution_row_count": len(distribution_rows),
                "public_ack_field_rows": [
                    {
                        "field_name": field.get("field_name") or "",
                        "field_key": field.get("field_key") or "",
                        "selected_value": public_ack_state.get("field_map", {}).get(field.get("field_key") or "", ""),
                    }
                    for field in public_ack_template_fields
                ],
                "public_ack_column_options": _reports_public_ack_column_options(public_ack_state.get("headers") or []),
                "waiting_hall_beneficiary_type_choices": [
                    ("", "All"),
                    (models.RecipientTypeChoices.DISTRICT, "District"),
                    (models.RecipientTypeChoices.PUBLIC, "Public"),
                    (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
                ],
                "waiting_hall_item_type_choices": [
                    (models.ItemTypeChoices.AID, "Aid"),
                    (models.ItemTypeChoices.ARTICLE, "Article"),
                    ("", "All"),
                ],
            }
        )
        return context


class ArticleListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.ARTICLE_MANAGEMENT
    permission_action = "view"
    model = models.Article
    template_name = "dashboard/article_list.html"
    context_object_name = "articles"

    def get(self, request, *args, **kwargs):
        if (request.GET.get("export") or "").strip().lower() == "csv":
            return self._export_csv()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        queryset = models.Article.objects.order_by("article_name")
        if q := self.request.GET.get("q"):
            queryset = queryset.filter(
                Q(article_name__icontains=q)
                | Q(article_name_tk__icontains=q)
                | Q(category__icontains=q)
                | Q(master_category__icontains=q)
                | Q(item_type__icontains=q)
            )
        if item_type := (self.request.GET.get("item_type") or "").strip():
            queryset = queryset.filter(item_type=item_type)
        if combo_filter := (self.request.GET.get("combo") or "").strip():
            if combo_filter == "combo":
                queryset = queryset.filter(combo=True)
            elif combo_filter == "separate":
                queryset = queryset.filter(combo=False)
        if category := (self.request.GET.get("category") or "").strip():
            queryset = queryset.filter(category=category)
        if master_category := (self.request.GET.get("master_category") or "").strip():
            queryset = queryset.filter(master_category=master_category)
        sort = (self.request.GET.get("sort") or "article_name").strip()
        direction = (self.request.GET.get("dir") or "asc").strip().lower()
        allowed_sorts = {
            "article_name": "article_name",
            "item_type": "item_type",
            "category": "category",
            "master_category": "master_category",
            "cost_per_unit": "cost_per_unit",
            "combo": "combo",
            "created_at": "created_at",
            "article_name_tk": "article_name_tk",
        }
        sort_field = allowed_sorts.get(sort, "article_name")
        if direction == "desc":
            sort_field = f"-{sort_field}"
        queryset = queryset.order_by(sort_field, "article_name")
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category_choices = (
            models.Article.objects.exclude(category__isnull=True)
            .exclude(category__exact="")
            .order_by("category")
            .values_list("category", flat=True)
            .distinct()
        )
        master_category_choices = (
            models.Article.objects.exclude(master_category__isnull=True)
            .exclude(master_category__exact="")
            .order_by("master_category")
            .values_list("master_category", flat=True)
            .distinct()
        )
        context.update(
            {
                "item_type_choices": models.ItemTypeChoices.choices,
                "combo_choices": [("combo", "Combo"), ("separate", "Separate")],
                "category_choices": category_choices,
                "master_category_choices": master_category_choices,
                "filters": {
                    "q": self.request.GET.get("q", ""),
                    "item_type": self.request.GET.get("item_type", ""),
                    "combo": self.request.GET.get("combo", ""),
                    "category": self.request.GET.get("category", ""),
                    "master_category": self.request.GET.get("master_category", ""),
                },
                "current_sort": (self.request.GET.get("sort") or "article_name").strip(),
                "current_dir": (self.request.GET.get("dir") or "asc").strip().lower(),
                "query_string_without_page": self._query_string_without_page(),
            }
        )
        return context

    def _query_string_without_page(self):
        params = self.request.GET.copy()
        return params.urlencode()

    def _export_csv(self):
        timestamp = timezone.localtime().strftime("%Y_%m_%d_%I_%M_%p")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="article-management_{timestamp}.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "Article Name",
                "Token Name",
                "Cost Per Unit",
                "Item Type",
                "Category",
                "Super Category",
                "Combo / Separate",
                "Status",
                "Created At",
                "Updated At",
            ]
        )
        for article in self.get_queryset():
            writer.writerow(
                [
                    article.article_name,
                    article.article_name_tk or "",
                    article.cost_per_unit,
                    article.get_item_type_display(),
                    article.category or "",
                    article.master_category or "",
                    "Combo" if article.combo else "Separate",
                    "Active" if article.is_active else "Inactive",
                    timezone.localtime(article.created_at).strftime("%d/%m/%Y %H:%M"),
                    timezone.localtime(article.updated_at).strftime("%d/%m/%Y %H:%M"),
                ]
            )
        return response


class ArticleCreateView(LoginRequiredMixin, WriteRoleMixin, CreateView):
    module_key = models.ModuleKeyChoices.ARTICLE_MANAGEMENT
    permission_action = "create_edit"
    model = models.Article
    form_class = ArticleForm
    template_name = "dashboard/article_form.html"
    success_url = reverse_lazy("ui:article-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["popup_mode"] = self.request.GET.get("popup") == "1"
        context["category_suggestions"] = (
            models.Article.objects.exclude(category__isnull=True)
            .exclude(category__exact="")
            .order_by("category")
            .values_list("category", flat=True)
            .distinct()
        )
        context["master_category_suggestions"] = (
            models.Article.objects.exclude(master_category__isnull=True)
            .exclude(master_category__exact="")
            .order_by("master_category")
            .values_list("master_category", flat=True)
            .distinct()
        )
        return context

    def form_valid(self, form):
        self.object = form.save()
        if self.request.GET.get("popup") == "1":
            payload = json.dumps(
                {
                    "id": self.object.id,
                    "article_name": self.object.article_name,
                    "cost_per_unit": str(self.object.cost_per_unit),
                    "item_type": self.object.item_type,
                }
            ).replace("</", "<\\/")
            return HttpResponse(
                f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>Article created</title></head>
<body style="font-family: Poppins, Segoe UI, Arial, sans-serif; padding: 24px;">
  <div>Article created. Returning to the application form…</div>
  <script>
    (function() {{
      var article = {payload};
      if (window.opener && !window.opener.closed && typeof window.opener.handleArticleCreated === "function") {{
        window.opener.handleArticleCreated(article);
        try {{ window.opener.focus(); }} catch (error) {{}}
        window.close();
        return;
      }}
      document.body.innerHTML = "<p>Article created. You can close this window now.</p>";
    }})();
  </script>
</body>
</html>"""
            )
        messages.success(self.request, "Article created.")
        return HttpResponseRedirect(self.get_success_url())


class ArticleUpdateView(LoginRequiredMixin, WriteRoleMixin, UpdateView):
    module_key = models.ModuleKeyChoices.ARTICLE_MANAGEMENT
    permission_action = "create_edit"
    model = models.Article
    form_class = ArticleForm
    template_name = "dashboard/article_form.html"
    success_url = reverse_lazy("ui:article-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["popup_mode"] = False
        context["category_suggestions"] = (
            models.Article.objects.exclude(category__isnull=True)
            .exclude(category__exact="")
            .order_by("category")
            .values_list("category", flat=True)
            .distinct()
        )
        context["master_category_suggestions"] = (
            models.Article.objects.exclude(master_category__isnull=True)
            .exclude(master_category__exact="")
            .order_by("master_category")
            .values_list("master_category", flat=True)
            .distinct()
        )
        return context

    def form_valid(self, form):
        messages.success(self.request, "Article updated.")
        return super().form_valid(form)


class AdminRequiredMixin(RoleRequiredMixin):
    allowed_roles = {"admin"}


class UserManagementListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.USER_MANAGEMENT
    permission_action = "view"
    model = models.AppUser
    template_name = "dashboard/user_management.html"
    context_object_name = "users"

    def get_queryset(self):
        queryset = models.AppUser.objects.select_related("created_by").order_by("first_name", "email")
        if q := (self.request.GET.get("q") or "").strip():
            queryset = queryset.filter(
                Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
                | Q(email__icontains=q)
                | Q(role__icontains=q)
                | Q(status__icontains=q)
            )
        if role := (self.request.GET.get("role") or "").strip():
            queryset = queryset.filter(role=role)
        if status := (self.request.GET.get("status") or "").strip():
            queryset = queryset.filter(status=status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filters"] = {
            "q": (self.request.GET.get("q") or "").strip(),
            "role": (self.request.GET.get("role") or "").strip(),
            "status": (self.request.GET.get("status") or "").strip(),
        }
        context["role_choices"] = models.RoleChoices.choices
        context["status_choices"] = models.StatusChoices.choices
        return context


class UserManagementCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    module_key = models.ModuleKeyChoices.USER_MANAGEMENT
    permission_action = "create_edit"
    model = models.AppUser
    form_class = AppUserCreateForm
    template_name = "dashboard/user_form.html"

    def get_success_url(self):
        return reverse("ui:user-list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "User created successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create User"
        context["submit_label"] = "Create"
        return context


class UserManagementUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    module_key = models.ModuleKeyChoices.USER_MANAGEMENT
    permission_action = "create_edit"
    model = models.AppUser
    form_class = AppUserUpdateForm
    template_name = "dashboard/user_form.html"

    def get_success_url(self):
        return reverse("ui:user-list")

    def form_valid(self, form):
        if self.object == self.request.user:
            if form.cleaned_data.get("status") != models.StatusChoices.ACTIVE:
                form.add_error("status", "You cannot deactivate your own account.")
                return self.form_invalid(form)
            if form.cleaned_data.get("role") != models.RoleChoices.ADMIN:
                form.add_error("role", "You cannot remove your own admin access.")
                return self.form_invalid(form)
        messages.success(self.request, "User updated successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Edit User"
        context["submit_label"] = "Save"
        return context


class UserManagementPasswordResetView(LoginRequiredMixin, AdminRequiredMixin, FormView):
    module_key = models.ModuleKeyChoices.USER_MANAGEMENT
    permission_action = "reset_password"
    form_class = AppUserPasswordResetForm
    template_name = "dashboard/user_password_reset.html"

    def dispatch(self, request, *args, **kwargs):
        self.target_user = get_object_or_404(models.AppUser, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse("ui:user-list")

    def form_valid(self, form):
        self.target_user.set_password(form.cleaned_data["password1"])
        self.target_user.save(update_fields=["password", "updated_at"])
        messages.success(self.request, f"Password reset for {self.target_user.display_name}.")
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Reset Password"
        context["target_user"] = self.target_user
        context["submit_label"] = "Reset Password"
        return context


class UserManagementDeleteView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.USER_MANAGEMENT
    permission_action = "delete"
    def post(self, request, *args, **kwargs):
        user = get_object_or_404(models.AppUser, pk=kwargs["pk"])
        if user == request.user:
            messages.error(request, "You cannot delete your own account.")
            return HttpResponseRedirect(reverse("ui:user-list"))
        user.delete()
        messages.success(request, "User deleted successfully.")
        return HttpResponseRedirect(reverse("ui:user-list"))


class ArticleDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    module_key = models.ModuleKeyChoices.ARTICLE_MANAGEMENT
    permission_action = "delete"
    model = models.Article
    template_name = "dashboard/article_confirm_delete.html"
    success_url = reverse_lazy("ui:article-list")

    def post(self, request, *args, **kwargs):
        messages.warning(self.request, "Article deleted.")
        return super().post(request, *args, **kwargs)


class MasterEntryView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/module_master_entry.html"

    def get(self, request, *args, **kwargs):
        export_scope = (request.GET.get("export_scope") or "").strip().lower()
        if export_scope:
            return _export_master_entry_csv(request, export_scope=export_scope)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        beneficiary_type = self.request.GET.get("type", "district")
        search_query = (self.request.GET.get("q") or "").strip()
        date_from = (self.request.GET.get("date_from") or "").strip()
        date_to = (self.request.GET.get("date_to") or "").strip()
        status_filter = (self.request.GET.get("status") or "").strip().lower()
        sort_by = (self.request.GET.get("sort") or "").strip()
        sort_dir = "asc" if (self.request.GET.get("dir") or "desc").lower() == "asc" else "desc"

        # IMPORTANT: This page can get very large (lots of rows + attachment metadata + hidden detail tables).
        # Only build the dataset that is currently selected, and lazy-load expanded row details.
        district_groups = []
        public_entries = []
        institution_groups = []

        context["beneficiary_type"] = beneficiary_type
        context["search_query"] = search_query
        context["date_from"] = date_from
        context["date_to"] = date_to
        context["status_filter"] = status_filter
        context["status_choices"] = [
            ("", "All Statuses"),
            (models.BeneficiaryStatusChoices.DRAFT, "Draft"),
            (models.BeneficiaryStatusChoices.SUBMITTED, "Submitted"),
        ]
        context["sort_by"] = sort_by
        context["sort_dir"] = sort_dir

        district_count = models.DistrictBeneficiaryEntry.objects.values("district_id").distinct().count()
        public_count = models.PublicBeneficiaryEntry.objects.count()
        institution_count = models.InstitutionsBeneficiaryEntry.objects.values("application_number").distinct().count()
        district_row_count = models.DistrictBeneficiaryEntry.objects.count()
        public_row_count = models.PublicBeneficiaryEntry.objects.count()
        institution_row_count = models.InstitutionsBeneficiaryEntry.objects.count()

        context["district_count"] = district_count
        context["public_count"] = public_count
        context["institution_count"] = institution_count
        context["total_material_rows"] = district_row_count + public_row_count + institution_row_count

        if beneficiary_type == "district":
            district_groups = _filter_sort_district_summaries(
                _build_district_entry_summaries(),
                search_query=search_query,
                date_from=date_from,
                date_to=date_to,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_dir=sort_dir,
            )
        elif beneficiary_type == "public":
            public_entries = _filter_sort_public_entries(
                models.PublicBeneficiaryEntry.objects.select_related("article").all(),
                search_query=search_query,
                date_from=date_from,
                date_to=date_to,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_dir=sort_dir,
            )
            public_entries = list(public_entries)
            public_attachment_latest, public_attachment_counts = _public_attachment_latest_and_counts(
                [entry.id for entry in public_entries]
            )
            for entry in public_entries:
                attachment = public_attachment_latest.get(entry.id)
                entry.attachment_id = attachment.id if attachment else None
                entry.attachment_preview_url = reverse("ui:application-attachment-download", kwargs={"attachment_id": attachment.id}) if attachment else ""
                entry.attachment_source = _attachment_preview_source(attachment)
                entry.attachment_title = _attachment_preview_title(attachment)
                entry.attachment_count = public_attachment_counts.get(entry.id, 0)
        elif beneficiary_type == "institutions":
            institution_groups = _filter_sort_institution_summaries(
                _build_institution_entry_summaries(),
                search_query=search_query,
                date_from=date_from,
                date_to=date_to,
                status_filter=status_filter,
                sort_by=sort_by,
                sort_dir=sort_dir,
            )

        context["district_groups"] = district_groups
        context["public_entries"] = public_entries
        context["institution_groups"] = institution_groups
        context["district_total_accrued"] = sum((row.get("total_accrued") or 0) for row in district_groups)
        context["public_total_accrued"] = sum((entry.total_amount or 0) for entry in public_entries)
        context["institution_total_accrued"] = sum((row.get("total_value") or 0) for row in institution_groups)
        context["public_submit_popup"] = self.request.session.pop("public_submit_popup", None)
        context["institution_submit_popup"] = self.request.session.pop("institution_submit_popup", None)
        return context


class DistrictMasterEntryInlineSummaryView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/partials/master_entry_district_summary.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        district = models.DistrictMaster.objects.get(pk=self.kwargs["district_id"], is_active=True)
        entries = list(
            models.DistrictBeneficiaryEntry.objects.filter(district=district).select_related("article").order_by("id")
        )
        context["district"] = district
        context["entries"] = entries
        context["internal_notes"] = entries[0].internal_notes if entries else ""
        return context


class PublicMasterEntryInlineSummaryView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/partials/master_entry_public_summary.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        entry = models.PublicBeneficiaryEntry.objects.select_related("article").get(pk=self.kwargs["pk"])
        history_summary = _public_history_summary(_public_history_matches(entry.aadhar_number))
        context["entry"] = entry
        context["history_summary"] = history_summary
        return context


class InstitutionsMasterEntryInlineSummaryView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/partials/master_entry_institution_summary.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application_number = self.kwargs["application_number"]
        entries = list(
            models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).select_related("article").order_by("id")
        )
        context["application_number"] = application_number
        context["entries"] = entries
        context["entry_header"] = entries[0] if entries else None
        context["internal_notes"] = entries[0].internal_notes if entries else ""
        return context


def _public_attachment_latest_and_counts(entry_ids):
    if not entry_ids:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
            public_entry_id__in=entry_ids,
        )
        .order_by("public_entry_id", "-created_at", "-id")
    )
    latest_map = {}
    count_map = {}
    for attachment in attachments:
        entry_id = attachment.public_entry_id
        count_map[entry_id] = count_map.get(entry_id, 0) + 1
        if entry_id not in latest_map:
            latest_map[entry_id] = attachment
    return latest_map, count_map


def _district_attachment_latest_and_counts(district_ids):
    if not district_ids:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
            district_id__in=district_ids,
        )
        .order_by("district_id", "-created_at", "-id")
    )
    latest_map = {}
    count_map = {}
    for attachment in attachments:
        district_id = attachment.district_id
        count_map[district_id] = count_map.get(district_id, 0) + 1
        if district_id not in latest_map:
            latest_map[district_id] = attachment
    return latest_map, count_map


def _institution_attachment_latest_and_counts(application_numbers):
    if not application_numbers:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
            institution_application_number__in=application_numbers,
        )
        .order_by("institution_application_number", "-created_at", "-id")
    )
    latest_map = {}
    count_map = {}
    for attachment in attachments:
        key = attachment.institution_application_number
        if not key:
            continue
        count_map[key] = count_map.get(key, 0) + 1
        if key not in latest_map:
            latest_map[key] = attachment
    return latest_map, count_map


class ApplicationAuditLogListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.AUDIT_LOGS
    permission_action = "view"
    model = models.AuditLog
    template_name = "dashboard/application_audit_logs.html"
    context_object_name = "audit_logs"
    paginate_by = 50

    def get_queryset(self):
        queryset = models.AuditLog.objects.select_related("user").filter(
            entity_type__in=[
                "district_application",
                "public_application",
                "institution_application",
            ]
        ).order_by("-created_at")
        q = (self.request.GET.get("q") or "").strip()
        application_type = (self.request.GET.get("application_type") or "").strip()
        user_id = (self.request.GET.get("user_id") or "").strip()
        date_from = (self.request.GET.get("date_from") or "").strip()
        date_to = (self.request.GET.get("date_to") or "").strip()
        if q:
            queryset = queryset.filter(
                Q(entity_type__icontains=q)
                | Q(entity_id__icontains=q)
                | Q(action_type__icontains=q)
                | Q(user__email__icontains=q)
                | Q(user__first_name__icontains=q)
                | Q(user__last_name__icontains=q)
            )
        if application_type:
            queryset = queryset.filter(entity_type=application_type)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_query"] = (self.request.GET.get("q") or "").strip()
        context["application_type"] = (self.request.GET.get("application_type") or "").strip()
        context["selected_user_id"] = (self.request.GET.get("user_id") or "").strip()
        context["date_from"] = (self.request.GET.get("date_from") or "").strip()
        context["date_to"] = (self.request.GET.get("date_to") or "").strip()
        context["application_type_choices"] = [
            ("district_application", "District Application"),
            ("public_application", "Public Application"),
            ("institution_application", "Institution Application"),
        ]
        context["audit_users"] = (
            models.AppUser.objects.filter(
                audit_logs__entity_type__in=[
                    "district_application",
                    "public_application",
                    "institution_application",
                ]
            )
            .distinct()
            .order_by("first_name", "email")
        )
        context["audit_rows"] = [
            {
                "log": log,
                "application_label": _application_audit_label(log),
                "change_lines": _application_audit_change_lines(log),
                "user_label": _audit_user_label(log.user),
            }
            for log in context["audit_logs"]
        ]
        return context


def _reconciliation_parse_decimal(value):
    raw = str(value or "").replace(",", "").strip()
    try:
        return Decimal(raw or "0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


EXPORT_COLUMNS = [
    "Application Number",
    "Beneficiary Name",
    "Requested Item",
    "Quantity",
    "Cost Per Unit",
    "Total Value",
    "Address",
    "Mobile",
    "Aadhar Number",
    "Name of Beneficiary",
    "Name of Institution",
    "Cheque / RTGS in Favour",
    "Handicapped Status",
    "Gender",
    "Gender Category",
    "Beneficiary Type",
    "Item Type",
    "Article Category",
    "Super Category Article",
    "Token Name",
    "Internal Notes",
    "Comments",
]

TOKEN_GENERATION_NUMERIC_HEADERS = {
    "Quantity",
    "Cost Per Unit",
    "Total Value",
    "Mobile",
    "Aadhar Number",
    "Waiting Hall Quantity",
    "Token Quantity",
    "Sequence No",
}

TOKEN_GENERATION_BENEFICIARY_ORDER = {
    "Institutions": 0,
    "Public": 1,
    "District": 2,
}

TOKEN_GENERATION_RENAME_MAP = {
    "I001 - Government Leprosy Centre,Chengalpattu.": "I001-Govt Leprosy Centre,CGL",
    "I002 - Athivakkam,Panchayat Union Primary School.": "I002-Athivakkam,Panchayat School",
    "I003 - Thirukazhukundram,Govt Girls Higher Secondary School.": "I003-Thirukazhukundram,Govt Girls School",
    "I004 - Acharapakkam,Govt Girls Higher Secondary School.": "I004-Acharapakkam,Govt Girls School",
    "I005 - Maduranthagam,District Educational Office.": "I005-Maduranthagam,District Edu Off",
    "I006 - Thozhupedu,Govt Higher Secondary School.": "I006-Thozhupedu,Govt School",
    "I007 - Kayappakkam,Government Higher Secondary School.": "I007-Kayappakkam,Government School",
    "I008 - Nolambur Government Higher SecondarySchool": "I008-Nolambur Government School",
    "I009 - Acharapakkam, Govt Boys Higher Secondary School.": "I009-Acharapakkam,Govt Boys School",
    "I010 - Cheyyur Govt Girls Higher Secondary School.": "I010-Cheyyur Govt Girls School",
    "I011 - Chunambedu Govt Higher Secondary School.": "I011-Chunambedu Govt School",
    "I012 - Avanippur Government Higher Secondary School": "I012-Avanippur Government School",
    "I013 - Polambakkam Govt Higher Secondary School.": "I013-Polambakkam Govt School",
}

TOKEN_GENERATION_ARTICLE_PRINT_EXCLUDES = {
    "Plant Sapling",
    "Provision items",
    "Aluminium holed rice strainer with Handle",
    "Goat(1 Pair)",
    "Ortho Caliper",
    "Fishing Net",
    "Grocery Items",
    "Artificial leg",
    "Cow & Calf",
}


def _request_audit_meta(request):
    return {
        "ip_address": request.META.get("REMOTE_ADDR"),
        "user_agent": request.META.get("HTTP_USER_AGENT", ""),
    }


def _attachment_upload_context(*, attachments=None, enabled=False, upload_url="", helper_text="Attachments can be added after the application is first saved."):
    return {
        "application_attachments": attachments or [],
        "attachment_upload_form": ApplicationAttachmentUploadForm(),
        "attachments_enabled": enabled,
        "attachment_upload_url": upload_url,
        "attachment_helper_text": helper_text,
        "attachment_constraints_text": (
            "Allowed files: PDF, JPG, JPEG, PNG, WEBP, DOC, DOCX, XLS, XLSX, CSV. "
            "Maximum file size: 10 MB. Maximum 2 files per application."
        ),
    }


def _prefixed_attachment_name(application_reference, uploaded_name, custom_name=""):
    prefix = (application_reference or "").strip()
    original_name = (uploaded_name or "").strip()
    chosen_name = (custom_name or "").strip() or original_name or "attachment"
    chosen_root, chosen_ext = os.path.splitext(chosen_name)
    _, original_ext = os.path.splitext(original_name)
    final_name = chosen_name if chosen_ext else f"{chosen_name}{original_ext}"
    if prefix:
        normalized_prefix = f"{prefix}_"
        if final_name.startswith(normalized_prefix):
            return final_name
        return f"{prefix}_{final_name}"
    return final_name


def _attachment_name_exists(queryset, final_name):
    normalized = (final_name or "").strip().lower()
    if not normalized:
        return False
    for existing_name in queryset.values_list("file_name", flat=True):
        if (existing_name or "").strip().lower() == normalized:
            return True
    return False


def _attachment_application_reference(attachment):
    if not attachment:
        return ""
    if attachment.application_type == models.ApplicationAttachmentTypeChoices.DISTRICT and attachment.district_id:
        return attachment.district.application_number or ""
    if attachment.application_type == models.ApplicationAttachmentTypeChoices.PUBLIC and attachment.public_entry_id:
        return attachment.public_entry.application_number or f"PUBLIC-{attachment.public_entry_id}"
    if attachment.application_type == models.ApplicationAttachmentTypeChoices.INSTITUTION:
        return attachment.institution_application_number or ""
    return ""


def _save_application_attachment(*, uploaded, display_name, application_type, uploaded_by, district=None, public_entry=None, institution_application_number=None):
    attachment_kwargs = {
        "application_type": application_type,
        "district": district,
        "public_entry": public_entry,
        "institution_application_number": institution_application_number,
        "file_name": display_name,
        "uploaded_by": uploaded_by,
    }
    application_reference = (
        (district.application_number if district else "")
        or (public_entry.application_number if public_entry else "")
        or institution_application_number
        or ""
    )
    if google_drive.is_configured():
        drive_file = google_drive.upload_application_attachment(
            content=uploaded.read(),
            display_name=display_name,
            content_type=str(getattr(uploaded, "content_type", "") or mimetypes.guess_type(display_name)[0] or "application/octet-stream"),
            application_type=application_type,
            application_reference=application_reference,
        )
        attachment_kwargs.update(
            {
                "drive_file_id": str(drive_file.get("file_id") or "").strip(),
                "drive_mime_type": str(drive_file.get("mime_type") or "").strip(),
                "drive_view_url": str(drive_file.get("view_url") or "").strip(),
            }
        )
    else:
        attachment_kwargs["file"] = uploaded
    return models.ApplicationAttachment.objects.create(**attachment_kwargs)


def _delete_application_attachment_file(attachment):
    try:
        if attachment.drive_file_id:
            google_drive.delete_file(attachment.drive_file_id)
        elif attachment.file:
            attachment.file.delete(save=False)
    except Exception:
        logger.exception("Failed to delete attachment file (id=%s). Continuing.", getattr(attachment, "id", None))


def _sync_drive_attachments_for_application(
    *,
    application_type,
    application_reference,
    district=None,
    public_entry=None,
    institution_application_number=None,
):
    if not google_drive.is_configured():
        return []
    reference = (application_reference or "").strip()
    if not reference:
        return []
    filters = {"application_type": application_type}
    if district is not None:
        filters["district"] = district
    if public_entry is not None:
        filters["public_entry"] = public_entry
    if institution_application_number is not None:
        filters["institution_application_number"] = institution_application_number
    attachments = list(
        models.ApplicationAttachment.objects.filter(**filters).select_related("uploaded_by").order_by("-created_at", "-id")
    )
    try:
        drive_files = google_drive.list_application_attachments(
            application_type=application_type,
            application_reference=reference,
        )
    except Exception:
        # Google Drive sync should never break data entry pages.
        # Return current DB attachments without syncing.
        logger.exception("Failed to sync Drive attachments for %s:%s", application_type, reference)
        return attachments
    if not drive_files:
        stale_drive_only_ids = [
            attachment.id
            for attachment in attachments
            if attachment.drive_file_id and not attachment.file
        ]
        if stale_drive_only_ids:
            models.ApplicationAttachment.objects.filter(id__in=stale_drive_only_ids).delete()
        return list(
            models.ApplicationAttachment.objects.filter(**filters).select_related("uploaded_by").order_by("-created_at", "-id")
        )

    def _normalized_name(value):
        return str(value or "").strip().casefold()

    existing_by_id = {
        str(attachment.drive_file_id or "").strip(): attachment
        for attachment in attachments
        if str(attachment.drive_file_id or "").strip()
    }
    available_by_name = {}
    for drive_file in drive_files:
        normalized_name = _normalized_name(drive_file.get("file_name"))
        if normalized_name and normalized_name not in available_by_name:
            available_by_name[normalized_name] = drive_file

    touched_ids = set()
    seen_drive_ids = {
        str(item.get("file_id") or "").strip()
        for item in drive_files
        if str(item.get("file_id") or "").strip()
    }
    for attachment in attachments:
        matched_drive_file = None
        drive_id = str(attachment.drive_file_id or "").strip()
        if drive_id and drive_id in existing_by_id and any(str(item.get("file_id") or "").strip() == drive_id for item in drive_files):
            matched_drive_file = next(
                (item for item in drive_files if str(item.get("file_id") or "").strip() == drive_id),
                None,
            )
        elif not attachment.file:
            matched_drive_file = available_by_name.get(_normalized_name(attachment.file_name))
        if not matched_drive_file:
            continue
        update_fields = []
        matched_id = str(matched_drive_file.get("file_id") or "").strip()
        if attachment.drive_file_id != matched_id:
            attachment.drive_file_id = matched_id
            update_fields.append("drive_file_id")
        matched_mime = str(matched_drive_file.get("mime_type") or "").strip()
        if attachment.drive_mime_type != matched_mime:
            attachment.drive_mime_type = matched_mime
            update_fields.append("drive_mime_type")
        matched_view = str(matched_drive_file.get("view_url") or "").strip()
        if attachment.drive_view_url != matched_view:
            attachment.drive_view_url = matched_view
            update_fields.append("drive_view_url")
        matched_name = str(matched_drive_file.get("file_name") or "").strip()
        if matched_name and attachment.file_name != matched_name:
            attachment.file_name = matched_name
            update_fields.append("file_name")
        if update_fields:
            attachment.save(update_fields=update_fields)
        touched_ids.add(matched_id)

    stale_attachment_ids = []
    for attachment in attachments:
        drive_id = str(attachment.drive_file_id or "").strip()
        if not drive_id:
            continue
        if attachment.file:
            continue
        if drive_id in seen_drive_ids:
            continue
        stale_attachment_ids.append(attachment.id)

    for drive_file in drive_files:
        matched_id = str(drive_file.get("file_id") or "").strip()
        if not matched_id or matched_id in touched_ids:
            continue
        existing_attachment = existing_by_id.get(matched_id)
        if existing_attachment:
            continue
        attachments.append(
            models.ApplicationAttachment.objects.create(
                application_type=application_type,
                district=district,
                public_entry=public_entry,
                institution_application_number=institution_application_number,
                file_name=str(drive_file.get("file_name") or "").strip(),
                drive_file_id=matched_id,
                drive_mime_type=str(drive_file.get("mime_type") or "").strip(),
                drive_view_url=str(drive_file.get("view_url") or "").strip(),
            )
        )

    if stale_attachment_ids:
        models.ApplicationAttachment.objects.filter(id__in=stale_attachment_ids).delete()

    return list(
        models.ApplicationAttachment.objects.filter(**filters).select_related("uploaded_by").order_by("-created_at", "-id")
    )


def _district_attachment_context(district):
    has_saved_application = bool(district and models.DistrictBeneficiaryEntry.objects.filter(district=district).exists())
    attachments = []
    upload_url = ""
    if has_saved_application:
        attachments = _sync_drive_attachments_for_application(
            application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
            application_reference=district.application_number,
            district=district,
        )
        upload_url = reverse("ui:district-attachment-upload", kwargs={"district_id": district.id})
    return _attachment_upload_context(
        attachments=attachments,
        enabled=has_saved_application,
        upload_url=upload_url,
        helper_text="Save the district application first to upload attachments." if not has_saved_application else "Upload files related to this district application. You can rename the file before upload.",
    )


def _public_attachment_context(entry):
    has_saved_application = bool(entry and entry.pk)
    attachments = []
    upload_url = ""
    if has_saved_application:
        attachments = _sync_drive_attachments_for_application(
            application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
            application_reference=entry.application_number or f"PUBLIC-{entry.pk}",
            public_entry=entry,
        )
        upload_url = reverse("ui:public-attachment-upload", kwargs={"pk": entry.pk})
    context = _attachment_upload_context(
        attachments=attachments,
        enabled=has_saved_application,
        upload_url=upload_url,
        helper_text="Save the public application first to upload attachments." if not has_saved_application else "Upload files related to this public application. You can rename the file before upload.",
    )
    context["entry_id"] = entry.pk if has_saved_application else None
    return context


def _institution_attachment_context(application_number):
    has_saved_application = bool(application_number)
    attachments = []
    upload_url = ""
    if has_saved_application:
        attachments = _sync_drive_attachments_for_application(
            application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
            application_reference=application_number,
            institution_application_number=application_number,
        )
        upload_url = reverse("ui:institution-attachment-upload", kwargs={"application_number": application_number})
    return _attachment_upload_context(
        attachments=attachments,
        enabled=has_saved_application,
        upload_url=upload_url,
        helper_text="Save the institution application first to upload attachments." if not has_saved_application else "Upload files related to this institution application. You can rename the file before upload.",
    )


def _district_attachment_preview_data(district_ids):
    if not district_ids:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
            district_id__in=district_ids,
        )
        .order_by("district_id", "-created_at", "-id")
    )
    preview_map = {}
    preview_lists = {}
    for attachment in attachments:
        if attachment.district_id not in preview_map:
            preview_map[attachment.district_id] = attachment
        preview_lists.setdefault(attachment.district_id, []).append(attachment)
    return preview_map, preview_lists


def _district_attachment_preview_map(district_ids):
    preview_map, _preview_lists = _district_attachment_preview_data(district_ids)
    return preview_map


def _district_attachment_preview_lists(district_ids):
    _preview_map, preview_lists = _district_attachment_preview_data(district_ids)
    return preview_lists


def _public_attachment_preview_data(entry_ids):
    if not entry_ids:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
            public_entry_id__in=entry_ids,
        )
        .order_by("public_entry_id", "-created_at", "-id")
    )
    preview_map = {}
    preview_lists = {}
    for attachment in attachments:
        if attachment.public_entry_id not in preview_map:
            preview_map[attachment.public_entry_id] = attachment
        preview_lists.setdefault(attachment.public_entry_id, []).append(attachment)
    return preview_map, preview_lists


def _public_attachment_preview_map(entry_ids):
    preview_map, _preview_lists = _public_attachment_preview_data(entry_ids)
    return preview_map


def _public_attachment_preview_lists(entry_ids):
    _preview_map, preview_lists = _public_attachment_preview_data(entry_ids)
    return preview_lists


def _institution_attachment_preview_data(application_numbers):
    if not application_numbers:
        return {}, {}
    attachments = (
        models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
            institution_application_number__in=application_numbers,
        )
        .order_by("institution_application_number", "-created_at", "-id")
    )
    preview_map = {}
    preview_lists = {}
    for attachment in attachments:
        key = attachment.institution_application_number
        if key and key not in preview_map:
            preview_map[key] = attachment
        if key:
            preview_lists.setdefault(key, []).append(attachment)
    return preview_map, preview_lists


def _institution_attachment_preview_map(application_numbers):
    preview_map, _preview_lists = _institution_attachment_preview_data(application_numbers)
    return preview_map


def _institution_attachment_preview_lists(application_numbers):
    _preview_map, preview_lists = _institution_attachment_preview_data(application_numbers)
    return preview_lists


def _attachment_preview_title(attachment):
    if not attachment:
        return ""
    if attachment.file_name:
        return attachment.file_name
    if attachment.drive_view_url:
        return "Attachment"
    if attachment.file:
        return os.path.basename(attachment.file.name)
    return ""


def _attachment_preview_payload(attachment):
    if not attachment:
        return None
    preview_url = reverse("ui:application-attachment-download", kwargs={"attachment_id": attachment.id})
    source_name = _attachment_preview_source(attachment)
    return {
        "id": attachment.id,
        "title": _attachment_preview_title(attachment),
        "preview_url": preview_url,
        "download_url": f"{preview_url}?download=1",
        "source": source_name,
    }


def _attachment_preview_source(attachment):
    if not attachment:
        return ""
    if attachment.file and attachment.file.name:
        return (attachment.file.name or "").lower()
    source_name = (attachment.file_name or "").lower()
    mime_type = (attachment.drive_mime_type or "").lower()
    if source_name and "." not in os.path.basename(source_name):
        if mime_type == "application/pdf":
            source_name = f"{source_name}.pdf"
        elif mime_type.startswith("image/"):
            ext = mime_type.split("/", 1)[1].strip()
            if ext == "jpeg":
                ext = "jpg"
            if ext:
                source_name = f"{source_name}.{ext}"
    return source_name


def _attachment_items_b64(items):
    payload = json.dumps(items, ensure_ascii=False)
    return base64.b64encode(payload.encode("utf-8")).decode("ascii")


def _district_audit_snapshot(district):
    entries = list(
        models.DistrictBeneficiaryEntry.objects.filter(district=district).select_related("article").order_by("id")
    )
    total_accrued = sum((entry.total_amount or 0) for entry in entries)
    return {
        "district_id": str(district.id),
        "application_number": district.application_number or "",
        "district_name": district.district_name or "",
        "president_name": district.president_name or "",
        "mobile_number": district.mobile_number or "",
        "allotted_budget": str(district.allotted_budget or 0),
        "status": entries[0].status if entries else "",
        "internal_notes": (entries[0].internal_notes or "") if entries else "",
        "total_accrued": str(total_accrued or 0),
        "item_count": len(entries),
        "items": [
            {
                "id": str(entry.id),
                "article_id": str(entry.article_id),
                "article_name": entry.article.article_name,
                "item_type": entry.article.item_type,
                "quantity": entry.quantity,
                "unit_cost": str(entry.article_cost_per_unit or 0),
                "total_amount": str(entry.total_amount or 0),
                "name_of_beneficiary": entry.name_of_beneficiary or "",
                "name_of_institution": entry.name_of_institution or "",
                "aadhar_number": entry.aadhar_number or "",
                "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
                "notes": entry.notes or "",
            }
            for entry in entries
        ],
    }


def _public_audit_snapshot(entry):
    return {
        "id": str(entry.id),
        "application_number": entry.application_number or "",
        "name": entry.name or "",
        "aadhar_number": entry.aadhar_number or "",
        "is_handicapped": entry.get_is_handicapped_display() if entry.is_handicapped else models.HandicappedStatusChoices.NO,
        "gender": entry.gender or "",
        "female_status": entry.female_status or "",
        "address": entry.address or "",
        "mobile": entry.mobile or "",
        "article_id": str(entry.article_id),
        "article_name": entry.article.article_name,
        "item_type": entry.article.item_type,
        "quantity": entry.quantity,
        "unit_cost": str(entry.article_cost_per_unit or 0),
        "total_amount": str(entry.total_amount or 0),
        "name_of_institution": entry.name_of_institution or "",
        "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
        "notes": entry.notes or "",
        "status": entry.status or "",
    }


def _institution_audit_snapshot(application_number):
    entries = list(
        models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).select_related("article").order_by("id")
    )
    if not entries:
        return {"application_number": application_number, "item_count": 0, "items": []}
    first = entries[0]
    total_value = sum((entry.total_amount or 0) for entry in entries)
    return {
        "application_number": application_number,
        "institution_name": first.institution_name or "",
        "institution_type": first.institution_type or "",
        "status": first.status or "",
        "address": first.address or "",
        "mobile": first.mobile or "",
        "internal_notes": first.internal_notes or "",
        "total_value": str(total_value or 0),
        "item_count": len(entries),
        "items": [
            {
                "id": str(entry.id),
                "article_id": str(entry.article_id),
                "article_name": entry.article.article_name,
                "item_type": entry.article.item_type,
                "quantity": entry.quantity,
                "unit_cost": str(entry.article_cost_per_unit or 0),
                "total_amount": str(entry.total_amount or 0),
                "name_of_beneficiary": entry.name_of_beneficiary or "",
                "name_of_institution": entry.name_of_institution or "",
                "aadhar_number": entry.aadhar_number or "",
                "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
                "notes": entry.notes or "",
            }
            for entry in entries
        ],
    }


AUDIT_FIELD_LABELS = {
    "district_name": "District",
    "president_name": "President",
    "mobile_number": "Mobile Number",
    "allotted_budget": "Allotted Budget",
    "name": "Name",
    "aadhar_number": "Aadhaar Number",
    "is_handicapped": "Disability Category",
    "gender": "Gender",
    "female_status": "Gender Category",
    "address": "Address",
    "mobile": "Mobile Number",
    "article_name": "Requested Item",
    "item_type": "Item Type",
    "quantity": "Quantity",
    "unit_cost": "Unit Price",
    "total_amount": "Total Value",
    "cheque_rtgs_in_favour": "Cheque / RTGS in Favour",
    "notes": "Comments",
    "internal_notes": "Internal Notes",
    "institution_name": "Institution Name",
    "institution_type": "Institution Type",
    "total_value": "Total Value",
}


def _application_audit_label(log):
    details = log.details or {}
    after = details.get("after") or {}
    before = details.get("before") or {}
    snapshot = after or before
    if log.entity_type == "district_application":
        return snapshot.get("application_number") or snapshot.get("district_name") or log.entity_id or "-"
    if log.entity_type == "public_application":
        application_number = snapshot.get("application_number") or log.entity_id or "-"
        name = snapshot.get("name") or ""
        return f"{application_number} - {name}".strip(" -")
    if log.entity_type == "institution_application":
        application_number = snapshot.get("application_number") or log.entity_id or "-"
        name = snapshot.get("institution_name") or ""
        return f"{application_number} - {name}".strip(" -")
    return log.entity_id or "-"


def _format_audit_value(value):
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _audit_user_label(user):
    if not user:
        return "System"
    if user.first_name:
        return user.first_name
    return user.email


def _audit_item_key(item):
    if item.get("id"):
        return ("id", str(item.get("id")))
    return ("article", item.get("article_id") or "", item.get("article_name") or "")


def _application_audit_change_lines(log):
    details = log.details or {}
    before = details.get("before") or {}
    after = details.get("after") or {}

    if log.action_type == models.ActionTypeChoices.CREATE:
        item_count = after.get("item_count")
        if item_count is not None:
            return [f"Created with {item_count} item(s)."]
        return ["Created application."]

    if log.action_type == models.ActionTypeChoices.DELETE:
        item_count = before.get("item_count")
        if item_count is not None:
            return [f"Deleted application with {item_count} item(s)."]
        return ["Deleted application."]

    change_lines = []
    for key, label in AUDIT_FIELD_LABELS.items():
        if key in {"article_name", "item_type", "quantity", "unit_cost", "total_amount", "notes"}:
            continue
        before_value = before.get(key)
        after_value = after.get(key)
        if before_value != after_value and (before_value is not None or after_value is not None):
            change_lines.append(f"{label}: {_format_audit_value(before_value)} -> {_format_audit_value(after_value)}")

    before_items = before.get("items") or []
    after_items = after.get("items") or []
    before_map = {_audit_item_key(item): item for item in before_items}
    after_map = {_audit_item_key(item): item for item in after_items}

    shared_keys = [key for key in before_map.keys() if key in after_map]
    added = [item for key, item in after_map.items() if key not in before_map]
    removed = [item for key, item in before_map.items() if key not in after_map]

    for key in shared_keys:
        before_item = before_map[key]
        after_item = after_map[key]
        item_changes = []
        for field in ["quantity", "unit_cost", "total_amount", "notes"]:
            if before_item.get(field) != after_item.get(field):
                item_changes.append(
                    f"{AUDIT_FIELD_LABELS[field]}: {_format_audit_value(before_item.get(field))} -> {_format_audit_value(after_item.get(field))}"
                )
        if item_changes:
            change_lines.append(f'{after_item.get("article_name")}: ' + "; ".join(item_changes))

    for item in added:
        change_lines.append(f'Added item: {item.get("article_name")} ({item.get("quantity")} x {_format_audit_value(item.get("unit_cost"))})')
    for item in removed:
        change_lines.append(f'Removed item: {item.get("article_name")}')

    return change_lines or ["No visible field changes recorded."]


def _master_entry_filters_from_request(request):
    return {
        "search_query": (request.GET.get("q") or "").strip(),
        "date_from": (request.GET.get("date_from") or "").strip(),
        "date_to": (request.GET.get("date_to") or "").strip(),
        "status_filter": (request.GET.get("status") or "").strip().lower(),
        "sort_by": (request.GET.get("sort") or "").strip(),
        "sort_dir": "asc" if (request.GET.get("dir") or "desc").lower() == "asc" else "desc",
        "beneficiary_type": (request.GET.get("type") or "district").strip(),
    }


def _decimal_to_csv(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        normalized = value.quantize(Decimal("0.01"))
        if normalized == normalized.to_integral():
            return str(int(normalized))
        return format(normalized.normalize(), "f")
    return str(value)


def _district_export_rows(filtered_summaries):
    district_ids = [row["district_id"] for row in filtered_summaries]
    if not district_ids:
        return []
    entries = models.DistrictBeneficiaryEntry.objects.select_related("district", "article").filter(district_id__in=district_ids).order_by("district__application_number", "created_at", "id")
    rows = []
    for entry in entries:
        rows.append({
            "Application Number": entry.district.application_number or entry.application_number or "",
            "Beneficiary Name": entry.district.district_name or "",
            "Requested Item": entry.article.article_name or "",
            "Quantity": str(entry.quantity or 0),
            "Cost Per Unit": _decimal_to_csv(entry.article_cost_per_unit),
            "Total Value": _decimal_to_csv(entry.total_amount),
            "Address": entry.district.district_name or "",
            "Mobile": entry.district.mobile_number or "",
            "Aadhar Number": entry.aadhar_number or "",
            "Name of Beneficiary": entry.name_of_beneficiary or "",
            "Name of Institution": entry.name_of_institution or "",
            "Cheque / RTGS in Favour": entry.cheque_rtgs_in_favour or "",
            "Handicapped Status": "",
            "Gender": "",
            "Gender Category": "",
            "Beneficiary Type": "District",
            "Item Type": entry.article.item_type or "",
            "Article Category": entry.article.category or "",
            "Super Category Article": entry.article.master_category or "",
            "Token Name": entry.article.article_name_tk or "",
            "Internal Notes": "",
            "Comments": entry.notes or "",
        })
    return rows


def _public_export_rows(filtered_entries):
    rows = []
    for entry in filtered_entries:
        rows.append({
            "Application Number": entry.application_number or "",
            "Beneficiary Name": entry.name or "",
            "Requested Item": entry.article.article_name or "",
            "Quantity": str(entry.quantity or 0),
            "Cost Per Unit": _decimal_to_csv(entry.article_cost_per_unit),
            "Total Value": _decimal_to_csv(entry.total_amount),
            "Address": entry.address or "",
            "Mobile": entry.mobile or "",
            "Aadhar Number": entry.aadhar_number or "",
            "Name of Beneficiary": entry.name or "",
            "Name of Institution": entry.name_of_institution or "",
            "Cheque / RTGS in Favour": entry.cheque_rtgs_in_favour or "",
            "Handicapped Status": entry.get_is_handicapped_display() if entry.is_handicapped else models.HandicappedStatusChoices.NO,
            "Gender": entry.gender or "",
            "Gender Category": entry.female_status or entry.gender or "",
            "Beneficiary Type": "Public",
            "Item Type": entry.article.item_type or "",
            "Article Category": entry.article.category or "",
            "Super Category Article": entry.article.master_category or "",
            "Token Name": entry.article.article_name_tk or "",
            "Internal Notes": "",
            "Comments": entry.notes or "",
        })
    return rows


def _institution_export_rows(filtered_summaries):
    application_numbers = [row["application_number"] for row in filtered_summaries]
    if not application_numbers:
        return []
    entries = models.InstitutionsBeneficiaryEntry.objects.select_related("article").filter(application_number__in=application_numbers).order_by("application_number", "created_at", "id")
    rows = []
    for entry in entries:
        rows.append({
            "Application Number": entry.application_number or "",
            "Beneficiary Name": entry.institution_name or "",
            "Requested Item": entry.article.article_name or "",
            "Quantity": str(entry.quantity or 0),
            "Cost Per Unit": _decimal_to_csv(entry.article_cost_per_unit),
            "Total Value": _decimal_to_csv(entry.total_amount),
            "Address": entry.address or "",
            "Mobile": entry.mobile or "",
            "Aadhar Number": entry.aadhar_number or "",
            "Name of Beneficiary": entry.name_of_beneficiary or "",
            "Name of Institution": entry.name_of_institution or "",
            "Cheque / RTGS in Favour": entry.cheque_rtgs_in_favour or "",
            "Handicapped Status": "",
            "Gender": "",
            "Gender Category": "",
            "Beneficiary Type": "Institutions",
            "Item Type": entry.article.item_type or "",
            "Article Category": entry.article.category or "",
            "Super Category Article": entry.article.master_category or "",
            "Token Name": entry.article.article_name_tk or "",
            "Internal Notes": entry.internal_notes or "",
            "Comments": entry.notes or "",
        })
    return rows


def _export_master_entry_csv(request, *, export_scope):
    filters = _master_entry_filters_from_request(request)
    district_rows = []
    public_rows = []
    institution_rows = []

    if export_scope in {"all", "district"}:
        filtered_district_summaries = _filter_sort_district_summaries(
            _build_district_entry_summaries(),
            search_query=filters["search_query"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            status_filter=filters["status_filter"],
            sort_by=filters["sort_by"],
            sort_dir=filters["sort_dir"],
        )
        district_rows = _district_export_rows(filtered_district_summaries)

    if export_scope in {"all", "public"}:
        filtered_public_entries = _filter_sort_public_entries(
            models.PublicBeneficiaryEntry.objects.select_related("article").all(),
            search_query=filters["search_query"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            status_filter=filters["status_filter"],
            sort_by=filters["sort_by"],
            sort_dir=filters["sort_dir"],
        )
        public_rows = _public_export_rows(filtered_public_entries)

    if export_scope in {"all", "institutions"}:
        filtered_institution_summaries = _filter_sort_institution_summaries(
            _build_institution_entry_summaries(),
            search_query=filters["search_query"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            status_filter=filters["status_filter"],
            sort_by=filters["sort_by"],
            sort_dir=filters["sort_dir"],
        )
        institution_rows = _institution_export_rows(filtered_institution_summaries)

    rows = district_rows + public_rows + institution_rows
    response = HttpResponse(content_type="text/csv")
    timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
    if export_scope == "all":
        has_non_submitted = (
            models.DistrictBeneficiaryEntry.objects.exclude(status=models.BeneficiaryStatusChoices.SUBMITTED).exists()
            or models.PublicBeneficiaryEntry.objects.exclude(status=models.BeneficiaryStatusChoices.SUBMITTED).exists()
            or models.InstitutionsBeneficiaryEntry.objects.exclude(status=models.BeneficiaryStatusChoices.SUBMITTED).exists()
        )
        status_label = "Draft" if has_non_submitted else "Submitted"
        filename = f"1_Master_Data_{status_label}_{timestamp}.csv"
    else:
        has_non_submitted = False
        scope_rows = []
        if export_scope == "district":
            scope_rows = models.DistrictBeneficiaryEntry.objects.all()
        elif export_scope == "public":
            scope_rows = models.PublicBeneficiaryEntry.objects.all()
        elif export_scope == "institutions":
            scope_rows = models.InstitutionsBeneficiaryEntry.objects.all()
        if scope_rows:
            has_non_submitted = scope_rows.exclude(status=models.BeneficiaryStatusChoices.SUBMITTED).exists()
        status_label = "Draft" if has_non_submitted else "Submitted"
        scope_label = export_scope.title()
        filename = f"1_Master_Data_{scope_label}_{status_label}_{timestamp}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.DictWriter(response, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return response


def _build_district_entry_summaries():
    entries = list(
        models.DistrictBeneficiaryEntry.objects.select_related("district", "article").order_by("-created_at")
    )
    grouped = {}
    for entry in entries:
        key = entry.district_id
        grouped.setdefault(key, []).append(entry)

    attachment_latest, attachment_counts = _district_attachment_latest_and_counts(list(grouped.keys()))
    summaries = []
    for district_id, district_entries in grouped.items():
        first = district_entries[0]
        total_accrued = sum((entry.total_amount or 0) for entry in district_entries)
        total_quantity = sum((entry.quantity or 0) for entry in district_entries)
        remaining = (first.district.allotted_budget or 0) - total_accrued
        attachment = attachment_latest.get(district_id)
        summaries.append(
            {
                "district_id": district_id,
                "application_number": first.district.application_number or first.application_number or "-",
                "district_name": first.district.district_name,
                "article_names": ", ".join(sorted({entry.article.article_name for entry in district_entries})),
                "article_count": len({entry.article_id for entry in district_entries if entry.article_id}),
                "total_quantity": total_quantity,
                "allotted_budget": first.district.allotted_budget or 0,
                "total_accrued": total_accrued,
                "remaining_fund": remaining,
                "status": first.status,
                "internal_notes": first.internal_notes or "",
                "created_at": max(entry.created_at for entry in district_entries),
                "attachment_id": attachment.id if attachment else None,
                "attachment_preview_url": reverse("ui:application-attachment-download", kwargs={"attachment_id": attachment.id}) if attachment else "",
                "attachment_source": _attachment_preview_source(attachment),
                "attachment_title": _attachment_preview_title(attachment),
                "attachment_count": attachment_counts.get(district_id, 0),
            }
        )
    summaries.sort(key=lambda row: row["application_number"])
    return summaries


def _filter_sort_district_summaries(summaries, *, search_query="", date_from="", date_to="", status_filter="", sort_by="", sort_dir="desc"):
    if search_query:
        query = search_query.lower()
        summaries = [
            row for row in summaries
            if query in (row["application_number"] or "").lower()
            or query in (row["district_name"] or "").lower()
            or query in (row.get("article_names") or "").lower()
            or query in (row.get("internal_notes") or "").lower()
            or query in (row.get("status") or "").lower()
            or any(
                query in str(item.get("article_name") or "").lower()
                or query in str(item.get("name_of_beneficiary") or "").lower()
                or query in str(item.get("name_of_institution") or "").lower()
                or query in str(item.get("aadhar_number") or "").lower()
                or query in str(item.get("notes") or "").lower()
                or query in str(item.get("cheque_rtgs_in_favour") or "").lower()
                for item in row.get("detail_items", [])
            )
        ]

    if date_from:
        summaries = [row for row in summaries if row["created_at"].date().isoformat() >= date_from]
    if date_to:
        summaries = [row for row in summaries if row["created_at"].date().isoformat() <= date_to]
    if status_filter:
        summaries = [row for row in summaries if (row.get("status") or "") == status_filter]

    reverse = sort_dir == "desc"
    sort_map = {
        "application_number": lambda row: row["application_number"] or "",
        "district_name": lambda row: row["district_name"] or "",
        "total_accrued": lambda row: row["total_accrued"] or 0,
        "remaining_fund": lambda row: row["remaining_fund"] or 0,
        "status": lambda row: row.get("status") or "",
        "created_at": lambda row: row["created_at"],
    }
    if sort_by in sort_map:
        summaries = sorted(summaries, key=sort_map[sort_by], reverse=reverse)
    return summaries


def _district_form_context(district=None, entries=None, errors=None):
    articles = list(models.Article.objects.filter(is_active=True).order_by("article_name"))
    districts_queryset = models.DistrictMaster.objects.filter(is_active=True)
    if district is None:
        used_district_ids = models.DistrictBeneficiaryEntry.objects.values_list("district_id", flat=True).distinct()
        districts_queryset = districts_queryset.exclude(id__in=used_district_ids)
    districts = list(districts_queryset.order_by("district_name"))
    return {
        "district_master_list": districts,
        "articles_master_list": articles,
        "selected_district": district,
        "entry_rows": entries or [],
        "form_errors": errors or [],
        "form_successes": [],
        "application_status": (entries[0]["status"] if entries and isinstance(entries[0], dict) and entries[0].get("status") else ""),
        "internal_notes": (entries[0].get("internal_notes", "") if entries and isinstance(entries[0], dict) else ""),
    }


def _parse_district_rows(post_data):
    entry_ids = post_data.getlist("entry_id")
    article_ids = post_data.getlist("article_id")
    quantities = post_data.getlist("quantity")
    unit_costs = post_data.getlist("unit_cost")
    notes_list = post_data.getlist("notes")
    name_of_beneficiary_list = post_data.getlist("name_of_beneficiary")
    name_of_institution_list = post_data.getlist("name_of_institution")
    aadhar_number_list = post_data.getlist("aadhar_number")
    cheque_rtgs_list = post_data.getlist("cheque_rtgs_in_favour")
    rows = []
    max_len = max(
        len(article_ids),
        len(quantities),
        len(unit_costs),
        len(notes_list),
        len(name_of_beneficiary_list),
        len(name_of_institution_list),
        len(aadhar_number_list),
        len(cheque_rtgs_list),
        0,
    )
    for idx in range(max_len):
        rows.append(
            {
                "entry_id": (entry_ids[idx] if idx < len(entry_ids) else "").strip(),
                "article_id": (article_ids[idx] if idx < len(article_ids) else "").strip(),
                "quantity": (quantities[idx] if idx < len(quantities) else "").strip(),
                "unit_cost": (unit_costs[idx] if idx < len(unit_costs) else "").strip(),
                "notes": (notes_list[idx] if idx < len(notes_list) else "").strip(),
                "name_of_beneficiary": (name_of_beneficiary_list[idx] if idx < len(name_of_beneficiary_list) else "").strip(),
                "name_of_institution": (name_of_institution_list[idx] if idx < len(name_of_institution_list) else "").strip(),
                "aadhar_number": (aadhar_number_list[idx] if idx < len(aadhar_number_list) else "").strip(),
                "cheque_rtgs_in_favour": (cheque_rtgs_list[idx] if idx < len(cheque_rtgs_list) else "").strip(),
            }
        )
    return [row for row in rows if any(row.values())]


def _validate_and_build_district_entries(district, raw_rows, *, internal_notes=""):
    errors = []
    built_rows = []
    seen_articles = set()

    if not raw_rows:
        errors.append("Add at least one article or aid item.")
        return built_rows, errors

    article_map = {
        str(article.id): article
        for article in models.Article.objects.filter(id__in=[row["article_id"] for row in raw_rows if row["article_id"]])
    }

    for index, row in enumerate(raw_rows, start=1):
        article = article_map.get(row["article_id"])
        if not article:
            errors.append(f"Row {index}: select a valid article.")
            continue

        try:
            quantity = int(row["quantity"])
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Row {index}: quantity must be greater than 0.")
            continue

        if article.item_type != models.ItemTypeChoices.AID and article.id in seen_articles:
            errors.append(f"Row {index}: {article.article_name} can be added only once.")
            continue

        raw_unit_cost = str(row.get("unit_cost") or "").strip()
        if raw_unit_cost:
            try:
                unit_cost = Decimal(raw_unit_cost)
                if unit_cost < 0:
                    raise ValueError
            except (InvalidOperation, TypeError, ValueError):
                errors.append(f"Row {index}: enter a valid price for {article.article_name}.")
                continue
        elif article.cost_per_unit and article.cost_per_unit > 0:
            unit_cost = article.cost_per_unit
        else:
            errors.append(f"Row {index}: enter a valid price for {article.article_name}.")
            continue

        total_amount = unit_cost * quantity
        built_rows.append(
            {
                "entry_id": int(row["entry_id"]) if str(row.get("entry_id") or "").strip().isdigit() else None,
                "district": district,
                "application_number": district.application_number,
                "article": article,
                "article_cost_per_unit": unit_cost,
                "quantity": quantity,
                "total_amount": total_amount,
                "name_of_beneficiary": row.get("name_of_beneficiary") or None,
                "name_of_institution": row["name_of_institution"] or None,
                "aadhar_number": row["aadhar_number"] or None,
                "cheque_rtgs_in_favour": row["cheque_rtgs_in_favour"] or None,
                "notes": row["notes"] or None,
                "internal_notes": internal_notes or None,
                "status": models.BeneficiaryStatusChoices.PENDING,
            }
        )
        if article.item_type != models.ItemTypeChoices.AID:
            seen_articles.add(article.id)

    return built_rows, errors


def _timestamp_conflict_token(value):
    if not value:
        return ""
    return timezone.localtime(value).isoformat()


def _district_conflict_token(district):
    if not district:
        return ""
    latest = (
        models.DistrictBeneficiaryEntry.objects.filter(district=district)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    return _timestamp_conflict_token(latest)


def _public_conflict_token(entry):
    if not entry:
        return ""
    return _timestamp_conflict_token(entry.updated_at)


def _institution_conflict_token(application_number):
    if not application_number:
        return ""
    latest = (
        models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number)
        .order_by("-updated_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    return _timestamp_conflict_token(latest)


def _conflict_message(label):
    return (
        f"This {label} was updated after you opened this page. "
        f"We stopped the save so newer changes are not overwritten. "
        f"Please review the latest version and then try again."
    )


class DistrictMasterEntryBaseView(LoginRequiredMixin, WriteRoleMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    template_name = "dashboard/master_entry_district_form.html"

    def get_district(self):
        district_id = self.kwargs.get("district_id")
        if district_id:
            return models.DistrictMaster.objects.get(pk=district_id, is_active=True)
        return None

    def _render_form(self, district=None, entries=None, errors=None):
        context = self.get_context_data(**_district_form_context(district=district, entries=entries, errors=errors))
        context.update(_district_attachment_context(district))
        context["conflict_token"] = _district_conflict_token(district)
        return self.render_to_response(context)

    def _save_entries(self, district, rows, *, replace=False):
        action = (self.request.POST.get("action") or "draft").strip().lower()
        internal_notes = (self.request.POST.get("internal_notes") or "").strip()
        target_status = models.BeneficiaryStatusChoices.SUBMITTED if action == "submit" else models.BeneficiaryStatusChoices.DRAFT
        built_rows, errors = _validate_and_build_district_entries(district, rows, internal_notes=internal_notes)
        for built in built_rows:
            built["status"] = target_status
        if errors:
            hydrated_rows = []
            article_lookup = {str(article.id): article for article in models.Article.objects.filter(is_active=True)}
            for row in rows:
                article = article_lookup.get(row["article_id"])
                hydrated_rows.append(
                    {
                        "article_id": row["article_id"],
                        "entry_id": row.get("entry_id", ""),
                        "quantity": row["quantity"],
                        "unit_cost": row["unit_cost"],
                        "notes": row["notes"],
                        "name_of_beneficiary": row.get("name_of_beneficiary", ""),
                        "name_of_institution": row.get("name_of_institution", ""),
                        "aadhar_number": row.get("aadhar_number", ""),
                        "cheque_rtgs_in_favour": row.get("cheque_rtgs_in_favour", ""),
                        "article_name": article.article_name if article else "",
                        "item_type": article.item_type if article else "",
                        "internal_notes": internal_notes,
                    }
                )
            return self._render_form(district=district, entries=hydrated_rows, errors=errors)

        with transaction.atomic():
            if replace:
                existing_entries = list(
                    models.DistrictBeneficiaryEntry.objects.filter(district=district).select_related("article").order_by("id")
                )
                previous_status = existing_entries[0].status if existing_entries else None
                before_snapshot = _district_audit_snapshot(district)
                _sync_district_entries(existing_entries, built_rows, self.request.user)
                after_snapshot = _district_audit_snapshot(district)
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.UPDATE,
                    entity_type="district_application",
                    entity_id=str(district.id),
                    details={"before": before_snapshot, "after": after_snapshot},
                    **_request_audit_meta(self.request),
                )
                if previous_status != target_status:
                    services.log_audit(
                        user=self.request.user,
                        action_type=models.ActionTypeChoices.STATUS_CHANGE,
                        entity_type="district_application",
                        entity_id=str(district.id),
                        details={"from": previous_status, "to": target_status},
                        **_request_audit_meta(self.request),
                    )
            else:
                for built in built_rows:
                    create_kwargs = {key: value for key, value in built.items() if key != "entry_id"}
                    models.DistrictBeneficiaryEntry.objects.create(created_by=self.request.user, **create_kwargs)
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.CREATE,
                    entity_type="district_application",
                    entity_id=str(district.id),
                    details={"after": _district_audit_snapshot(district)},
                    **_request_audit_meta(self.request),
                )
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.STATUS_CHANGE,
                    entity_type="district_application",
                    entity_id=str(district.id),
                    details={"from": None, "to": target_status},
                    **_request_audit_meta(self.request),
                )
        return None


class DistrictMasterEntryCreateView(DistrictMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        return self._render_form()

    def post(self, request, *args, **kwargs):
        district_id = request.POST.get("district_id")
        if not district_id:
            return self._render_form(errors=["Select a district."])
        try:
            district = models.DistrictMaster.objects.get(pk=district_id, is_active=True)
        except models.DistrictMaster.DoesNotExist:
            return self._render_form(errors=["Select a valid district."])

        if models.DistrictBeneficiaryEntry.objects.filter(district=district).exists():
            return self._render_form(district=district, errors=["This district already has an entry. Use modify instead."])

        rows = _parse_district_rows(request.POST)
        response = self._save_entries(district, rows, replace=False)
        if response is not None:
            return response
        action = (request.POST.get("action") or "draft").strip().lower()
        if action == "submit":
            messages.success(request, "District application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry"))
        else:
            messages.success(request, "District application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))


class DistrictMasterEntryUpdateView(DistrictMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        district = self.get_district()
        entries = list(models.DistrictBeneficiaryEntry.objects.filter(district=district).select_related("article").order_by("id"))
        if entries and entries[0].status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This district application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry"))
        hydrated = [
            {
                "article_id": str(entry.article_id),
                "entry_id": str(entry.id),
                "quantity": entry.quantity,
                "unit_cost": entry.article_cost_per_unit,
                "notes": entry.notes or "",
                "name_of_beneficiary": entry.name_of_beneficiary or "",
                "name_of_institution": entry.name_of_institution or "",
                "aadhar_number": entry.aadhar_number or "",
                "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
                "article_name": entry.article.article_name,
                "item_type": entry.article.item_type,
                "status": entry.status,
                "internal_notes": entry.internal_notes or "",
            }
            for entry in entries
        ]
        return self._render_form(district=district, entries=hydrated)

    def post(self, request, *args, **kwargs):
        district = self.get_district()
        locked_entries = list(models.DistrictBeneficiaryEntry.objects.filter(district=district).order_by("id"))
        if locked_entries and locked_entries[0].status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This district application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry"))
        submitted_conflict_token = request.POST.get("_conflict_token", "")
        current_conflict_token = _district_conflict_token(district)
        if submitted_conflict_token and current_conflict_token and submitted_conflict_token != current_conflict_token:
            messages.error(request, _conflict_message("district application"), extra_tags="persistent")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))
        rows = _parse_district_rows(request.POST)
        response = self._save_entries(district, rows, replace=True)
        if response is not None:
            return response
        action = (request.POST.get("action") or "draft").strip().lower()
        if action == "submit":
            messages.success(request, "District application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry"))
        else:
            messages.success(request, "District application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))


class DistrictMasterEntryDetailView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/master_entry_district_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        district = models.DistrictMaster.objects.get(pk=self.kwargs["district_id"], is_active=True)
        entries = list(models.DistrictBeneficiaryEntry.objects.filter(district=district).select_related("article").order_by("id"))
        current_sort = (self.request.GET.get("sort") or "article_name").strip()
        current_dir = "asc" if (self.request.GET.get("dir") or "asc").lower() == "asc" else "desc"
        entries = _sort_application_detail_entries(entries, current_sort, current_dir)
        total_accrued = sum((entry.total_amount or 0) for entry in entries)
        total_quantity = sum((entry.quantity or 0) for entry in entries)
        status = entries[0].status if entries else ""
        context.update(
            {
                "district": district,
                "entries": entries,
                "total_accrued": total_accrued,
                "total_quantity": total_quantity,
                "remaining_fund": (district.allotted_budget or 0) - total_accrued,
                "application_status": status,
                "current_sort": current_sort,
                "current_dir": current_dir,
                "sort_querystrings": _application_detail_sort_querystrings(current_sort, current_dir),
            }
        )
        context.update(_district_attachment_context(district))
        return context


class DistrictMasterEntryDeleteView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "delete"
    def post(self, request, *args, **kwargs):
        district = models.DistrictMaster.objects.get(pk=kwargs["district_id"], is_active=True)
        snapshot = _district_audit_snapshot(district)
        models.DistrictBeneficiaryEntry.objects.filter(district=district).delete()
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.DELETE,
            entity_type="district_application",
            entity_id=str(district.id),
            details={"before": snapshot},
            **_request_audit_meta(request),
        )
        messages.warning(request, "District entry deleted.")
        return HttpResponseRedirect(reverse("ui:master-entry"))


class DistrictMasterEntryReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "reopen"
    def post(self, request, *args, **kwargs):
        district = models.DistrictMaster.objects.get(pk=kwargs["district_id"], is_active=True)
        models.DistrictBeneficiaryEntry.objects.filter(district=district).update(status=models.BeneficiaryStatusChoices.DRAFT)
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="district_application",
            entity_id=str(district.id),
            details={"from": models.BeneficiaryStatusChoices.SUBMITTED, "to": models.BeneficiaryStatusChoices.DRAFT},
            **_request_audit_meta(request),
        )
        messages.success(request, "District application reopened as draft.")
        return HttpResponseRedirect(reverse("ui:master-entry"))


def _public_history_matches(aadhar_number):
    if not aadhar_number:
        return []
    return list(
        models.PublicBeneficiaryHistory.objects.filter(
            Q(aadhar_number=aadhar_number) | Q(aadhar_number_sp=aadhar_number)
        ).order_by("-year", "name")[:10]
    )


def _public_history_summary(history_matches):
    items = []
    for row in history_matches or []:
        year = getattr(row, "year", "") or "-"
        article = getattr(row, "article_name", "") or "-"
        items.append(f"{year}: {article}")
    return "; ".join(items)


def _public_current_match(aadhar_number, *, exclude_pk=None):
    if not aadhar_number:
        return None
    queryset = models.PublicBeneficiaryEntry.objects.select_related("article").filter(aadhar_number=aadhar_number)
    if exclude_pk:
        queryset = queryset.exclude(pk=exclude_pk)
    return queryset.order_by("-created_at").first()


def _public_form_context(entry=None, form_data=None, history_matches=None, current_match=None, warnings=None, errors=None, successes=None, allow_duplicate_save=False):
    entry = entry or {}
    return {
        "public_entry": entry,
        "public_form_data": form_data or {},
        "history_matches": history_matches or [],
        "history_summary": _public_history_summary(history_matches or []),
        "current_match": current_match,
        "form_warnings": warnings or [],
        "form_errors": errors or [],
        "form_successes": [],
        "allow_duplicate_save": allow_duplicate_save,
        "articles_master_list": list(models.Article.objects.filter(is_active=True).order_by("article_name")),
        "gender_choices": models.GenderChoices.choices,
        "female_status_choices": models.FemaleStatusChoices.choices,
        "disability_category_choices": models.DisabilityCategoryChoices.choices,
        "female_status_descriptions": FEMALE_STATUS_DESCRIPTIONS,
        "application_status": getattr(entry, "status", ""),
    }


def _build_public_form_data(post_data):
    disability_category = (post_data.get("disability_category") or "").strip()
    is_handicapped = post_data.get("is_handicapped", "")
    if disability_category:
        is_handicapped = "true"
    return {
        "aadhar_number": (post_data.get("aadhar_number") or "").strip(),
        "name": (post_data.get("name") or "").strip(),
        "is_handicapped": is_handicapped,
        "disability_category": disability_category,
        "gender": (post_data.get("gender") or "").strip(),
        "female_status": (post_data.get("female_status") or "").strip(),
        "address": (post_data.get("address") or "").strip(),
        "mobile": (post_data.get("mobile") or "").strip(),
        "article_id": (post_data.get("article_id") or "").strip(),
        "article_cost_per_unit": (post_data.get("article_cost_per_unit") or "").strip(),
        "quantity": (post_data.get("quantity") or "").strip(),
        "name_of_institution": (post_data.get("name_of_institution") or "").strip(),
        "cheque_rtgs_in_favour": (post_data.get("cheque_rtgs_in_favour") or "").strip(),
        "notes": (post_data.get("notes") or "").strip(),
    }


def _validate_public_form(form_data, *, require_complete=True):
    errors = []
    article = None

    aadhar_number = form_data["aadhar_number"]
    if require_complete:
        if not (aadhar_number.isdigit() and len(aadhar_number) == 12):
            errors.append("Aadhaar number must be a valid 12-digit number.")
    elif aadhar_number and not (aadhar_number.isdigit() and len(aadhar_number) == 12):
        errors.append("Aadhaar number must be a valid 12-digit number.")

    if require_complete and not form_data["name"]:
        errors.append("Name is required.")

    if require_complete and form_data["is_handicapped"] not in {"true", "false"}:
        errors.append("Handicapped status is required.")

    if form_data["is_handicapped"] == "true":
        valid_disability_values = {value for value, _label in models.DisabilityCategoryChoices.choices}
        if require_complete and not form_data["disability_category"]:
            errors.append("Disability category is required when handicapped is Yes.")
        elif form_data["disability_category"] and form_data["disability_category"] not in valid_disability_values:
            errors.append("Select a valid disability category.")

    if require_complete and not form_data["gender"]:
        errors.append("Gender is required.")
    elif form_data["gender"] == models.GenderChoices.FEMALE and require_complete and not form_data["female_status"]:
        errors.append("Gender Category is required when gender is Female.")

    if require_complete and not form_data["address"]:
        errors.append("Address is required.")

    if require_complete and not form_data["mobile"]:
        errors.append("Mobile number is required.")
    elif form_data["mobile"]:
        mobile_numbers = [value.strip() for value in form_data["mobile"].split("&") if value.strip()]
        if not mobile_numbers and require_complete:
            errors.append("Mobile number is required.")
        elif mobile_numbers and any((not number.isdigit()) or len(number) != 10 for number in mobile_numbers):
            errors.append("Each mobile number must be exactly 10 digits.")

    if require_complete and not form_data["article_id"]:
        errors.append("Select an article or aid.")
    elif form_data["article_id"]:
        try:
            article = models.Article.objects.get(pk=form_data["article_id"], is_active=True)
        except models.Article.DoesNotExist:
            errors.append("Select a valid article or aid.")

    quantity = 1
    quantity_raw = form_data["quantity"]
    if require_complete or form_data["article_id"] or quantity_raw:
        try:
            quantity = int(quantity_raw or 0)
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append("Quantity must be greater than 0.")
            quantity = 0

    unit_cost = None
    if article:
        raw_unit_cost = str(form_data.get("article_cost_per_unit") or "").strip()
        if raw_unit_cost:
            try:
                unit_cost = Decimal(raw_unit_cost)
                if unit_cost < 0:
                    raise ValueError
            except (InvalidOperation, TypeError, ValueError):
                errors.append("Enter a valid cost per unit.")
        elif article.cost_per_unit and article.cost_per_unit > 0:
            unit_cost = article.cost_per_unit
        else:
            errors.append("Enter a valid cost per unit.")

    return article, quantity, unit_cost, errors


def _build_institution_entry_summaries():
    entries = list(
        models.InstitutionsBeneficiaryEntry.objects.select_related("article").order_by("-created_at")
    )
    grouped = {}
    for entry in entries:
        key = entry.application_number or str(entry.pk)
        grouped.setdefault(key, []).append(entry)

    attachment_latest, attachment_counts = _institution_attachment_latest_and_counts(list(grouped.keys()))
    summaries = []
    for application_number, group_entries in grouped.items():
        first = group_entries[0]
        attachment = attachment_latest.get(application_number)
        summaries.append(
            {
                "application_number": application_number,
                "institution_name": first.institution_name,
                "institution_type": first.get_institution_type_display(),
                "article_names": ", ".join(sorted({entry.article.article_name for entry in group_entries})),
                "article_count": len({entry.article_id for entry in group_entries if entry.article_id}),
                "total_quantity": sum((row.quantity or 0) for row in group_entries),
                "total_value": sum((row.total_amount or 0) for row in group_entries),
                "status": first.status,
                "internal_notes": first.internal_notes or "",
                "created_at": max(row.created_at for row in group_entries),
                "address": first.address,
                "mobile": first.mobile,
                "attachment_id": attachment.id if attachment else None,
                "attachment_preview_url": reverse("ui:application-attachment-download", kwargs={"attachment_id": attachment.id}) if attachment else "",
                "attachment_source": _attachment_preview_source(attachment),
                "attachment_title": _attachment_preview_title(attachment),
                "attachment_count": attachment_counts.get(application_number, 0),
            }
        )
    summaries.sort(key=lambda row: row["application_number"])
    return summaries


def _filter_sort_public_entries(queryset, *, search_query="", date_from="", date_to="", status_filter="", sort_by="", sort_dir="desc"):
    if search_query:
        query_filter = (
            Q(application_number__icontains=search_query)
            | Q(name__icontains=search_query)
            | Q(aadhar_number__icontains=search_query)
            | Q(address__icontains=search_query)
            | Q(mobile__icontains=search_query)
            | Q(article__article_name__icontains=search_query)
            | Q(name_of_institution__icontains=search_query)
            | Q(notes__icontains=search_query)
            | Q(cheque_rtgs_in_favour__icontains=search_query)
        )
        queryset = queryset.filter(query_filter)
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    ordering_map = {
        "application_number": "application_number",
        "name": "name",
        "aadhar_number": "aadhar_number",
        "total_amount": "total_amount",
        "status": "status",
        "created_at": "created_at",
    }
    ordering = ordering_map.get(sort_by, "created_at")
    if sort_dir == "desc":
        ordering = f"-{ordering}"
    return queryset.order_by(ordering)


def _filter_sort_institution_summaries(summaries, *, search_query="", date_from="", date_to="", status_filter="", sort_by="", sort_dir="desc"):
    if search_query:
        query = search_query.lower()
        summaries = [
            row for row in summaries
            if query in (row["application_number"] or "").lower()
            or query in (row["institution_name"] or "").lower()
            or query in (row.get("institution_type") or "").lower()
            or query in (row.get("article_names") or "").lower()
            or query in (row.get("internal_notes") or "").lower()
            or query in (row.get("address") or "").lower()
            or query in (row.get("mobile") or "").lower()
            or query in (row.get("status") or "").lower()
            or any(
                query in str(item.get("article_name") or "").lower()
                or query in str(item.get("name_of_beneficiary") or "").lower()
                or query in str(item.get("name_of_institution") or "").lower()
                or query in str(item.get("aadhar_number") or "").lower()
                or query in str(item.get("notes") or "").lower()
                or query in str(item.get("cheque_rtgs_in_favour") or "").lower()
                for item in row.get("detail_items", [])
            )
        ]

    if date_from:
        summaries = [row for row in summaries if row["created_at"].date().isoformat() >= date_from]
    if date_to:
        summaries = [row for row in summaries if row["created_at"].date().isoformat() <= date_to]
    if status_filter:
        summaries = [row for row in summaries if (row.get("status") or "") == status_filter]

    reverse = sort_dir == "desc"
    sort_map = {
        "application_number": lambda row: row["application_number"] or "",
        "institution_name": lambda row: row["institution_name"] or "",
        "total_value": lambda row: row["total_value"] or 0,
        "status": lambda row: row.get("status") or "",
        "created_at": lambda row: row["created_at"],
    }
    if sort_by in sort_map:
        summaries = sorted(summaries, key=sort_map[sort_by], reverse=reverse)
    return summaries


def _application_detail_sort_querystrings(current_sort: str, current_dir: str):
    allowed_columns = (
        "article_name",
        "item_type",
        "quantity",
        "unit_price",
        "total_amount",
        "name_of_beneficiary",
        "name_of_institution",
        "aadhar_number",
        "notes",
        "cheque_rtgs_in_favour",
        "updated_at",
    )
    return {
        column: f"?sort={column}&dir={'desc' if current_sort == column and current_dir == 'asc' else 'asc'}"
        for column in allowed_columns
    }


def _sort_application_detail_entries(entries, sort_key: str, sort_dir: str):
    sort_map = {
        "article_name": lambda entry: (entry.article.article_name or "").casefold(),
        "item_type": lambda entry: (entry.article.item_type or "").casefold(),
        "quantity": lambda entry: entry.quantity or 0,
        "unit_price": lambda entry: entry.article_cost_per_unit or Decimal("0"),
        "total_amount": lambda entry: entry.total_amount or Decimal("0"),
        "name_of_beneficiary": lambda entry: (getattr(entry, "name_of_beneficiary", "") or "").casefold(),
        "name_of_institution": lambda entry: (getattr(entry, "name_of_institution", "") or "").casefold(),
        "aadhar_number": lambda entry: (getattr(entry, "aadhar_number", "") or "").casefold(),
        "notes": lambda entry: (getattr(entry, "notes", "") or "").casefold(),
        "cheque_rtgs_in_favour": lambda entry: (getattr(entry, "cheque_rtgs_in_favour", "") or "").casefold(),
        "updated_at": lambda entry: entry.updated_at or timezone.now(),
    }
    selected_key = sort_map.get(sort_key, sort_map["article_name"])
    return sorted(
        entries,
        key=lambda entry: (selected_key(entry), (entry.article.article_name or "").casefold(), getattr(entry, "id", 0)),
        reverse=(sort_dir == "desc"),
    )


def _institution_form_context(form_data=None, rows=None, errors=None, application_number=None):
    return {
        "institution_form_data": form_data or {},
        "institution_rows": rows or [],
        "form_errors": errors or [],
        "articles_master_list": list(models.Article.objects.filter(is_active=True).order_by("article_name")),
        "institution_type_choices": models.InstitutionTypeChoices.choices,
        "institution_application_number": application_number,
        "application_status": (rows[0]["status"] if rows and isinstance(rows[0], dict) and rows[0].get("status") else ""),
        "internal_notes": (form_data or {}).get("internal_notes", ""),
    }


def _build_institution_form_data(post_data):
    return {
        "institution_name": (post_data.get("institution_name") or "").strip(),
        "institution_type": (post_data.get("institution_type") or "").strip(),
        "address": (post_data.get("address") or "").strip(),
        "mobile": (post_data.get("mobile") or "").strip(),
        "internal_notes": (post_data.get("internal_notes") or "").strip(),
    }


def _validate_institution_rows(raw_rows, *, require_complete=True, internal_notes=""):
    errors = []
    built_rows = []
    seen_articles = set()
    if not raw_rows:
        errors.append("Add at least one article or aid item.")
        return built_rows, errors

    article_map = {
        str(article.id): article
        for article in models.Article.objects.filter(id__in=[row["article_id"] for row in raw_rows if row["article_id"]])
    }

    for index, row in enumerate(raw_rows, start=1):
        article = article_map.get(row["article_id"])
        if not article:
            errors.append(f"Row {index}: select a valid article.")
            continue

        try:
            quantity = int(row["quantity"])
            if quantity <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"Row {index}: quantity must be greater than 0.")
            continue

        if article.item_type != models.ItemTypeChoices.AID and article.id in seen_articles:
            errors.append(f"Row {index}: {article.article_name} can be added only once.")
            continue

        raw_unit_cost = str(row.get("unit_cost") or "").strip()
        if raw_unit_cost:
            try:
                unit_cost = Decimal(raw_unit_cost)
                if unit_cost < 0:
                    raise ValueError
            except (InvalidOperation, TypeError, ValueError):
                errors.append(f"Row {index}: enter a valid price for {article.article_name}.")
                continue
        elif article.cost_per_unit and article.cost_per_unit > 0:
            unit_cost = article.cost_per_unit
        else:
            errors.append(f"Row {index}: enter a valid price for {article.article_name}.")
            continue

        built_rows.append(
            {
                "entry_id": int(row["entry_id"]) if str(row.get("entry_id") or "").strip().isdigit() else None,
                "article": article,
                "article_cost_per_unit": unit_cost,
                "quantity": quantity,
                "total_amount": unit_cost * quantity,
                "name_of_beneficiary": row.get("name_of_beneficiary") or None,
                "name_of_institution": row.get("name_of_institution") or None,
                "aadhar_number": row.get("aadhar_number") or None,
                "cheque_rtgs_in_favour": row.get("cheque_rtgs_in_favour") or None,
                "notes": row["notes"] or None,
                "internal_notes": internal_notes or None,
            }
        )
        if article.item_type != models.ItemTypeChoices.AID:
            seen_articles.add(article.id)
    return built_rows, errors


def _sync_district_entries(existing_entries, built_rows, user):
    by_id = {entry.id: entry for entry in existing_entries}
    by_article = {}
    for entry in existing_entries:
        by_article.setdefault(entry.article_id, []).append(entry)

    used_ids = set()
    for built in built_rows:
        match = None
        entry_id = built.get("entry_id")
        if entry_id:
            candidate = by_id.get(entry_id)
            if candidate and candidate.id not in used_ids:
                match = candidate
        if match is None:
            candidates = by_article.get(built["article"].id, [])
            match = next((entry for entry in candidates if entry.id not in used_ids), None)
        if match is None:
            create_kwargs = {key: value for key, value in built.items() if key != "entry_id"}
            models.DistrictBeneficiaryEntry.objects.create(created_by=user, **create_kwargs)
            continue

        changed = False
        if match.application_number != built["application_number"]:
            match.application_number = built["application_number"]
            changed = True
        if match.article_id != built["article"].id:
            match.article = built["article"]
            changed = True
        if match.article_cost_per_unit != built["article_cost_per_unit"]:
            match.article_cost_per_unit = built["article_cost_per_unit"]
            changed = True
        if match.quantity != built["quantity"]:
            match.quantity = built["quantity"]
            changed = True
        if match.total_amount != built["total_amount"]:
            match.total_amount = built["total_amount"]
            changed = True
        if match.name_of_beneficiary != built.get("name_of_beneficiary"):
            match.name_of_beneficiary = built.get("name_of_beneficiary")
            changed = True
        if match.name_of_institution != built.get("name_of_institution"):
            match.name_of_institution = built.get("name_of_institution")
            changed = True
        if match.aadhar_number != built.get("aadhar_number"):
            match.aadhar_number = built.get("aadhar_number")
            changed = True
        if match.cheque_rtgs_in_favour != built.get("cheque_rtgs_in_favour"):
            match.cheque_rtgs_in_favour = built.get("cheque_rtgs_in_favour")
            changed = True
        if match.notes != built["notes"]:
            match.notes = built["notes"]
            changed = True
        if match.internal_notes != built.get("internal_notes"):
            match.internal_notes = built.get("internal_notes")
            changed = True
        if match.status != built["status"]:
            match.status = built["status"]
            changed = True
        if not match.created_by_id:
            match.created_by = user
            changed = True
        if changed:
            match.save()
        used_ids.add(match.id)

    for entry in existing_entries:
        if entry.id not in used_ids:
            entry.delete()


def _sync_institution_entries(existing_entries, built_rows, user, *, application_number, form_data):
    by_id = {entry.id: entry for entry in existing_entries}
    by_article = {}
    for entry in existing_entries:
        by_article.setdefault(entry.article_id, []).append(entry)

    used_ids = set()
    for built in built_rows:
        match = None
        entry_id = built.get("entry_id")
        if entry_id:
            candidate = by_id.get(entry_id)
            if candidate and candidate.id not in used_ids:
                match = candidate
        if match is None:
            candidates = by_article.get(built["article"].id, [])
            match = next((entry for entry in candidates if entry.id not in used_ids), None)
        if match is None:
            create_kwargs = {key: value for key, value in built.items() if key != "entry_id"}
            models.InstitutionsBeneficiaryEntry.objects.create(
                created_by=user,
                application_number=application_number,
                institution_name=form_data["institution_name"],
                institution_type=form_data["institution_type"],
                address=form_data["address"] or None,
                mobile=form_data["mobile"] or None,
                **create_kwargs,
            )
            continue

        changed = False
        if match.application_number != application_number:
            match.application_number = application_number
            changed = True
        if match.institution_name != form_data["institution_name"]:
            match.institution_name = form_data["institution_name"]
            changed = True
        if match.institution_type != form_data["institution_type"]:
            match.institution_type = form_data["institution_type"]
            changed = True
        address = form_data["address"] or None
        if match.address != address:
            match.address = address
            changed = True
        mobile = form_data["mobile"] or None
        if match.mobile != mobile:
            match.mobile = mobile
            changed = True
        if match.article_id != built["article"].id:
            match.article = built["article"]
            changed = True
        if match.article_cost_per_unit != built["article_cost_per_unit"]:
            match.article_cost_per_unit = built["article_cost_per_unit"]
            changed = True
        if match.quantity != built["quantity"]:
            match.quantity = built["quantity"]
            changed = True
        if match.total_amount != built["total_amount"]:
            match.total_amount = built["total_amount"]
            changed = True
        if match.name_of_beneficiary != built.get("name_of_beneficiary"):
            match.name_of_beneficiary = built.get("name_of_beneficiary")
            changed = True
        if match.name_of_institution != built.get("name_of_institution"):
            match.name_of_institution = built.get("name_of_institution")
            changed = True
        if match.aadhar_number != built.get("aadhar_number"):
            match.aadhar_number = built.get("aadhar_number")
            changed = True
        if match.cheque_rtgs_in_favour != built.get("cheque_rtgs_in_favour"):
            match.cheque_rtgs_in_favour = built.get("cheque_rtgs_in_favour")
            changed = True
        if match.notes != built["notes"]:
            match.notes = built["notes"]
            changed = True
        if match.internal_notes != built.get("internal_notes"):
            match.internal_notes = built.get("internal_notes")
            changed = True
        if match.status != built["status"]:
            match.status = built["status"]
            changed = True
        if not match.created_by_id:
            match.created_by = user
            changed = True
        if changed:
            match.save()
        used_ids.add(match.id)

    for entry in existing_entries:
        if entry.id not in used_ids:
            entry.delete()


class PublicMasterEntryBaseView(LoginRequiredMixin, WriteRoleMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    template_name = "dashboard/master_entry_public_form.html"

    def get_entry(self):
        pk = self.kwargs.get("pk")
        if pk:
            return models.PublicBeneficiaryEntry.objects.select_related("article").get(pk=pk)
        return None

    def _render_form(self, *, entry=None, form_data=None, history_matches=None, current_match=None, warnings=None, errors=None, successes=None, allow_duplicate_save=False):
        context = self.get_context_data(
            **_public_form_context(
                entry=entry,
                form_data=form_data,
                history_matches=history_matches,
                current_match=current_match,
                warnings=warnings,
                errors=errors,
                successes=successes,
                allow_duplicate_save=allow_duplicate_save,
            )
        )
        context.update(_public_attachment_context(entry))
        context["conflict_token"] = _public_conflict_token(entry)
        return self.render_to_response(context)

    def _save_entry(self, *, entry=None, form_data=None):
        action = (self.request.POST.get("action") or "draft").strip().lower()
        target_status = models.BeneficiaryStatusChoices.SUBMITTED if action == "submit" else models.BeneficiaryStatusChoices.DRAFT
        require_complete = target_status == models.BeneficiaryStatusChoices.SUBMITTED
        article, quantity, unit_cost, errors = _validate_public_form(form_data, require_complete=require_complete)
        if not require_complete and article is None:
            errors.append("Select an article or aid before saving draft.")
        history_matches = _public_history_matches(form_data["aadhar_number"])
        current_match = _public_current_match(form_data["aadhar_number"], exclude_pk=getattr(entry, "pk", None))
        warnings = ["This Aadhaar number exists in past beneficiary history."] if history_matches else []
        if current_match:
            warnings.append("Duplicate found in current applications.")
        if errors:
            return self._render_form(
                entry=entry,
                form_data=form_data,
                history_matches=history_matches,
                current_match=current_match,
                warnings=warnings,
                errors=errors,
            ), None

        if current_match:
            return self._render_form(
                entry=entry,
                form_data=form_data,
                history_matches=history_matches,
                current_match=current_match,
                warnings=warnings + ["Duplicate found in current applications. Review the existing record below and use Modify if you want to update it."],
            ), None

        if entry is None:
            entry = models.PublicBeneficiaryEntry(application_number=f"DRAFT-PUB-{uuid.uuid4().hex[:12].upper()}")
            action_type = models.ActionTypeChoices.CREATE
            before_snapshot = None
            previous_status = None
        else:
            action_type = models.ActionTypeChoices.UPDATE
            before_snapshot = _public_audit_snapshot(entry)
            previous_status = entry.status

        entry.name = form_data["name"]
        entry.aadhar_number = form_data["aadhar_number"]
        entry.is_handicapped = form_data["disability_category"] or models.HandicappedStatusChoices.NO
        entry.gender = form_data["gender"]
        entry.female_status = form_data["female_status"] or None
        entry.address = form_data["address"] or None
        entry.mobile = form_data["mobile"]
        entry.article = article
        entry.article_cost_per_unit = unit_cost
        entry.quantity = quantity
        entry.total_amount = unit_cost * quantity
        entry.name_of_institution = form_data["name_of_institution"] or None
        entry.cheque_rtgs_in_favour = form_data["cheque_rtgs_in_favour"] or None
        entry.notes = form_data["notes"] or None
        if target_status == models.BeneficiaryStatusChoices.SUBMITTED and (not entry.application_number or str(entry.application_number).startswith("DRAFT-PUB-")):
            entry.application_number = services.next_public_application_number()
        entry.status = target_status
        if not entry.created_by_id:
            entry.created_by = self.request.user
        entry.save()
        services.log_audit(
            user=self.request.user,
            action_type=action_type,
            entity_type="public_application",
            entity_id=str(entry.id),
            details={"before": before_snapshot, "after": _public_audit_snapshot(entry)},
            **_request_audit_meta(self.request),
        )
        if previous_status != target_status:
            services.log_audit(
                user=self.request.user,
                action_type=models.ActionTypeChoices.STATUS_CHANGE,
                entity_type="public_application",
                entity_id=str(entry.id),
                details={"from": previous_status, "to": target_status},
                **_request_audit_meta(self.request),
            )
        return None, entry


class PublicMasterEntryCreateView(PublicMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        return self._render_form(form_data={"quantity": "1"})

    def post(self, request, *args, **kwargs):
        form_data = _build_public_form_data(request.POST)
        if request.POST.get("action") == "verify":
            if not (form_data["aadhar_number"].isdigit() and len(form_data["aadhar_number"]) == 12):
                return self._render_form(
                    form_data=form_data,
                    errors=["Aadhaar number must be a valid 12-digit number."],
                )
            history_matches = _public_history_matches(form_data["aadhar_number"])
            current_match = _public_current_match(form_data["aadhar_number"])
            warnings = ["This Aadhaar number exists in past beneficiary history."] if history_matches else []
            successes = ["Verification passed."] if not history_matches and not current_match else []
            if current_match:
                warnings.append("Duplicate found in current applications.")
            return self._render_form(
                form_data=form_data,
                history_matches=history_matches,
                current_match=current_match,
                warnings=warnings,
                successes=successes,
                allow_duplicate_save=bool(current_match),
            )

        response, saved_entry = self._save_entry(form_data=form_data)
        if response is not None:
            return response
        action = (request.POST.get("action") or "draft").strip().lower()
        if action == "submit":
            request.session["public_submit_popup"] = {"application_number": saved_entry.application_number, "name": saved_entry.name}
            messages.success(request, "Public application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")
        else:
            messages.success(request, "Public application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": saved_entry.pk}))


class PublicMasterEntryUpdateView(PublicMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        entry = self.get_entry()
        if entry.status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This public application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")
        return self._render_form(
            entry=entry,
            form_data={
                "aadhar_number": entry.aadhar_number,
                "name": entry.name,
                "is_handicapped": "true" if entry.is_handicapped and entry.is_handicapped != models.HandicappedStatusChoices.NO else "false",
                "disability_category": entry.is_handicapped if entry.is_handicapped and entry.is_handicapped != models.HandicappedStatusChoices.NO else "",
                "gender": entry.gender or "",
                "female_status": entry.female_status or "",
                "address": entry.address or "",
                "mobile": entry.mobile or "",
                "article_id": str(entry.article_id),
                "article_cost_per_unit": str(entry.article_cost_per_unit),
                "quantity": str(entry.quantity),
                "name_of_institution": entry.name_of_institution or "",
                "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
                "notes": entry.notes or "",
            },
            history_matches=_public_history_matches(entry.aadhar_number),
        )

    def post(self, request, *args, **kwargs):
        entry = self.get_entry()
        if entry.status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This public application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")
        submitted_conflict_token = request.POST.get("_conflict_token", "")
        current_conflict_token = _public_conflict_token(entry)
        if submitted_conflict_token and current_conflict_token and submitted_conflict_token != current_conflict_token:
            messages.error(request, _conflict_message("public application"), extra_tags="persistent")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))
        form_data = _build_public_form_data(request.POST)
        if request.POST.get("action") == "verify":
            if not (form_data["aadhar_number"].isdigit() and len(form_data["aadhar_number"]) == 12):
                return self._render_form(
                    entry=entry,
                    form_data=form_data,
                    errors=["Aadhaar number must be a valid 12-digit number."],
                )
            history_matches = _public_history_matches(form_data["aadhar_number"])
            current_match = _public_current_match(form_data["aadhar_number"], exclude_pk=entry.pk)
            warnings = ["This Aadhaar number exists in past beneficiary history."] if history_matches else []
            successes = ["Verification passed."] if not history_matches and not current_match else []
            if current_match:
                warnings.append("Duplicate found in current applications.")
            return self._render_form(
                entry=entry,
                form_data=form_data,
                history_matches=history_matches,
                current_match=current_match,
                warnings=warnings,
                successes=successes,
                allow_duplicate_save=bool(current_match),
            )

        response, saved_entry = self._save_entry(entry=entry, form_data=form_data)
        if response is not None:
            return response
        action = (request.POST.get("action") or "draft").strip().lower()
        if action == "submit":
            request.session["public_submit_popup"] = {"application_number": saved_entry.application_number, "name": saved_entry.name}
            messages.success(request, "Public application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")
        else:
            messages.success(request, "Public application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": saved_entry.pk}))


class PublicMasterEntryDetailView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/master_entry_public_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        entry = models.PublicBeneficiaryEntry.objects.select_related("article").get(pk=self.kwargs["pk"])
        context["entry"] = entry
        context["history_matches"] = _public_history_matches(entry.aadhar_number)
        current_sort = (self.request.GET.get("sort") or "article_name").strip()
        current_dir = "asc" if (self.request.GET.get("dir") or "asc").lower() == "asc" else "desc"
        detail_entries = _sort_application_detail_entries([entry], current_sort, current_dir)
        context["detail_entries"] = detail_entries
        context["current_sort"] = current_sort
        context["current_dir"] = current_dir
        context["sort_querystrings"] = _application_detail_sort_querystrings(current_sort, current_dir)
        context.update(_public_attachment_context(entry))
        return context


class PublicMasterEntryDeleteView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "delete"
    def post(self, request, *args, **kwargs):
        entry = models.PublicBeneficiaryEntry.objects.get(pk=kwargs["pk"])
        snapshot = _public_audit_snapshot(entry)
        entry.delete()
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.DELETE,
            entity_type="public_application",
            entity_id=str(kwargs["pk"]),
            details={"before": snapshot},
            **_request_audit_meta(request),
        )
        messages.warning(request, "Public entry deleted.")
        return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")


class PublicMasterEntryReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "reopen"
    def post(self, request, *args, **kwargs):
        entry = models.PublicBeneficiaryEntry.objects.get(pk=kwargs["pk"])
        previous_status = entry.status
        entry.status = models.BeneficiaryStatusChoices.DRAFT
        entry.save(update_fields=["status"])
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="public_application",
            entity_id=str(entry.id),
            details={"from": previous_status, "to": models.BeneficiaryStatusChoices.DRAFT},
            **_request_audit_meta(request),
        )
        messages.success(request, "Public application reopened as draft.")
        return HttpResponseRedirect(reverse("ui:master-entry") + "?type=public")


class InstitutionsMasterEntryBaseView(LoginRequiredMixin, WriteRoleMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    template_name = "dashboard/master_entry_institution_form.html"

    def _render_form(self, *, form_data=None, rows=None, errors=None, application_number=None):
        context = self.get_context_data(
            **_institution_form_context(
                form_data=form_data,
                rows=rows,
                errors=errors,
                application_number=application_number,
            )
        )
        context.update(_institution_attachment_context(application_number))
        context["conflict_token"] = _institution_conflict_token(application_number)
        return self.render_to_response(context)

    def _save_group(self, *, application_number=None, form_data=None, raw_rows=None, replace=False):
        action = (self.request.POST.get("action") or "draft").strip().lower()
        require_complete = action == "submit"
        errors = []
        if require_complete and not form_data["institution_name"]:
            errors.append("Institution name is required.")
        if require_complete and not form_data["institution_type"]:
            errors.append("Institution type is required.")
        if require_complete and not form_data["address"]:
            errors.append("Address is required.")
        if require_complete and not form_data["mobile"]:
            errors.append("Mobile number is required.")
        elif form_data["mobile"]:
            mobile_numbers = [value.strip() for value in form_data["mobile"].split("&") if value.strip()]
            if mobile_numbers and any((not number.isdigit()) or len(number) != 10 for number in mobile_numbers):
                errors.append("Each mobile number must be exactly 10 digits.")

        built_rows, row_errors = _validate_institution_rows(
            raw_rows,
            require_complete=require_complete,
            internal_notes=form_data["internal_notes"],
        )
        errors.extend(row_errors)

        if errors:
            article_lookup = {str(article.id): article for article in models.Article.objects.filter(is_active=True)}
            hydrated_rows = []
            for row in raw_rows:
                article = article_lookup.get(row["article_id"])
                hydrated_rows.append(
                    {
                        "article_id": row["article_id"],
                        "entry_id": row.get("entry_id", ""),
                        "quantity": row["quantity"],
                        "unit_cost": row["unit_cost"],
                        "notes": row["notes"],
                        "name_of_beneficiary": row.get("name_of_beneficiary", ""),
                        "name_of_institution": row.get("name_of_institution", ""),
                        "aadhar_number": row.get("aadhar_number", ""),
                        "cheque_rtgs_in_favour": row.get("cheque_rtgs_in_favour", ""),
                        "article_name": article.article_name if article else "",
                        "item_type": article.item_type if article else "",
                        "internal_notes": form_data["internal_notes"],
                    }
                )
            return self._render_form(
                form_data=form_data,
                rows=hydrated_rows,
                errors=errors,
                application_number=application_number,
            )

        source_application_number = application_number
        is_draft_placeholder = bool(application_number and application_number.startswith("DRAFT-INS-"))
        if action == "submit":
            if not application_number or is_draft_placeholder:
                application_number = services.next_institution_application_number()
        elif not application_number:
            application_number = f"DRAFT-INS-{uuid.uuid4().hex[:12].upper()}"
        target_status = models.BeneficiaryStatusChoices.SUBMITTED if action == "submit" else models.BeneficiaryStatusChoices.DRAFT
        for built in built_rows:
            built["status"] = target_status

        with transaction.atomic():
            if replace:
                lookup_application_number = source_application_number or application_number
                before_snapshot = _institution_audit_snapshot(lookup_application_number)
                existing_entries = list(
                    models.InstitutionsBeneficiaryEntry.objects.filter(application_number=lookup_application_number).select_related("article").order_by("id")
                )
                previous_status = existing_entries[0].status if existing_entries else None
                _sync_institution_entries(
                    existing_entries,
                    built_rows,
                    self.request.user,
                    application_number=application_number,
                    form_data=form_data,
                )
                if lookup_application_number != application_number:
                    models.ApplicationAttachment.objects.filter(
                        application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
                        institution_application_number=lookup_application_number,
                    ).update(institution_application_number=application_number)
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.UPDATE,
                    entity_type="institution_application",
                    entity_id=application_number,
                    details={"before": before_snapshot, "after": _institution_audit_snapshot(application_number)},
                    **_request_audit_meta(self.request),
                )
                if previous_status != target_status:
                    services.log_audit(
                        user=self.request.user,
                        action_type=models.ActionTypeChoices.STATUS_CHANGE,
                        entity_type="institution_application",
                        entity_id=application_number,
                        details={"from": previous_status, "to": target_status},
                        **_request_audit_meta(self.request),
                    )
            else:
                for built in built_rows:
                    create_kwargs = {key: value for key, value in built.items() if key != "entry_id"}
                    models.InstitutionsBeneficiaryEntry.objects.create(
                        created_by=self.request.user,
                        application_number=application_number,
                        institution_name=form_data["institution_name"],
                        institution_type=form_data["institution_type"],
                        address=form_data["address"] or None,
                        mobile=form_data["mobile"] or None,
                        **create_kwargs,
                    )
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.CREATE,
                    entity_type="institution_application",
                    entity_id=application_number,
                    details={"after": _institution_audit_snapshot(application_number)},
                    **_request_audit_meta(self.request),
                )
                services.log_audit(
                    user=self.request.user,
                    action_type=models.ActionTypeChoices.STATUS_CHANGE,
                    entity_type="institution_application",
                    entity_id=application_number,
                    details={"from": None, "to": target_status},
                    **_request_audit_meta(self.request),
                )
        self._saved_institution_application_number = application_number
        return None


class InstitutionsMasterEntryCreateView(InstitutionsMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        return self._render_form()

    def post(self, request, *args, **kwargs):
        form_data = _build_institution_form_data(request.POST)
        rows = _parse_district_rows(request.POST)
        action = (request.POST.get("action") or "draft").strip().lower()
        response = self._save_group(form_data=form_data, raw_rows=rows)
        if response is not None:
            return response
        saved_application_number = getattr(self, "_saved_institution_application_number", None)
        if action == "submit":
            popup_entry = models.InstitutionsBeneficiaryEntry.objects.filter(application_number=saved_application_number).order_by("-updated_at").first()
            if popup_entry:
                request.session["institution_submit_popup"] = {
                    "application_number": popup_entry.application_number,
                    "name": popup_entry.institution_name,
                }
            messages.success(request, "Institution application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")
        else:
            messages.success(request, "Institution application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": saved_application_number}))


class InstitutionsMasterEntryUpdateView(InstitutionsMasterEntryBaseView):
    def get(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        entries = list(
            models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).select_related("article").order_by("id")
        )
        if entries and entries[0].status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This institution application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")
        first = entries[0]
        rows = [
            {
                "article_id": str(entry.article_id),
                "entry_id": str(entry.id),
                "quantity": entry.quantity,
                "unit_cost": entry.article_cost_per_unit,
                "notes": entry.notes or "",
                "name_of_beneficiary": entry.name_of_beneficiary or "",
                "name_of_institution": entry.name_of_institution or "",
                "aadhar_number": entry.aadhar_number or "",
                "cheque_rtgs_in_favour": entry.cheque_rtgs_in_favour or "",
                "article_name": entry.article.article_name,
                "item_type": entry.article.item_type,
                "status": entry.status,
                "internal_notes": entry.internal_notes or "",
            }
            for entry in entries
        ]
        form_data = {
            "institution_name": first.institution_name,
            "institution_type": first.institution_type,
            "address": first.address or "",
            "mobile": first.mobile or "",
            "internal_notes": first.internal_notes or "",
        }
        return self._render_form(form_data=form_data, rows=rows, application_number=application_number)

    def post(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        existing_entries = list(models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).order_by("id"))
        if existing_entries and existing_entries[0].status == models.BeneficiaryStatusChoices.SUBMITTED:
            messages.error(request, "This institution application is submitted and locked. Reopen it first.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")
        submitted_conflict_token = request.POST.get("_conflict_token", "")
        current_conflict_token = _institution_conflict_token(application_number)
        if submitted_conflict_token and current_conflict_token and submitted_conflict_token != current_conflict_token:
            messages.error(request, _conflict_message("institution application"), extra_tags="persistent")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))
        form_data = _build_institution_form_data(request.POST)
        rows = _parse_district_rows(request.POST)
        action = (request.POST.get("action") or "draft").strip().lower()
        response = self._save_group(application_number=application_number, form_data=form_data, raw_rows=rows, replace=True)
        if response is not None:
            return response
        saved_application_number = getattr(self, "_saved_institution_application_number", application_number)
        if action == "submit":
            popup_entry = models.InstitutionsBeneficiaryEntry.objects.filter(application_number=saved_application_number).order_by("-updated_at").first()
            if popup_entry:
                request.session["institution_submit_popup"] = {
                    "application_number": popup_entry.application_number,
                    "name": popup_entry.institution_name,
                }
            messages.success(request, "Institution application submitted.")
            return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")
        else:
            messages.success(request, "Institution application saved as draft.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": saved_application_number}))


class InstitutionsMasterEntryDetailView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    template_name = "dashboard/master_entry_institution_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        application_number = self.kwargs["application_number"]
        entries = list(
            models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).select_related("article").order_by("id")
        )
        current_sort = (self.request.GET.get("sort") or "article_name").strip()
        current_dir = "asc" if (self.request.GET.get("dir") or "asc").lower() == "asc" else "desc"
        entries = _sort_application_detail_entries(entries, current_sort, current_dir)
        first = entries[0]
        context.update(
            {
                "application_number": application_number,
                "entry_header": first,
                "entries": entries,
                "total_quantity": sum((row.quantity or 0) for row in entries),
                "total_value": sum((row.total_amount or 0) for row in entries),
                "application_status": first.status,
                "current_sort": current_sort,
                "current_dir": current_dir,
                "sort_querystrings": _application_detail_sort_querystrings(current_sort, current_dir),
            }
        )
        context.update(_institution_attachment_context(application_number))
        return context


class InstitutionsMasterEntryDeleteView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "delete"
    def post(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        snapshot = _institution_audit_snapshot(application_number)
        models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).delete()
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.DELETE,
            entity_type="institution_application",
            entity_id=application_number,
            details={"before": snapshot},
            **_request_audit_meta(request),
        )
        messages.warning(request, "Institution entry deleted.")
        return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")


class InstitutionsMasterEntryReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "reopen"
    def post(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).update(
            status=models.BeneficiaryStatusChoices.DRAFT
        )
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="institution_application",
            entity_id=application_number,
            details={"from": models.BeneficiaryStatusChoices.SUBMITTED, "to": models.BeneficiaryStatusChoices.DRAFT},
            **_request_audit_meta(request),
        )
        messages.success(request, "Institution application reopened as draft.")
        return HttpResponseRedirect(reverse("ui:master-entry") + "?type=institutions")


class ApplicationAttachmentDownloadView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "view"
    def get(self, request, *args, **kwargs):
        attachment = get_object_or_404(models.ApplicationAttachment.objects.select_related("uploaded_by"), pk=kwargs["attachment_id"])
        if not attachment.file and not attachment.drive_file_id:
            raise Http404("File not found.")
        stored_name = attachment.file_name or (os.path.basename(attachment.file.name) if attachment.file else "attachment")
        application_reference = _attachment_application_reference(attachment)
        display_name = _prefixed_attachment_name(application_reference, stored_name, attachment.file_name or stored_name)
        as_attachment = (request.GET.get("download") or "").strip() == "1"
        content_type, _ = mimetypes.guess_type(stored_name)
        content_type = attachment.drive_mime_type or content_type or "application/octet-stream"
        if attachment.drive_file_id:
            file_bytes = google_drive.download_file(attachment.drive_file_id)
            if as_attachment:
                display_root, display_ext = os.path.splitext(display_name)
                _, stored_ext = os.path.splitext(stored_name)
                download_name = display_name if display_ext else f"{display_name}{stored_ext}"
                response = HttpResponse(file_bytes, content_type=content_type)
                response["Content-Disposition"] = f'attachment; filename="{download_name}"'
                return response
            response = HttpResponse(file_bytes, content_type=content_type)
            response["Content-Disposition"] = "inline"
            return response
        if as_attachment:
            display_root, display_ext = os.path.splitext(display_name)
            _, stored_ext = os.path.splitext(stored_name)
            download_name = display_name if display_ext else f"{display_name}{stored_ext}"
            return FileResponse(
                attachment.file.open("rb"),
                as_attachment=True,
                filename=download_name,
                content_type=content_type,
            )
        response = FileResponse(
            attachment.file.open("rb"),
            as_attachment=False,
            content_type=content_type,
        )
        response["Content-Disposition"] = "inline"
        return response


class DistrictApplicationAttachmentUploadView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        district = get_object_or_404(models.DistrictMaster, pk=kwargs["district_id"], is_active=True)
        form = ApplicationAttachmentUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "; ".join(form.errors.get("file", []) + form.errors.get("file_name", [])) or "Choose a valid file before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))

        attachments_qs = models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
            district=district,
        )
        existing_count = attachments_qs.count()
        if existing_count >= ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION:
            messages.error(request, f"Maximum {ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION} files are allowed for one application.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))

        uploaded = form.cleaned_data["file"]
        display_name = _prefixed_attachment_name(district.application_number, uploaded.name, form.cleaned_data.get("file_name") or "")
        if _attachment_name_exists(attachments_qs, display_name):
            messages.error(request, "A file with this name already exists for this application. Please rename it before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))
        try:
            _save_application_attachment(
                uploaded=uploaded,
                display_name=display_name,
                application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
                district=district,
                uploaded_by=request.user,
            )
        except Exception:
            logger.exception("Attachment upload failed for district=%s", district.id)
            messages.error(request, "Attachment upload failed. Please check Google Drive configuration and try again.")
            return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))
        messages.success(request, "Attachment uploaded.")
        return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))


class DistrictApplicationAttachmentDeleteView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        district = get_object_or_404(models.DistrictMaster, pk=kwargs["district_id"], is_active=True)
        attachment = get_object_or_404(
            models.ApplicationAttachment,
            pk=kwargs["attachment_id"],
            application_type=models.ApplicationAttachmentTypeChoices.DISTRICT,
            district=district,
        )
        _delete_application_attachment_file(attachment)
        attachment.delete()
        messages.success(request, "Attachment removed.")
        return HttpResponseRedirect(reverse("ui:master-entry-district-edit", kwargs={"district_id": district.id}))


class PublicApplicationAttachmentUploadView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        entry = get_object_or_404(models.PublicBeneficiaryEntry, pk=kwargs["pk"])
        form = ApplicationAttachmentUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "; ".join(form.errors.get("file", []) + form.errors.get("file_name", [])) or "Choose a valid file before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))

        attachments_qs = models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
            public_entry=entry,
        )
        existing_count = attachments_qs.count()
        if existing_count >= ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION:
            messages.error(request, f"Maximum {ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION} files are allowed for one application.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))

        uploaded = form.cleaned_data["file"]
        application_reference = entry.application_number or f"PUBLIC-{entry.pk}"
        display_name = _prefixed_attachment_name(application_reference, uploaded.name, form.cleaned_data.get("file_name") or "")
        if _attachment_name_exists(attachments_qs, display_name):
            messages.error(request, "A file with this name already exists for this application. Please rename it before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))
        try:
            _save_application_attachment(
                uploaded=uploaded,
                display_name=display_name,
                application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
                public_entry=entry,
                uploaded_by=request.user,
            )
        except Exception:
            logger.exception("Attachment upload failed for public_entry=%s", entry.pk)
            messages.error(request, "Attachment upload failed. Please check Google Drive configuration and try again.")
            return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))
        messages.success(request, "Attachment uploaded.")
        return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))


class PublicApplicationAttachmentDeleteView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        entry = get_object_or_404(models.PublicBeneficiaryEntry, pk=kwargs["pk"])
        attachment = get_object_or_404(
            models.ApplicationAttachment,
            pk=kwargs["attachment_id"],
            application_type=models.ApplicationAttachmentTypeChoices.PUBLIC,
            public_entry=entry,
        )
        _delete_application_attachment_file(attachment)
        attachment.delete()
        messages.success(request, "Attachment removed.")
        return HttpResponseRedirect(reverse("ui:master-entry-public-edit", kwargs={"pk": entry.pk}))


class InstitutionApplicationAttachmentUploadView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        if not models.InstitutionsBeneficiaryEntry.objects.filter(application_number=application_number).exists():
            raise Http404("Institution application not found.")
        form = ApplicationAttachmentUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "; ".join(form.errors.get("file", []) + form.errors.get("file_name", [])) or "Choose a valid file before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))

        attachments_qs = models.ApplicationAttachment.objects.filter(
            application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
            institution_application_number=application_number,
        )
        existing_count = attachments_qs.count()
        if existing_count >= ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION:
            messages.error(request, f"Maximum {ApplicationAttachmentUploadForm.MAX_FILES_PER_APPLICATION} files are allowed for one application.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))

        uploaded = form.cleaned_data["file"]
        display_name = _prefixed_attachment_name(application_number, uploaded.name, form.cleaned_data.get("file_name") or "")
        if _attachment_name_exists(attachments_qs, display_name):
            messages.error(request, "A file with this name already exists for this application. Please rename it before uploading.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))
        try:
            _save_application_attachment(
                uploaded=uploaded,
                display_name=display_name,
                application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
                institution_application_number=application_number,
                uploaded_by=request.user,
            )
        except Exception:
            logger.exception("Attachment upload failed for institution=%s", application_number)
            messages.error(request, "Attachment upload failed. Please check Google Drive configuration and try again.")
            return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))
        messages.success(request, "Attachment uploaded.")
        return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))


class InstitutionApplicationAttachmentDeleteView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.APPLICATION_ENTRY
    permission_action = "create_edit"
    def post(self, request, *args, **kwargs):
        application_number = kwargs["application_number"]
        attachment = get_object_or_404(
            models.ApplicationAttachment,
            pk=kwargs["attachment_id"],
            application_type=models.ApplicationAttachmentTypeChoices.INSTITUTION,
            institution_application_number=application_number,
        )
        _delete_application_attachment_file(attachment)
        attachment.delete()
        messages.success(request, "Attachment removed.")
        return HttpResponseRedirect(reverse("ui:master-entry-institution-edit", kwargs={"application_number": application_number}))


def _split_order_article_names(article: models.Article | None) -> tuple[list[str], bool]:
    if not article or not article.article_name:
        return [], False
    raw_name = article.article_name.strip()
    if not raw_name:
        return [], False
    has_plus = "+" in raw_name
    if has_plus:
        parts = [part.strip() for part in raw_name.split("+") if part.strip()]
    else:
        parts = [raw_name]
    return parts, bool(article.combo or has_plus)


def _ensure_order_summary_row(rows_map, order_name: str, article: models.Article | None, combo_related: bool):
    key = order_name.casefold()
    if key not in rows_map:
        rows_map[key] = {
            "row_key": key,
            "article_name": order_name,
            "item_type": article.item_type if article else "",
            "category": article.category if article else "",
            "master_category": article.master_category if article else "",
            "combo_related": combo_related,
            "total_quantity": 0,
            "quantity_ordered": 0,
            "quantity_received": 0,
            "quantity_pending": 0,
            "total_value": Decimal("0"),
            "breakdown": {"district": 0, "public": 0, "institutions": 0},
            "source_items": set(),
            "beneficiaries": [],
            "statuses": set(),
        }
    row = rows_map[key]
    if article:
        row["item_type"] = row["item_type"] or article.item_type
        row["category"] = row["category"] or (article.category or "")
        row["master_category"] = row["master_category"] or (article.master_category or "")
    row["combo_related"] = row["combo_related"] or combo_related
    return row


def _build_order_management_rows():
    rows_map = {}

    district_entries = (
        models.DistrictBeneficiaryEntry.objects.select_related("district", "article")
        .order_by("application_number", "district__district_name", "created_at")
    )
    for entry in district_entries:
        parts, combo_related = _split_order_article_names(entry.article)
        if not parts:
            continue
        split_count = Decimal(len(parts))
        value_share = (entry.total_amount or Decimal("0")) / split_count if split_count else Decimal("0")
        for part in parts:
            row = _ensure_order_summary_row(rows_map, part, entry.article, combo_related)
            row["total_quantity"] += entry.quantity
            row["total_value"] += value_share
            row["breakdown"]["district"] += entry.quantity
            row["source_items"].add(entry.article.article_name)
            row["statuses"].add(entry.status)
            row["beneficiaries"].append(
                {
                    "beneficiary_type": "District",
                    "application_number": entry.application_number or "",
                    "beneficiary_name": entry.district.district_name,
                    "quantity": entry.quantity,
                    "source_item": entry.article.article_name,
                    "notes": entry.notes or "",
                    "status": entry.status,
                    "item_type": entry.article.item_type,
                    "linked_fund_request_status": getattr(entry.fund_request, "status", "") if entry.fund_request_id else "",
                    "created_at": entry.created_at,
                }
            )

    public_entries = (
        models.PublicBeneficiaryEntry.objects.select_related("article")
        .order_by("application_number", "name", "created_at")
    )
    for entry in public_entries:
        parts, combo_related = _split_order_article_names(entry.article)
        if not parts:
            continue
        split_count = Decimal(len(parts))
        value_share = (entry.total_amount or Decimal("0")) / split_count if split_count else Decimal("0")
        for part in parts:
            row = _ensure_order_summary_row(rows_map, part, entry.article, combo_related)
            row["total_quantity"] += entry.quantity
            row["total_value"] += value_share
            row["breakdown"]["public"] += entry.quantity
            row["source_items"].add(entry.article.article_name)
            row["statuses"].add(entry.status)
            row["beneficiaries"].append(
                {
                    "beneficiary_type": "Public",
                    "application_number": entry.application_number or "",
                    "beneficiary_name": entry.name,
                    "quantity": entry.quantity,
                    "source_item": entry.article.article_name,
                    "notes": entry.notes or "",
                    "status": entry.status,
                    "item_type": entry.article.item_type,
                    "linked_fund_request_status": getattr(entry.fund_request, "status", "") if entry.fund_request_id else "",
                    "created_at": entry.created_at,
                }
            )

    institution_entries = (
        models.InstitutionsBeneficiaryEntry.objects.select_related("article")
        .order_by("application_number", "institution_name", "created_at")
    )
    for entry in institution_entries:
        parts, combo_related = _split_order_article_names(entry.article)
        if not parts:
            continue
        split_count = Decimal(len(parts))
        value_share = (entry.total_amount or Decimal("0")) / split_count if split_count else Decimal("0")
        for part in parts:
            row = _ensure_order_summary_row(rows_map, part, entry.article, combo_related)
            row["total_quantity"] += entry.quantity
            row["total_value"] += value_share
            row["breakdown"]["institutions"] += entry.quantity
            row["source_items"].add(entry.article.article_name)
            row["statuses"].add(entry.status)
            row["beneficiaries"].append(
                {
                    "beneficiary_type": "Institutions",
                    "application_number": entry.application_number or "",
                    "beneficiary_name": entry.institution_name,
                    "quantity": entry.quantity,
                    "source_item": entry.article.article_name,
                    "notes": entry.notes or "",
                    "status": entry.status,
                    "item_type": entry.article.item_type,
                    "linked_fund_request_status": getattr(entry.fund_request, "status", "") if entry.fund_request_id else "",
                    "created_at": entry.created_at,
                }
            )

    order_entries = (
        models.OrderEntry.objects.select_related("article")
        .exclude(status=models.OrderStatusChoices.CANCELLED)
        .order_by("article__article_name", "order_date", "created_at")
    )
    for entry in order_entries:
        parts, combo_related = _split_order_article_names(entry.article)
        if not parts:
            continue
        for part in parts:
            row = _ensure_order_summary_row(rows_map, part, entry.article, combo_related)
            if entry.status in {models.OrderStatusChoices.ORDERED, models.OrderStatusChoices.RECEIVED}:
                row["quantity_ordered"] += entry.quantity_ordered
            if entry.status == models.OrderStatusChoices.RECEIVED:
                row["quantity_received"] += entry.quantity_ordered

    rows = []
    for index, row in enumerate(sorted(rows_map.values(), key=lambda item: item["article_name"].casefold()), start=1):
        if models.BeneficiaryStatusChoices.SUBMITTED in row["statuses"]:
            row["source_status"] = models.BeneficiaryStatusChoices.SUBMITTED
        elif models.BeneficiaryStatusChoices.DRAFT in row["statuses"]:
            row["source_status"] = models.BeneficiaryStatusChoices.DRAFT
        else:
            row["source_status"] = ""
        quantity_gap = row["total_quantity"] - row["quantity_ordered"]
        row["quantity_pending"] = quantity_gap if quantity_gap > 0 else 0
        row["quantity_excess"] = abs(quantity_gap) if quantity_gap < 0 else 0
        row["source_items_display"] = ", ".join(sorted(row["source_items"]))
        row["beneficiaries"] = sorted(
            row["beneficiaries"],
            key=lambda item: (
                item["beneficiary_type"],
                item.get("created_at") or timezone.now(),
                item["application_number"],
                item["beneficiary_name"].casefold(),
                item["source_item"].casefold(),
            ),
        )
        allocation_status_priority = {
            models.BeneficiaryStatusChoices.SUBMITTED: 0,
            models.BeneficiaryStatusChoices.DRAFT: 1,
        }
        beneficiaries_for_allocation = sorted(
            row["beneficiaries"],
            key=lambda item: (
                allocation_status_priority.get(item.get("application_status") or item.get("status") or "", 99),
                item["beneficiary_type"],
                item.get("created_at") or timezone.now(),
                item["application_number"],
                item["beneficiary_name"].casefold(),
                item["source_item"].casefold(),
            ),
        )
        remaining_ordered = int(row["quantity_ordered"] or 0)
        for item in beneficiaries_for_allocation:
            quantity = int(item.get("quantity") or 0)
            source_status = item.get("status") or ""
            item["application_status"] = source_status or ""
            if item.get("item_type") == models.ItemTypeChoices.AID:
                linked_status = item.get("linked_fund_request_status") or ""
                item["ordered_quantity"] = quantity if linked_status == models.FundRequestStatusChoices.SUBMITTED else 0
                item["order_status"] = "Fund Raised" if item["ordered_quantity"] >= quantity and quantity > 0 else "No"
                continue
            if remaining_ordered <= 0:
                item["ordered_quantity"] = 0
                item["order_status"] = "No"
                continue
            if remaining_ordered >= quantity:
                item["ordered_quantity"] = quantity
                item["order_status"] = "Fund Raised"
                remaining_ordered -= quantity
            else:
                item["ordered_quantity"] = remaining_ordered
                item["order_status"] = "Partial"
                remaining_ordered = 0
        row["order_statuses"] = {item.get("order_status") or "" for item in row["beneficiaries"]}
        row["beneficiary_names_display"] = ", ".join(
            f'{item["beneficiary_type"]}: {item["beneficiary_name"]}' for item in row["beneficiaries"]
        )
        row["row_id"] = f"order-row-{index}"
        rows.append(row)
    return rows


def _order_management_metrics(rows):
    metrics = []
    for item_type in models.ItemTypeChoices:
        typed_rows = [row for row in rows if row["item_type"] == item_type.value]
        metrics.append(
            {
                "label": item_type.label,
                "count": len(typed_rows),
                "needed": sum(row["total_quantity"] for row in typed_rows),
                "ordered": sum(row["quantity_ordered"] for row in typed_rows),
                "pending": sum(row["quantity_pending"] for row in typed_rows),
                "excess": sum(row["quantity_excess"] for row in typed_rows),
            }
        )
    return metrics


def _partition_order_management_rows_by_application_status(rows, application_status_filter):
    if not application_status_filter:
        return rows

    status_priority = [
        models.BeneficiaryStatusChoices.SUBMITTED,
        models.BeneficiaryStatusChoices.DRAFT,
    ]

    partitioned_rows = []
    for row in rows:
        all_beneficiaries = row["beneficiaries"]
        beneficiaries = [
            item for item in all_beneficiaries
            if (item.get("application_status") or item.get("status") or "") == application_status_filter
        ]
        if not beneficiaries:
            continue

        filtered_row = dict(row)
        filtered_row["beneficiaries"] = beneficiaries
        filtered_row["breakdown"] = {"district": 0, "public": 0, "institutions": 0}
        filtered_row["total_quantity"] = 0
        filtered_row["quantity_ordered"] = 0
        filtered_row["quantity_pending"] = 0
        filtered_row["quantity_excess"] = 0
        filtered_row["statuses"] = {application_status_filter}
        filtered_row["source_status"] = application_status_filter
        filtered_row["order_statuses"] = {item.get("order_status") or "" for item in beneficiaries}

        quantities_by_status = {status: 0 for status in status_priority}
        for item in all_beneficiaries:
            item_status = (item.get("application_status") or item.get("status") or "").strip()
            if item_status in quantities_by_status:
                quantities_by_status[item_status] += int(item.get("quantity") or 0)

        filtered_total_quantity = sum(int(item.get("quantity") or 0) for item in beneficiaries)
        ordered_pool = int(row.get("quantity_ordered") or 0)

        if application_status_filter == models.BeneficiaryStatusChoices.SUBMITTED:
            ordered_for_filtered = min(ordered_pool, filtered_total_quantity)
        elif application_status_filter == models.BeneficiaryStatusChoices.DRAFT:
            ordered_after_submitted = max(
                ordered_pool - quantities_by_status.get(models.BeneficiaryStatusChoices.SUBMITTED, 0),
                0,
            )
            ordered_for_filtered = min(ordered_after_submitted, filtered_total_quantity)
        else:
            ordered_reserved_for_prior = 0
            for status in status_priority:
                if status == application_status_filter:
                    break
                ordered_reserved_for_prior += quantities_by_status.get(status, 0)
            ordered_for_filtered = min(max(ordered_pool - ordered_reserved_for_prior, 0), filtered_total_quantity)

        filtered_row["quantity_ordered"] = ordered_for_filtered
        filtered_row["quantity_pending"] = max(filtered_total_quantity - ordered_for_filtered, 0)

        for item in beneficiaries:
            quantity = int(item.get("quantity") or 0)
            filtered_row["total_quantity"] += quantity
            if item.get("beneficiary_type") == "District":
                filtered_row["breakdown"]["district"] += quantity
            elif item.get("beneficiary_type") == "Public":
                filtered_row["breakdown"]["public"] += quantity
            else:
                filtered_row["breakdown"]["institutions"] += quantity
        filtered_row["beneficiary_names_display"] = ", ".join(
            f'{item["beneficiary_type"]}: {item["beneficiary_name"]}' for item in beneficiaries
        )
        partitioned_rows.append(filtered_row)

    return partitioned_rows


def _partition_order_management_rows_by_order_status(rows, order_status_filter):
    if not order_status_filter:
        return rows

    partitioned_rows = []
    for row in rows:
        beneficiaries = [
            item for item in row["beneficiaries"]
            if (item.get("order_status") or "") == order_status_filter
        ]
        if not beneficiaries:
            continue

        filtered_row = dict(row)
        filtered_row["beneficiaries"] = beneficiaries
        filtered_row["breakdown"] = {"district": 0, "public": 0, "institutions": 0}
        filtered_row["total_quantity"] = 0
        filtered_row["quantity_ordered"] = 0
        filtered_row["quantity_pending"] = 0
        filtered_row["quantity_excess"] = 0
        filtered_row["order_statuses"] = {order_status_filter}

        statuses = {item.get("status") or "" for item in beneficiaries if item.get("status")}
        filtered_row["statuses"] = statuses
        if models.BeneficiaryStatusChoices.SUBMITTED in statuses:
            filtered_row["source_status"] = models.BeneficiaryStatusChoices.SUBMITTED
        elif models.BeneficiaryStatusChoices.DRAFT in statuses:
            filtered_row["source_status"] = models.BeneficiaryStatusChoices.DRAFT
        else:
            filtered_row["source_status"] = ""

        for item in beneficiaries:
            quantity = int(item.get("quantity") or 0)
            filtered_row["total_quantity"] += quantity
            if item.get("beneficiary_type") == "District":
                filtered_row["breakdown"]["district"] += quantity
            elif item.get("beneficiary_type") == "Public":
                filtered_row["breakdown"]["public"] += quantity
            else:
                filtered_row["breakdown"]["institutions"] += quantity

            ordered_quantity = int(item.get("ordered_quantity") or 0)
            if order_status_filter == "Fund Raised":
                filtered_row["quantity_ordered"] += ordered_quantity
            elif order_status_filter == "Partial":
                filtered_row["quantity_ordered"] += ordered_quantity
                filtered_row["quantity_pending"] += max(quantity - ordered_quantity, 0)
            else:
                filtered_row["quantity_pending"] += quantity

        filtered_row["beneficiary_names_display"] = ", ".join(
            f'{item["beneficiary_type"]}: {item["beneficiary_name"]}' for item in beneficiaries
        )
        partitioned_rows.append(filtered_row)

    return partitioned_rows


class OrderManagementView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.INVENTORY_PLANNING
    permission_action = "view"
    template_name = "dashboard/order_management.html"

    def get(self, request, *args, **kwargs):
        if (request.GET.get("export") or "").strip().lower() == "csv":
            return self._export_csv()
        return super().get(request, *args, **kwargs)

    def _get_filtered_rows(self):
        rows = _build_order_management_rows()
        status_filter = self.request.GET.get("status")
        if status_filter is None:
            status_filter = ""
        status_filter = status_filter.strip()
        order_status_filter = (self.request.GET.get("order_status") or "").strip()
        combo_filter = (self.request.GET.get("combo") or "").strip()
        balance_filter = (self.request.GET.get("balance") or "").strip()
        q = (self.request.GET.get("q") or "").strip().casefold()
        item_type = (self.request.GET.get("item_type") or "").strip()
        rows = _partition_order_management_rows_by_application_status(rows, status_filter)
        rows = _partition_order_management_rows_by_order_status(rows, order_status_filter)
        if combo_filter == "combo":
            rows = [row for row in rows if row["combo_related"]]
        elif combo_filter == "separate":
            rows = [row for row in rows if not row["combo_related"]]
        if q:
            rows = [
                row for row in rows
                if q in row["article_name"].casefold()
                or q in row["source_items_display"].casefold()
                or q in row["beneficiary_names_display"].casefold()
                or any(
                    q in item["application_number"].casefold()
                    or q in item["beneficiary_name"].casefold()
                    for item in row["beneficiaries"]
                )
            ]
        if item_type:
            rows = [row for row in rows if row["item_type"] == item_type]
        if balance_filter == "pending":
            rows = [row for row in rows if row["quantity_pending"] > 0]
        elif balance_filter == "excess":
            rows = [row for row in rows if row["quantity_excess"] > 0]

        sort = (self.request.GET.get("sort") or "article_name").strip()
        direction = (self.request.GET.get("dir") or "asc").strip().lower()
        allowed_sorts = {
            "article_name": lambda row: row["article_name"].casefold(),
            "total_quantity": lambda row: row["total_quantity"],
            "quantity_ordered": lambda row: row["quantity_ordered"],
            "quantity_pending": lambda row: row["quantity_pending"],
            "district": lambda row: row["breakdown"]["district"],
            "public": lambda row: row["breakdown"]["public"],
            "institutions": lambda row: row["breakdown"]["institutions"],
        }
        sort_key = allowed_sorts.get(sort, allowed_sorts["article_name"])
        rows = sorted(rows, key=sort_key, reverse=(direction == "desc"))
        previous_combo_related = None
        for index, row in enumerate(rows):
            row["group_break_before"] = index > 0 and previous_combo_related != row["combo_related"]
            previous_combo_related = row["combo_related"]
        return rows

    def _query_string_without_page(self):
        params = self.request.GET.copy()
        return params.urlencode()

    def _query_string_without(self, *keys):
        params = self.request.GET.copy()
        for key in keys:
            if key in params:
                del params[key]
        return params.urlencode()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_rows = _build_order_management_rows()
        rows = self._get_filtered_rows()
        article_rows = [row for row in all_rows if row["item_type"] == models.ItemTypeChoices.ARTICLE]
        aid_rows = [row for row in all_rows if row["item_type"] == models.ItemTypeChoices.AID]
        project_rows = [row for row in all_rows if row["item_type"] == models.ItemTypeChoices.PROJECT]
        context.update(
            {
                "rows": rows,
                "metrics": _order_management_metrics(all_rows),
                "item_type_choices": models.ItemTypeChoices.choices,
                "status_choices": [
                    ("", "All"),
                    (models.BeneficiaryStatusChoices.DRAFT, "Draft"),
                    (models.BeneficiaryStatusChoices.SUBMITTED, "Submitted"),
                ],
                "combo_choices": [
                    ("", "All"),
                    ("separate", "Separate"),
                    ("combo", "Combo / Split"),
                ],
                "order_status_choices": [
                    ("", "All"),
                    ("Fund Raised", "Fund Raised"),
                    ("Partial", "Partial"),
                    ("No", "No"),
                ],
                "balance_choices": [
                    ("", "All"),
                    ("pending", "Pending"),
                    ("excess", "Excess"),
                ],
                "filters": {
                    "q": self.request.GET.get("q", ""),
                    "item_type": self.request.GET.get("item_type", ""),
                    "status": (self.request.GET.get("status") if self.request.GET.get("status") is not None else ""),
                    "order_status": self.request.GET.get("order_status", ""),
                    "combo": self.request.GET.get("combo", ""),
                    "balance": self.request.GET.get("balance", ""),
                },
                "current_sort": (self.request.GET.get("sort") or "article_name").strip(),
                "current_dir": (self.request.GET.get("dir") or "asc").strip().lower(),
                "query_string_without_page": self._query_string_without_page(),
                "query_string_without_sort": self._query_string_without("sort", "dir", "export"),
                "article_count": len(article_rows),
                "aid_count": len(aid_rows),
                "project_count": len(project_rows),
                "all_item_count": len(all_rows),
                "filtered_item_count": len(rows),
            }
        )
        return context

    def _export_csv(self):
        rows = self._get_filtered_rows()
        export_mode = (self.request.GET.get("export_mode") or "overview").strip().lower()
        if export_mode not in {"overview", "complete"}:
            export_mode = "overview"
        timestamp = timezone.localtime().strftime("%Y_%m_%d_%I_%M_%p")
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="inventory-planning-{export_mode}_{timestamp}.csv"'
        writer = csv.writer(response)
        if export_mode == "complete":
            writer.writerow(
                [
                    "Order Item",
                    "Item Type",
                    "Combo / Separate",
                    "Application Status Filtered",
                    "Source Items",
                    "Total Quantity Needed",
                    "Quantity Ordered",
                    "Quantity Pending",
                    "Quantity Excess",
                    "District",
                    "Public",
                    "Institutions & Others",
                    "Beneficiary Type",
                    "Application Number",
                    "Beneficiary",
                    "Beneficiary Quantity",
                    "Application Status",
                    "Order Status",
                    "Source Item",
                    "Notes",
                ]
            )
            for row in rows:
                beneficiaries = row["beneficiaries"] or [None]
                for item in beneficiaries:
                    writer.writerow(
                        [
                            row["article_name"],
                            row["item_type"],
                            "Combo / Split" if row["combo_related"] else "Separate",
                            row.get("source_status", ""),
                            row["source_items_display"],
                            row["total_quantity"],
                            row["quantity_ordered"],
                            row["quantity_pending"],
                            row["quantity_excess"],
                            row["breakdown"]["district"],
                            row["breakdown"]["public"],
                            row["breakdown"]["institutions"],
                            item["beneficiary_type"] if item else "",
                            item["application_number"] if item else "",
                            item["beneficiary_name"] if item else "",
                            item["quantity"] if item else "",
                            (item["application_status"].title() if item and item.get("application_status") else ""),
                            item["order_status"] if item else "",
                            item["source_item"] if item else "",
                            item["notes"] if item else "",
                        ]
                    )
        else:
            writer.writerow(
                [
                    "Order Item",
                    "Item Type",
                    "Combo / Separate",
                    "Application Status Filtered",
                    "Source Items",
                    "Total Quantity Needed",
                    "Quantity Ordered",
                    "Quantity Pending",
                    "Quantity Excess",
                    "District",
                    "Public",
                    "Institutions & Others",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row["article_name"],
                        row["item_type"],
                        "Combo / Split" if row["combo_related"] else "Separate",
                        row.get("source_status", ""),
                        row["source_items_display"],
                        row["total_quantity"],
                        row["quantity_ordered"],
                        row["quantity_pending"],
                        row["quantity_excess"],
                        row["breakdown"]["district"],
                        row["breakdown"]["public"],
                        row["breakdown"]["institutions"],
                    ]
                )
        return response


def _is_editable_by_user(user, fr):
    if not user or not user.is_authenticated:
        return False
    if user.role not in {"admin", "editor"}:
        return False
    return fr.status == models.FundRequestStatusChoices.DRAFT


def _is_editable_purchase_order(user, purchase_order):
    if not user or not user.is_authenticated:
        return False
    if user.role not in {"admin", "editor"}:
        return False
    return purchase_order.status == models.FundRequestStatusChoices.DRAFT


def _fund_request_recipient_display_name(recipient) -> str:
    source_entry_id = getattr(recipient, "source_entry_id", None)
    beneficiary_type = getattr(recipient, "beneficiary_type", None)
    source_entry = None
    if source_entry_id and beneficiary_type:
        if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
            source_entry = models.DistrictBeneficiaryEntry.objects.select_related("district").filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                district_name = str(getattr(source_entry.district, "district_name", "") or "").strip()
                if application_number and district_name:
                    return f"{application_number} - {district_name}"
                return district_name or application_number or "-"
        elif beneficiary_type == models.RecipientTypeChoices.PUBLIC:
            source_entry = models.PublicBeneficiaryEntry.objects.filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                public_name = str(source_entry.name or "").strip()
                if application_number and public_name:
                    return f"{application_number} - {public_name}"
                return public_name or application_number or "-"
        elif beneficiary_type in {
            models.RecipientTypeChoices.INSTITUTIONS,
            models.RecipientTypeChoices.OTHERS,
        }:
            source_entry = models.InstitutionsBeneficiaryEntry.objects.filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                institution_name = str(source_entry.institution_name or "").strip()
                if application_number and institution_name:
                    return f"{application_number} - {institution_name}"
                return institution_name or application_number or "-"
    return (
        str(getattr(recipient, "recipient_name", "") or "").strip()
        or str(getattr(recipient, "name_of_beneficiary", "") or "").strip()
        or str(getattr(recipient, "name_of_institution", "") or "").strip()
        or str(getattr(recipient, "beneficiary", "") or "").strip()
        or "-"
    )


def _fund_request_article_beneficiary_display(article_item) -> str:
    article = getattr(article_item, "article", None)
    if not article:
        return str(getattr(article_item, "beneficiary", "") or "").strip() or "-"

    labels = []
    if article.district_entries.exists():
        labels.append("District")
    if article.public_entries.exists():
        labels.append("Public")
    if article.institution_entries.exists():
        labels.append("Institutions")

    if len(labels) == 3:
        return "All beneficiaries"
    if labels:
        return " & ".join(labels)
    return str(getattr(article_item, "beneficiary", "") or "").strip() or "-"


def _normalize_vendor_group_payload(raw_group):
    if not isinstance(raw_group, dict):
        return None
    key = str(raw_group.get("key") or "").strip()
    if not key:
        return None
    payload = {
        "key": key,
        "vendor_name": str(raw_group.get("vendor_name") or "").strip(),
        "gst_no": str(raw_group.get("gst_no") or "").strip(),
        "vendor_address": str(raw_group.get("vendor_address") or "").strip(),
        "vendor_city": str(raw_group.get("vendor_city") or "").strip(),
        "vendor_state": str(raw_group.get("vendor_state") or "").strip(),
        "vendor_pincode": str(raw_group.get("vendor_pincode") or "").strip(),
        "cheque_in_favour": str(raw_group.get("cheque_in_favour") or "").strip(),
    }
    return payload


def _build_vendor_groups_from_articles(article_rows):
    groups = []
    row_key_map = {}
    seen = {}
    for article in article_rows:
        signature = (
            str(article.vendor_name or "").strip(),
            str(article.gst_no or "").strip(),
            str(article.vendor_address or "").strip(),
            str(article.vendor_city or "").strip(),
            str(article.vendor_state or "").strip(),
            str(article.vendor_pincode or "").strip(),
            str(article.cheque_in_favour or "").strip(),
        )
        if not any(signature):
            continue
        key = seen.get(signature)
        if not key:
            key = f"vendor-{len(groups) + 1}"
            seen[signature] = key
            groups.append(
                {
                    "key": key,
                    "vendor_name": signature[0],
                    "gst_no": signature[1],
                    "vendor_address": signature[2],
                    "vendor_city": signature[3],
                    "vendor_state": signature[4],
                    "vendor_pincode": signature[5],
                    "cheque_in_favour": signature[6],
                }
            )
        if getattr(article, "pk", None):
            row_key_map[str(article.pk)] = key
    return groups, row_key_map


FundRequestRecipientFormSet = inlineformset_factory(
    models.FundRequest,
    models.FundRequestRecipient,
    form=FundRequestRecipientForm,
    extra=0,
    can_delete=True,
)

FundRequestArticleFormSet = inlineformset_factory(
    models.FundRequest,
    models.FundRequestArticle,
    form=FundRequestArticleForm,
    extra=0,
    can_delete=True,
)

PurchaseOrderItemFormSet = inlineformset_factory(
    models.PurchaseOrder,
    models.PurchaseOrderItem,
    form=PurchaseOrderItemForm,
    extra=0,
    can_delete=True,
)


class FundRequestListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"
    model = models.FundRequest
    template_name = "dashboard/fund_request_list.html"
    context_object_name = "fund_requests"
    paginate_by = 20

    def get(self, request, *args, **kwargs):
        if (request.GET.get("export") or "").strip().lower() in {"xlsx", "export"}:
            return self._export_xlsx()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        queryset = (
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles", "documents")
        )
        self._sort_key = (self.request.GET.get("sort") or "created_at").strip()
        self._sort_dir = (self.request.GET.get("dir") or "desc").strip().lower()
        if q := (self.request.GET.get("q") or "").strip():
            fund_request_number_q = Q(fund_request_number__icontains=q)
            parsed_request_number = models.parse_fund_request_sequence(q)
            if parsed_request_number is not None:
                alternate_terms = {
                    models.format_fund_request_number(f"FR-{parsed_request_number}"),
                    f"FR-{parsed_request_number}",
                    f"FR{parsed_request_number}",
                }
                for term in alternate_terms:
                    fund_request_number_q |= Q(fund_request_number__icontains=term)
            matching_ids = (
                models.FundRequest.objects.select_related("created_by")
                .annotate(
                    total_amount_text=Cast("total_amount", output_field=CharField()),
                    created_at_text=Cast("created_at", output_field=CharField()),
                    recipient_source_entry_id_text=Cast("recipients__source_entry_id", output_field=CharField()),
                    recipient_fund_requested_text=Cast("recipients__fund_requested", output_field=CharField()),
                    article_sl_no_text=Cast("articles__sl_no", output_field=CharField()),
                    article_quantity_text=Cast("articles__quantity", output_field=CharField()),
                    article_unit_price_text=Cast("articles__unit_price", output_field=CharField()),
                    article_price_including_gst_text=Cast("articles__price_including_gst", output_field=CharField()),
                    article_value_text=Cast("articles__value", output_field=CharField()),
                    article_cumulative_text=Cast("articles__cumulative", output_field=CharField()),
                    document_generated_at_text=Cast("documents__generated_at", output_field=CharField()),
                )
                .filter(
                    fund_request_number_q
                    | Q(fund_request_type__icontains=q)
                    | Q(status__icontains=q)
                    | Q(aid_type__icontains=q)
                    | Q(notes__icontains=q)
                    | Q(total_amount_text__icontains=q)
                    | Q(created_at_text__icontains=q)
                    | Q(gst_number__icontains=q)
                    | Q(supplier_name__icontains=q)
                    | Q(supplier_address__icontains=q)
                    | Q(supplier_city__icontains=q)
                    | Q(supplier_state__icontains=q)
                    | Q(supplier_pincode__icontains=q)
                    | Q(purchase_order_number__icontains=q)
                    | Q(created_by__email__icontains=q)
                    | Q(created_by__first_name__icontains=q)
                    | Q(created_by__last_name__icontains=q)
                    | Q(recipients__recipient_name__icontains=q)
                    | Q(recipients__name_of_beneficiary__icontains=q)
                    | Q(recipients__name_of_institution__icontains=q)
                    | Q(recipients__beneficiary_type__icontains=q)
                    | Q(recipients__beneficiary__icontains=q)
                    | Q(recipients__details__icontains=q)
                    | Q(recipients__address__icontains=q)
                    | Q(recipients__cheque_in_favour__icontains=q)
                    | Q(recipients__cheque_no__icontains=q)
                    | Q(recipients__notes__icontains=q)
                    | Q(recipients__district_name__icontains=q)
                    | Q(recipient_source_entry_id_text__icontains=q)
                    | Q(recipient_fund_requested_text__icontains=q)
                    | Q(recipients__aadhar_number__icontains=q)
                    | Q(articles__article_name__icontains=q)
                    | Q(articles__beneficiary__icontains=q)
                    | Q(articles__vendor_name__icontains=q)
                    | Q(articles__gst_no__icontains=q)
                    | Q(articles__vendor_address__icontains=q)
                    | Q(articles__vendor_city__icontains=q)
                    | Q(articles__vendor_state__icontains=q)
                    | Q(articles__vendor_pincode__icontains=q)
                    | Q(articles__cheque_in_favour__icontains=q)
                    | Q(articles__cheque_no__icontains=q)
                    | Q(articles__supplier_article_name__icontains=q)
                    | Q(articles__description__icontains=q)
                    | Q(article_sl_no_text__icontains=q)
                    | Q(article_quantity_text__icontains=q)
                    | Q(article_unit_price_text__icontains=q)
                    | Q(article_price_including_gst_text__icontains=q)
                    | Q(article_value_text__icontains=q)
                    | Q(article_cumulative_text__icontains=q)
                    | Q(documents__file_name__icontains=q)
                    | Q(documents__file_path__icontains=q)
                    | Q(documents__document_type__icontains=q)
                    | Q(document_generated_at_text__icontains=q)
                    | Q(documents__generated_by__email__icontains=q)
                    | Q(documents__generated_by__first_name__icontains=q)
                    | Q(documents__generated_by__last_name__icontains=q)
                )
                .values_list("pk", flat=True)
                .distinct()
            )
            queryset = queryset.filter(pk__in=matching_ids)
        if request_type := (self.request.GET.get("request_type") or "").strip():
            queryset = queryset.filter(fund_request_type=request_type)
        if status := (self.request.GET.get("status") or "").strip():
            queryset = queryset.filter(status=status)
        if supplier := (self.request.GET.get("supplier") or "").strip():
            queryset = queryset.filter(supplier_name__icontains=supplier)
        sort_fields = {
            "fund_request_number": "fund_request_number",
            "fund_request_type": "fund_request_type",
            "item_type": "aid_type",
            "total_amount": "total_amount",
            "status": "status",
            "created_at": "created_at",
            "supplier_name": "supplier_name",
        }
        sort_field = sort_fields.get(self._sort_key, "created_at")
        sort_prefix = "" if self._sort_dir == "asc" else "-"
        if sort_field == "fund_request_number":
            queryset = queryset.order_by(f"{sort_prefix}created_at", f"{sort_prefix}id")
            queryset = sorted(
                queryset,
                key=lambda fr: (
                    models.parse_fund_request_sequence(fr.fund_request_number) is None,
                    models.parse_fund_request_sequence(fr.fund_request_number) or 0,
                    fr.created_at,
                    fr.id,
                ),
                reverse=(self._sort_dir == "desc"),
            )
            return queryset
        queryset = queryset.order_by(f"{sort_prefix}{sort_field}", "-id")
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for fr in context["fund_requests"]:
            recipients = list(fr.recipients.all())
            articles = list(fr.articles.all())
            for recipient in recipients:
                recipient.display_name = _fund_request_recipient_display_name(recipient)
                recipient.beneficiary_display = _fund_request_recipient_display_name(recipient)
            fr.district_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT
            )
            fr.public_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC
            )
            fr.institutions_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.INSTITUTIONS
            )
            fr.others_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.OTHERS
            )
            fr.article_total_quantity = sum(int(article.quantity or 0) for article in articles)
        current_sort = getattr(self, "_sort_key", "created_at")
        current_dir = getattr(self, "_sort_dir", "desc")
        def build_sort_params(column):
            params = self.request.GET.copy()
            params.pop("page", None)
            next_dir = "asc"
            if current_sort == column and current_dir == "asc":
                next_dir = "desc"
            params["sort"] = column
            params["dir"] = next_dir
            return params.urlencode()
        context["request_type_choices"] = models.FundRequestTypeChoices.choices
        context["status_choices"] = [
            (models.FundRequestStatusChoices.DRAFT, "Draft"),
            (models.FundRequestStatusChoices.SUBMITTED, "Submitted"),
        ]
        context["filters"] = {
            "q": self.request.GET.get("q", ""),
            "request_type": self.request.GET.get("request_type", ""),
            "status": self.request.GET.get("status", ""),
            "supplier": self.request.GET.get("supplier", ""),
        }
        context["current_sort"] = current_sort
        context["current_dir"] = current_dir
        context["sort_querystrings"] = {
            "fund_request_number": build_sort_params("fund_request_number"),
            "fund_request_type": build_sort_params("fund_request_type"),
            "item_type": build_sort_params("item_type"),
            "total_amount": build_sort_params("total_amount"),
            "status": build_sort_params("status"),
            "created_at": build_sort_params("created_at"),
        }
        return context

    def _event_birthday_number(self, event_year: int) -> int:
        return max(event_year - 1940, 1)

    def _beneficiary_display_for_export(self, recipient, fund_request_type):
        if fund_request_type == models.FundRequestTypeChoices.ARTICLE:
            return ""
        if not recipient:
            return ""
        return _fund_request_recipient_display_name(recipient)

    def _export_xlsx(self):
        export_status = (self.request.GET.get("export_status") or "").strip().lower()
        queryset = list(
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles")
            .order_by("fund_request_number", "created_at", "id")
        )
        if q := (self.request.GET.get("q") or "").strip():
            queryset = [
                fr for fr in queryset
                if q.lower() in str(fr.fund_request_number or "").lower()
                or q.lower() in str(fr.formatted_fund_request_number or "").lower()
                or q.lower() in str(fr.aid_type or "").lower()
                or q.lower() in str(fr.supplier_name or "").lower()
            ]
        if request_type := (self.request.GET.get("request_type") or "").strip():
            queryset = [fr for fr in queryset if fr.fund_request_type == request_type]
        if export_status in {
            models.FundRequestStatusChoices.DRAFT,
            models.FundRequestStatusChoices.SUBMITTED,
        }:
            queryset = [fr for fr in queryset if fr.status == export_status]

        def _fr_sort_key(fr):
            sequence = models.parse_fund_request_sequence(fr.fund_request_number)
            if sequence is not None:
                return (0, sequence)
            raw = str(fr.fund_request_number or "").strip()
            return (1, raw)

        queryset.sort(key=_fr_sort_key)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Fund Requests"

        thin = Side(style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        total_fill = PatternFill("solid", fgColor="D3D3D3")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        event_year = timezone.localdate().year
        birthday_number = self._event_birthday_number(event_year)

        current_row = 1
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = "OMSAKTHI"
        cell.font = Font(size=10, bold=True)
        cell.alignment = center
        current_row += 1

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = (
            f"MASM Makkal Nala Pani Payment Request Details for Distribution on the eve of "
            f"{birthday_number}th Birthday Celebrations of"
        )
        cell.font = Font(size=12, bold=True)
        cell.alignment = center
        current_row += 1

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = f"His Holiness AMMA at Melmaruvathur on 03.03.{event_year}"
        cell.font = Font(size=12, bold=True)
        cell.alignment = center
        current_row += 2

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        label = "Payment Request - MASTER LIST"
        if export_status == models.FundRequestStatusChoices.DRAFT:
            label = "Payment Request - DRAFT MASTER LIST"
        elif export_status == models.FundRequestStatusChoices.SUBMITTED:
            label = "Payment Request - SUBMITTED MASTER LIST"
        cell.value = label
        cell.font = Font(size=14, bold=True)
        cell.alignment = center
        current_row += 2

        headers = [
            "FUND REQ NO.",
            "Request Type",
            "Beneficiary",
            "Name of Beneficiary/Article",
            "Name of Institution/Article",
            "Vendor Name",
            "GST / Aadhaar Number",
            "Details",
            "Units",
            "Price incl GST",
            "Value",
            "Fund Request Value",
            "CHEQUE (OR) RTGS IN FAVOUR",
            "CHEQUE NO.",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(current_row, idx)
            cell.value = header
            cell.font = Font(size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        current_row += 1

        all_rows = []
        fr_value_map = {}
        for fr in queryset:
            if fr.fund_request_type == models.FundRequestTypeChoices.AID:
                fr_total = Decimal("0")
                for recipient in fr.recipients.all():
                    amount = Decimal(str(recipient.fund_requested or 0))
                    fr_total += amount
                    all_rows.append(
                        {
                            "fr_id": fr.id,
                            "fund_request_number": fr.formatted_fund_request_number or "",
                            "request_type": fr.aid_type or "Aid",
                            "beneficiary": self._beneficiary_display_for_export(recipient, fr.fund_request_type),
                            "name_beneficiary_article": recipient.recipient_name or recipient.name_of_beneficiary or "",
                            "name_institution_article": recipient.name_of_institution or "",
                            "vendor_name": "",
                            "gst_aadhar": recipient.aadhar_number or "",
                            "details": recipient.notes or recipient.details or "",
                            "units": 1,
                            "price_incl_gst": float(amount),
                            "value": float(amount),
                            "fund_request_value": 0,
                            "cheque_in_favour": recipient.cheque_in_favour or "",
                            "cheque_no": recipient.cheque_no or "",
                        }
                    )
                fr_value_map[fr.id] = float(fr_total)
            else:
                fr_total = Decimal("0")
                for article in fr.articles.all():
                    line_value = Decimal(str(article.value or 0))
                    fr_total += line_value
                    all_rows.append(
                        {
                            "fr_id": fr.id,
                            "fund_request_number": fr.formatted_fund_request_number or "",
                            "request_type": "Article",
                            "beneficiary": _fund_request_article_beneficiary_display(article),
                            "name_beneficiary_article": article.article_name or "",
                            "name_institution_article": article.supplier_article_name or "",
                            "vendor_name": article.vendor_name or "",
                            "gst_aadhar": article.gst_no or fr.gst_number or "",
                            "details": "",
                            "units": article.quantity or 0,
                            "price_incl_gst": float(article.price_including_gst or article.unit_price or 0),
                            "value": float(line_value),
                            "fund_request_value": 0,
                            "cheque_in_favour": article.cheque_in_favour or "",
                            "cheque_no": article.cheque_no or "",
                        }
                    )
                fr_value_map[fr.id] = float(fr_total)

        for row in all_rows:
            row["fund_request_value"] = fr_value_map.get(row["fr_id"], 0)

        fr_groups = {}
        current_fr_id = None
        group_start = current_row
        for row in all_rows:
            if row["fr_id"] != current_fr_id:
                if current_fr_id is not None:
                    fr_groups[current_fr_id] = (group_start, current_row - 1)
                current_fr_id = row["fr_id"]
                group_start = current_row

            values = [
                row["fund_request_number"],
                row["request_type"],
                row["beneficiary"],
                row["name_beneficiary_article"],
                row["name_institution_article"],
                row["vendor_name"],
                row["gst_aadhar"],
                row["details"],
                row["units"],
                row["price_incl_gst"],
                row["value"],
                row["fund_request_value"],
                row["cheque_in_favour"],
                row["cheque_no"],
            ]
            for idx, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, idx)
                cell.value = value
                cell.border = border
                cell.alignment = right if idx in {9, 10, 11, 12} else left
            current_row += 1
        if current_fr_id is not None:
            fr_groups[current_fr_id] = (group_start, current_row - 1)

        for _fr_id, (start_row, end_row) in fr_groups.items():
            if end_row > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=12, end_row=end_row, end_column=12)
                worksheet.cell(start_row, 12).alignment = right

        grand_total = sum(fr_value_map.values())
        total_row = current_row
        for col in range(1, 15):
            cell = worksheet.cell(total_row, col)
            cell.fill = total_fill
            cell.border = border
        worksheet.cell(total_row, 1).value = "TOTAL"
        worksheet.cell(total_row, 1).font = Font(size=11, bold=True)
        worksheet.cell(total_row, 1).alignment = left
        worksheet.cell(total_row, 12).value = grand_total
        worksheet.cell(total_row, 12).font = Font(size=11, bold=True)
        worksheet.cell(total_row, 12).alignment = right

        widths = {
            1: 18, 2: 18, 3: 20, 4: 24, 5: 24, 6: 22, 7: 18, 8: 28,
            9: 12, 10: 15, 11: 15, 12: 18, 13: 25, 14: 15,
        }
        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        date_stamp = timezone.localtime().strftime("%Y-%m-%d")
        suffix = export_status or "all"
        if export_status == models.FundRequestStatusChoices.SUBMITTED:
            filename = f"Fund_Request_Sub_{date_stamp}.xlsx"
        elif export_status == models.FundRequestStatusChoices.DRAFT:
            filename = f"Fund_Request_dft_{date_stamp}.xlsx"
        else:
            filename = f"Fund_Request_{suffix}_{date_stamp}.xlsx"
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response



def _fund_request_aid_type_choices():
    names = set()
    names.update(
        filter(
            None,
            models.DistrictBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list('article__article_name', flat=True),
        )
    )
    names.update(
        filter(
            None,
            models.PublicBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list('article__article_name', flat=True),
        )
    )
    names.update(
        filter(
            None,
            models.InstitutionsBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list('article__article_name', flat=True),
        )
    )
    return sorted(names)


def _fund_request_article_choices(current_fund_request=None):
    all_rows = [
        row
        for row in _build_order_management_rows()
        if row['item_type'] == models.ItemTypeChoices.ARTICLE
    ]
    rows_by_name = {
        str(row['article_name']).casefold(): dict(row)
        for row in all_rows
    }
    rows = [row for row in rows_by_name.values() if row['quantity_pending'] > 0]
    article_map = {
        article.article_name.casefold(): article
        for article in models.Article.objects.filter(item_type=models.ItemTypeChoices.ARTICLE)
    }
    choices = []
    for row in rows:
        article = article_map.get(row['article_name'].casefold())
        choices.append(
            {
                'name': row['article_name'],
                'label': f"{row['article_name']} (Pending: {row['quantity_pending']})",
                'article_id': article.id if article else '',
                'default_price': str(article.cost_per_unit if article else 0),
                'pending_qty': int(row['quantity_pending'] or 0),
            }
        )
    return choices


def _aid_entry_queryset(aid_type, beneficiary_type):
    filters = {
        'article__item_type': models.ItemTypeChoices.AID,
        'article__article_name__iexact': aid_type,
    }
    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        return models.DistrictBeneficiaryEntry.objects.select_related('district', 'article', 'fund_request').filter(**filters)
    if beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        return models.PublicBeneficiaryEntry.objects.select_related('article', 'fund_request').filter(**filters)
    return models.InstitutionsBeneficiaryEntry.objects.select_related('article', 'fund_request').filter(**filters)


def _build_aid_option_payload(entry, beneficiary_type):
    amount = float(entry.total_amount or 0)
    amount_display = format(amount, ".2f").rstrip("0").rstrip(".")
    details_display = (entry.notes or "").strip() or "-"
    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        beneficiary_name = entry.district.district_name
        return {
            'source_entry_id': entry.pk,
            'application_number': entry.application_number or '',
            'display_text': f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
            'recipient_name': beneficiary_name,
            'name_of_beneficiary': entry.name_of_beneficiary or '',
            'name_of_institution': entry.name_of_institution or '',
            'details': entry.notes or '',
            'fund_requested': amount,
            'aadhar_number': entry.aadhar_number or '',
            'cheque_in_favour': entry.cheque_rtgs_in_favour or '',
            'district_name': beneficiary_name,
            'source_item': entry.article.article_name,
        }
    if beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        beneficiary_name = entry.name
        return {
            'source_entry_id': entry.pk,
            'application_number': entry.application_number or '',
            'display_text': f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
            'recipient_name': beneficiary_name,
            'name_of_beneficiary': beneficiary_name,
            'name_of_institution': entry.name_of_institution or '',
            'details': entry.notes or '',
            'fund_requested': amount,
            'aadhar_number': entry.aadhar_number or '',
            'cheque_in_favour': entry.cheque_rtgs_in_favour or '',
            'district_name': '',
            'source_item': entry.article.article_name,
        }
    beneficiary_name = entry.institution_name
    return {
        'source_entry_id': entry.pk,
        'application_number': entry.application_number or '',
        'display_text': f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
        'recipient_name': beneficiary_name,
        'name_of_beneficiary': entry.name_of_beneficiary or '',
        'name_of_institution': entry.name_of_institution or beneficiary_name,
        'details': entry.notes or '',
        'fund_requested': amount,
        'aadhar_number': entry.aadhar_number or '',
        'cheque_in_favour': entry.cheque_rtgs_in_favour or '',
        'district_name': '',
        'source_item': entry.article.article_name,
    }


def _get_aid_beneficiary_options(aid_type, beneficiary_type, current_fund_request=None):
    aid_type = (aid_type or '').strip()
    blocked = set()
    options = []
    if not aid_type or beneficiary_type not in {
        models.RecipientTypeChoices.DISTRICT,
        models.RecipientTypeChoices.PUBLIC,
        models.RecipientTypeChoices.INSTITUTIONS,
    }:
        return options, []

    for entry in _aid_entry_queryset(aid_type, beneficiary_type).order_by('application_number', 'created_at'):
        if entry.fund_request_id and (not current_fund_request or entry.fund_request_id != current_fund_request.id):
            label = entry.fund_request.formatted_fund_request_number if entry.fund_request and entry.fund_request.fund_request_number else f'Draft #{entry.fund_request_id}'
            if entry.fund_request:
                blocked.add(f"{label} ({entry.fund_request.get_status_display()})")
            else:
                blocked.add(label)
            continue
        options.append(_build_aid_option_payload(entry, beneficiary_type))
    return options, sorted(blocked)


def _get_aid_available_beneficiary_type_choices(aid_type, current_fund_request=None):
    results = []
    for value, label in [
        (models.RecipientTypeChoices.DISTRICT, 'District'),
        (models.RecipientTypeChoices.PUBLIC, 'Public'),
        (models.RecipientTypeChoices.INSTITUTIONS, 'Institutions'),
    ]:
        options, _blocked = _get_aid_beneficiary_options(aid_type, value, current_fund_request=current_fund_request)
        if options:
            results.append({'value': value, 'label': label, 'count': len(options)})
    return results


class FundRequestAidOptionsView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = 'create_edit'

    def get(self, request, *args, **kwargs):
        aid_type = request.GET.get('aid_type') or ''
        beneficiary_type = request.GET.get('beneficiary_type') or ''
        current_id = request.GET.get('fund_request_id') or ''
        current_fund_request = None
        if current_id:
            current_fund_request = models.FundRequest.objects.filter(pk=current_id).first()
        options_by_type = {}
        blocked_by_type = {}
        for type_key in [
            models.RecipientTypeChoices.DISTRICT,
            models.RecipientTypeChoices.PUBLIC,
            models.RecipientTypeChoices.INSTITUTIONS,
        ]:
            type_options, type_blocked = _get_aid_beneficiary_options(aid_type, type_key, current_fund_request=current_fund_request)
            options_by_type[str(type_key)] = type_options
            blocked_by_type[str(type_key)] = type_blocked
        payload = {
            'available_types': _get_aid_available_beneficiary_type_choices(aid_type, current_fund_request=current_fund_request),
            'options': options_by_type.get(str(beneficiary_type), []) if beneficiary_type else [],
            'blocked': blocked_by_type.get(str(beneficiary_type), []) if beneficiary_type else [],
            'options_by_type': options_by_type,
            'blocked_by_type': blocked_by_type,
        }
        return JsonResponse(payload)


class FundRequestCreateUpdateMixin(WriteRoleMixin):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = 'create_edit'
    form_class = FundRequestForm
    template_name = 'dashboard/fund_request_form.html'
    model = models.FundRequest
    success_url = reverse_lazy('ui:fund-request-list')

    def _build_formsets(self, instance: models.FundRequest | None = None):
        recipient_formset = FundRequestRecipientFormSet(self.request.POST or None, prefix='recipients', instance=instance)
        article_formset = FundRequestArticleFormSet(self.request.POST or None, prefix='articles', instance=instance)
        return recipient_formset, article_formset

    def _can_edit(self, fr: models.FundRequest):
        return _is_editable_by_user(self.request.user, fr)

    def is_purchase_order_mode(self):
        return False

    def dispatch(self, request, *args, **kwargs):
        self.object = getattr(self, 'object', None)
        if self.object and not self._can_edit(self.object):
            messages.error(request, 'Submitted fund requests must be reopened before editing.')
            return HttpResponseRedirect(reverse('ui:fund-request-detail', args=[self.object.pk]))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recipient_formset'] = kwargs.get('recipient_formset', None) or self._build_formsets(self.object)[0]
        context['article_formset'] = kwargs.get('article_formset', None) or self._build_formsets(self.object)[1]
        context['status_choices'] = models.FundRequestStatusChoices.choices
        context['aid_type_choices'] = _fund_request_aid_type_choices()
        context['article_request_choices_json'] = json.dumps(_fund_request_article_choices(self.object))
        context['current_fund_request_id'] = getattr(self.object, 'pk', '') or ''
        context['purchase_order_mode'] = self.is_purchase_order_mode()
        context['back_url'] = reverse('ui:purchase-order-list') if self.is_purchase_order_mode() else reverse('ui:fund-request-list')
        vendor_groups = []
        article_vendor_group_keys = {}
        if self.request.method == 'POST':
            raw_groups = self.request.POST.get('vendor_groups_json', '[]')
            try:
                parsed_groups = json.loads(raw_groups or '[]')
            except (TypeError, ValueError):
                parsed_groups = []
            vendor_groups = [group for group in (_normalize_vendor_group_payload(item) for item in parsed_groups) if group]
            for form in context['article_formset'].forms:
                article_vendor_group_keys[form.prefix] = str(self.request.POST.get(f'{form.prefix}-vendor_group_key') or '').strip()
        elif self.object:
            vendor_groups, row_key_map = _build_vendor_groups_from_articles(self.object.articles.all())
            for form in context['article_formset'].forms:
                instance_pk = getattr(getattr(form, 'instance', None), 'pk', None)
                article_vendor_group_keys[form.prefix] = row_key_map.get(str(instance_pk), '')
        context['vendor_groups_json'] = json.dumps(vendor_groups)
        context['article_vendor_group_keys_json'] = json.dumps(article_vendor_group_keys)
        return context

    def _parse_vendor_groups_from_request(self):
        raw_groups = self.request.POST.get('vendor_groups_json', '[]')
        try:
            parsed_groups = json.loads(raw_groups or '[]')
        except (TypeError, ValueError):
            parsed_groups = []
        return {
            group['key']: group
            for group in (_normalize_vendor_group_payload(item) for item in parsed_groups)
            if group
        }

    def _apply_article_vendor_summary(self, fr, article_formset, vendor_groups):
        if fr.fund_request_type != models.FundRequestTypeChoices.ARTICLE or self.is_purchase_order_mode():
            return
        active_keys = []
        for form in article_formset.forms:
            if not getattr(form, 'cleaned_data', None) or form.cleaned_data.get('DELETE', False):
                continue
            group_key = str(self.request.POST.get(f'{form.prefix}-vendor_group_key') or '').strip()
            if group_key and group_key in vendor_groups:
                active_keys.append(group_key)
        active_keys = list(dict.fromkeys(active_keys))
        if len(active_keys) == 1:
            group = vendor_groups[active_keys[0]]
            fr.supplier_name = group['vendor_name']
            fr.gst_number = group['gst_no']
            fr.supplier_address = group['vendor_address']
            fr.supplier_city = group['vendor_city']
            fr.supplier_state = group['vendor_state']
            fr.supplier_pincode = group['vendor_pincode']
        elif len(active_keys) > 1:
            fr.supplier_name = 'Multiple Vendors'
            fr.gst_number = ''
            fr.supplier_address = ''
            fr.supplier_city = ''
            fr.supplier_state = ''
            fr.supplier_pincode = ''

    def _collect_totals(self, instance: models.FundRequest):
        for article in instance.articles.all():
            article.recompute_totals(unit_price=article.unit_price, quantity=article.quantity)
        services.sync_fund_request_totals(instance)

    def _resolve_article_record(self, article_name: str):
        article_name = (article_name or '').strip()
        if not article_name:
            return None
        article = models.Article.objects.filter(article_name__iexact=article_name).first()
        if article:
            return article
        return models.Article.objects.create(
            article_name=article_name,
            cost_per_unit=0,
            item_type=models.ItemTypeChoices.ARTICLE,
            combo=True,
            is_active=False,
        )

    def _link_aid_sources(self, fr: models.FundRequest):
        models.DistrictBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        models.PublicBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        models.InstitutionsBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        if fr.fund_request_type != models.FundRequestTypeChoices.AID:
            return
        for recipient in fr.recipients.exclude(source_entry_id__isnull=True).exclude(source_entry_id__exact=0):
            if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT:
                models.DistrictBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)
            elif recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC:
                models.PublicBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)
            elif recipient.beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
                models.InstitutionsBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)

    def _is_aid_source_available(self, beneficiary_type, source_entry_id, current_fund_request=None):
        if not source_entry_id or not beneficiary_type:
            return False, 'Select a valid beneficiary.'
        entry = None
        if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
            entry = models.DistrictBeneficiaryEntry.objects.select_related('fund_request').filter(pk=source_entry_id).first()
        elif beneficiary_type == models.RecipientTypeChoices.PUBLIC:
            entry = models.PublicBeneficiaryEntry.objects.select_related('fund_request').filter(pk=source_entry_id).first()
        elif beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
            entry = models.InstitutionsBeneficiaryEntry.objects.select_related('fund_request').filter(pk=source_entry_id).first()
        if not entry:
            return False, 'This beneficiary is no longer available.'
        if entry.fund_request_id and (not current_fund_request or entry.fund_request_id != current_fund_request.id):
            label = entry.fund_request.formatted_fund_request_number if entry.fund_request and entry.fund_request.fund_request_number else f'Draft #{entry.fund_request_id}'
            if entry.fund_request:
                return False, f'Already present in {label} ({entry.fund_request.get_status_display()}).'
            return False, f'Already present in {label}.'
        return True, ''

    def _validate_fund_request_formsets(self, fr, action, recipient_formset, article_formset):
        is_valid = True
        if fr.fund_request_type == models.FundRequestTypeChoices.AID:
            if action == 'submit' and not (fr.aid_type or '').strip():
                return False
            active_forms = [form for form in recipient_formset.forms if form.cleaned_data and not form.cleaned_data.get('DELETE', False)]
            if action == 'submit' and not active_forms:
                recipient_formset._non_form_errors = recipient_formset.error_class(['Add at least one recipient.'])
                return False
            seen_source_keys = {}
            for form in active_forms:
                beneficiary_type = form.cleaned_data.get('beneficiary_type')
                source_entry_id = form.cleaned_data.get('source_entry_id')
                if beneficiary_type and source_entry_id:
                    source_key = (str(beneficiary_type), str(source_entry_id))
                    if source_key in seen_source_keys:
                        form.add_error('beneficiary', 'This recipient is already added in the same fund request.')
                        seen_source_keys[source_key].add_error('beneficiary', 'This recipient is already added in the same fund request.')
                        is_valid = False
                    else:
                        seen_source_keys[source_key] = form
                if source_entry_id:
                    ok, message = self._is_aid_source_available(
                        beneficiary_type,
                        source_entry_id,
                        current_fund_request=fr if getattr(fr, 'pk', None) else None,
                    )
                    if not ok:
                        form.add_error('beneficiary', message)
                        is_valid = False
                if action == 'submit':
                    required_fields = ['beneficiary_type', 'beneficiary', 'source_entry_id', 'fund_requested', 'name_of_beneficiary', 'name_of_institution', 'details', 'cheque_in_favour']
                    if beneficiary_type != models.RecipientTypeChoices.DISTRICT:
                        required_fields.append('aadhar_number')
                    for field_name in required_fields:
                        value = form.cleaned_data.get(field_name)
                        if value in (None, '', 0, '0'):
                            form.add_error(field_name, 'Required for submit.')
                            is_valid = False
        else:
            vendor_groups = self._parse_vendor_groups_from_request()
            active_forms = [form for form in article_formset.forms if form.cleaned_data and not form.cleaned_data.get('DELETE', False)]
            if action == 'submit' and not active_forms:
                article_formset._non_form_errors = article_formset.error_class(['Add at least one item.'])
                return False
            if action == 'submit':
                for form in active_forms:
                    required_fields = ['article_name', 'quantity', 'unit_price']
                    for field_name in required_fields:
                        value = form.cleaned_data.get(field_name)
                        if value in (None, '', 0, '0'):
                            form.add_error(field_name, 'Required for submit.')
                            is_valid = False
                    if not self.is_purchase_order_mode():
                        group_key = str(self.request.POST.get(f'{form.prefix}-vendor_group_key') or '').strip()
                        if not group_key or group_key not in vendor_groups:
                            form.add_error('article_name', 'Select a vendor.')
                            is_valid = False
                            continue
                        group = vendor_groups[group_key]
                        for value in [
                            group.get('vendor_name'),
                            group.get('gst_no'),
                            group.get('vendor_address'),
                            group.get('vendor_city'),
                            group.get('vendor_state'),
                            group.get('vendor_pincode'),
                            group.get('cheque_in_favour'),
                        ]:
                            if not str(value or '').strip():
                                form.add_error('article_name', 'Complete the selected vendor details.')
                                is_valid = False
                                break
        return is_valid

    def _validate_article_header_fields(self, form, fr, action):
        if action != 'submit' or fr.fund_request_type != models.FundRequestTypeChoices.ARTICLE or not self.is_purchase_order_mode():
            return True
        valid = True
        labels = [
            ('supplier_name', 'Vendor Name' if self.is_purchase_order_mode() else 'Supplier Name'),
            ('supplier_address', 'Vendor Address' if self.is_purchase_order_mode() else 'Address'),
            ('supplier_city', 'City'),
            ('supplier_state', 'State'),
            ('supplier_pincode', 'Pincode'),
        ]
        for field_name, label in labels:
            value = getattr(fr, field_name, None)
            if not str(value or '').strip():
                form.add_error(field_name, f'{label} is required for submit.')
                valid = False
        return valid

    def _set_fund_request_status(self, fr, action):
        if action == 'submit':
            fr.status = models.FundRequestStatusChoices.SUBMITTED
        else:
            fr.status = models.FundRequestStatusChoices.DRAFT

    def form_valid(self, form):
        action = self.request.POST.get('action', 'draft')
        if action == 'submit' and self.object and self.object.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(self.request, 'Only draft fund requests can be submitted.')
            return HttpResponseRedirect(self.get_success_url())

        fr = form.save(commit=False)
        if self.is_purchase_order_mode():
            fr.fund_request_type = models.FundRequestTypeChoices.ARTICLE
            fr.aid_type = None
        if self.object and self.object.fund_request_number:
            fr.fund_request_number = self.object.fund_request_number
        if self.object and self.object.purchase_order_number:
            fr.purchase_order_number = self.object.purchase_order_number
        if not fr.created_by:
            fr.created_by = self.request.user
        self._set_fund_request_status(fr, action)

        recipient_formset, article_formset = self._build_formsets(fr)
        formsets_ok = recipient_formset.is_valid() and article_formset.is_valid()

        header_ok = True
        if fr.fund_request_type == models.FundRequestTypeChoices.AID and action == 'submit' and not (fr.aid_type or '').strip():
            form.add_error('aid_type', 'Select the aid type before submit.')
            header_ok = False
        if formsets_ok and fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE:
            self._apply_article_vendor_summary(fr, article_formset, self._parse_vendor_groups_from_request())
        header_ok = self._validate_article_header_fields(form, fr, action) and header_ok

        if not header_ok or not formsets_ok or not self._validate_fund_request_formsets(fr, action, recipient_formset, article_formset):
            messages.error(self.request, 'Please fix errors in recipients/articles before saving.')
            return self.render_to_response(
                self.get_context_data(
                    form=form,
                    recipient_formset=recipient_formset,
                    article_formset=article_formset,
                )
            )

        try:
            with transaction.atomic():
                if action == 'submit' and not fr.fund_request_number:
                    fr.fund_request_number = services.next_fund_request_number()
                if action == 'submit' and fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE and not fr.purchase_order_number:
                    fr.purchase_order_number = services.next_purchase_order_number()
                fr.save()

                recipient_formset.instance = fr
                article_formset.instance = fr

                for deleted_form in getattr(recipient_formset, "deleted_forms", []):
                    deleted_instance = getattr(deleted_form, "instance", None)
                    if deleted_instance and deleted_instance.pk:
                        deleted_instance.delete()
                recipient_instances = recipient_formset.save(commit=False)
                for recipient in recipient_instances:
                    recipient.fund_request = fr
                    if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT:
                        recipient.recipient_name = (
                            recipient.name_of_beneficiary
                            or recipient.name_of_institution
                            or recipient.district_name
                            or recipient.recipient_name
                            or recipient.beneficiary
                            or 'Recipient'
                        )
                    elif recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC:
                        recipient.recipient_name = recipient.name_of_beneficiary or recipient.recipient_name or recipient.beneficiary or 'Recipient'
                    elif recipient.beneficiary_type in {
                        models.RecipientTypeChoices.INSTITUTIONS,
                        models.RecipientTypeChoices.OTHERS,
                    }:
                        recipient.recipient_name = (
                            recipient.name_of_institution
                            or recipient.name_of_beneficiary
                            or recipient.recipient_name
                            or recipient.beneficiary
                            or 'Recipient'
                        )
                    else:
                        recipient.recipient_name = recipient.recipient_name or recipient.name_of_beneficiary or recipient.name_of_institution or recipient.beneficiary or 'Recipient'
                    recipient.save()

                for deleted_form in getattr(article_formset, "deleted_forms", []):
                    deleted_instance = getattr(deleted_form, "instance", None)
                    if deleted_instance and deleted_instance.pk:
                        deleted_instance.delete()
                vendor_groups = self._parse_vendor_groups_from_request()
                for article_form in article_formset.forms:
                    if not getattr(article_form, "cleaned_data", None) or article_form.cleaned_data.get("DELETE", False):
                        continue
                    article = article_form.save(commit=False)
                    article.fund_request = fr
                    if not article.article_id:
                        article.article = self._resolve_article_record(article.article_name)
                    if article.article and not article.article_name:
                        article.article_name = article.article.article_name
                    if self.is_purchase_order_mode():
                        article.vendor_name = fr.supplier_name
                        article.gst_no = fr.gst_number
                        article.vendor_address = fr.supplier_address
                        article.vendor_city = fr.supplier_city
                        article.vendor_state = fr.supplier_state
                        article.vendor_pincode = fr.supplier_pincode
                    else:
                        group_key = str(self.request.POST.get(f'{article_form.prefix}-vendor_group_key') or '').strip()
                        group = vendor_groups.get(group_key, {})
                        article.vendor_name = group.get('vendor_name', '')
                        article.gst_no = group.get('gst_no', '')
                        article.vendor_address = group.get('vendor_address', '')
                        article.vendor_city = group.get('vendor_city', '')
                        article.vendor_state = group.get('vendor_state', '')
                        article.vendor_pincode = group.get('vendor_pincode', '')
                        article.cheque_in_favour = group.get('cheque_in_favour', '')
                    article.unit_price = article.unit_price or 0
                    article.price_including_gst = article.unit_price * (article.quantity or 0)
                    article.value = article.price_including_gst
                    article.cumulative = article.value
                    article.save()

                self._link_aid_sources(fr)
                self._collect_totals(fr)
                services.sync_order_entries_from_fund_request(fr, actor=self.request.user)
                self.object = fr

                if action == 'submit':
                    services.log_audit(
                        user=self.request.user,
                        action_type=models.ActionTypeChoices.STATUS_CHANGE,
                        entity_type='fund_request',
                        entity_id=str(fr.id),
                        details={'status': models.FundRequestStatusChoices.SUBMITTED},
                        ip_address=self.request.META.get('REMOTE_ADDR'),
                        user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
                    )
                    messages.success(self.request, 'Fund request submitted.')
                else:
                    messages.success(self.request, 'Fund request saved as draft.')
        except IntegrityError:
            form.add_error(None, 'Fund request number already exists. Please try submitting again.')
            return self.render_to_response(
                self.get_context_data(
                    form=form,
                    recipient_formset=recipient_formset,
                    article_formset=article_formset,
                )
            )
        return HttpResponseRedirect(self.get_success_url())


class FundRequestCreateView(LoginRequiredMixin, FundRequestCreateUpdateMixin, CreateView):
    pass


class FundRequestUpdateView(LoginRequiredMixin, FundRequestCreateUpdateMixin, UpdateView):
    pass


class FundRequestDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"
    model = models.FundRequest
    template_name = "dashboard/fund_request_detail.html"
    context_object_name = "fund_request"

    def get_queryset(self):
        return (
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["media_url"] = settings.MEDIA_URL
        context["can_edit"] = _is_editable_by_user(self.request.user, self.object)
        context["can_reopen"] = self.request.user.role == "admin" and self.object.status == models.FundRequestStatusChoices.SUBMITTED
        context["can_delete"] = self.request.user.role == "admin" and self.request.user.has_module_permission(
            models.ModuleKeyChoices.ORDER_FUND_REQUEST,
            "delete",
        )
        context["back_url"] = reverse("ui:fund-request-list")
        recipient_rows = list(self.object.recipients.all())
        for recipient in recipient_rows:
            recipient.beneficiary_display = _fund_request_recipient_display_name(recipient)
        context["recipient_rows"] = recipient_rows
        article_rows = list(self.object.articles.all())
        for article in article_rows:
            article.beneficiary_display = _fund_request_article_beneficiary_display(article)
        context["article_rows"] = article_rows
        return context


class FundRequestPDFView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"

    def get(self, request, pk):
        fund_request = get_object_or_404(
            models.FundRequest.objects.select_related("created_by").prefetch_related("recipients", "articles"),
            pk=pk,
        )
        pdf_buffer = services.generate_fund_request_pdf(fund_request)
        filename_base = fund_request.formatted_fund_request_number or f"FR-DRAFT-{fund_request.pk}"
        response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename_base}.pdf"'
        return response


class FundRequestDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "delete"
    model = models.FundRequest
    template_name = "dashboard/fund_request_confirm_delete.html"
    success_url = reverse_lazy("ui:fund-request-list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        models.OrderEntry.objects.filter(fund_request=self.object).delete()
        messages.warning(self.request, "Fund request deleted.")
        return super().post(request, *args, **kwargs)


class FundRequestSubmitView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "submit"
    allowed_roles = {"admin", "editor"}

    def post(self, request, pk):
        fr = models.FundRequest.objects.get(pk=pk)
        if not _is_editable_by_user(request.user, fr) or request.user.role == "viewer":
            return HttpResponse("Forbidden", status=403)
        if fr.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(request, "Only draft fund requests can be submitted.")
            return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))
        fr.status = models.FundRequestStatusChoices.SUBMITTED
        if not fr.fund_request_number:
            fr.fund_request_number = services.next_fund_request_number()
        update_fields = ["status", "fund_request_number"]
        if fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE and not fr.purchase_order_number:
            fr.purchase_order_number = services.next_purchase_order_number()
            update_fields.append("purchase_order_number")
        fr.save(update_fields=update_fields)
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="fund_request",
            entity_id=str(fr.id),
            details={"status": models.FundRequestStatusChoices.SUBMITTED},
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
        )
        services.sync_fund_request_totals(fr)
        services.sync_order_entries_from_fund_request(fr, actor=request.user)
        messages.success(request, "Fund request submitted.")
        return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))


class FundRequestReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "reopen"

    def post(self, request, pk):
        fr = get_object_or_404(models.FundRequest, pk=pk)
        if fr.status != models.FundRequestStatusChoices.SUBMITTED:
            messages.error(request, "Only submitted fund requests can be reopened.")
            return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))
        previous_status = fr.status
        fr.status = models.FundRequestStatusChoices.DRAFT
        fr.save(update_fields=["status"])
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="fund_request",
            entity_id=str(fr.id),
            details={"from": previous_status, "to": models.FundRequestStatusChoices.DRAFT},
            **_request_audit_meta(request),
        )
        services.sync_fund_request_totals(fr)
        services.sync_order_entries_from_fund_request(fr, actor=request.user)
        messages.success(request, "Fund request reopened as draft.")
        return HttpResponseRedirect(reverse("ui:fund-request-edit", args=[fr.pk]))


class FundRequestDocumentUploadView(LoginRequiredMixin, WriteRoleMixin, FormView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "create_edit"
    template_name = "dashboard/fund_request_upload_document.html"
    form_class = FundRequestDocumentUploadForm

    def dispatch(self, request, *args, **kwargs):
        self.fund_request = models.FundRequest.objects.get(pk=kwargs["pk"])
        if not _is_editable_by_user(request.user, self.fund_request):
            return HttpResponse("Forbidden", status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse("ui:fund-request-detail", args=[self.fund_request.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["fund_request"] = self.fund_request
        return context

    def form_valid(self, form):
        uploaded_file = form.cleaned_data["file"]
        upload_dir = os.path.join("fund-request-docs", str(self.fund_request.pk))
        relative_path = os.path.join(upload_dir, uploaded_file.name)
        from django.core.files.storage import default_storage

        stored_path = default_storage.save(relative_path, uploaded_file)
        models.FundRequestDocument.objects.create(
            fund_request=self.fund_request,
            document_type=form.cleaned_data["document_type"],
            file_path=stored_path,
            file_name=uploaded_file.name,
            generated_by=self.request.user,
        )
        messages.success(self.request, "Document uploaded.")
        return HttpResponseRedirect(self.get_success_url())


def _purchase_order_sequence(value):
    raw = str(value or "").strip().upper()
    prefix = "MASM/MNP"
    if not raw.startswith(prefix):
        return None
    suffix = raw[len(prefix):]
    if len(suffix) != 5 or not suffix.isdigit():
        return None
    return int(suffix[:3])


def _purchase_order_article_choices():
    return [
        {
            'name': article.article_name,
            'label': article.article_name,
            'article_id': article.id,
            'default_price': str(article.cost_per_unit or 0),
            'pending_qty': 0,
        }
        for article in models.Article.objects.filter(
            item_type=models.ItemTypeChoices.ARTICLE,
            is_active=True,
        ).order_by('article_name')
    ]


class PurchaseOrderListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "view"
    model = models.PurchaseOrder
    template_name = "dashboard/purchase_order_list.html"
    context_object_name = "purchase_orders"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            models.PurchaseOrder.objects
            .select_related("created_by")
            .prefetch_related("items", "items__article")
        )
        self._sort_key = (self.request.GET.get("sort") or "created_at").strip()
        self._sort_dir = (self.request.GET.get("dir") or "desc").strip().lower()
        if q := (self.request.GET.get("q") or "").strip():
            matching_ids = (
                models.PurchaseOrder.objects
                .annotate(
                    total_amount_text=Cast("total_amount", output_field=CharField()),
                    created_at_text=Cast("created_at", output_field=CharField()),
                    item_quantity_text=Cast("items__quantity", output_field=CharField()),
                    item_unit_price_text=Cast("items__unit_price", output_field=CharField()),
                    item_total_text=Cast("items__total_value", output_field=CharField()),
                )
                .filter(
                    Q(purchase_order_number__icontains=q)
                    | Q(vendor_name__icontains=q)
                    | Q(vendor_address__icontains=q)
                    | Q(vendor_city__icontains=q)
                    | Q(vendor_state__icontains=q)
                    | Q(vendor_pincode__icontains=q)
                    | Q(comments__icontains=q)
                    | Q(status__icontains=q)
                    | Q(total_amount_text__icontains=q)
                    | Q(created_at_text__icontains=q)
                    | Q(created_by__email__icontains=q)
                    | Q(created_by__first_name__icontains=q)
                    | Q(created_by__last_name__icontains=q)
                    | Q(items__article_name__icontains=q)
                    | Q(items__supplier_article_name__icontains=q)
                    | Q(items__description__icontains=q)
                    | Q(item_quantity_text__icontains=q)
                    | Q(item_unit_price_text__icontains=q)
                    | Q(item_total_text__icontains=q)
                )
                .values_list("pk", flat=True)
                .distinct()
            )
            queryset = queryset.filter(pk__in=matching_ids)

        if status := (self.request.GET.get("status") or "").strip():
            queryset = queryset.filter(status=status)

        sort_field_map = {
            "purchase_order_number": "purchase_order_number",
            "vendor_name": "vendor_name",
            "vendor_city": "vendor_city",
            "total_amount": "total_amount",
            "status": "status",
            "created_at": "created_at",
        }
        sort_field = sort_field_map.get(self._sort_key, "created_at")
        sort_prefix = "" if self._sort_dir == "asc" else "-"
        if sort_field == "purchase_order_number":
            rows = list(queryset.order_by(f"{sort_prefix}created_at", f"{sort_prefix}id"))
            rows.sort(
                key=lambda po: (
                    _purchase_order_sequence(po.purchase_order_number) is None,
                    _purchase_order_sequence(po.purchase_order_number) or 0,
                    po.created_at,
                    po.id,
                ),
                reverse=(self._sort_dir == "desc"),
            )
            return rows
        return queryset.order_by(f"{sort_prefix}{sort_field}", "-id")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_sort = getattr(self, "_sort_key", "created_at")
        current_dir = getattr(self, "_sort_dir", "desc")

        for purchase_order in context["purchase_orders"]:
            purchase_order.item_count = purchase_order.items.count()
            purchase_order.items_total_quantity = sum(int(item.quantity or 0) for item in purchase_order.items.all())
            purchase_order.display_comments = (purchase_order.comments or "").strip() or models.PURCHASE_ORDER_DEFAULT_COMMENTS

        def build_sort_params(column):
            params = self.request.GET.copy()
            params.pop("page", None)
            next_dir = "asc"
            if current_sort == column and current_dir == "asc":
                next_dir = "desc"
            params["sort"] = column
            params["dir"] = next_dir
            return params.urlencode()

        context["status_choices"] = [
            ("", "All"),
            (models.FundRequestStatusChoices.DRAFT, "Draft"),
            (models.FundRequestStatusChoices.SUBMITTED, "Submitted"),
        ]
        context["filters"] = {
            "q": self.request.GET.get("q", ""),
            "status": self.request.GET.get("status", ""),
        }
        context["current_sort"] = current_sort
        context["current_dir"] = current_dir
        context["sort_querystrings"] = {
            "purchase_order_number": build_sort_params("purchase_order_number"),
            "vendor_name": build_sort_params("vendor_name"),
            "total_amount": build_sort_params("total_amount"),
            "status": build_sort_params("status"),
            "created_at": build_sort_params("created_at"),
        }
        context["can_create_edit"] = self.request.user.has_module_permission(models.ModuleKeyChoices.PURCHASE_ORDER, "create_edit")
        context["can_submit"] = self.request.user.has_module_permission(models.ModuleKeyChoices.PURCHASE_ORDER, "submit")
        context["can_reopen"] = self.request.user.has_module_permission(models.ModuleKeyChoices.PURCHASE_ORDER, "reopen")
        context["can_delete"] = self.request.user.has_module_permission(models.ModuleKeyChoices.PURCHASE_ORDER, "delete")
        return context


class PurchaseOrderCreateUpdateMixin(WriteRoleMixin):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "create_edit"
    form_class = PurchaseOrderForm
    template_name = "dashboard/purchase_order_form.html"
    model = models.PurchaseOrder
    success_url = reverse_lazy("ui:purchase-order-list")

    def _build_formset(self, instance: models.PurchaseOrder | None = None):
        return PurchaseOrderItemFormSet(self.request.POST or None, prefix="items", instance=instance)

    def _can_edit(self, purchase_order: models.PurchaseOrder):
        return _is_editable_purchase_order(self.request.user, purchase_order)

    def dispatch(self, request, *args, **kwargs):
        self.object = getattr(self, "object", None)
        if self.object and not self._can_edit(self.object):
            messages.error(request, "Submitted purchase orders must be reopened before editing.")
            return HttpResponseRedirect(reverse("ui:purchase-order-list"))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["item_formset"] = kwargs.get("item_formset", None) or self._build_formset(self.object)
        context["article_request_choices_json"] = json.dumps(_purchase_order_article_choices())
        return context

    def _resolve_article_record(self, article_name: str):
        article_name = (article_name or "").strip()
        if not article_name:
            return None
        return models.Article.objects.filter(article_name__iexact=article_name).first()

    def _validate_header_fields(self, form, purchase_order, action):
        if action != "submit":
            return True
        valid = True
        for field_name, label in [
            ("vendor_name", "Vendor Name"),
            ("vendor_address", "Vendor Address"),
            ("vendor_city", "City"),
            ("vendor_state", "State"),
            ("vendor_pincode", "Pincode"),
        ]:
            value = getattr(purchase_order, field_name, None)
            if not str(value or "").strip():
                form.add_error(field_name, f"{label} is required for submit.")
                valid = False
        return valid

    def _validate_item_formset(self, item_formset, action):
        active_forms = [form for form in item_formset.forms if form.cleaned_data and not form.cleaned_data.get("DELETE", False)]
        if action == "submit" and not active_forms:
            item_formset._non_form_errors = item_formset.error_class(["Add at least one item."])
            return False
        valid = True
        if action == "submit":
            for form in active_forms:
                for field_name in ["article_name", "supplier_article_name", "description", "quantity", "unit_price"]:
                    value = form.cleaned_data.get(field_name)
                    if value in (None, "", 0, "0"):
                        form.add_error(field_name, "Required for submit.")
                        valid = False
        return valid

    def form_valid(self, form):
        action = self.request.POST.get("action", "draft")
        if action == "submit" and self.object and self.object.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(self.request, "Only draft purchase orders can be submitted.")
            return HttpResponseRedirect(self.get_success_url())

        purchase_order = form.save(commit=False)
        if self.object and self.object.purchase_order_number:
            purchase_order.purchase_order_number = self.object.purchase_order_number
        if not purchase_order.created_by:
            purchase_order.created_by = self.request.user
        purchase_order.status = (
            models.FundRequestStatusChoices.SUBMITTED
            if action == "submit"
            else models.FundRequestStatusChoices.DRAFT
        )

        header_ok = self._validate_header_fields(form, purchase_order, action)
        item_formset = self._build_formset(purchase_order)
        formsets_ok = item_formset.is_valid()
        items_ok = self._validate_item_formset(item_formset, action)
        if not header_ok or not formsets_ok or not items_ok:
            messages.error(self.request, "Please fix errors in purchase order fields before saving.")
            return self.render_to_response(self.get_context_data(form=form, item_formset=item_formset))

        with transaction.atomic():
            if action == "submit" and not purchase_order.purchase_order_number:
                purchase_order.purchase_order_number = services.next_purchase_order_number()
            purchase_order.save()

            item_formset.instance = purchase_order
            for deleted_form in getattr(item_formset, "deleted_forms", []):
                deleted_instance = getattr(deleted_form, "instance", None)
                if deleted_instance and deleted_instance.pk:
                    deleted_instance.delete()
            item_instances = item_formset.save(commit=False)
            for item in item_instances:
                item.purchase_order = purchase_order
                if not item.article_id:
                    item.article = self._resolve_article_record(item.article_name)
                if item.article and not item.article_name:
                    item.article_name = item.article.article_name
                item.quantity = item.quantity or 0
                item.unit_price = item.unit_price or 0
                item.total_value = (item.unit_price or 0) * (item.quantity or 0)
                item.save()

            services.sync_purchase_order_totals(purchase_order)
            self.object = purchase_order

        messages.success(
            self.request,
            "Purchase order submitted." if action == "submit" else "Purchase order saved as draft.",
        )
        return HttpResponseRedirect(self.get_success_url())


class PurchaseOrderCreateView(LoginRequiredMixin, PurchaseOrderCreateUpdateMixin, CreateView):
    pass


class PurchaseOrderUpdateView(LoginRequiredMixin, PurchaseOrderCreateUpdateMixin, UpdateView):
    def get_queryset(self):
        return models.PurchaseOrder.objects.all()


class PurchaseOrderPDFView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "view"

    def get(self, request, pk):
        purchase_order = get_object_or_404(
            models.PurchaseOrder.objects.select_related("created_by").prefetch_related("items"),
            pk=pk,
        )
        services.ensure_purchase_order_number(purchase_order)
        pdf_buffer = services.generate_purchase_order_pdf(purchase_order)
        filename_base = purchase_order.purchase_order_number or f"PO-DRAFT-{purchase_order.pk}"
        response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename_base}.pdf"'
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        return response


class PurchaseOrderSubmitView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "submit"
    allowed_roles = {"admin", "editor"}

    def post(self, request, pk):
        purchase_order = get_object_or_404(models.PurchaseOrder, pk=pk)
        if not _is_editable_purchase_order(request.user, purchase_order) or request.user.role == "viewer":
            return HttpResponse("Forbidden", status=403)
        if purchase_order.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(request, "Only draft purchase orders can be submitted.")
            return HttpResponseRedirect(reverse("ui:purchase-order-list"))
        purchase_order.status = models.FundRequestStatusChoices.SUBMITTED
        if not purchase_order.purchase_order_number:
            purchase_order.purchase_order_number = services.next_purchase_order_number()
        purchase_order.save(update_fields=["status", "purchase_order_number"])
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="purchase_order",
            entity_id=str(purchase_order.id),
            details={"status": models.FundRequestStatusChoices.SUBMITTED},
            **_request_audit_meta(request),
        )
        services.sync_purchase_order_totals(purchase_order)
        messages.success(request, "Purchase order submitted.")
        return HttpResponseRedirect(reverse("ui:purchase-order-list"))


class PurchaseOrderReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "reopen"

    def post(self, request, pk):
        purchase_order = get_object_or_404(models.PurchaseOrder, pk=pk)
        if purchase_order.status != models.FundRequestStatusChoices.SUBMITTED:
            messages.error(request, "Only submitted purchase orders can be reopened.")
            return HttpResponseRedirect(reverse("ui:purchase-order-list"))
        previous_status = purchase_order.status
        purchase_order.status = models.FundRequestStatusChoices.DRAFT
        purchase_order.save(update_fields=["status"])
        services.log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="purchase_order",
            entity_id=str(purchase_order.id),
            details={"from": previous_status, "to": models.FundRequestStatusChoices.DRAFT},
            **_request_audit_meta(request),
        )
        messages.success(request, "Purchase order reopened as draft.")
        return HttpResponseRedirect(reverse("ui:purchase-order-edit", args=[purchase_order.pk]))


class PurchaseOrderDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    module_key = models.ModuleKeyChoices.PURCHASE_ORDER
    permission_action = "delete"
    model = models.PurchaseOrder
    template_name = "dashboard/purchase_order_confirm_delete.html"
    success_url = reverse_lazy("ui:purchase-order-list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.warning(self.request, "Purchase order deleted.")
        return super().post(request, *args, **kwargs)


class AidRecipientTemplateDownloadView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.BASE_FILES
    permission_action = "view"

    def get(self, request, *args, **kwargs):
        headers = [
            "beneficiary_type",
            "application_number",
            "beneficiary",
            "fund_requested",
            "details",
            "cheque_rtgs_in_favour",
        ]
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="aid_recipients_template.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        return response


class MasterDataBaseView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.BASE_FILES
    permission_action = "view"
    template_name = "dashboard/module_master_data.html"
    data_key = None
    page_title = "Base Files"
    upload_help = "Upload a CSV file to refresh the stored records."

    def dispatch(self, request, *args, **kwargs):
        if not self.data_key:
            raise ValueError("MasterDataBaseView requires data_key")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "active_data_key": self.data_key,
                "page_title": self.page_title,
                "upload_help": self.upload_help,
                "upload_form": kwargs.get("upload_form") or MasterDataUploadForm(),
                "records": kwargs.get("records") if kwargs.get("records") is not None else self.get_records(),
                "summary": self.get_summary(),
                "replace_supported": self.data_key == "history",
            }
        )
        return context

    def get_records(self):
        if self.data_key == "districts":
            return list(models.DistrictMaster.objects.order_by("district_name")[:100])
        if self.data_key == "articles":
            return list(models.Article.objects.order_by("article_name")[:100])
        return list(models.PublicBeneficiaryHistory.objects.order_by("-year", "-created_at")[:100])

    def get_summary(self):
        return {
            "districts": models.DistrictMaster.objects.count(),
            "articles": models.Article.objects.count(),
            "history": models.PublicBeneficiaryHistory.objects.count(),
        }

    def post(self, request, *args, **kwargs):
        if not request.user.has_module_permission(self.module_key, "upload_replace"):
            messages.error(request, "You do not have permission to upload base files.")
            return HttpResponseRedirect(request.path)

        form = MasterDataUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return self.render_to_response(self.get_context_data(upload_form=form))

        uploaded = form.cleaned_data["file"]
        replace_existing = bool(form.cleaned_data.get("replace_existing"))
        inserted, updated = self.import_rows(uploaded, replace_existing=replace_existing)
        messages.success(request, f"{self.page_title} import complete. inserted={inserted}, updated={updated}")
        return HttpResponseRedirect(request.path)

    def import_rows(self, uploaded_file, *, replace_existing=False):
        if self.data_key == "districts":
            return _import_district_master_csv(uploaded_file)
        if self.data_key == "articles":
            return _import_article_master_csv(uploaded_file)
        return _import_public_history_csv(uploaded_file, replace_existing=replace_existing)


class MasterDataDistrictView(MasterDataBaseView):
    data_key = "districts"
    page_title = "District Master"
    upload_help = "Upload the yearly district president file. The district code, president, budget, and mobile number will be updated in place."


class MasterDataArticleView(MasterDataBaseView):
    data_key = "articles"
    page_title = "Article Price List"
    upload_help = "Upload the latest article price list. Existing article names are updated; new article names are added."


class MasterDataHistoryView(MasterDataBaseView):
    data_key = "history"
    page_title = "Past Beneficiary History"
    upload_help = "Upload the past district/public beneficiary file. This is used for Aadhaar warnings and reference checks during entry."


def _csv_reader_from_upload(uploaded_file):
    uploaded_file.seek(0)
    return csv.DictReader(io.StringIO(uploaded_file.read().decode("utf-8-sig")))


def _tabular_rows_from_upload(uploaded_file):
    name = str(getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        raw_headers = list(values[0] or [])
        headers = [str(header or "").strip() for header in raw_headers]
        rows = []
        for row_values in values[1:]:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell_value = row_values[index] if index < len(row_values) else ""
                row[header] = "" if cell_value is None else str(cell_value)
            if any(str(value or "").strip() for value in row.values()):
                rows.append(row)
        return headers, rows
    reader = _csv_reader_from_upload(uploaded_file)
    return list(reader.fieldnames or []), list(reader)


PHASE2_MASTER_REQUIRED_HEADERS = [
    "Application Number",
    "Beneficiary Name",
    "Requested Item",
    "Quantity",
    "Beneficiary Type",
    "Item Type",
]


def _phase2_parse_number(value):
    raw = str(value or "").replace(",", "").strip()
    try:
        number = int(Decimal(raw or "0"))
    except (InvalidOperation, ValueError):
        return 0
    return max(number, 0)


def _phase2_normalize_text(value):
    return " ".join(str(value or "").strip().lower().split())


def _phase2_active_or_latest_session():
    return (
        models.EventSession.objects.order_by("-is_active", "-event_year", "session_name", "-id").first()
    )


def _phase2_get_or_create_default_session():
    session = _phase2_active_or_latest_session()
    if session:
        return session
    year = timezone.localdate().year
    return models.EventSession.objects.create(
        session_name=f"{year} Event",
        event_year=year,
        is_active=True,
    )


def _phase2_selected_session(request):
    session_id = (request.GET.get("session") or request.POST.get("session") or "").strip()
    if session_id:
        try:
            return models.EventSession.objects.get(pk=session_id)
        except (ValueError, models.EventSession.DoesNotExist):
            return _phase2_active_or_latest_session()
    return _phase2_active_or_latest_session()


def _phase2_session_querystring(request, session):
    params = request.GET.copy()
    params["session"] = str(session.pk)
    return params.urlencode()


def _phase2_redirect_url(request, view_name, session=None, *, filter_keys=None):
    params = request.GET.copy()
    if filter_keys is not None:
        filtered = params.__class__(mutable=True)
        for key in filter_keys:
            if key in params:
                values = params.getlist(key)
                if values:
                    filtered.setlist(key, values)
        params = filtered
    if session:
        params["session"] = str(session.pk)
    elif "session" in params:
        del params["session"]
    query = params.urlencode()
    base_url = reverse(view_name)
    return f"{base_url}?{query}" if query else base_url


def _phase2_url_with_extra_params(request, view_name, session=None, *, filter_keys=None, extra_params=None):
    params = request.GET.copy()
    if filter_keys is not None:
        filtered = params.__class__(mutable=True)
        for key in filter_keys:
            if key in params:
                values = params.getlist(key)
                if values:
                    filtered.setlist(key, values)
        params = filtered
    if session:
        params["session"] = str(session.pk)
    if extra_params:
        for key, value in extra_params.items():
            if value in (None, ""):
                params.pop(key, None)
            else:
                params[key] = value
    query = params.urlencode()
    base_url = reverse(view_name)
    return f"{base_url}?{query}" if query else base_url


def _phase2_distinct_beneficiary_key(application_number, beneficiary_name, beneficiary_type):
    return (
        _phase2_normalize_text(beneficiary_type),
        str(application_number or "").strip(),
        str(beneficiary_name or "").strip(),
    )


def _phase2_reconciliation_snapshot(*, source_rows, grouped_rows, total_value_getter):
    source_unique_items = set()
    grouped_unique_items = set()
    source_beneficiaries = {}
    grouped_beneficiaries = {}
    source_quantity_total = 0
    grouped_quantity_total = 0
    source_value_total = 0
    grouped_value_total = 0

    for row in source_rows:
        beneficiary_type = str(row.get("Beneficiary Type") or "").strip()
        application_number = str(row.get("Application Number") or "").strip()
        beneficiary_name = str(row.get("Beneficiary Name") or "").strip()
        requested_item = str(row.get("Requested Item") or "").strip()
        quantity = _phase2_parse_number(row.get("Quantity"))
        source_quantity_total += quantity
        source_value_total += total_value_getter(row)
        if requested_item:
            source_unique_items.add(requested_item)
        if beneficiary_type or application_number or beneficiary_name:
            source_beneficiaries.setdefault(beneficiary_type or "Unknown", set()).add(
                _phase2_distinct_beneficiary_key(application_number, beneficiary_name, beneficiary_type)
            )

    for row in grouped_rows:
        beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        application_number = str(row.get("application_number") or "").strip()
        beneficiary_name = str(row.get("beneficiary_name") or "").strip()
        requested_item = str(row.get("requested_item") or "").strip()
        quantity = int(row.get("quantity") or 0)
        grouped_quantity_total += quantity
        grouped_value_total += total_value_getter(row)
        if requested_item:
            grouped_unique_items.add(requested_item)
        if beneficiary_type or application_number or beneficiary_name:
            grouped_beneficiaries.setdefault(beneficiary_type or "Unknown", set()).add(
                _phase2_distinct_beneficiary_key(application_number, beneficiary_name, beneficiary_type)
            )

    beneficiary_labels = [
        models.RecipientTypeChoices.DISTRICT,
        models.RecipientTypeChoices.PUBLIC,
        models.RecipientTypeChoices.INSTITUTIONS,
        models.RecipientTypeChoices.OTHERS,
    ]
    beneficiary_metrics = []
    for label in beneficiary_labels:
        source_count = len(source_beneficiaries.get(label, set()))
        grouped_count = len(grouped_beneficiaries.get(label, set()))
        if source_count or grouped_count:
            beneficiary_metrics.append(
                {
                    "label": f"{label} Beneficiaries",
                    "source": source_count,
                    "grouped": grouped_count,
                }
            )

    return {
        "source_row_count": len(source_rows),
        "grouped_row_count": len(grouped_rows),
        "source_quantity_total": source_quantity_total,
        "grouped_quantity_total": grouped_quantity_total,
        "source_total_value": source_value_total,
        "grouped_total_value": grouped_value_total,
        "source_unique_items": len(source_unique_items),
        "grouped_unique_items": len(grouped_unique_items),
        "beneficiary_metrics": beneficiary_metrics,
    }


def _phase2_reconciliation_checks(reconciliation_snapshot):
    if not reconciliation_snapshot:
        return []
    checks = [
        {
            "label": "Quantity",
            "matched": reconciliation_snapshot.get("source_quantity_total", 0) == reconciliation_snapshot.get("grouped_quantity_total", 0),
            "source": reconciliation_snapshot.get("source_quantity_total", 0),
            "grouped": reconciliation_snapshot.get("grouped_quantity_total", 0),
        },
        {
            "label": "Total Value",
            "matched": reconciliation_snapshot.get("source_total_value", 0) == reconciliation_snapshot.get("grouped_total_value", 0),
            "source": reconciliation_snapshot.get("source_total_value", 0),
            "grouped": reconciliation_snapshot.get("grouped_total_value", 0),
        },
        {
            "label": "Unique Items",
            "matched": reconciliation_snapshot.get("source_unique_items", 0) == reconciliation_snapshot.get("grouped_unique_items", 0),
            "source": reconciliation_snapshot.get("source_unique_items", 0),
            "grouped": reconciliation_snapshot.get("grouped_unique_items", 0),
        },
    ]
    for metric in reconciliation_snapshot.get("beneficiary_metrics") or []:
        checks.append(
            {
                "label": metric.get("label") or "Beneficiaries",
                "matched": metric.get("source", 0) == metric.get("grouped", 0),
                "source": metric.get("source", 0),
                "grouped": metric.get("grouped", 0),
            }
        )
    return checks


def _phase2_split_reconciliation(rows, pending_waiting_by_id=None):
    pending_waiting_by_id = pending_waiting_by_id or {}
    total_quantity = 0
    total_waiting = 0
    total_token = 0
    row_mismatch_count = 0
    row_count = 0

    for row in rows:
        row_count += 1
        quantity = int(row.quantity or 0)
        total_quantity += quantity
        if str(row.id) in pending_waiting_by_id:
            waiting = pending_waiting_by_id[str(row.id)]
            waiting = max(min(int(waiting or 0), quantity), 0)
            token = max(quantity - waiting, 0)
        else:
            waiting = int(row.waiting_hall_quantity or 0)
            token = int(row.token_quantity or 0)
        total_waiting += waiting
        total_token += token
        if waiting + token != quantity:
            row_mismatch_count += 1

    return {
        "row_count": row_count,
        "row_mismatch_count": row_mismatch_count,
        "rowwise_matched": row_mismatch_count == 0,
        "total_quantity": total_quantity,
        "total_waiting": total_waiting,
        "total_token": total_token,
        "overall_matched": (total_waiting + total_token) == total_quantity,
    }


def _phase2_group_key(application_number, beneficiary_name, district, requested_item):
    return (
        str(application_number or "").strip(),
        str(beneficiary_name or "").strip(),
        str(district or "").strip(),
        str(requested_item or "").strip(),
    )


def _phase2_row_identity_key(data):
    if hasattr(data, "master_row"):
        master_row = getattr(data, "master_row", None) or {}
        master_headers = getattr(data, "master_headers", None) or []
    else:
        master_row = (data.get("master_row") or {}) if isinstance(data, dict) else {}
        master_headers = (data.get("master_headers") or []) if isinstance(data, dict) else []

    if master_row:
        headers = list(master_headers) if master_headers else list(master_row.keys())
        filtered_headers = [
            header
            for header in headers
            if _phase2_normalize_text(header) not in {"waiting hall quantity", "token quantity", "sequence no", "sequence list"}
        ]
        return (
            "master_row",
            tuple(
                (header, str(master_row.get(header, "") or "").strip())
                for header in filtered_headers
            ),
        )

    if hasattr(data, "application_number"):
        application_number = getattr(data, "application_number", "")
        beneficiary_name = getattr(data, "beneficiary_name", "")
        district = getattr(data, "district", "")
        requested_item = getattr(data, "requested_item", "")
        quantity = getattr(data, "quantity", 0)
        beneficiary_type = getattr(data, "beneficiary_type", "")
        item_type = getattr(data, "item_type", "")
        comments = getattr(data, "comments", "")
    else:
        application_number = data.get("application_number", "") if isinstance(data, dict) else ""
        beneficiary_name = data.get("beneficiary_name", "") if isinstance(data, dict) else ""
        district = data.get("district", "") if isinstance(data, dict) else ""
        requested_item = data.get("requested_item", "") if isinstance(data, dict) else ""
        quantity = data.get("quantity", 0) if isinstance(data, dict) else 0
        beneficiary_type = data.get("beneficiary_type", "") if isinstance(data, dict) else ""
        item_type = data.get("item_type", "") if isinstance(data, dict) else ""
        comments = data.get("comments", "") if isinstance(data, dict) else ""

    return (
        "fallback",
        str(application_number or "").strip(),
        str(beneficiary_name or "").strip(),
        str(district or "").strip(),
        str(requested_item or "").strip(),
        int(quantity or 0),
        str(beneficiary_type or "").strip(),
        str(item_type or "").strip(),
        str(comments or "").strip(),
    )


def _phase2_row_identity_candidates(data):
    primary = _phase2_row_identity_key(data)
    candidates = [primary]

    if hasattr(data, "application_number"):
        application_number = getattr(data, "application_number", "")
        beneficiary_name = getattr(data, "beneficiary_name", "")
        district = getattr(data, "district", "")
        requested_item = getattr(data, "requested_item", "")
        quantity = getattr(data, "quantity", 0)
        beneficiary_type = getattr(data, "beneficiary_type", "")
        item_type = getattr(data, "item_type", "")
        comments = getattr(data, "comments", "")
    else:
        application_number = data.get("application_number", "") if isinstance(data, dict) else ""
        beneficiary_name = data.get("beneficiary_name", "") if isinstance(data, dict) else ""
        district = data.get("district", "") if isinstance(data, dict) else ""
        requested_item = data.get("requested_item", "") if isinstance(data, dict) else ""
        quantity = data.get("quantity", 0) if isinstance(data, dict) else 0
        beneficiary_type = data.get("beneficiary_type", "") if isinstance(data, dict) else ""
        item_type = data.get("item_type", "") if isinstance(data, dict) else ""
        comments = data.get("comments", "") if isinstance(data, dict) else ""

    fallback = (
        "fallback",
        str(application_number or "").strip(),
        str(beneficiary_name or "").strip(),
        str(district or "").strip(),
        str(requested_item or "").strip(),
        int(quantity or 0),
        str(beneficiary_type or "").strip(),
        str(item_type or "").strip(),
        str(comments or "").strip(),
    )
    structural = (
        "structural",
        str(application_number or "").strip(),
        str(beneficiary_name or "").strip(),
        str(district or "").strip(),
        str(requested_item or "").strip(),
    )
    if structural not in candidates:
        candidates.append(structural)
    if fallback not in candidates:
        candidates.append(fallback)
    return candidates


def _phase2_preview_row_from_upload(row):
    return {
        "beneficiary_type": str(row.get("beneficiary_type") or "").strip(),
        "application_number": str(row.get("application_number") or "").strip(),
        "beneficiary_name": str(row.get("beneficiary_name") or "").strip(),
        "requested_item": str(row.get("requested_item") or "").strip(),
    }


def _phase2_preview_row_from_existing(row):
    return {
        "beneficiary_type": str(row.beneficiary_type or "").strip(),
        "application_number": str(row.application_number or "").strip(),
        "beneficiary_name": str(row.beneficiary_name or "").strip(),
        "requested_item": str(row.requested_item or "").strip(),
    }


def _phase2_preserve_existing_split_state(session, upload_rows):
    existing_rows = list(models.SeatAllocationRow.objects.filter(session=session))
    existing_map = {}
    for row in existing_rows:
        for key in _phase2_row_identity_candidates(row):
            existing_map.setdefault(key, []).append(row)
    preserved_count = 0
    matched_existing_ids = set()
    matched_count = 0

    for row in upload_rows:
        existing = None
        for key in _phase2_row_identity_candidates(row):
            existing_candidates = existing_map.get(key) or []
            while existing_candidates and str(existing_candidates[0].id) in matched_existing_ids:
                existing_candidates.pop(0)
            if existing_candidates:
                existing = existing_candidates.pop(0)
                break
        if not existing:
            continue
        matched_existing_ids.add(str(existing.id))
        quantity = int(row.get("quantity") or 0)
        preserved_waiting = min(int(existing.waiting_hall_quantity or 0), quantity)
        row["waiting_hall_quantity"] = preserved_waiting
        row["token_quantity"] = max(quantity - preserved_waiting, 0)
        row["sequence_no"] = existing.sequence_no
        preserved_count += 1
        matched_count += 1

    removed_count = max(len(existing_rows) - len(matched_existing_ids), 0)
    new_count = max(len(upload_rows) - matched_count, 0)
    return {
        "rows": upload_rows,
        "preserved_count": preserved_count,
        "new_count": new_count,
        "removed_count": removed_count,
    }


def _phase2_preview_sync_state(session, upload_rows):
    existing_rows = list(models.SeatAllocationRow.objects.filter(session=session))
    existing_map = {}
    for row in existing_rows:
        for key in _phase2_row_identity_candidates(row):
            existing_map.setdefault(key, []).append(row)

    preserved_count = 0
    added_rows = []
    matched_existing_ids = set()
    for row in upload_rows:
        matched = False
        for key in _phase2_row_identity_candidates(row):
            existing_candidates = existing_map.get(key) or []
            while existing_candidates and str(existing_candidates[0].id) in matched_existing_ids:
                existing_candidates.pop(0)
            if existing_candidates:
                matched_existing_ids.add(str(existing_candidates.pop(0).id))
                matched = True
                break
        if matched:
            preserved_count += 1
            continue
        if len(added_rows) < 8:
            added_rows.append(_phase2_preview_row_from_upload(row))

    removed_rows = []
    unmatched_existing = [row for row in existing_rows if str(row.id) not in matched_existing_ids]
    removed_count = len(unmatched_existing)
    for row in unmatched_existing[:8]:
        removed_rows.append(_phase2_preview_row_from_existing(row))

    new_count = max(len(upload_rows) - preserved_count, 0)
    return {
        "preserved_count": preserved_count,
        "new_count": new_count,
        "removed_count": removed_count,
        "added_rows": added_rows,
        "removed_rows": removed_rows,
    }


def _phase2_master_change_state(session, upload_rows):
    existing_rows = list(models.SeatAllocationRow.objects.filter(session=session))
    existing_map = {}
    for row in existing_rows:
        for key in _phase2_row_identity_candidates(row):
            existing_map.setdefault(key, []).append(row)

    new_count = 0
    removed_count = 0
    updated_count = 0
    updated_rows = []
    matched_existing_ids = set()

    field_labels = {
        "Application Number": "Application Number",
        "Beneficiary Name": "Beneficiary Name",
        "Requested Item": "Requested Item",
        "Quantity": "Quantity",
        "Cost Per Unit": "Cost Per Unit",
        "Total Value": "Total Value",
        "Beneficiary Type": "Beneficiary Type",
        "Item Type": "Item Type",
        "Comments": "Comments",
        "Aadhar Number": "Aadhaar Number",
        "Name of Beneficiary": "Name of Beneficiary",
        "Name of Institution": "Name of Institution",
        "Cheque / RTGS in Favour": "Cheque / RTGS in Favour",
    }

    for incoming in upload_rows:
        matched = False
        matched_row = None
        for key in _phase2_row_identity_candidates(incoming):
            existing_candidates = existing_map.get(key) or []
            while existing_candidates and str(existing_candidates[0].id) in matched_existing_ids:
                existing_candidates.pop(0)
            if existing_candidates:
                matched_row = existing_candidates.pop(0)
                matched_existing_ids.add(str(matched_row.id))
                matched = True
                break
        if matched:
            incoming_master = incoming.get("master_row") or {}
            existing_master = matched_row.master_row or {}
            changes = []
            for header, label in field_labels.items():
                incoming_value = str(incoming_master.get(header, "") or "").strip()
                existing_value = str(existing_master.get(header, "") or "").strip()
                if incoming_value != existing_value:
                    changes.append(
                        f"{label} {existing_value or '-'} -> {incoming_value or '-'}"
                    )
            if changes:
                updated_count += 1
                if len(updated_rows) < 8:
                    updated_rows.append(
                        {
                            "beneficiary_type": str(incoming.get("beneficiary_type") or "").strip(),
                            "application_number": str(incoming.get("application_number") or "").strip(),
                            "beneficiary_name": str(incoming.get("beneficiary_name") or "").strip(),
                            "requested_item": str(incoming.get("requested_item") or "").strip(),
                            "changes": changes[:3],
                        }
                    )
            continue
        new_count += 1
        if len(updated_rows) < 8:
            updated_rows.append(
                {
                    "beneficiary_type": str(incoming.get("beneficiary_type") or "").strip(),
                    "application_number": str(incoming.get("application_number") or "").strip(),
                    "beneficiary_name": str(incoming.get("beneficiary_name") or "").strip(),
                    "requested_item": str(incoming.get("requested_item") or "").strip(),
                    "changes": ["New or changed row"],
                }
            )

    removed_count = max(len(existing_rows) - len(matched_existing_ids), 0)

    return {
        "has_changes": bool(new_count or removed_count or updated_count),
        "new_count": new_count,
        "removed_count": removed_count,
        "updated_count": updated_count,
        "updated_rows": updated_rows,
    }


def _phase2_export_rows(rows, *, include_sequence=False):
    rows = list(rows)
    if not rows:
        return [], []

    base_headers = []
    for row in rows:
        if row.master_headers:
            base_headers = list(row.master_headers)
            break
    if not base_headers:
        base_headers = [
            "Application Number",
            "Beneficiary Name",
            "Requested Item",
            "Quantity",
            "Beneficiary Type",
            "Item Type",
            "Comments",
        ]

    filtered_headers = [
        header for header in base_headers
        if _phase2_normalize_text(header) not in {"waiting hall quantity", "token quantity", "sequence no"}
    ]
    export_headers = [*filtered_headers, "Waiting Hall Quantity", "Token Quantity"]
    if include_sequence:
        export_headers.append("Sequence No")
    export_rows = []
    for row in rows:
        export_row = {}
        for header in filtered_headers:
            export_row[header] = (row.master_row or {}).get(header, "")
        export_row["Waiting Hall Quantity"] = row.waiting_hall_quantity
        export_row["Token Quantity"] = row.token_quantity
        if include_sequence:
            export_row["Sequence No"] = row.sequence_no or ""
        export_rows.append(export_row)
    return export_rows, export_headers


def _sequence_project_rows(rows, headers):
    projected_rows = []
    for row in rows:
        projected_rows.append({header: row.get(header, "") for header in headers})
    return projected_rows


def _sequence_row_key(row, headers):
    return tuple("" if row.get(header) is None else str(row.get(header)) for header in headers)


def _sequence_row_label(row):
    return " / ".join(
        [
            str(row.get("Application Number") or "-").strip() or "-",
            str(row.get("Beneficiary Name") or "-").strip() or "-",
            str(row.get("Requested Item") or "-").strip() or "-",
        ]
    )


def _sequence_prepare_export_row(row, *, sequence_no=None):
    prepared = dict(row or {})
    application_number = str(prepared.get("Application Number") or "").strip()
    beneficiary_name = str(prepared.get("Beneficiary Name") or "").strip()
    beneficiary_type = str(prepared.get("Beneficiary Type") or "").strip()

    if beneficiary_type == "District":
        prepared["Names"] = beneficiary_name
    elif application_number and beneficiary_name:
        prepared["Names"] = f"{application_number} - {beneficiary_name}"
    else:
        prepared["Names"] = application_number or beneficiary_name

    if beneficiary_type == "Public":
        prepared["R_Names"] = "AA_Public"
    elif beneficiary_type == "Institutions":
        prepared["R_Names"] = "A_Institutions"
    else:
        prepared["R_Names"] = beneficiary_name

    if sequence_no is not None:
        prepared["Sequence No"] = sequence_no
    elif "Sequence No" not in prepared:
        prepared["Sequence No"] = ""

    for header in list(prepared.keys()):
        value = prepared.get(header)
        if value is None or (isinstance(value, str) and not value.strip()):
            prepared[header] = "0" if header in TOKEN_GENERATION_NUMERIC_HEADERS else "N/A"
    return prepared


def _sequence_prepare_export_rows(rows):
    return [_sequence_prepare_export_row(row) for row in rows]


def _sequence_final_headers(seat_headers):
    headers = []
    inserted_names = False
    for header in seat_headers:
        headers.append(header)
        if header == "Application Number":
            headers.append("Names")
            inserted_names = True
    if not inserted_names:
        headers.append("Names")
    headers.extend(["Sequence No", "R_Names"])
    return headers


def _phase2_unique_headers(headers):
    unique_headers = []
    seen = set()
    for header in list(headers or []):
        key = str(header or "")
        if key in seen:
            continue
        seen.add(key)
        unique_headers.append(key)
    return unique_headers


def _token_generation_empty_value_summary(rows, headers):
    entries = []
    for header in headers:
        empty_count = 0
        for row in rows:
            value = row.get(header, "")
            if value is None or (isinstance(value, str) and not value.strip()):
                empty_count += 1
        if empty_count:
            entries.append(
                {
                    "column": header,
                    "count": empty_count,
                    "fill_value": "0" if header in TOKEN_GENERATION_NUMERIC_HEADERS else "N/A",
                }
            )
    return entries


def _token_generation_invalid_value_summary(rows, headers=None):
    checks = [
        ("Quantity", "Quantity"),
        ("Cost Per Unit", "Cost Per Unit"),
        ("Total Value", "Total Value"),
    ]
    header_set = {str(header or "").strip() for header in (headers or []) if str(header or "").strip()}
    entries = []
    for header, label in checks:
        if header_set and header not in header_set:
            continue
        invalid_count = 0
        for row in rows:
            value = _reconciliation_parse_decimal(row.get(header))
            if value < Decimal("1"):
                invalid_count += 1
        if invalid_count:
            entries.append(
                {
                    "column": header,
                    "label": label,
                    "count": invalid_count,
                }
            )
    return entries


def _token_generation_sequence_no(value):
    parsed = _phase2_parse_number(value)
    if parsed is None:
        return 10**9
    try:
        return int(parsed)
    except (TypeError, ValueError):
        return 10**9


def _token_generation_sort_order(row):
    beneficiary_type = str(row.get("Beneficiary Type") or "").strip()
    requested_item = str(row.get("Requested Item") or "")
    application_number = str(row.get("Application Number") or "").strip()
    handicapped_status = str(row.get("Handicapped Status") or "").strip().lower()
    names_value = str(row.get("Names") or "").strip()
    is_public_laptop = beneficiary_type == "Public" and "laptop" in requested_item.lower()
    p116_top = 0 if is_public_laptop and application_number == "P116" else 1
    handicap_first = 0 if handicapped_status == "yes" else 1
    return (
        _token_generation_sequence_no(row.get("Sequence No")),
        TOKEN_GENERATION_BENEFICIARY_ORDER.get(beneficiary_type, 99),
        p116_top,
        handicap_first,
        names_value.lower(),
        str(row.get("Requested Item") or "").strip().lower(),
        application_number.lower(),
    )


def _token_generation_apply_names_cleanup(row):
    names_value = str(row.get("Names") or "").strip()
    if names_value in TOKEN_GENERATION_RENAME_MAP:
        row["Names"] = TOKEN_GENERATION_RENAME_MAP[names_value]
    token_name = str(row.get("Token Name") or "").strip()
    if token_name == "Wet Grinder Floor 2L":
        row["Token Name"] = "Wet Grinder FLR 2L"
    return row


def _token_generation_token_print_flag(row):
    token_quantity = max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0)
    item_type = str(row.get("Item Type") or "").strip()
    requested_item = str(row.get("Requested Item") or "").strip()
    flag = 1 if item_type == models.ItemTypeChoices.ARTICLE and requested_item not in TOKEN_GENERATION_ARTICLE_PRINT_EXCLUDES else 0
    if token_quantity <= 0:
        flag = 0
    row["Token Print for ARTL"] = str(flag)
    return row


def _token_generation_apply_token_ranges(rows):
    running_end = 0
    for row in rows:
        token_quantity = max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0)
        if token_quantity > 0:
            start_token = running_end + 1
            end_token = running_end + token_quantity
            running_end = end_token
        else:
            start_token = 0
            end_token = 0
        row["Start Token No"] = str(start_token)
        row["End Token No"] = str(end_token)
    return rows


def _token_generation_quality_checks(rows):
    sequence_to_items = {}
    item_to_sequences = {}
    token_quantity_total = 0
    printable_token_total = 0
    zero_token_rows = 0
    duplicate_rows = 0
    duplicate_example = None
    row_counter = Counter()

    for row in rows:
        row_key = tuple((key, "" if row.get(key) is None else str(row.get(key))) for key in sorted(row.keys()))
        row_counter[row_key] += 1
        sequence_no = _phase2_parse_number(row.get("Sequence No"))
        requested_item = str(row.get("Requested Item") or "").strip()
        if sequence_no:
            sequence_to_items.setdefault(int(sequence_no), set()).add(requested_item)
        if requested_item:
            item_to_sequences.setdefault(requested_item, set())
            if sequence_no:
                item_to_sequences[requested_item].add(int(sequence_no))

        token_quantity = max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0)
        token_quantity_total += token_quantity
        if token_quantity == 0:
            zero_token_rows += 1
        if str(row.get("Token Print for ARTL") or "").strip() == "1":
            printable_token_total += token_quantity

    sequence_item_conflicts = [
        {
            "sequence_no": sequence_no,
            "items": sorted(item for item in items if item),
        }
        for sequence_no, items in sorted(sequence_to_items.items())
        if len({item for item in items if item}) > 1
    ]
    article_sequence_conflicts = [
        {
            "requested_item": requested_item,
            "sequence_numbers": sorted(sequence_numbers),
        }
        for requested_item, sequence_numbers in sorted(item_to_sequences.items())
        if len(sequence_numbers) > 1
    ]

    sequence_numbers = sorted(
        int(number)
        for number in {
            _phase2_parse_number(row.get("Sequence No"))
            for row in rows
        }
        if number
    )
    max_sequence = max(sequence_numbers) if sequence_numbers else 0
    missing_sequences = [number for number in range(1, max_sequence + 1) if number not in set(sequence_numbers)]

    for row_key, count in row_counter.items():
        if count > 1:
            duplicate_rows += count - 1
            if duplicate_example is None:
                row_map = dict(row_key)
                duplicate_example = _sequence_row_label(row_map)

    return {
        "sequence_item_conflicts": sequence_item_conflicts,
        "article_sequence_conflicts": article_sequence_conflicts,
        "missing_sequences": missing_sequences,
        "token_quantity_total": token_quantity_total,
        "printable_token_total": printable_token_total,
        "zero_token_rows": zero_token_rows,
        "duplicate_rows": duplicate_rows,
        "duplicate_example": duplicate_example or "",
    }


def _sequence_exact_compare(*, left_rows, left_headers, right_rows, right_headers, matched_label, mismatch_label):
    if list(left_headers) != list(right_headers):
        return {
            "matched": False,
            "details": (
                f"{mismatch_label}. Column mismatch: final has {len(left_headers)} column(s), "
                f"expected {len(right_headers)} column(s)."
            ),
        }

    left_counter = Counter(_sequence_row_key(row, left_headers) for row in left_rows)
    right_counter = Counter(_sequence_row_key(row, right_headers) for row in right_rows)
    if left_counter == right_counter:
        return {
            "matched": True,
            "details": matched_label,
        }

    missing_counter = right_counter - left_counter
    extra_counter = left_counter - right_counter
    parts = []
    if missing_counter:
        missing_key = next(iter(missing_counter))
        missing_row = {header: missing_key[idx] for idx, header in enumerate(right_headers)}
        parts.append(
            f"missing {sum(missing_counter.values())} row(s), e.g. {_sequence_row_label(missing_row)}"
        )
    if extra_counter:
        extra_key = next(iter(extra_counter))
        extra_row = {header: extra_key[idx] for idx, header in enumerate(left_headers)}
        parts.append(
            f"extra {sum(extra_counter.values())} row(s), e.g. {_sequence_row_label(extra_row)}"
        )

    return {
        "matched": False,
        "details": f"{mismatch_label}. " + " | ".join(parts),
    }


def _sequence_final_export_rows(session, sequence_map):
    seat_rows = list(
        models.SeatAllocationRow.objects.filter(session=session).order_by(
            "sort_order",
            "requested_item",
            "application_number",
            "id",
        )
    )
    seat_export_rows, seat_export_headers = _phase2_export_rows(seat_rows, include_sequence=False)
    final_export_rows = []
    for row in seat_export_rows:
        item_name = str(row.get("Requested Item") or "").strip()
        final_row = dict(row)
        final_row["Sequence No"] = sequence_map.get(item_name, "")
        final_export_rows.append(final_row)
    return {
        "seat_rows": seat_export_rows,
        "seat_headers": seat_export_headers,
        "final_rows": final_export_rows,
        "final_headers": [*seat_export_headers, "Sequence No"],
    }


def _sequence_map_from_seat_allocation(session):
    sequence_map = {}
    for item_name, sequence_no in (
        models.SeatAllocationRow.objects.filter(session=session)
        .exclude(sequence_no__isnull=True)
        .values_list("requested_item", "sequence_no")
    ):
        item_name = str(item_name or "").strip()
        if item_name and sequence_no and item_name not in sequence_map:
            sequence_map[item_name] = int(sequence_no)
    return sequence_map


def _token_generation_headers(base_headers):
    headers = []
    inserted_names = False
    base_headers = _phase2_unique_headers(base_headers)
    for header in list(base_headers or []):
        headers.append(header)
        if header == "Application Number" and "Names" not in headers:
            headers.append("Names")
            inserted_names = True
    if not inserted_names and "Names" not in headers:
        headers.append("Names")
    if "Token Print for ARTL" not in headers:
        headers.append("Token Print for ARTL")
    return _phase2_unique_headers(headers)


def _token_generation_generated_headers(base_headers):
    headers = _phase2_unique_headers(base_headers)
    if "Start Token No" not in headers:
        headers.append("Start Token No")
    if "End Token No" not in headers:
        headers.append("End Token No")
    return _phase2_unique_headers(headers)


def _token_generation_prepare_row(row):
    prepared = dict(row or {})
    application_number = str(prepared.get("Application Number") or "").strip()
    beneficiary_name = str(prepared.get("Beneficiary Name") or "").strip()
    beneficiary_type = str(prepared.get("Beneficiary Type") or "").strip()
    if beneficiary_type == "District":
        prepared["Names"] = beneficiary_name
    elif application_number and beneficiary_name:
        prepared["Names"] = f"{application_number} - {beneficiary_name}"
    else:
        prepared["Names"] = application_number or beneficiary_name
    for header in list(prepared.keys()):
        value = prepared.get(header)
        if value is None or (isinstance(value, str) and not value.strip()):
            prepared[header] = "0" if header in TOKEN_GENERATION_NUMERIC_HEADERS else "N/A"
    prepared = _token_generation_apply_names_cleanup(prepared)
    prepared = _token_generation_token_print_flag(prepared)
    return prepared


def _token_generation_prepare_dataset(rows, headers):
    prepared_headers = _token_generation_headers(headers)
    prepared_rows = [_token_generation_prepare_row(row) for row in rows]
    prepared_rows = _token_generation_sort_dataset(prepared_rows)
    blank_summary = _token_generation_empty_value_summary(rows, prepared_headers)
    quality_checks = _token_generation_quality_checks(prepared_rows)
    return {
        "headers": prepared_headers,
        "rows": prepared_rows,
        "blank_summary": blank_summary,
        "quality_checks": quality_checks,
    }


def _token_generation_sort_dataset(rows):
    sorted_rows = [dict(row) for row in rows]
    sorted_rows.sort(key=_token_generation_sort_order)
    return sorted_rows


def _token_generation_generate_dataset(rows, headers):
    generated_rows = _token_generation_sort_dataset(rows)
    generated_rows = _token_generation_apply_token_ranges(generated_rows)
    return {
        "headers": _token_generation_generated_headers(headers),
        "rows": generated_rows,
    }


def _token_generation_is_sorted(rows):
    if not rows:
        return False
    return [dict(row) for row in rows] == _token_generation_sort_dataset(rows)


def _token_generation_is_generated(rows):
    if not rows:
        return False
    has_any_token = False
    for row in rows:
        token_quantity = max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0)
        start_token = _phase2_parse_number(row.get("Start Token No"))
        end_token = _phase2_parse_number(row.get("End Token No"))
        if token_quantity > 0:
            has_any_token = True
            if not start_token or not end_token:
                return False
        else:
            if (start_token or 0) != 0 or (end_token or 0) != 0:
                return False
    return has_any_token


def _token_generation_edit_candidates(rows, *, length_limit):
    unique_names = []
    unique_token_names = []
    seen_names = set()
    seen_token_names = set()
    for row in rows:
        names_value = str(row.get("Names") or "").strip()
        token_name_value = str(row.get("Token Name") or "").strip()
        if len(names_value) > length_limit and names_value not in seen_names:
            seen_names.add(names_value)
            unique_names.append({"value": names_value, "length": len(names_value)})
        if len(token_name_value) > length_limit and token_name_value not in seen_token_names:
            seen_token_names.add(token_name_value)
            unique_token_names.append({"value": token_name_value, "length": len(token_name_value)})
    return {
        "names": unique_names,
        "token_names": unique_token_names,
    }


def _token_generation_parse_rule_lines(raw_value):
    lines = []
    seen = set()
    for line in str(raw_value or "").splitlines():
        value = line.strip()
        normalized = value.lower()
        if value and normalized not in seen:
            seen.add(normalized)
            lines.append(value)
    return lines


def _token_generation_filter_state(request):
    return {
        "application_number": str(request.GET.get("filter_application_number") or "").strip(),
        "beneficiary_name": str(request.GET.get("filter_beneficiary_name") or "").strip(),
        "beneficiary_type": str(request.GET.get("filter_beneficiary_type") or "").strip(),
        "requested_item": str(request.GET.get("filter_requested_item") or "").strip(),
        "item_type": str(request.GET.get("filter_item_type") or "").strip(),
        "comments": str(request.GET.get("filter_comments") or "").strip(),
    }


def _token_generation_filter_rows(rows, filters):
    entries = []
    for index, row in enumerate(rows):
        application_number = str(row.get("Application Number") or "").strip()
        beneficiary_name = str(row.get("Beneficiary Name") or "").strip()
        beneficiary_type = str(row.get("Beneficiary Type") or "").strip()
        requested_item = str(row.get("Requested Item") or "").strip()
        item_type = str(row.get("Item Type") or "").strip()
        comments = str(row.get("Comments") or "").strip()

        if filters["application_number"] and filters["application_number"].lower() not in application_number.lower():
            continue
        if filters["beneficiary_name"] and filters["beneficiary_name"].lower() not in beneficiary_name.lower():
            continue
        if filters["beneficiary_type"] and filters["beneficiary_type"].lower() != beneficiary_type.lower():
            continue
        if filters["requested_item"] and filters["requested_item"].lower() not in requested_item.lower():
            continue
        if filters["item_type"] and filters["item_type"].lower() != item_type.lower():
            continue
        if filters["comments"] and filters["comments"].lower() not in comments.lower():
            continue

        entries.append(
            {
                "row_index": index,
                "application_number": application_number,
                "beneficiary_name": beneficiary_name,
                "beneficiary_type": beneficiary_type,
                "requested_item": requested_item,
                "item_type": item_type,
                "comments": comments,
            }
        )
    return entries


def _token_generation_has_active_filters(filters):
    return any(str(value or "").strip() for value in (filters or {}).values())


def _token_generation_article_toggle_rows(rows):
    article_rows = {}
    for row in rows:
        if str(row.get("Item Type") or "").strip() != models.ItemTypeChoices.ARTICLE:
            continue
        requested_item = str(row.get("Requested Item") or "").strip()
        if not requested_item:
            continue
        article_rows.setdefault(requested_item, []).append(row)

    entries = []
    for requested_item in sorted(article_rows.keys(), key=str.lower):
        item_rows = article_rows[requested_item]
        skip_label = all(str(row.get("Token Print for ARTL") or "").strip() == "0" for row in item_rows)
        token_total = sum(max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0) for row in item_rows)
        entries.append(
            {
                "requested_item": requested_item,
                "token_total": token_total,
                "skip_label": skip_label,
            }
        )
    return entries


def _token_generation_source_dataset(session):
    seat_rows = models.SeatAllocationRow.objects.filter(session=session).order_by(
        F("sequence_no").asc(nulls_last=True),
        "sort_order",
        "requested_item",
        "application_number",
        "id",
    )
    source_rows, source_headers = _phase2_export_rows(seat_rows, include_sequence=True)
    return {
        "headers": source_headers,
        "rows": source_rows,
        "blank_summary": _token_generation_empty_value_summary(source_rows, source_headers),
        "quality_checks": _token_generation_quality_checks(source_rows),
    }


def _token_generation_store_dataset(*, session, dataset, source_name, user):
    models.TokenGenerationRow.objects.filter(session=session).delete()
    rows = dataset["rows"]
    headers = _phase2_unique_headers(dataset["headers"])
    models.TokenGenerationRow.objects.bulk_create(
        [
            models.TokenGenerationRow(
                session=session,
                source_file_name=source_name,
                application_number=str(row.get("Application Number") or "").strip() or None,
                beneficiary_name=str(row.get("Beneficiary Name") or "").strip() or None,
                requested_item=str(row.get("Requested Item") or "").strip() or None,
                beneficiary_type=str(row.get("Beneficiary Type") or "").strip() or None,
                sequence_no=_phase2_parse_number(row.get("Sequence No")) or None,
                start_token_no=_phase2_parse_number(row.get("Start Token No")) or 0,
                end_token_no=_phase2_parse_number(row.get("End Token No")) or 0,
                row_data=row,
                headers=headers,
                sort_order=index + 1,
                created_by=user,
                updated_by=user,
            )
            for index, row in enumerate(rows)
        ]
    )


def _token_generation_saved_dataset(session):
    rows = list(
        models.TokenGenerationRow.objects.filter(session=session).order_by(
            "sort_order",
            F("sequence_no").asc(nulls_last=True),
            "requested_item",
            "application_number",
            "id",
        )
    )
    if not rows:
        return {"headers": [], "rows": [], "source_name": "", "saved_at": None}
    headers = _phase2_unique_headers(rows[0].headers or [])
    prepared_rows = []
    for row in rows:
        row_data = dict(row.row_data or {})
        row_data["Application Number"] = row.application_number or row_data.get("Application Number") or ""
        row_data["Beneficiary Name"] = row.beneficiary_name or row_data.get("Beneficiary Name") or ""
        row_data["Requested Item"] = row.requested_item or row_data.get("Requested Item") or ""
        row_data["Beneficiary Type"] = row.beneficiary_type or row_data.get("Beneficiary Type") or ""
        row_data["Sequence No"] = row.sequence_no if row.sequence_no is not None else row_data.get("Sequence No") or ""
        row_data["Start Token No"] = row.start_token_no if row.start_token_no is not None else row_data.get("Start Token No") or 0
        row_data["End Token No"] = row.end_token_no if row.end_token_no is not None else row_data.get("End Token No") or 0
        existing_token_quantity = row_data.get("Token Quantity")
        if existing_token_quantity in {None, ""}:
            existing_token_quantity = max(
                (row.end_token_no or 0) - (row.start_token_no or 0) + 1,
                0,
            )
        row_data["Token Quantity"] = existing_token_quantity
        prepared_rows.append(row_data)
    return {
        "headers": headers,
        "rows": prepared_rows,
        "source_name": rows[0].source_file_name or "",
        "saved_at": rows[0].updated_at,
    }


def _token_generation_stage_state(request, session):
    state_map = request.session.get("token_generation_stage_state", {})
    return dict(state_map.get(str(session.pk), {}))


def _token_generation_set_stage_state(request, session, **updates):
    state_map = dict(request.session.get("token_generation_stage_state", {}))
    session_state = dict(state_map.get(str(session.pk), {}))
    session_state.update(updates)
    state_map[str(session.pk)] = session_state
    request.session["token_generation_stage_state"] = state_map
    request.session.modified = True


def _token_generation_latest_source_marker(session):
    latest_source_updated_at = (
        models.SeatAllocationRow.objects.filter(session=session)
        .order_by("-updated_at", "-created_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    if not latest_source_updated_at:
        return ""
    return timezone.localtime(latest_source_updated_at).isoformat()


def _token_generation_sync_required(request, session, dataset=None):
    dataset = dataset or _token_generation_saved_dataset(session)
    if not dataset["rows"]:
        return False
    source_name = str(dataset.get("source_name") or "")
    if not source_name.startswith("Synced from Sequence List"):
        return False
    stage_state = _token_generation_stage_state(request, session)
    return str(stage_state.get("source_sync_marker") or "") != _token_generation_latest_source_marker(session)


LABELS_DEFAULT_2L_ITEMS = [
    "Tiffen Set",
    "Tiffen Set + Alu Idli Box + MS Stove 2 Burner",
    "Tiffen Set + MS Stove 2 Burner",
    "Tiffen Set + Tea Can 10 Ltrs SS",
    "Push Cart Without Top",
    "Push Cart With Top",
    "Push Cart With Top + Alu Idli Box + MS Stove 2 Burner",
    "Office Table 4 X 2",
    "S Type Chair",
    "Steel Cupboard 6 1/2'",
]


def _labels_saved_dataset(session):
    rows = list(
        models.LabelGenerationRow.objects.filter(session=session).order_by(
            "sort_order",
            F("sequence_no").asc(nulls_last=True),
            "requested_item",
            "application_number",
            "id",
        )
    )
    if not rows:
        return {"headers": [], "rows": [], "source_name": "", "saved_at": None}
    headers = _phase2_unique_headers(rows[0].headers or [])
    return {
        "headers": headers,
        "rows": [dict(row.row_data or {}) for row in rows],
        "source_name": rows[0].source_file_name or "",
        "saved_at": rows[0].updated_at,
    }


def _labels_store_dataset(*, session, dataset, source_name, user):
    rows = [dict(row) for row in list(dataset.get("rows") or [])]
    headers = _phase2_unique_headers(dataset.get("headers") or [])
    with transaction.atomic():
        models.LabelGenerationRow.objects.filter(session=session).delete()
        models.LabelGenerationRow.objects.bulk_create(
            [
                models.LabelGenerationRow(
                    session=session,
                    source_file_name=source_name or "",
                    application_number=str(row.get("Application Number") or "").strip() or None,
                    beneficiary_name=str(row.get("Beneficiary Name") or "").strip() or None,
                    requested_item=str(row.get("Requested Item") or "").strip() or None,
                    beneficiary_type=str(row.get("Beneficiary Type") or "").strip() or None,
                    sequence_no=_phase2_parse_number(row.get("Sequence No")) or None,
                    start_token_no=_phase2_parse_number(row.get("Start Token No")) or 0,
                    end_token_no=_phase2_parse_number(row.get("End Token No")) or 0,
                    row_data=row,
                    headers=headers,
                    sort_order=index + 1,
                    created_by=user,
                    updated_by=user,
                )
                for index, row in enumerate(rows)
            ]
        )


def _labels_source_dataset(session):
    dataset = _token_generation_saved_dataset(session)
    return _labels_normalize_dataset({
        "headers": _phase2_unique_headers(dataset.get("headers") or []),
        "rows": [dict(row) for row in list(dataset.get("rows") or [])],
    })


def _labels_stage_state(request, session):
    state_map = request.session.get("labels_stage_state", {})
    session_state = dict(state_map.get(str(session.pk), {}))
    selected_items = session_state.get("large_items")
    if not isinstance(selected_items, list):
        selected_items = list(LABELS_DEFAULT_2L_ITEMS)
    session_state["large_items"] = selected_items
    session_state["large_items_saved"] = bool(session_state.get("large_items_saved"))
    return session_state


def _labels_set_stage_state(request, session, **updates):
    state_map = dict(request.session.get("labels_stage_state", {}))
    session_state = dict(state_map.get(str(session.pk), {}))
    session_state.update(updates)
    state_map[str(session.pk)] = session_state
    request.session["labels_stage_state"] = state_map
    request.session.modified = True


def _labels_latest_source_marker(session):
    latest_source_updated_at = (
        models.TokenGenerationRow.objects.filter(session=session)
        .order_by("-updated_at", "-created_at")
        .values_list("updated_at", flat=True)
        .first()
    )
    if not latest_source_updated_at:
        return ""
    return timezone.localtime(latest_source_updated_at).isoformat()


def _labels_sync_required(request, session, dataset=None):
    dataset = dataset or _labels_saved_dataset(session)
    if not dataset["rows"]:
        return False
    source_name = str(dataset.get("source_name") or "")
    if not source_name.startswith("Synced from Token Generation"):
        return False
    stage_state = _labels_stage_state(request, session)
    return str(stage_state.get("source_sync_marker") or "") != _labels_latest_source_marker(session)


def _labels_has_generated_tokens(rows):
    return bool(rows) and all(
        _phase2_parse_number(row.get("Start Token No")) is not None
        and _phase2_parse_number(row.get("End Token No")) is not None
        for row in rows
    )


def _labels_available_requested_items(rows):
    items = []
    seen = set()
    for row in rows:
        requested_item = str(row.get("Requested Item") or "").strip()
        if not requested_item or requested_item in seen:
            continue
        seen.add(requested_item)
        items.append(requested_item)
    return items


def _labels_expand_entries(rows, *, row_filter=None, group_by=None, sort_key=None):
    filtered_rows = []
    for row in rows:
        if row_filter and not row_filter(row):
            continue
        filtered_rows.append(dict(row))
    if sort_key:
        filtered_rows.sort(key=sort_key)

    entries = []
    for row in filtered_rows:
        start = _phase2_parse_number(row.get("Start Token No")) or 0
        end = _phase2_parse_number(row.get("End Token No")) or 0
        if start <= 0 or end < start:
            continue
        name_value = str(row.get("Names") or row.get("Beneficiary Name") or "").strip()
        article_value = str(row.get("Token Name") or row.get("Requested Item") or "").strip()
        group_value = group_by(row) if group_by else ""
        for token in range(start, end + 1):
            entries.append(
                {
                    "token": str(token),
                    "name": name_value,
                    "article": article_value,
                    "group": str(group_value or ""),
                }
            )
    return entries


def _labels_audit_download(rows, *, download_kind, large_items):
    large_items = set(large_items or [])

    def token_qty(row):
        return max(_phase2_parse_number(row.get("Token Quantity")) or 0, 0)

    def is_printable_article(row):
        return token_qty(row) > 0 and (_phase2_parse_number(row.get("Token Print for ARTL")) or 0) != 0

    def sort_by_name_and_start(row):
        return (
            str(row.get("Names") or "").strip(),
            _phase2_parse_number(row.get("Start Token No")) or 0,
        )

    def sort_by_item_and_start(row):
        return (
            str(row.get("Token Name") or row.get("Requested Item") or "").strip(),
            str(row.get("Names") or "").strip(),
            _phase2_parse_number(row.get("Start Token No")) or 0,
        )

    row_filter = None
    group_by = None
    sort_key = None
    labels_per_page = 12

    if download_kind in {"article_12l_separate", "article_12l_continuous"}:
        row_filter = lambda row: is_printable_article(row) and str(row.get("Requested Item") or "").strip() not in large_items
        if download_kind == "article_12l_separate":
            group_by = lambda row: str(row.get("Token Name") or row.get("Requested Item") or "").strip()
    elif download_kind == "article_2l_continuous":
        row_filter = lambda row: token_qty(row) > 0 and str(row.get("Requested Item") or "").strip() in large_items
        labels_per_page = 2
    elif download_kind in {"district_separate", "district_continuous"}:
        row_filter = lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "District"
        sort_key = sort_by_name_and_start
        if download_kind == "district_separate":
            group_by = lambda row: str(row.get("Names") or "").strip()
    elif download_kind == "institution":
        row_filter = lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "Institutions"
        sort_key = sort_by_name_and_start
    elif download_kind == "public":
        row_filter = lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "Public"
        sort_key = sort_by_item_and_start
    elif download_kind in {"chair_separate", "chair_continuous"}:
        if download_kind == "chair_separate":
            row_filter = lambda row: token_qty(row) > 0
            group_by = lambda row: str(row.get("Token Name") or row.get("Requested Item") or "").strip()
        else:
            row_filter = lambda row: token_qty(row) > 0
    else:
        return {
            "ready": False,
            "status_label": "Needs Review",
            "status_class": "bad",
            "reason": "Unknown label download type.",
            "included_rows": 0,
            "expected_labels": 0,
            "actual_labels": 0,
            "first_token": 0,
            "last_token": 0,
            "duplicate_tokens": 0,
            "invalid_range_rows": 0,
            "missing_labels": 0,
            "page_count": 0,
        }

    filtered_rows = [dict(row) for row in rows if not row_filter or row_filter(row)]
    expected_labels = sum(token_qty(row) for row in filtered_rows)
    invalid_range_rows = 0
    for row in filtered_rows:
        start = _phase2_parse_number(row.get("Start Token No")) or 0
        end = _phase2_parse_number(row.get("End Token No")) or 0
        qty = token_qty(row)
        if qty <= 0:
            continue
        if start <= 0 or end < start or (end - start + 1) != qty:
            invalid_range_rows += 1

    entries = _labels_expand_entries(filtered_rows, group_by=group_by, sort_key=sort_key)
    tokens = [int(entry["token"]) for entry in entries if str(entry.get("token") or "").strip().isdigit()]
    actual_labels = len(entries)
    duplicate_tokens = max(actual_labels - len(set(tokens)), 0)
    missing_labels = max(expected_labels - actual_labels, 0)
    page_count = 0
    if actual_labels:
        if group_by:
            grouped_counts = {}
            for entry in entries:
                group_key = str(entry.get("group") or "")
                grouped_counts[group_key] = grouped_counts.get(group_key, 0) + 1
            page_count = sum(((count - 1) // labels_per_page) + 1 for count in grouped_counts.values() if count > 0)
        else:
            page_count = ((actual_labels - 1) // labels_per_page) + 1
    ready = expected_labels > 0 and duplicate_tokens == 0 and invalid_range_rows == 0 and missing_labels == 0
    reason = ""
    status_label = "Data Ready"
    status_class = "ok"
    if expected_labels <= 0:
        status_label = "No Matching Rows"
        status_class = "neutral"
        reason = "No matching token rows are currently available for this label type."
    elif duplicate_tokens > 0:
        status_label = "Needs Review"
        status_class = "bad"
        reason = f"{duplicate_tokens} duplicate token number(s) found."
    elif invalid_range_rows > 0:
        status_label = "Needs Review"
        status_class = "bad"
        reason = f"{invalid_range_rows} row(s) have invalid token ranges."
    elif missing_labels > 0:
        status_label = "Needs Review"
        status_class = "bad"
        reason = f"{missing_labels} expected label(s) are missing from token ranges."

    return {
        "ready": ready,
        "status_label": status_label,
        "status_class": status_class,
        "reason": reason,
        "included_rows": len(filtered_rows),
        "expected_labels": expected_labels,
        "actual_labels": actual_labels,
        "first_token": min(tokens) if tokens else 0,
        "last_token": max(tokens) if tokens else 0,
        "duplicate_tokens": duplicate_tokens,
        "invalid_range_rows": invalid_range_rows,
        "missing_labels": missing_labels,
        "page_count": page_count,
    }


def _labels_normalize_row(row):
    normalized = dict(row or {})
    if "Token Name" not in normalized or not str(normalized.get("Token Name") or "").strip():
        legacy_token_name = str(normalized.get("Requested Item Tk") or "").strip()
        normalized["Token Name"] = legacy_token_name or str(normalized.get("Requested Item") or "").strip()
    if "Start Token No" not in normalized:
        normalized["Start Token No"] = normalized.get("Start Token No.", "")
    if "End Token No" not in normalized:
        normalized["End Token No"] = normalized.get("End Token No.", "")
    if "Names" not in normalized or not str(normalized.get("Names") or "").strip():
        application_number = str(normalized.get("Application Number") or "").strip()
        beneficiary_name = str(normalized.get("Beneficiary Name") or "").strip()
        beneficiary_type = str(normalized.get("Beneficiary Type") or "").strip()
        if beneficiary_type == "District":
            normalized["Names"] = beneficiary_name
        else:
            normalized["Names"] = f"{application_number} - {beneficiary_name}".strip(" -") if (application_number or beneficiary_name) else ""
    start_token = _phase2_parse_number(normalized.get("Start Token No"))
    end_token = _phase2_parse_number(normalized.get("End Token No"))
    if "Token Quantity" not in normalized or not str(normalized.get("Token Quantity") or "").strip():
        normalized["Token Quantity"] = str(max(end_token - start_token + 1, 0) if start_token and end_token >= start_token else 0)
    if "Token Print for ARTL" not in normalized or not str(normalized.get("Token Print for ARTL") or "").strip():
        normalized["Token Print for ARTL"] = "1" if (_phase2_parse_number(normalized.get("Token Quantity")) or 0) > 0 else "0"
    return normalized


def _labels_normalize_headers(headers, rows):
    normalized_headers = []
    for header in _phase2_unique_headers(headers):
        if header == "Requested Item Tk":
            if "Token Name" not in normalized_headers:
                normalized_headers.append("Token Name")
            continue
        if header == "Start Token No.":
            if "Start Token No" not in normalized_headers:
                normalized_headers.append("Start Token No")
            continue
        if header == "End Token No.":
            if "End Token No" not in normalized_headers:
                normalized_headers.append("End Token No")
            continue
        if header not in normalized_headers:
            normalized_headers.append(header)
    for required_header in ["Names", "Token Name", "Token Quantity", "Token Print for ARTL", "Start Token No", "End Token No"]:
        if any(required_header in row for row in rows) and required_header not in normalized_headers:
            normalized_headers.append(required_header)
    return _phase2_unique_headers(normalized_headers)


def _labels_normalize_dataset(dataset):
    rows = [_labels_normalize_row(row) for row in list(dataset.get("rows") or [])]
    headers = _labels_normalize_headers(dataset.get("headers") or [], rows)
    return {
        "headers": headers,
        "rows": rows,
    }


def _labels_download_filename(prefix):
    timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
    return f"{prefix}_{timestamp}.pdf"


def _sequence_seat_allocation_integrity(session):
    source_rows = _phase2_master_export_rows()
    seat_rows = list(models.SeatAllocationRow.objects.filter(session=session))
    grouped_rows = [
        {
            "application_number": row.application_number or "",
            "beneficiary_name": row.beneficiary_name or "",
            "district": row.district or "",
            "requested_item": row.requested_item or "",
            "quantity": int(row.quantity or 0),
            "waiting_hall_quantity": int(row.waiting_hall_quantity or 0),
            "token_quantity": int(row.token_quantity or 0),
            "beneficiary_type": row.beneficiary_type or "",
            "item_type": row.item_type or "",
            "comments": row.comments or "",
            "master_row": row.master_row or {},
            "master_headers": row.master_headers or [],
        }
        for row in seat_rows
    ]
    snapshot = _phase2_reconciliation_snapshot(
        source_rows=source_rows,
        grouped_rows=grouped_rows,
        total_value_getter=lambda row: _phase2_parse_number(
            row.get("Total Value")
            if isinstance(row, dict) and "Total Value" in row
            else (row.get("master_row") or {}).get("Total Value")
        ),
    )
    return {
        "snapshot": snapshot,
        "checks": _phase2_reconciliation_checks(snapshot),
        "split_check": _phase2_split_reconciliation(seat_rows),
    }


def _phase2_build_rows_from_master_export_rows(rows, *, source_file_name):
    headers = list(rows[0].keys()) if rows else []
    order = 0
    source_rows_for_reconciliation = []
    raw_rows = []
    for source_row in rows:
        application_number = str(source_row.get("Application Number") or "").strip()
        beneficiary_name = str(source_row.get("Beneficiary Name") or "").strip()
        requested_item = str(source_row.get("Requested Item") or "").strip()
        beneficiary_type = str(source_row.get("Beneficiary Type") or "").strip()
        item_type = str(source_row.get("Item Type") or "").strip()
        comments = str(source_row.get("Comments") or "").strip()
        quantity = _phase2_parse_number(source_row.get("Quantity"))
        if not (application_number or beneficiary_name or requested_item or quantity):
            continue
        source_rows_for_reconciliation.append(source_row)
        district = beneficiary_name if _phase2_normalize_text(beneficiary_type) == "district" else "Non-District"
        master_row = {header: source_row.get(header, "") for header in headers}
        order += 1
        raw_rows.append(
            {
                "source_file_name": source_file_name,
                "application_number": application_number,
                "beneficiary_name": beneficiary_name,
                "district": district,
                "requested_item": requested_item,
                "quantity": quantity,
                "waiting_hall_quantity": 0,
                "token_quantity": quantity,
                "beneficiary_type": beneficiary_type,
                "item_type": item_type,
                "comments": comments,
                "master_row": master_row,
                "master_headers": headers,
                "sort_order": order,
            }
        )
    reconciliation_snapshot = _phase2_reconciliation_snapshot(
        source_rows=source_rows_for_reconciliation,
        grouped_rows=raw_rows,
        total_value_getter=lambda row: _phase2_parse_number(
            row.get("Total Value")
            if isinstance(row, dict) and "Total Value" in row
            else (row.get("master_row") or {}).get("Total Value")
        ),
    )
    return {
        "rows": raw_rows,
        "headers": headers,
        **reconciliation_snapshot,
        "reconciliation_snapshot": reconciliation_snapshot,
    }


def _phase2_master_export_rows():
    district_rows = _district_export_rows(_build_district_entry_summaries())
    public_rows = _public_export_rows(models.PublicBeneficiaryEntry.objects.select_related("article").all())
    institution_rows = _institution_export_rows(_build_institution_entry_summaries())
    return district_rows + public_rows + institution_rows


def _phase2_replace_session_rows(session, upload_rows, *, source_file_name, user, reconciliation=None):
    with transaction.atomic():
        models.SeatAllocationRow.objects.filter(session=session).delete()
        for row in upload_rows:
            models.SeatAllocationRow.objects.create(
                session=session,
                source_file_name=source_file_name,
                application_number=row["application_number"],
                beneficiary_name=row["beneficiary_name"],
                district=row["district"],
                requested_item=row["requested_item"],
                quantity=row["quantity"],
                waiting_hall_quantity=row["waiting_hall_quantity"],
                token_quantity=row["token_quantity"],
                beneficiary_type=row["beneficiary_type"],
                item_type=row["item_type"],
                comments=row["comments"],
                master_row=row["master_row"],
                master_headers=row["master_headers"],
                sort_order=row["sort_order"],
                sequence_no=row.get("sequence_no"),
                created_by=user,
                updated_by=user,
            )
        if reconciliation is None:
            reconciliation = {}
        session.phase2_source_name = source_file_name
        session.phase2_source_row_count = int(reconciliation.get("source_row_count") or len(upload_rows))
        session.phase2_grouped_row_count = int(reconciliation.get("grouped_row_count") or len(upload_rows))
        session.phase2_source_quantity_total = int(
            reconciliation.get("source_quantity_total")
            if reconciliation.get("source_quantity_total") is not None
            else sum(int(row.get("quantity") or 0) for row in upload_rows)
        )
        session.phase2_grouped_quantity_total = int(
            reconciliation.get("grouped_quantity_total")
            if reconciliation.get("grouped_quantity_total") is not None
            else sum(int(row.get("quantity") or 0) for row in upload_rows)
        )
        session.phase2_reconciliation_snapshot = reconciliation.get("reconciliation_snapshot") or {}
        session.save(
            update_fields=[
                "phase2_source_name",
                "phase2_source_row_count",
                "phase2_grouped_row_count",
                "phase2_source_quantity_total",
                "phase2_grouped_quantity_total",
                "phase2_reconciliation_snapshot",
                "updated_at",
            ]
        )


def _phase2_build_upload_rows(uploaded_file):
    headers, uploaded_rows = _tabular_rows_from_upload(uploaded_file)
    if not headers:
        raise ValueError("Uploaded file is empty.")

    normalized_headers = {_phase2_normalize_text(header): header for header in headers}
    missing = [header for header in PHASE2_MASTER_REQUIRED_HEADERS if _phase2_normalize_text(header) not in normalized_headers]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    quantity_header = normalized_headers[_phase2_normalize_text("Quantity")]
    application_header = normalized_headers[_phase2_normalize_text("Application Number")]
    beneficiary_header = normalized_headers[_phase2_normalize_text("Beneficiary Name")]
    requested_item_header = normalized_headers[_phase2_normalize_text("Requested Item")]
    beneficiary_type_header = normalized_headers[_phase2_normalize_text("Beneficiary Type")]
    item_type_header = normalized_headers[_phase2_normalize_text("Item Type")]
    comments_header = normalized_headers.get(_phase2_normalize_text("Comments"))
    total_value_header = normalized_headers.get(_phase2_normalize_text("Total Value"))
    waiting_hall_header = normalized_headers.get(_phase2_normalize_text("Waiting Hall Quantity"))
    token_header = normalized_headers.get(_phase2_normalize_text("Token Quantity"))

    order = 0
    source_rows_for_reconciliation = []
    raw_rows = []
    for source_row in uploaded_rows:
        application_number = str(source_row.get(application_header) or "").strip()
        beneficiary_name = str(source_row.get(beneficiary_header) or "").strip()
        requested_item = str(source_row.get(requested_item_header) or "").strip()
        beneficiary_type = str(source_row.get(beneficiary_type_header) or "").strip()
        item_type = str(source_row.get(item_type_header) or "").strip()
        comments = str(source_row.get(comments_header) or "").strip() if comments_header else ""
        quantity = _phase2_parse_number(source_row.get(quantity_header))
        if not (application_number or beneficiary_name or requested_item or quantity):
            continue
        waiting_hall_quantity = _phase2_parse_number(source_row.get(waiting_hall_header)) if waiting_hall_header else None
        token_quantity = _phase2_parse_number(source_row.get(token_header)) if token_header else None
        if waiting_hall_quantity is None and token_quantity is None:
            waiting_hall_quantity = 0
            token_quantity = quantity
        elif waiting_hall_quantity is None:
            token_quantity = max(min(token_quantity, quantity), 0)
            waiting_hall_quantity = max(quantity - token_quantity, 0)
        else:
            waiting_hall_quantity = max(min(waiting_hall_quantity, quantity), 0)
            token_quantity = max(quantity - waiting_hall_quantity, 0)
        normalized_source_row = {header: source_row.get(header, "") for header in headers}
        source_rows_for_reconciliation.append(
            {
                "Application Number": normalized_source_row.get(application_header, ""),
                "Beneficiary Name": normalized_source_row.get(beneficiary_header, ""),
                "Requested Item": normalized_source_row.get(requested_item_header, ""),
                "Quantity": normalized_source_row.get(quantity_header, ""),
                "Beneficiary Type": normalized_source_row.get(beneficiary_type_header, ""),
                "Item Type": normalized_source_row.get(item_type_header, ""),
                "Total Value": normalized_source_row.get(total_value_header, "") if total_value_header else "",
            }
        )

        district = beneficiary_name if _phase2_normalize_text(beneficiary_type) == "district" else "Non-District"
        master_row = normalized_source_row
        order += 1
        raw_rows.append(
            {
                "application_number": application_number,
                "beneficiary_name": beneficiary_name,
                "district": district,
                "requested_item": requested_item,
                "quantity": quantity,
                "waiting_hall_quantity": waiting_hall_quantity,
                "token_quantity": token_quantity,
                "beneficiary_type": beneficiary_type,
                "item_type": item_type,
                "comments": comments,
                "master_row": master_row,
                "master_headers": headers,
                "sort_order": order,
            }
        )
    reconciliation_snapshot = _phase2_reconciliation_snapshot(
        source_rows=source_rows_for_reconciliation,
        grouped_rows=raw_rows,
        total_value_getter=lambda row: _phase2_parse_number(
            row.get("Total Value")
            if isinstance(row, dict) and "Total Value" in row
            else (row.get("master_row") or {}).get(total_value_header, "")
        ),
    )
    return {
        "rows": raw_rows,
        "headers": headers,
        **reconciliation_snapshot,
        "reconciliation_snapshot": reconciliation_snapshot,
    }


class SeatAllocationListView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.SEAT_ALLOCATION
    permission_action = "view"
    template_name = "dashboard/seat_allocation_list.html"
    preserved_filter_keys = ("q", "beneficiary_type", "district_filter", "item_filter", "sort", "dir")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        submit_result = self.request.session.pop("seat_allocation_submit_result", None)
        session = kwargs.get("selected_session") or _phase2_selected_session(self.request)
        all_rows = []
        if session:
            all_rows = list(
                models.SeatAllocationRow.objects.filter(session=session).order_by(
                    "sort_order", "district", "requested_item", "application_number", "id"
                )
            )

        beneficiary_type_filter = (self.request.GET.get("beneficiary_type") or "all").strip()
        q = (self.request.GET.get("q") or "").strip().lower()
        district_filter = (self.request.GET.get("district_filter") or "all").strip()
        item_filter = (self.request.GET.get("item_filter") or "all").strip()
        sort_key = (self.request.GET.get("sort") or "").strip()
        sort_dir = "asc" if (self.request.GET.get("dir") or "asc").lower() == "asc" else "desc"

        normalized_type = _phase2_normalize_text(beneficiary_type_filter)
        is_district_type = normalized_type == "district"
        is_institution_like_type = normalized_type in {"institutions", "others"}

        type_scoped_rows = [
            row for row in all_rows
            if beneficiary_type_filter == "all"
            or (normalized_type == "institutions" and _phase2_normalize_text(row.beneficiary_type) in {"institution", "institutions", "instn"})
            or _phase2_normalize_text(row.beneficiary_type) == normalized_type
        ]
        district_options_source = type_scoped_rows
        if item_filter != "all":
            district_options_source = [row for row in district_options_source if row.requested_item == item_filter]
        if is_district_type:
            district_options = sorted({row.district for row in district_options_source if row.district})
        elif is_institution_like_type:
            district_options = sorted({row.beneficiary_name for row in district_options_source if row.beneficiary_name})
        else:
            district_options = []

        article_options_source = type_scoped_rows
        if district_filter != "all":
            if is_district_type:
                article_options_source = [row for row in article_options_source if row.district == district_filter]
            elif is_institution_like_type:
                article_options_source = [row for row in article_options_source if row.beneficiary_name == district_filter]
        article_options = sorted({row.requested_item for row in article_options_source if row.requested_item})

        rows = type_scoped_rows
        if district_filter != "all":
            if is_district_type:
                rows = [row for row in rows if row.district == district_filter]
            elif is_institution_like_type:
                rows = [row for row in rows if row.beneficiary_name == district_filter]
        if item_filter != "all":
            rows = [row for row in rows if row.requested_item == item_filter]
        if q:
            rows = [
                row for row in rows
                if q in str(row.application_number or "").lower()
                or q in str(row.beneficiary_name or "").lower()
                or q in str(row.district or "").lower()
                or q in str(row.requested_item or "").lower()
                or q in str(row.comments or "").lower()
            ]

        sort_map = {
            "district": lambda row: row.district or "",
            "application_number": lambda row: row.application_number or "",
            "beneficiary_name": lambda row: row.beneficiary_name or "",
            "requested_item": lambda row: row.requested_item or "",
            "quantity": lambda row: int(row.quantity or 0),
            "waiting_hall_quantity": lambda row: int(row.waiting_hall_quantity or 0),
            "token_quantity": lambda row: int(row.token_quantity or 0),
        }
        if sort_key in sort_map:
            rows = sorted(rows, key=sort_map[sort_key], reverse=(sort_dir == "desc"))

        totals = {
            "quantity": sum(int(row.quantity or 0) for row in rows),
            "waiting_hall_quantity": sum(int(row.waiting_hall_quantity or 0) for row in rows),
            "token_quantity": sum(int(row.token_quantity or 0) for row in rows),
        }
        overall_totals = {
            "quantity": sum(int(row.quantity or 0) for row in all_rows),
            "waiting_hall_quantity": sum(int(row.waiting_hall_quantity or 0) for row in all_rows),
            "token_quantity": sum(int(row.token_quantity or 0) for row in all_rows),
        }
        all_rows_count = len(all_rows)
        seat_unique_item_count = len({str(row.requested_item or "").strip() for row in all_rows if str(row.requested_item or "").strip()})
        last_saved_at = None
        if all_rows:
            last_saved_at = max((row.updated_at or row.created_at) for row in all_rows)
        elif session:
            last_saved_at = session.updated_at
        if last_saved_at:
            last_saved_at = timezone.localtime(last_saved_at)
        sync_preview = None
        if session and self.request.GET.get("preview_sync") == "1":
            source_rows = _phase2_master_export_rows()
            preview_upload_result = _phase2_build_rows_from_master_export_rows(source_rows, source_file_name="master-entry-db")
            sync_preview = {
                **_phase2_preview_sync_state(session, preview_upload_result["rows"]),
                "source_row_count": preview_upload_result["source_row_count"],
                "grouped_row_count": preview_upload_result["grouped_row_count"],
            }
        master_change_state = None
        if session:
            source_rows = _phase2_master_export_rows()
            current_master_result = _phase2_build_rows_from_master_export_rows(source_rows, source_file_name="master-entry-db")
            master_change_state = _phase2_master_change_state(session, current_master_result["rows"])
        reconciliation_snapshot = (session.phase2_reconciliation_snapshot or {}) if session else {}
        reconciliation_checks = _phase2_reconciliation_checks(reconciliation_snapshot) if session else []
        def build_sort_params(column):
            params = self.request.GET.copy()
            params["session"] = str(session.pk) if session else ""
            params["sort"] = column
            params["dir"] = "desc" if sort_key == column and sort_dir == "asc" else "asc"
            return params.urlencode()
        context.update(
            {
                "page_title": "Seat Allocation",
                "selected_session": session,
                "event_sessions": list(models.EventSession.objects.order_by("-is_active", "-event_year", "session_name")),
                "seat_rows": rows,
                "totals": totals,
                "overall_totals": overall_totals,
                "percentages": {
                    "waiting_hall": round((totals["waiting_hall_quantity"] / totals["quantity"]) * 100, 1) if totals["quantity"] else 0,
                    "token": round((totals["token_quantity"] / totals["quantity"]) * 100, 1) if totals["quantity"] else 0,
                },
                "all_rows_count": all_rows_count,
                "seat_unique_item_count": seat_unique_item_count,
                "last_saved_at": last_saved_at,
                "filters": {
                    "q": self.request.GET.get("q", ""),
                    "beneficiary_type": beneficiary_type_filter,
                    "district_filter": district_filter,
                    "item_filter": item_filter,
                    "sort": sort_key,
                    "dir": sort_dir,
                },
                "district_options": district_options,
                "article_options": article_options,
                "beneficiary_type_choices": [
                    (models.RecipientTypeChoices.DISTRICT, "District"),
                    (models.RecipientTypeChoices.PUBLIC, "Public"),
                    (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
                    (models.RecipientTypeChoices.OTHERS, "Others"),
                    ("all", "All Types"),
                ],
                "is_district_type": is_district_type,
                "is_institution_like_type": is_institution_like_type,
                "sort_querystrings": {
                    "district": build_sort_params("district"),
                    "application_number": build_sort_params("application_number"),
                    "beneficiary_name": build_sort_params("beneficiary_name"),
                    "requested_item": build_sort_params("requested_item"),
                    "quantity": build_sort_params("quantity"),
                    "waiting_hall_quantity": build_sort_params("waiting_hall_quantity"),
                    "token_quantity": build_sort_params("token_quantity"),
                },
                "current_sort": sort_key,
                "current_dir": sort_dir,
                "can_create_edit": self.request.user.has_module_permission(models.ModuleKeyChoices.SEAT_ALLOCATION, "create_edit"),
                "can_export": self.request.user.has_module_permission(models.ModuleKeyChoices.SEAT_ALLOCATION, "export"),
                "can_upload_replace": self.request.user.has_module_permission(models.ModuleKeyChoices.SEAT_ALLOCATION, "upload_replace"),
                "reconciliation": {
                    "source_name": session.phase2_source_name if session else "",
                    "checks": reconciliation_checks,
                    "all_matched": bool(reconciliation_checks) and all(check["matched"] for check in reconciliation_checks),
                    "last_checked_at": timezone.localtime(session.updated_at) if session and reconciliation_snapshot else None,
                },
                "sync_preview": sync_preview,
                "master_change_state": master_change_state,
                "preview_sync_url": _phase2_url_with_extra_params(
                    self.request,
                    "ui:seat-allocation-list",
                    session,
                    filter_keys=self.preserved_filter_keys,
                    extra_params={"preview_sync": "1"},
                ) if session else reverse("ui:seat-allocation-list"),
                "clear_preview_url": _phase2_url_with_extra_params(
                    self.request,
                    "ui:seat-allocation-list",
                    session,
                    filter_keys=self.preserved_filter_keys,
                    extra_params={"preview_sync": None},
                ) if session else reverse("ui:seat-allocation-list"),
                "submit_result": submit_result,
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if request.GET.get("export") and session and request.user.has_module_permission(self.module_key, "export"):
            export_rows, export_headers = _phase2_export_rows(
                models.SeatAllocationRow.objects.filter(session=session).order_by(
                    "sort_order", "district", "requested_item", "application_number", "id"
                ),
                include_sequence=False,
            )
            response = HttpResponse(content_type="text/csv")
            timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
            response["Content-Disposition"] = f'attachment; filename="2_Master_Data_Seat_{timestamp}.csv"'
            writer = csv.DictWriter(response, fieldnames=export_headers)
            writer.writeheader()
            for row in export_rows:
                writer.writerow(row)
            return response
        return self.render_to_response(self.get_context_data(selected_session=session))

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip()
        session = _phase2_selected_session(request)

        if action == "create_session":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to create sessions.")
                return HttpResponseRedirect(reverse("ui:seat-allocation-list"))
            session_name = (request.POST.get("session_name") or "").strip()
            event_year = _phase2_parse_number(request.POST.get("event_year")) or timezone.localdate().year
            if not session_name:
                messages.error(request, "Enter a session name.")
                return HttpResponseRedirect(reverse("ui:seat-allocation-list"))
            session = models.EventSession.objects.create(
                session_name=session_name,
                event_year=event_year,
                is_active=bool(request.POST.get("is_active")),
                notes=(request.POST.get("notes") or "").strip(),
            )
            messages.success(request, f'Session "{session.session_name}" created.')
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if not session and action in {"upload_csv", "use_existing", "save_splits", "submit_splits", "reset_splits", "reset_filtered_splits", "bulk_waiting_full", "bulk_waiting_zero"}:
            session = _phase2_get_or_create_default_session()

        if action == "activate_session":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to activate sessions.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            session.is_active = True
            session.save(update_fields=["is_active", "updated_at"])
            messages.success(request, f'{session.session_name} is now the active event session.')
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if action == "upload_csv":
            if not request.user.has_module_permission(self.module_key, "upload_replace"):
                messages.error(request, "You do not have permission to upload seat allocation data.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            uploaded = request.FILES.get("file")
            if not uploaded:
                messages.error(request, "Choose a master-entry export CSV or Excel file.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            try:
                upload_result = _phase2_build_upload_rows(uploaded)
            except ValueError as exc:
                messages.error(request, str(exc))
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            upload_rows = upload_result["rows"]
            _phase2_replace_session_rows(
                session,
                upload_rows,
                source_file_name=uploaded.name,
                user=request.user,
                reconciliation=upload_result,
            )
            messages.success(
                request,
                f"Uploaded {upload_result['source_row_count']} master row(s) into {session.session_name} and created "
                f"{upload_result['grouped_row_count']} seat-allocation working row(s).",
            )
            return HttpResponseRedirect(
                _phase2_url_with_extra_params(
                    request,
                    "ui:seat-allocation-list",
                    session,
                    filter_keys=self.preserved_filter_keys,
                    extra_params={"preview_sync": None},
                )
            )

        if action == "use_existing":
            if not request.user.has_module_permission(self.module_key, "upload_replace"):
                messages.error(request, "You do not have permission to load existing master-entry data.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            source_rows = _phase2_master_export_rows()
            upload_result = _phase2_build_rows_from_master_export_rows(source_rows, source_file_name="master-entry-db")
            preserve_result = _phase2_preserve_existing_split_state(session, upload_result["rows"])
            upload_rows = preserve_result["rows"]
            _phase2_replace_session_rows(
                session,
                upload_rows,
                source_file_name="master-entry-db",
                user=request.user,
                reconciliation=upload_result,
            )
            messages.success(
                request,
                f"Synced {upload_result['source_row_count']} master row(s) into {upload_result['grouped_row_count']} seat-allocation working row(s). "
                f"Preserved {preserve_result['preserved_count']} existing split row(s), added {preserve_result['new_count']}, removed {preserve_result['removed_count']}.",
            )
            if preserve_result["removed_count"]:
                messages.warning(
                    request,
                    f"Master sync removed {preserve_result['removed_count']} seat-allocation row(s) that no longer exist in master data.",
                )
            return HttpResponseRedirect(
                _phase2_url_with_extra_params(
                    request,
                    "ui:seat-allocation-list",
                    session,
                    filter_keys=self.preserved_filter_keys,
                    extra_params={"preview_sync": None},
                )
            )

        if action == "save_splits":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to update seat allocation rows.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            row_ids = request.POST.getlist("row_id")
            waiting_values = request.POST.getlist("waiting_hall_quantity")
            updated_count = 0
            rows_by_id = {
                str(row.id): row
                for row in models.SeatAllocationRow.objects.filter(session=session, id__in=row_ids)
            }
            with transaction.atomic():
                for idx, row_id in enumerate(row_ids):
                    row = rows_by_id.get(str(row_id))
                    if not row:
                        continue
                    waiting_hall_quantity = _phase2_parse_number(waiting_values[idx] if idx < len(waiting_values) else 0)
                    waiting_hall_quantity = min(waiting_hall_quantity, int(row.quantity or 0))
                    row.waiting_hall_quantity = waiting_hall_quantity
                    row.token_quantity = max(int(row.quantity or 0) - waiting_hall_quantity, 0)
                    row.updated_by = request.user
                    row.save(update_fields=["waiting_hall_quantity", "token_quantity", "updated_by", "updated_at"])
                    updated_count += 1
            messages.success(request, f"Saved split values for {updated_count} row(s).")
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if action == "submit_splits":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to validate seat allocation rows.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            row_ids = request.POST.getlist("row_id")
            waiting_values = request.POST.getlist("waiting_hall_quantity")
            pending_waiting_by_id = {}
            for idx, row_id in enumerate(row_ids):
                raw_value = waiting_values[idx] if idx < len(waiting_values) else 0
                pending_waiting_by_id[str(row_id)] = _phase2_parse_number(raw_value)
            rows_by_id = {
                str(row.id): row
                for row in models.SeatAllocationRow.objects.filter(session=session, id__in=row_ids)
            }
            with transaction.atomic():
                for idx, row_id in enumerate(row_ids):
                    row = rows_by_id.get(str(row_id))
                    if not row:
                        continue
                    waiting_hall_quantity = _phase2_parse_number(waiting_values[idx] if idx < len(waiting_values) else 0)
                    waiting_hall_quantity = min(waiting_hall_quantity, int(row.quantity or 0))
                    row.waiting_hall_quantity = waiting_hall_quantity
                    row.token_quantity = max(int(row.quantity or 0) - waiting_hall_quantity, 0)
                    row.updated_by = request.user
                    row.save(update_fields=["waiting_hall_quantity", "token_quantity", "updated_by", "updated_at"])

            all_rows = list(models.SeatAllocationRow.objects.filter(session=session))
            source_checks = _phase2_reconciliation_checks(session.phase2_reconciliation_snapshot or {})
            split_check = _phase2_split_reconciliation(all_rows)
            source_ok = bool(source_checks) and all(check["matched"] for check in source_checks)
            split_ok = split_check["rowwise_matched"] and split_check["overall_matched"]

            submit_checks = []
            for check in source_checks:
                submit_checks.append(
                    {
                        "label": check["label"],
                        "matched": check["matched"],
                        "details": "Matched" if check["matched"] else f"Master {check['source']}, Working Copy {check['grouped']}",
                    }
                )
            submit_checks.append(
                {
                    "label": "Row-wise split quantities",
                    "matched": split_check["rowwise_matched"],
                    "details": (
                        f"All {split_check['row_count']} row(s) matched"
                        if split_check["rowwise_matched"]
                        else f"{split_check['row_mismatch_count']} row(s) have Quantity != Waiting Hall + Token"
                    ),
                }
            )
            submit_checks.append(
                {
                    "label": "Overall split quantities",
                    "matched": split_check["overall_matched"],
                    "details": f"Waiting Hall {split_check['total_waiting']}, Token {split_check['total_token']}, Total Quantity {split_check['total_quantity']}",
                }
            )

            request.session["seat_allocation_submit_result"] = {
                "all_matched": source_ok and split_ok,
                "checks": submit_checks,
                "checked_at_display": timezone.localtime().strftime("%d/%m/%Y %I:%M %p"),
            }
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if action == "reset_splits":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to reset split values.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            models.SeatAllocationRow.objects.filter(session=session).update(
                waiting_hall_quantity=0,
                token_quantity=F("quantity"),
                updated_by=request.user,
                updated_at=timezone.now(),
            )
            messages.success(request, "Waiting hall and token quantities were reset.")
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if action == "reset_filtered_splits":
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to reset split values.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            row_ids = request.POST.getlist("visible_row_id")
            rows = list(models.SeatAllocationRow.objects.filter(session=session, id__in=row_ids))
            with transaction.atomic():
                for row in rows:
                    row.waiting_hall_quantity = 0
                    row.token_quantity = int(row.quantity or 0)
                    row.updated_by = request.user
                    row.save(update_fields=["waiting_hall_quantity", "token_quantity", "updated_by", "updated_at"])
            messages.success(request, f"Reset split values for {len(rows)} filtered row(s).")
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        if action in {"bulk_waiting_full", "bulk_waiting_zero"}:
            if not request.user.has_module_permission(self.module_key, "create_edit"):
                messages.error(request, "You do not have permission to bulk update split values.")
                return HttpResponseRedirect(
                    _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
                )
            row_ids = request.POST.getlist("visible_row_id")
            rows = list(models.SeatAllocationRow.objects.filter(session=session, id__in=row_ids))
            with transaction.atomic():
                for row in rows:
                    waiting_value = int(row.quantity or 0) if action == "bulk_waiting_full" else 0
                    row.waiting_hall_quantity = waiting_value
                    row.token_quantity = max(int(row.quantity or 0) - waiting_value, 0)
                    row.updated_by = request.user
                    row.save(update_fields=["waiting_hall_quantity", "token_quantity", "updated_by", "updated_at"])
            messages.success(
                request,
                f'Updated {len(rows)} filtered row(s): Waiting Hall set to {"full quantity" if action == "bulk_waiting_full" else "0"}.',
            )
            return HttpResponseRedirect(
                _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
            )

        return HttpResponseRedirect(
            _phase2_redirect_url(request, "ui:seat-allocation-list", session, filter_keys=self.preserved_filter_keys)
        )


class SequenceListView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.SEQUENCE_LIST
    permission_action = "view"
    template_name = "dashboard/sequence_list.html"

    def _master_items(self):
        grouped = {}
        for row in _phase2_master_export_rows():
            item = str(row.get("Requested Item") or "").strip()
            if not item:
                continue
            grouped.setdefault(
                item,
                {
                    "item": item,
                    "category": str(row.get("Article Category") or row.get("Category") or "").strip() or "Uncategorized",
                },
            )
        return list(grouped.values())

    def _uploaded_source_rows_from_file(self, uploaded_file):
        headers, uploaded_rows = _tabular_rows_from_upload(uploaded_file)
        if not headers:
            raise ValueError("Uploaded file is empty.")

        normalized_headers = {_phase2_normalize_text(header): header for header in headers if str(header or "").strip()}

        def find_header(candidates):
            for candidate in candidates:
                match = normalized_headers.get(_phase2_normalize_text(candidate))
                if match:
                    return match
            return None

        item_header = find_header(["Requested Item", "Item", "Article", "Article Name", "requested_item"])
        token_header = find_header(["Token Quantity", "Token Qty", "token_quantity", "Token"])
        sequence_header = find_header(["Sequence No", "Sequence List", "sequence_no", "sequence list"])
        if not item_header:
            raise ValueError("Uploaded file missing item column. Expected Requested Item / Item / Article.")

        source_rows = []
        for source_row in uploaded_rows:
            requested_item = str(source_row.get(item_header) or "").strip()
            if not requested_item:
                continue
            source_rows.append(
                {
                    "requested_item": requested_item,
                    "token_quantity": _phase2_parse_number(source_row.get(token_header)) if token_header else 0,
                    "sequence_no": (_phase2_parse_number(source_row.get(sequence_header)) or None) if sequence_header else None,
                    "row_map": {header: source_row.get(header, "") for header in headers},
                    "headers": headers,
                }
            )
        return {"headers": headers, "rows": source_rows}

    def _source_rows(self, session):
        rows = list(
            models.SeatAllocationRow.objects.filter(session=session).order_by(
                F("sequence_no").asc(nulls_last=True),
                "sort_order",
                "requested_item",
                "application_number",
                "id",
            )
        )
        payload = []
        for row in rows:
            base_headers = [
                header
                for header in (row.master_headers or [])
                if _phase2_normalize_text(header) not in {"waiting hall quantity", "token quantity", "sequence no", "sequence list"}
            ]
            headers = [*base_headers, "Waiting Hall Quantity", "Token Quantity"]
            row_map = {}
            master_row = row.master_row or {}
            for header in base_headers:
                row_map[header] = master_row.get(header, "")
            row_map["Waiting Hall Quantity"] = int(row.waiting_hall_quantity or 0)
            row_map["Token Quantity"] = int(row.token_quantity or 0)
            payload.append(
                {
                    "source_id": str(row.id),
                    "requested_item": row.requested_item or "",
                    "token_quantity": int(row.token_quantity or 0),
                    "sequence_no": int(row.sequence_no) if row.sequence_no else None,
                    "row_map": row_map,
                    "headers": headers,
                }
            )
        return payload

    def _saved_sequence_items(self, session):
        return list(
            models.SequenceListItem.objects.filter(session=session)
            .order_by("sequence_no", "sort_order", "item_name")
            .values("item_name", "sequence_no", "sort_order")
        )

    def _seat_allocation_grouped_rows(self, session):
        rows = list(models.SeatAllocationRow.objects.filter(session=session))
        return [
            {
                "application_number": row.application_number or "",
                "beneficiary_name": row.beneficiary_name or "",
                "district": row.district or "",
                "requested_item": row.requested_item or "",
                "quantity": int(row.quantity or 0),
                "waiting_hall_quantity": int(row.waiting_hall_quantity or 0),
                "token_quantity": int(row.token_quantity or 0),
                "beneficiary_type": row.beneficiary_type or "",
                "item_type": row.item_type or "",
                "comments": row.comments or "",
                "master_row": row.master_row or {},
                "master_headers": row.master_headers or [],
            }
            for row in rows
        ]

    def _sequence_reconciliation_state(self, session):
        master_item_names = {
            str(row.get("item") or "").strip()
            for row in self._master_items()
            if str(row.get("item") or "").strip()
        }
        seat_allocation_items = {
            str(item).strip()
            for item in models.SeatAllocationRow.objects.filter(session=session).values_list("requested_item", flat=True)
            if str(item).strip()
        }
        saved_sequence_items = {
            str(item).strip()
            for item in models.SequenceListItem.objects.filter(session=session).values_list("item_name", flat=True)
            if str(item).strip()
        }
        return {
            "master_items": master_item_names,
            "seat_allocation_items": seat_allocation_items,
            "saved_sequence_items": saved_sequence_items,
            "missing_in_seat_allocation": sorted(master_item_names - seat_allocation_items),
            "extra_in_seat_allocation": sorted(seat_allocation_items - master_item_names),
            "new_to_sequence": sorted(master_item_names - saved_sequence_items),
            "extra_in_sequence": sorted(saved_sequence_items - master_item_names),
        }

    def _sequence_submit_result(self, *, reconciliation_state, sequence_item_names, sequence_number_count, exact_checks, seat_integrity):
        master_item_count = len(reconciliation_state["master_items"])
        seat_item_count = len(reconciliation_state["seat_allocation_items"])
        sequence_item_count = len(sequence_item_names)
        missing_in_seat = reconciliation_state["missing_in_seat_allocation"]
        extra_in_seat = reconciliation_state["extra_in_seat_allocation"]
        unassigned_master = sorted(reconciliation_state["master_items"] - sequence_item_names)
        extra_in_sequence = sorted(sequence_item_names - reconciliation_state["master_items"])
        split_check = seat_integrity["split_check"]
        base_integrity_ok = (
            all(check["matched"] for check in seat_integrity["checks"])
            and split_check["rowwise_matched"]
            and split_check["overall_matched"]
            and exact_checks["master_match"]["matched"]
            and exact_checks["seat_match"]["matched"]
        )

        checks = [
            {
                "label": "Final data integrity",
                "matched": base_integrity_ok,
                "details": (
                    "Master Data, Seat Allocation, split totals, and final sequence output all matched."
                    if base_integrity_ok
                    else "One or more Master Data / Seat Allocation / split / final output checks failed."
                ),
            },
            *[
                {
                    "label": f"Master Data vs Seat Allocation {check['label']}",
                    "matched": check["matched"],
                    "details": (
                        f"Matched {check['source']}"
                        if check["matched"]
                        else f"Master {check['source']}, Seat Allocation {check['grouped']}"
                    ),
                }
                for check in seat_integrity["checks"]
            ],
            {
                "label": "Seat Allocation row-wise split",
                "matched": split_check["rowwise_matched"],
                "details": (
                    f"All {split_check['row_count']} row(s) matched"
                    if split_check["rowwise_matched"]
                    else f"{split_check['row_mismatch_count']} row(s) have Quantity != Waiting Hall + Token"
                ),
            },
            {
                "label": "Seat Allocation overall split",
                "matched": split_check["overall_matched"],
                "details": (
                    f"Waiting Hall {split_check['total_waiting']}, Token {split_check['total_token']}, Total Quantity {split_check['total_quantity']}"
                ),
            },
            {
                "label": "Final Sequence vs Master Data",
                "matched": exact_checks["master_match"]["matched"],
                "details": exact_checks["master_match"]["details"],
            },
            {
                "label": "Final Sequence vs Seat Allocation",
                "matched": exact_checks["seat_match"]["matched"],
                "details": exact_checks["seat_match"]["details"],
            },
            {
                "label": "Master Data items vs Sequence List",
                "matched": not unassigned_master and not extra_in_sequence,
                "details": (
                    f"Matched {master_item_count} item(s)"
                    if not unassigned_master and not extra_in_sequence
                    else " | ".join(
                        part
                        for part in [
                            (
                                f"Unassigned master item(s): {', '.join(unassigned_master[:5])}"
                                + (f" and {len(unassigned_master) - 5} more" if len(unassigned_master) > 5 else "")
                            ) if unassigned_master else "",
                            (
                                f"Removed/stale sequence item(s): {', '.join(extra_in_sequence[:5])}"
                                + (f" and {len(extra_in_sequence) - 5} more" if len(extra_in_sequence) > 5 else "")
                            ) if extra_in_sequence else "",
                        ]
                        if part
                    ),
                ),
            },
            {
                "label": "Item uniqueness",
                "matched": sequence_item_count == len(sequence_item_names),
                "details": f"{len(sequence_item_names)} unique item(s) in Sequence List",
            },
            {
                "label": "Sequence number uniqueness",
                "matched": sequence_number_count == sequence_item_count,
                "details": (
                    f"{sequence_number_count} unique sequence number(s) for {sequence_item_count} item(s)"
                    if sequence_number_count == sequence_item_count
                    else f"{sequence_number_count} unique sequence number(s) for {sequence_item_count} item(s)"
                ),
            },
            {
                "label": "Sequence coverage on Seat Allocation",
                "matched": seat_item_count == sequence_item_count and not missing_in_seat and not extra_in_seat,
                "details": (
                    f"Sequence will apply to all {seat_item_count} Seat Allocation item(s)"
                    if seat_item_count == sequence_item_count and not missing_in_seat and not extra_in_seat
                    else f"Seat Allocation items: {seat_item_count}, Sequence items: {sequence_item_count}"
                ),
            },
        ]
        return {
            "all_matched": all(check["matched"] for check in checks),
            "checks": checks,
            "checked_at_display": timezone.localtime().strftime("%d/%m/%Y %I:%M %p"),
        }

    def _item_summaries(self, session):
        rows = list(
            models.SeatAllocationRow.objects.filter(session=session).order_by(
                "sequence_no", "requested_item", "application_number", "id"
            )
        )
        include_only_token = (self.request.GET.get("token_only") or "").strip() == "1"
        if include_only_token:
            rows = [row for row in rows if int(row.token_quantity or 0) > 0]

        q = (self.request.GET.get("q") or "").strip().lower()
        grouped = {}
        for row in rows:
            item = (row.requested_item or "").strip()
            if not item:
                continue
            summary = grouped.setdefault(
                item,
                {
                    "item": item,
                    "beneficiary_type": row.beneficiary_type or "",
                    "item_type": row.item_type or "",
                    "row_count": 0,
                    "quantity": 0,
                    "waiting_hall_quantity": 0,
                    "token_quantity": 0,
                    "sequence_no": row.sequence_no,
                },
            )
            summary["row_count"] += 1
            summary["quantity"] += int(row.quantity or 0)
            summary["waiting_hall_quantity"] += int(row.waiting_hall_quantity or 0)
            summary["token_quantity"] += int(row.token_quantity or 0)
            if summary["sequence_no"] in {None, ""} and row.sequence_no:
                summary["sequence_no"] = row.sequence_no

        items = list(grouped.values())
        if q:
            items = [
                item for item in items
                if q in str(item["item"]).lower()
                or q in str(item["beneficiary_type"]).lower()
                or q in str(item["item_type"]).lower()
            ]
        items.sort(key=lambda item: (item["sequence_no"] is None, item["sequence_no"] or 999999, item["item"].lower()))
        return items

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = kwargs.get("selected_session") or _phase2_selected_session(self.request)
        items = self._item_summaries(session) if session else []
        submit_result = self.request.session.pop("sequence_submit_result", None)
        context.update(
            {
                "page_title": "Sequence List",
                "selected_session": session,
                "event_sessions": list(models.EventSession.objects.order_by("-is_active", "-event_year", "session_name")),
                "sequence_items": items,
                "sequence_source_rows": self._source_rows(session) if session else [],
                "sequence_saved_items": self._saved_sequence_items(session) if session else [],
                "sequence_master_items": self._master_items(),
                "sequence_source_name": session.phase2_source_name if session else "",
                "sequence_source_is_uploaded": bool(session and session.phase2_source_name and session.phase2_source_name != "master-entry-db"),
                "sequence_default_items": SEQUENCE_DEFAULT_ITEMS,
                "filters": {
                    "q": self.request.GET.get("q", ""),
                    "token_only": self.request.GET.get("token_only", ""),
                    "start_from": self.request.GET.get("start_from", "1"),
                },
                "can_create_edit": self.request.user.has_module_permission(models.ModuleKeyChoices.SEQUENCE_LIST, "create_edit"),
                "can_export": self.request.user.has_module_permission(models.ModuleKeyChoices.SEQUENCE_LIST, "export"),
                "sequence_reconciliation_state": self._sequence_reconciliation_state(session) if session else {},
                "submit_result": submit_result,
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if request.GET.get("export") and session and request.user.has_module_permission(self.module_key, "export"):
            export_rows, export_headers = _phase2_export_rows(
                models.SeatAllocationRow.objects.filter(session=session).order_by(
                    F("sequence_no").asc(nulls_last=True),
                    "sort_order",
                    "requested_item",
                    "application_number",
                    "id",
                ),
                include_sequence=True,
            )
            response = HttpResponse(content_type="text/csv")
            timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
            response["Content-Disposition"] = f'attachment; filename="3_Master_Data_Seq_{timestamp}.csv"'
            writer = csv.DictWriter(response, fieldnames=export_headers)
            writer.writeheader()
            for row in export_rows:
                writer.writerow(row)
            return response
        return self.render_to_response(self.get_context_data(selected_session=session))

    def post(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if not session:
            messages.error(request, "Create or select a session first.")
            return HttpResponseRedirect(reverse("ui:sequence-list"))
        if not request.user.has_module_permission(self.module_key, "create_edit"):
            messages.error(request, "You do not have permission to update sequence rows.")
            return HttpResponseRedirect(f'{reverse("ui:sequence-list")}?session={session.pk}')

        action = (request.POST.get("action") or "").strip()
        start_from = _phase2_parse_number(request.POST.get("start_from")) or 1

        if action == "parse_upload":
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                return JsonResponse({"error": "Choose a CSV or Excel file to upload."}, status=400)
            try:
                upload_result = self._uploaded_source_rows_from_file(uploaded_file)
            except ValueError as exc:
                return JsonResponse({"error": str(exc)}, status=400)
            return JsonResponse(
                {
                    "rows": upload_result["rows"],
                    "file_name": uploaded_file.name,
                    "source_label": "Loaded Uploaded File",
                }
            )

        if action == "auto_assign":
            items = self._item_summaries(session)
            existing_sequences = {int(item["sequence_no"]) for item in items if item["sequence_no"]}
            next_sequence = start_from
            for item in items:
                if item["sequence_no"]:
                    continue
                while next_sequence in existing_sequences:
                    next_sequence += 1
                models.SeatAllocationRow.objects.filter(session=session, requested_item=item["item"]).update(
                    sequence_no=next_sequence,
                    updated_by=request.user,
                    updated_at=timezone.now(),
                )
                existing_sequences.add(next_sequence)
                next_sequence += 1
            messages.success(request, "Blank sequence numbers were assigned.")
            return HttpResponseRedirect(f'{reverse("ui:sequence-list")}?session={session.pk}')

        if action == "save_sequences":
            item_names = request.POST.getlist("item_name")
            sequence_values = request.POST.getlist("sequence_no")
            attempted_sequence_names = {str(item_name or "").strip() for item_name in item_names if str(item_name or "").strip()}
            attempted_sequence_numbers = {
                _phase2_parse_number(raw_value)
                for raw_value in sequence_values
                if str(raw_value).strip() and _phase2_parse_number(raw_value)
            }
            reconciliation_state = self._sequence_reconciliation_state(session)
            seat_integrity = _sequence_seat_allocation_integrity(session)
            submitted_rows = []
            seen_items = set()
            seen_sequences = set()
            validation_errors = []

            for idx, item_name in enumerate(item_names):
                item_name = str(item_name or "").strip()
                raw_value = sequence_values[idx] if idx < len(sequence_values) else ""
                if not item_name:
                    continue
                if item_name in seen_items:
                    validation_errors.append(f"Duplicate item in sequence list: {item_name}.")
                    continue
                sequence_no = _phase2_parse_number(raw_value) if str(raw_value).strip() else None
                if not sequence_no:
                    validation_errors.append(f"Missing sequence number for {item_name}.")
                    continue
                if sequence_no in seen_sequences:
                    validation_errors.append(f"Sequence number {sequence_no} is used more than once.")
                    continue
                seen_items.add(item_name)
                seen_sequences.add(sequence_no)
                submitted_rows.append(
                    {
                        "item_name": item_name,
                        "sequence_no": sequence_no,
                        "sort_order": idx + 1,
                    }
                )

            sequence_map = {row["item_name"]: row["sequence_no"] for row in submitted_rows}
            final_export = _sequence_final_export_rows(session, sequence_map)
            master_export_rows = _phase2_master_export_rows()
            exact_checks = {
                "master_match": _sequence_exact_compare(
                    left_rows=_sequence_project_rows(final_export["final_rows"], EXPORT_COLUMNS),
                    left_headers=EXPORT_COLUMNS,
                    right_rows=_sequence_project_rows(master_export_rows, EXPORT_COLUMNS),
                    right_headers=EXPORT_COLUMNS,
                    matched_label=f"Matched {len(master_export_rows)} row(s) across {len(EXPORT_COLUMNS)} column(s)",
                    mismatch_label="Master Data comparison failed",
                ),
                "seat_match": _sequence_exact_compare(
                    left_rows=_sequence_project_rows(final_export["final_rows"], final_export["seat_headers"]),
                    left_headers=final_export["seat_headers"],
                    right_rows=final_export["seat_rows"],
                    right_headers=final_export["seat_headers"],
                    matched_label=f"Matched {len(final_export['seat_rows'])} row(s) across {len(final_export['seat_headers'])} column(s)",
                    mismatch_label="Seat Allocation comparison failed",
                ),
            }

            if validation_errors:
                request.session["sequence_submit_result"] = {
                    "all_matched": False,
                    "checks": [
                        *[
                            {
                                "label": f"Master Data vs Seat Allocation {check['label']}",
                                "matched": check["matched"],
                                "details": (
                                    f"Matched {check['source']}"
                                    if check["matched"]
                                    else f"Master {check['source']}, Seat Allocation {check['grouped']}"
                                ),
                            }
                            for check in seat_integrity["checks"]
                        ],
                        {
                            "label": "Seat Allocation row-wise split",
                            "matched": seat_integrity["split_check"]["rowwise_matched"],
                            "details": (
                                f"All {seat_integrity['split_check']['row_count']} row(s) matched"
                                if seat_integrity["split_check"]["rowwise_matched"]
                                else f"{seat_integrity['split_check']['row_mismatch_count']} row(s) have Quantity != Waiting Hall + Token"
                            ),
                        },
                        {
                            "label": "Seat Allocation overall split",
                            "matched": seat_integrity["split_check"]["overall_matched"],
                            "details": (
                                f"Waiting Hall {seat_integrity['split_check']['total_waiting']}, Token {seat_integrity['split_check']['total_token']}, Total Quantity {seat_integrity['split_check']['total_quantity']}"
                            ),
                        },
                        {
                            "label": "Final Sequence vs Master Data",
                            "matched": exact_checks["master_match"]["matched"],
                            "details": exact_checks["master_match"]["details"],
                        },
                        {
                            "label": "Final Sequence vs Seat Allocation",
                            "matched": exact_checks["seat_match"]["matched"],
                            "details": exact_checks["seat_match"]["details"],
                        },
                        {
                            "label": "Sequence validation",
                            "matched": False,
                            "details": validation_errors[0],
                        }
                    ],
                    "checked_at_display": timezone.localtime().strftime("%d/%m/%Y %I:%M %p"),
                }
                return HttpResponseRedirect(f'{reverse("ui:sequence-list")}?session={session.pk}')

            with transaction.atomic():
                master_item_names = reconciliation_state["master_items"]
                sequence_item_names = {row["item_name"] for row in submitted_rows}
                unassigned_master_items = sorted(master_item_names - sequence_item_names)
                extra_sequence_items = sorted(sequence_item_names - master_item_names)
                submit_result = self._sequence_submit_result(
                    reconciliation_state=reconciliation_state,
                    sequence_item_names=sequence_item_names,
                    sequence_number_count=len(seen_sequences),
                    exact_checks=exact_checks,
                    seat_integrity=seat_integrity,
                )

                models.SequenceListItem.objects.filter(session=session).delete()
                models.SequenceListItem.objects.bulk_create(
                    [
                        models.SequenceListItem(
                            session=session,
                            item_name=row["item_name"],
                            sequence_no=row["sequence_no"],
                            sort_order=row["sort_order"],
                            created_by=request.user,
                            updated_by=request.user,
                        )
                        for row in submitted_rows
                    ]
                )

                source_is_uploaded = bool(session.phase2_source_name and session.phase2_source_name != "master-entry-db")
                allow_apply = source_is_uploaded or (submit_result["all_matched"] and not unassigned_master_items and not extra_sequence_items)
                if allow_apply:
                    models.SeatAllocationRow.objects.filter(session=session).update(
                        sequence_no=None,
                        updated_by=request.user,
                        updated_at=timezone.now(),
                    )
                    for row in submitted_rows:
                        models.SeatAllocationRow.objects.filter(
                            session=session,
                            requested_item=row["item_name"],
                        ).update(
                            sequence_no=row["sequence_no"],
                            updated_by=request.user,
                            updated_at=timezone.now(),
                        )
                    if source_is_uploaded and not submit_result["all_matched"]:
                        messages.warning(
                            request,
                            "Sequence list was applied to the uploaded Seat Allocation working copy. Current Master Data reconciliation still shows differences.",
                        )
                else:
                    messages.warning(
                        request,
                        "Sequence list was saved, but it was not applied to Seat Allocation because reconciliation is not clean yet.",
                    )
                request.session["sequence_submit_result"] = submit_result
            return HttpResponseRedirect(f'{reverse("ui:sequence-list")}?session={session.pk}')

        return HttpResponseRedirect(f'{reverse("ui:sequence-list")}?session={session.pk}')


class TokenGenerationView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.TOKEN_GENERATION
    permission_action = "view"
    template_name = "dashboard/token_generation.html"

    def _token_open_section(self, session, filter_state):
        requested = str(self.request.GET.get("open_section") or "").strip()
        if requested:
            return requested
        stored = self.request.session.pop("token_generation_open_section", None)
        if stored:
            return stored
        if _token_generation_has_active_filters(filter_state):
            return "transformations"
        return ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = kwargs.get("selected_session") or _phase2_selected_session(self.request)
        dataset = _token_generation_saved_dataset(session) if session else {"headers": [], "rows": [], "source_name": "", "saved_at": None}
        rows = dataset["rows"]
        stage_state = _token_generation_stage_state(self.request, session) if session else {}
        filter_state = _token_generation_filter_state(self.request)
        open_section = self._token_open_section(session, filter_state)
        name_length_limit = _phase2_parse_number(self.request.GET.get("name_limit")) or 25
        preview_rows = rows[:10]
        preview_table = [
            [row.get(header, "") for header in dataset["headers"]]
            for row in preview_rows
        ]
        blank_summary = _token_generation_empty_value_summary(rows, dataset["headers"]) if rows else []
        quality_checks = _token_generation_quality_checks(rows) if rows else {
            "sequence_item_conflicts": [],
            "article_sequence_conflicts": [],
            "missing_sequences": [],
            "token_quantity_total": 0,
            "printable_token_total": 0,
            "zero_token_rows": 0,
        }
        prep_validation_summary = self.request.session.pop("token_generation_prep_validation_summary", None)
        token_is_prepared = bool(rows) and all("Names" in row for row in rows)
        is_sorted = _token_generation_is_sorted(rows)
        is_generated = _token_generation_is_generated(rows)
        token_step3_saved = bool(stage_state.get("token_print_saved")) and token_is_prepared
        generated_start_tokens = [
            _phase2_parse_number(row.get("Start Token No"))
            for row in rows
            if (_phase2_parse_number(row.get("Start Token No")) or 0) > 0
        ]
        generated_end_tokens = [
            _phase2_parse_number(row.get("End Token No"))
            for row in rows
            if (_phase2_parse_number(row.get("End Token No")) or 0) > 0
        ]
        token_start_no = min(generated_start_tokens) if generated_start_tokens else 0
        token_end_no = max(generated_end_tokens) if generated_end_tokens else 0
        prep_checks = [
            {"label": "Names column created", "passed": token_is_prepared and all("Names" in row for row in rows)},
            {"label": "Empty values filled", "passed": token_is_prepared},
            {"label": "Sorted by sequence, beneficiary type, handicapped status, and names", "passed": token_is_prepared and is_sorted},
            {"label": "Quantity, Cost Per Unit, and Total Value are all greater than 0", "passed": token_is_prepared and not prep_validation_summary},
            {"label": "Duplicate row check complete", "passed": token_is_prepared},
        ]
        edit_candidates = _token_generation_edit_candidates(rows, length_limit=name_length_limit) if token_is_prepared else []
        article_toggle_rows = _token_generation_article_toggle_rows(rows) if token_is_prepared else []
        filtered_rows = _token_generation_filter_rows(rows, filter_state) if rows else []
        context.update(
            {
                "page_title": "Token Generation",
                "selected_session": session,
                "token_rows": preview_table,
                "token_headers": dataset["headers"],
                "token_row_count": len(rows),
                "token_source_name": dataset["source_name"],
                "token_saved_at_display": timezone.localtime(dataset["saved_at"]).strftime("%d/%m/%Y %I:%M %p") if dataset["saved_at"] else "",
                "token_blank_summary": blank_summary,
                "token_prep_validation_summary": prep_validation_summary,
                "token_quality_checks": quality_checks,
                "token_column_count": len(dataset["headers"]),
                "token_prep_checks": prep_checks,
                "token_is_prepared": token_is_prepared,
                "token_is_sorted": is_sorted,
                "token_is_generated": is_generated,
                "token_step3_saved": token_step3_saved,
                "token_start_no": token_start_no,
                "token_end_no": token_end_no,
                "token_edit_candidates": edit_candidates,
                "token_name_length_limit": int(name_length_limit),
                "token_article_toggles": article_toggle_rows,
                "token_filter_state": filter_state,
                "token_filtered_rows": filtered_rows,
                "token_open_section": open_section,
                "token_sync_required": _token_generation_sync_required(self.request, session, dataset) if session else False,
                "can_create_edit": self.request.user.has_module_permission(self.module_key, "create_edit"),
                "can_export": self.request.user.has_module_permission(self.module_key, "export"),
                "can_upload_replace": self.request.user.has_module_permission(self.module_key, "upload_replace"),
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if request.GET.get("export") and session and request.user.has_module_permission(self.module_key, "export"):
            dataset = _token_generation_saved_dataset(session)
            if not dataset["rows"]:
                messages.warning(request, "Sync or upload token data first.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            response = HttpResponse(content_type="text/csv")
            timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
            response["Content-Disposition"] = f'attachment; filename="4_Master_Data_Token_{timestamp}.csv"'
            writer = csv.DictWriter(response, fieldnames=dataset["headers"])
            writer.writeheader()
            for row in dataset["rows"]:
                writer.writerow(row)
            return response
        return self.render_to_response(self.get_context_data(selected_session=session))

    def post(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if not session:
            messages.error(request, "Create or select a session first.")
            return HttpResponseRedirect(reverse("ui:token-generation"))
        if not request.user.has_module_permission(self.module_key, "create_edit"):
            messages.error(request, "You do not have permission to update token data.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        action = (request.POST.get("action") or "").strip()
        if action == "sync_data":
            dataset = _token_generation_source_dataset(session)
            _token_generation_store_dataset(
                session=session,
                dataset=dataset,
                source_name=f"Synced from Sequence List ({session.session_name})",
                user=request.user,
            )
            _token_generation_set_stage_state(
                request,
                session,
                token_print_saved=False,
                source_sync_marker=_token_generation_latest_source_marker(session),
            )
            request.session["token_generation_open_section"] = ""
            messages.success(
                request,
                (
                    f"Token Generation synced {len(dataset['rows'])} row(s). "
                    f"Previous-stage data loaded successfully."
                ),
            )
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action == "upload_csv":
            if not request.user.has_module_permission(self.module_key, "upload_replace"):
                messages.error(request, "You do not have permission to upload token data.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Choose a CSV or Excel file to upload.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            source_headers, source_rows = _tabular_rows_from_upload(uploaded_file)
            if not source_headers:
                messages.error(request, "Uploaded file is empty.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            dataset = {
                "headers": source_headers,
                "rows": source_rows,
                "blank_summary": _token_generation_empty_value_summary(source_rows, source_headers),
                "quality_checks": _token_generation_quality_checks(source_rows),
            }
            _token_generation_store_dataset(
                session=session,
                dataset=dataset,
                source_name=uploaded_file.name,
                user=request.user,
            )
            _token_generation_set_stage_state(
                request,
                session,
                token_print_saved=False,
                source_sync_marker="",
            )
            request.session["token_generation_open_section"] = ""
            messages.success(
                request,
                (
                    f"Uploaded {len(dataset['rows'])} row(s) into Token Generation. "
                    f"Previous-stage data loaded successfully."
                ),
            )
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        saved_dataset = _token_generation_saved_dataset(session)
        if not saved_dataset["rows"]:
            messages.warning(request, "Sync or upload token data first.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
        if _token_generation_sync_required(request, session, saved_dataset):
            messages.warning(request, "Click Sync Data first to load the latest Sequence List data before continuing.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action == "exclude_selected_rows":
            selected_indices = {
                _phase2_parse_number(value)
                for value in request.POST.getlist("selected_row_index")
                if str(value).strip()
            }
            selected_indices = {int(value) for value in selected_indices if value is not None}
            original_rows = [dict(row) for row in saved_dataset["rows"]]
            filtered_rows = [
                row for index, row in enumerate(original_rows)
                if index not in selected_indices
            ]
            removed_count = len(original_rows) - len(filtered_rows)
            filtered_headers = [
                header for header in list(saved_dataset["headers"] or [])
                if header not in {"Start Token No", "End Token No"}
            ]
            for row in filtered_rows:
                row.pop("Start Token No", None)
                row.pop("End Token No", None)
            saved_dataset["headers"] = filtered_headers
            saved_dataset["rows"] = filtered_rows
            _token_generation_store_dataset(
                session=session,
                dataset=saved_dataset,
                source_name=saved_dataset["source_name"] or f"Rule-filtered Token Data ({session.session_name})",
                user=request.user,
            )
            _token_generation_set_stage_state(request, session, token_print_saved=False)
            request.session["token_generation_open_section"] = "transformations"
            messages.success(
                request,
                f"Transformation rules applied. Removed {removed_count} selected row(s). Source data is unchanged.",
            )
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action == "run_data_prep":
            prepared_dataset = _token_generation_prepare_dataset(saved_dataset["rows"], saved_dataset["headers"])
            invalid_value_summary = _token_generation_invalid_value_summary(prepared_dataset["rows"], saved_dataset["headers"])
            if invalid_value_summary:
                request.session["token_generation_prep_validation_summary"] = invalid_value_summary
                request.session["token_generation_open_section"] = "step1"
                issue_text = ", ".join(
                    f"{entry['label']} ({entry['count']})"
                    for entry in invalid_value_summary
                )
                messages.error(
                    request,
                    f"Step 1 data prep is blocked. After filling empty values, values below 1 were found in: {issue_text}. Please fix the source data first.",
                )
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            _token_generation_store_dataset(
                session=session,
                dataset=prepared_dataset,
                source_name=saved_dataset["source_name"] or f"Prepared Token Data ({session.session_name})",
                user=request.user,
            )
            _token_generation_set_stage_state(request, session, token_print_saved=False)
            request.session["token_generation_open_section"] = "step1"
            messages.success(request, "Step 1 data prep completed. Names, blanks, sorting, and duplicate checks are ready.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action in {"sort_data", "save_adjustments"}:
            if not (bool(saved_dataset["rows"]) and all("Names" in row for row in saved_dataset["rows"])):
                request.session["token_generation_open_section"] = "step2"
                messages.warning(request, "Run Step 1 data prep before saving name changes.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            rows = [dict(row) for row in saved_dataset["rows"]]
            replacements_applied = 0
            for original_name, replacement_name in request.POST.items():
                if original_name.startswith("replace_name__"):
                    source = original_name.replace("replace_name__", "", 1)
                    source = source.replace("__SLASH__", "/")
                    replacement = (replacement_name or "").strip()
                    if replacement:
                        for row in rows:
                            if str(row.get("Names") or "") == source:
                                row["Names"] = replacement
                                replacements_applied += 1
                if original_name.startswith("replace_token_name__"):
                    source = original_name.replace("replace_token_name__", "", 1)
                    source = source.replace("__SLASH__", "/")
                    replacement = (replacement_name or "").strip()
                    if replacement:
                        for row in rows:
                            if str(row.get("Token Name") or "") == source:
                                row["Token Name"] = replacement
                                replacements_applied += 1
            saved_dataset["rows"] = _token_generation_sort_dataset(rows)
            _token_generation_store_dataset(
                session=session,
                dataset=saved_dataset,
                source_name=saved_dataset["source_name"] or f"Sorted Token Data ({session.session_name})",
                user=request.user,
            )
            _token_generation_set_stage_state(request, session, token_print_saved=False)
            request.session["token_generation_open_section"] = "step2"
            messages.success(request, f"Step 2 adjustments saved. {replacements_applied} replacement(s) were applied.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action == "save_token_print":
            if not (bool(saved_dataset["rows"]) and all("Names" in row for row in saved_dataset["rows"])):
                request.session["token_generation_open_section"] = "step3"
                messages.warning(request, "Run Step 1 data prep before updating token print settings.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            rows = [dict(row) for row in saved_dataset["rows"]]
            skip_items = {
                value.strip()
                for value in request.POST.getlist("skip_label_item")
                if str(value).strip()
            }
            for row in rows:
                requested_item = str(row.get("Requested Item") or "").strip()
                if requested_item in skip_items:
                    row["Token Print for ARTL"] = "0"
                else:
                    _token_generation_token_print_flag(row)
            saved_dataset["rows"] = rows
            _token_generation_store_dataset(
                session=session,
                dataset=saved_dataset,
                source_name=saved_dataset["source_name"] or f"Token Print Updated ({session.session_name})",
                user=request.user,
            )
            _token_generation_set_stage_state(request, session, token_print_saved=True)
            request.session["token_generation_open_section"] = "step3"
            messages.success(request, "Step 3 token print settings were saved.")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        if action == "generate_tokens":
            rows = saved_dataset["rows"]
            if not (bool(rows) and all("Names" in row for row in rows)):
                request.session["token_generation_open_section"] = "step4"
                messages.warning(request, "Run Step 1 data prep before generating token numbers.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            if not _token_generation_is_sorted(rows):
                request.session["token_generation_open_section"] = "step4"
                messages.warning(request, "Complete Step 2 adjustments before generating token numbers.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            if not _token_generation_stage_state(request, session).get("token_print_saved"):
                request.session["token_generation_open_section"] = "step4"
                messages.warning(request, "Complete Step 3 token print review before generating token numbers.")
                return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')
            generated_dataset = _token_generation_generate_dataset(rows, saved_dataset["headers"])
            saved_dataset["rows"] = generated_dataset["rows"]
            saved_dataset["headers"] = generated_dataset["headers"]
            _token_generation_store_dataset(
                session=session,
                dataset=saved_dataset,
                source_name=saved_dataset["source_name"] or f"Generated Token Data ({session.session_name})",
                user=request.user,
            )
            request.session["token_generation_open_section"] = "step4"
            start_no = min(
                (_phase2_parse_number(row.get("Start Token No")) or 0)
                for row in saved_dataset["rows"]
                if (_phase2_parse_number(row.get("Start Token No")) or 0) > 0
            ) if saved_dataset["rows"] else 0
            end_no = max(
                (_phase2_parse_number(row.get("End Token No")) or 0)
                for row in saved_dataset["rows"]
                if (_phase2_parse_number(row.get("End Token No")) or 0) > 0
            ) if saved_dataset["rows"] else 0
            messages.success(request, f"Token numbers generated. Starting: {start_no}, Ending: {end_no}")
            return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')

        return HttpResponseRedirect(f'{reverse("ui:token-generation")}?session={session.pk}')


class LabelGenerationView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    module_key = models.ModuleKeyChoices.LABELS
    permission_action = "view"
    template_name = "dashboard/labels.html"

    def _build_download_response(self, request, session, download_kind):
        dataset = _labels_saved_dataset(session)
        rows = dataset["rows"]
        if not rows:
            messages.warning(request, "Sync or upload label data first.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
        if _labels_sync_required(request, session, dataset):
            messages.warning(request, "Click Sync Data first to load the latest Token Generation data before downloading labels.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
        if not _labels_has_generated_tokens(rows):
            messages.warning(request, "Generate token numbers in Token Generation before downloading labels.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        stage_state = _labels_stage_state(request, session)
        large_items = set(stage_state["large_items"])
        audit = _labels_audit_download(rows, download_kind=download_kind, large_items=large_items)
        if not audit["ready"]:
            if audit["expected_labels"] <= 0:
                messages.warning(request, audit["reason"] or "No matching labels are available for this download yet.")
            else:
                messages.warning(
                    request,
                    (
                        f"Label check failed for this download. Expected labels: {audit['expected_labels']}, "
                        f"generated labels: {audit['actual_labels']}, duplicate tokens: {audit['duplicate_tokens']}, "
                        f"invalid ranges: {audit['invalid_range_rows']}."
                    ),
                )
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        def token_qty(row):
            return (_phase2_parse_number(row.get("Token Quantity")) or 0)

        def is_printable_article(row):
            return token_qty(row) > 0 and (_phase2_parse_number(row.get("Token Print for ARTL")) or 0) != 0

        def sort_by_name_and_start(row):
            return (
                str(row.get("Names") or "").strip(),
                _phase2_parse_number(row.get("Start Token No")) or 0,
            )

        def sort_by_item_and_start(row):
            return (
                str(row.get("Token Name") or row.get("Requested Item") or "").strip(),
                str(row.get("Names") or "").strip(),
                _phase2_parse_number(row.get("Start Token No")) or 0,
            )

        label_buffer = None
        filename = None

        if download_kind == "article_12l_separate":
            if not stage_state.get("large_items_saved"):
                messages.warning(request, "Select 2L Items first before downloading article labels.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: is_printable_article(row) and str(row.get("Requested Item") or "").strip() not in large_items,
                group_by=lambda row: str(row.get("Token Name") or row.get("Requested Item") or "").strip(),
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="separate")
            filename = _labels_download_filename("1_Article_Labels_S")
        elif download_kind == "article_12l_continuous":
            if not stage_state.get("large_items_saved"):
                messages.warning(request, "Select 2L Items first before downloading article labels.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: is_printable_article(row) and str(row.get("Requested Item") or "").strip() not in large_items,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="continuous")
            filename = _labels_download_filename("1_Article_Labels_C")
        elif download_kind == "article_2l_continuous":
            if not stage_state.get("large_items_saved"):
                messages.warning(request, "Select 2L Items first before downloading article labels.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0 and str(row.get("Requested Item") or "").strip() in large_items,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="2L", mode="continuous")
            filename = _labels_download_filename("2_Article_2L_Labels")
        elif download_kind == "district_separate":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "District",
                group_by=lambda row: str(row.get("Names") or "").strip(),
                sort_key=sort_by_name_and_start,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="separate")
            filename = _labels_download_filename("3_District_Labels_S")
        elif download_kind == "district_continuous":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "District",
                sort_key=sort_by_name_and_start,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="continuous")
            filename = _labels_download_filename("3_District_Labels_C")
        elif download_kind == "institution":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "Institutions",
                sort_key=sort_by_name_and_start,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="continuous")
            filename = _labels_download_filename("5_Institution_Labels")
        elif download_kind == "public":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0 and str(row.get("Beneficiary Type") or "").strip() == "Public",
                sort_key=sort_by_item_and_start,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="continuous")
            filename = _labels_download_filename("5_Public_Labels")
        elif download_kind == "chair_separate":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0,
                group_by=lambda row: str(row.get("Token Name") or row.get("Requested Item") or "").strip(),
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="separate")
            filename = _labels_download_filename("Chair_Labels_S")
        elif download_kind == "chair_continuous":
            entries = _labels_expand_entries(
                rows,
                row_filter=lambda row: token_qty(row) > 0,
            )
            label_buffer = services.generate_mnp_labels_pdf(entries, layout="12L", mode="continuous")
            filename = _labels_download_filename("Chair_Labels_C")
        else:
            messages.warning(request, "Unknown label download requested.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        response = HttpResponse(label_buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = kwargs.get("selected_session") or _phase2_selected_session(self.request)
        dataset = _labels_saved_dataset(session) if session else {"headers": [], "rows": [], "source_name": "", "saved_at": None}
        rows = dataset["rows"]
        stage_state = _labels_stage_state(self.request, session) if session else {"large_items": list(LABELS_DEFAULT_2L_ITEMS)}
        available_items = _labels_available_requested_items(rows)
        large_items = stage_state["large_items"] if session else list(LABELS_DEFAULT_2L_ITEMS)
        label_audits = {
            "article_12l_continuous": _labels_audit_download(rows, download_kind="article_12l_continuous", large_items=large_items) if rows else {},
            "article_12l_separate": _labels_audit_download(rows, download_kind="article_12l_separate", large_items=large_items) if rows else {},
            "article_2l_continuous": _labels_audit_download(rows, download_kind="article_2l_continuous", large_items=large_items) if rows else {},
            "district_continuous": _labels_audit_download(rows, download_kind="district_continuous", large_items=large_items) if rows else {},
            "district_separate": _labels_audit_download(rows, download_kind="district_separate", large_items=large_items) if rows else {},
            "institution": _labels_audit_download(rows, download_kind="institution", large_items=large_items) if rows else {},
            "public": _labels_audit_download(rows, download_kind="public", large_items=large_items) if rows else {},
            "chair_continuous": _labels_audit_download(rows, download_kind="chair_continuous", large_items=large_items) if rows else {},
            "chair_separate": _labels_audit_download(rows, download_kind="chair_separate", large_items=large_items) if rows else {},
        }
        context.update(
            {
                "page_title": "Labels",
                "selected_session": session,
                "label_row_count": len(rows),
                "label_column_count": len(dataset["headers"]),
                "label_saved_at_display": timezone.localtime(dataset["saved_at"]).strftime("%d/%m/%Y %I:%M %p") if dataset["saved_at"] else "",
                "label_source_name": str(dataset["source_name"] or "").replace(" (Event)", ""),
                "label_has_generated_tokens": _labels_has_generated_tokens(rows),
                "label_sync_required": _labels_sync_required(self.request, session, dataset) if session else False,
                "label_large_items_saved": bool(stage_state.get("large_items_saved")),
                "label_large_items": available_items,
                "label_selected_large_items": stage_state["large_items"],
                "label_audits": label_audits,
                "can_create_edit": self.request.user.has_module_permission(self.module_key, "create_edit"),
                "can_export": self.request.user.has_module_permission(self.module_key, "export"),
                "can_upload_replace": self.request.user.has_module_permission(self.module_key, "upload_replace"),
            }
        )
        return context

    def get(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if request.GET.get("download") and session and request.user.has_module_permission(self.module_key, "export"):
            return self._build_download_response(request, session, str(request.GET.get("download") or "").strip())
        return self.render_to_response(self.get_context_data(selected_session=session))

    def post(self, request, *args, **kwargs):
        session = _phase2_selected_session(request)
        if not session:
            messages.error(request, "Create or select a session first.")
            return HttpResponseRedirect(reverse("ui:labels"))
        if not request.user.has_module_permission(self.module_key, "create_edit"):
            messages.error(request, "You do not have permission to update label data.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        action = (request.POST.get("action") or "").strip()
        if action == "sync_data":
            dataset = _labels_source_dataset(session)
            _labels_store_dataset(
                session=session,
                dataset=dataset,
                source_name="Synced from Token Generation",
                user=request.user,
            )
            _labels_set_stage_state(
                request,
                session,
                source_sync_marker=_labels_latest_source_marker(session),
                large_items_saved=False,
            )
            messages.success(request, f"Labels synced {len(dataset['rows'])} row(s) from Token Generation.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        if action == "upload_csv":
            if not request.user.has_module_permission(self.module_key, "upload_replace"):
                messages.error(request, "You do not have permission to upload label data.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                messages.error(request, "Choose a CSV or Excel file to upload.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')
            dataset = {
                "headers": [],
                "rows": [],
            }
            dataset["headers"], dataset["rows"] = _tabular_rows_from_upload(uploaded_file)
            dataset = _labels_normalize_dataset(dataset)
            _labels_store_dataset(
                session=session,
                dataset=dataset,
                source_name=uploaded_file.name,
                user=request.user,
            )
            _labels_set_stage_state(
                request,
                session,
                source_sync_marker="",
                large_items_saved=False,
            )
            messages.success(request, f"Uploaded {len(dataset['rows'])} row(s) into Labels.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        if action == "save_large_items":
            selected_large_items = [
                value.strip()
                for value in request.POST.getlist("large_label_item")
                if str(value).strip()
            ]
            _labels_set_stage_state(request, session, large_items=selected_large_items, large_items_saved=True)
            messages.success(request, "2L article label selections saved.")
            return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

        if action == "download_custom_labels":
            if not request.user.has_module_permission(self.module_key, "export"):
                messages.error(request, "You do not have permission to download labels.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

            def _parse_custom_bulk(raw_text, layout):
                default_font_size = 72 if layout == "A4" else 54 if layout == "2L" else 30
                entries = []
                for raw_line in str(raw_text or "").splitlines():
                    line = str(raw_line or "").strip()
                    if not line:
                        continue
                    parts = [part.strip() for part in line.split(",")]
                    if len(parts) < 2:
                        continue
                    text_value = parts[0]
                    count_value = _phase2_parse_number(parts[1]) or 0
                    font_size_value = _phase2_parse_number(parts[2]) if len(parts) >= 3 else default_font_size
                    font_size_value = max(8, min(int(font_size_value or default_font_size), 120))
                    line_spacing_value = _phase2_parse_number(parts[3]) if len(parts) >= 4 else None
                    if line_spacing_value is not None:
                        line_spacing_value = max(0, min(int(line_spacing_value), 120))
                    if text_value and count_value > 0:
                        entries.append(
                            {
                                "text": text_value,
                                "count": count_value,
                                "font_size": font_size_value,
                                "line_spacing": line_spacing_value,
                            }
                        )
                return entries

            custom_bulks = request.POST.getlist("custom_label_bulk")
            custom_layouts = request.POST.getlist("custom_label_layout")
            groups = []
            for index, raw_bulk in enumerate(custom_bulks):
                layout_value = str(custom_layouts[index] if index < len(custom_layouts) else "12L" or "12L").strip().upper()
                if layout_value not in {"12L", "2L", "A4"}:
                    layout_value = "12L"
                entries = _parse_custom_bulk(raw_bulk, layout_value)
                if entries:
                    groups.append({"layout": layout_value, "entries": entries})

            if not groups:
                messages.error(request, "Paste at least one custom label row as Label text,count or Label text,count,font size,line spacing.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

            if len(groups) == 1:
                label_buffer = services.generate_mnp_custom_labels_pdf(groups[0]["entries"], layout=groups[0]["layout"])
                response = HttpResponse(label_buffer.getvalue(), content_type="application/pdf")
                single_filename = _labels_download_filename("Custom_Labels_%s" % groups[0]["layout"])
                response["Content-Disposition"] = f'attachment; filename="{single_filename}"'
                return response

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as bundle:
                for index, group in enumerate(groups, start=1):
                    label_buffer = services.generate_mnp_custom_labels_pdf(group["entries"], layout=group["layout"])
                    filename = _labels_download_filename(f"Custom_Labels_{index}_{group['layout']}")
                    bundle.writestr(filename, label_buffer.getvalue())
            zip_buffer.seek(0)
            timestamp = timezone.localtime().strftime("%d_%b_%y_%H_%M")
            response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
            response["Content-Disposition"] = f'attachment; filename="Custom_Labels_{timestamp}.zip"'
            return response

        if action == "preview_custom_labels":
            if not request.user.has_module_permission(self.module_key, "export"):
                messages.error(request, "You do not have permission to preview labels.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

            def _parse_custom_bulk(raw_text, layout):
                default_font_size = 72 if layout == "A4" else 54 if layout == "2L" else 30
                entries = []
                for raw_line in str(raw_text or "").splitlines():
                    line = str(raw_line or "").strip()
                    if not line:
                        continue
                    parts = [part.strip() for part in line.split(",")]
                    if len(parts) < 2:
                        continue
                    text_value = parts[0]
                    count_value = _phase2_parse_number(parts[1]) or 0
                    font_size_value = _phase2_parse_number(parts[2]) if len(parts) >= 3 else default_font_size
                    font_size_value = max(8, min(int(font_size_value or default_font_size), 120))
                    line_spacing_value = _phase2_parse_number(parts[3]) if len(parts) >= 4 else None
                    if line_spacing_value is not None:
                        line_spacing_value = max(0, min(int(line_spacing_value), 120))
                    if text_value and count_value > 0:
                        entries.append(
                            {
                                "text": text_value,
                                "count": count_value,
                                "font_size": font_size_value,
                                "line_spacing": line_spacing_value,
                            }
                        )
                return entries

            custom_bulks = request.POST.getlist("custom_label_bulk")
            custom_layouts = request.POST.getlist("custom_label_layout")
            groups = []
            for index, raw_bulk in enumerate(custom_bulks):
                layout_value = str(custom_layouts[index] if index < len(custom_layouts) else "12L" or "12L").strip().upper()
                if layout_value not in {"12L", "2L", "A4"}:
                    layout_value = "12L"
                entries = _parse_custom_bulk(raw_bulk, layout_value)
                if entries:
                    groups.append({"layout": layout_value, "entries": entries})

            if not groups:
                messages.error(request, "Paste at least one custom label row as Label text,count or Label text,count,font size,line spacing.")
                return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')

            if len(groups) == 1:
                label_buffer = services.generate_mnp_custom_labels_pdf(groups[0]["entries"], layout=groups[0]["layout"])
                response = HttpResponse(label_buffer.getvalue(), content_type="application/pdf")
                response["Content-Disposition"] = 'inline; filename="Custom_Labels_Preview.pdf"'
                return response

            merged_writer = PdfWriter()
            for group in groups:
                label_buffer = services.generate_mnp_custom_labels_pdf(group["entries"], layout=group["layout"])
                group_reader = PdfReader(io.BytesIO(label_buffer.getvalue()))
                for page in group_reader.pages:
                    merged_writer.add_page(page)
            merged_output = io.BytesIO()
            merged_writer.write(merged_output)
            merged_output.seek(0)
            response = HttpResponse(merged_output.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = 'inline; filename="Custom_Labels_Preview.pdf"'
            return response

        return HttpResponseRedirect(f'{reverse("ui:labels")}?session={session.pk}')


def _import_district_master_csv(uploaded_file):
    inserted = 0
    updated = 0
    for row in _csv_reader_from_upload(uploaded_file):
        district_name = (row.get("district_name") or "").strip()
        if not district_name:
            continue
        budget_raw = (row.get("allotted_budget") or "0").strip().replace(",", "")
        allotted_budget = Decimal(budget_raw or "0")
        _, created = models.DistrictMaster.objects.update_or_create(
            district_name=district_name,
            defaults={
                "application_number": (row.get("application_number") or "").strip(),
                "allotted_budget": allotted_budget,
                "president_name": (row.get("president_name") or "").strip(),
                "mobile_number": (row.get("mobile_number") or "").strip(),
                "is_active": True,
            },
        )
        inserted += int(created)
        updated += int(not created)
    return inserted, updated


def _import_article_master_csv(uploaded_file):
    inserted = 0
    updated = 0
    valid_item_types = {choice for choice, _ in models.ItemTypeChoices.choices}
    for row in _csv_reader_from_upload(uploaded_file):
        article_name = (row.get("article_name") or "").strip()
        if not article_name:
            continue
        cost_raw = (row.get("cost_per_unit") or "0").strip().replace(",", "")
        cost_per_unit = Decimal(cost_raw or "0")
        item_type = (row.get("item_type") or models.ItemTypeChoices.ARTICLE).strip()
        if item_type not in valid_item_types:
            item_type = models.ItemTypeChoices.ARTICLE
        is_active_value = (row.get("is_active") or "").strip().lower()
        is_active = is_active_value in {"active", "true", "1", "yes"}
        _, created = models.Article.objects.update_or_create(
            article_name=article_name,
            defaults={
                "cost_per_unit": cost_per_unit,
                "item_type": item_type,
                "category": (row.get("category") or "").strip() or None,
                "master_category": (row.get("master_category") or "").strip() or None,
                "is_active": is_active,
            },
        )
        inserted += int(created)
        updated += int(not created)
    return inserted, updated


def _import_public_history_csv(uploaded_file, *, replace_existing=False):
    inserted = 0
    if replace_existing:
        models.PublicBeneficiaryHistory.objects.all().delete()
    for row in _csv_reader_from_upload(uploaded_file):
        year_raw = (row.get("year") or "").strip()
        if not year_raw:
            continue
        models.PublicBeneficiaryHistory.objects.create(
            aadhar_number=(row.get("aadhar_number") or "").strip(),
            name=(row.get("name") or "").strip(),
            year=int(year_raw),
            article_name=(row.get("article_name") or "").strip() or None,
            application_number=(row.get("application_number") or "").strip() or None,
            comments=(row.get("comments") or "").strip() or None,
            is_handicapped=_parse_bool(row.get("is_handicapped")),
            address=(row.get("address") or "").strip() or None,
            mobile=(row.get("mobile") or "").strip() or None,
            aadhar_number_sp=(row.get("aadhar_number_sp") or "").strip() or None,
            is_selected=_parse_bool(row.get("is_selected")),
            category=(row.get("category") or "").strip() or None,
        )
        inserted += 1
    return inserted, 0


def _parse_bool(value):
    text = str(value or "").strip().lower()
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return None
