from __future__ import annotations

"""Views and helper functions for the order fund request workflow."""

import io
import json
import os
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import IntegrityError, transaction
from django.db.models import CharField, Q
from django.db.models.functions import Cast
from django.forms import inlineformset_factory
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, FormView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core import models
from core.order_fund_request import services
from core.order_fund_request.forms import (
    FundRequestArticleForm,
    FundRequestDocumentUploadForm,
    FundRequestForm,
    FundRequestRecipientForm,
)
from core.purchase_order import services as purchase_order_services
from core.shared.inventory import build_order_management_rows
from core.shared.audit import get_request_audit_meta
from core.shared.audit import log_audit
from core.shared.permissions import AdminRequiredMixin, RoleRequiredMixin, WriteRoleMixin


def _is_editable_by_user(user, fr):
    if not user or not user.is_authenticated:
        return False
    if user.role not in {"admin", "editor"}:
        return False
    return fr.status == models.FundRequestStatusChoices.DRAFT


def _fund_request_recipient_display_name(recipient) -> str:
    source_entry_id = getattr(recipient, "source_entry_id", None)
    beneficiary_type = getattr(recipient, "beneficiary_type", None)
    source_entry = None
    if source_entry_id and beneficiary_type:
        if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
            source_entry = models.DistrictBeneficiaryEntry.objects.select_related("district").filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                district_name = str(getattr(source_entry.district, "district_name", "") or "").strip()
                if application_number and district_name:
                    return f"{application_number} - {district_name}"
                return district_name or application_number or "-"
        elif beneficiary_type == models.RecipientTypeChoices.PUBLIC:
            source_entry = models.PublicBeneficiaryEntry.objects.filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                public_name = str(source_entry.name or "").strip()
                if application_number and public_name:
                    return f"{application_number} - {public_name}"
                return public_name or application_number or "-"
        elif beneficiary_type in {
            models.RecipientTypeChoices.INSTITUTIONS,
            models.RecipientTypeChoices.OTHERS,
        }:
            source_entry = models.InstitutionsBeneficiaryEntry.objects.filter(pk=source_entry_id).first()
            if source_entry:
                application_number = str(source_entry.application_number or "").strip()
                institution_name = str(source_entry.institution_name or "").strip()
                if application_number and institution_name:
                    return f"{application_number} - {institution_name}"
                return institution_name or application_number or "-"
    return (
        str(getattr(recipient, "recipient_name", "") or "").strip()
        or str(getattr(recipient, "name_of_beneficiary", "") or "").strip()
        or str(getattr(recipient, "name_of_institution", "") or "").strip()
        or str(getattr(recipient, "beneficiary", "") or "").strip()
        or "-"
    )


def _fund_request_article_beneficiary_display(article_item) -> str:
    article = getattr(article_item, "article", None)
    if not article:
        return str(getattr(article_item, "beneficiary", "") or "").strip() or "-"

    labels = []
    if article.district_entries.exists():
        labels.append("District")
    if article.public_entries.exists():
        labels.append("Public")
    if article.institution_entries.exists():
        labels.append("Institutions")

    if len(labels) == 3:
        return "All beneficiaries"
    if labels:
        return " & ".join(labels)
    return str(getattr(article_item, "beneficiary", "") or "").strip() or "-"


def _normalize_vendor_group_payload(raw_group):
    if not isinstance(raw_group, dict):
        return None
    key = str(raw_group.get("key") or "").strip()
    if not key:
        return None
    vendor_id = raw_group.get("vendor_id")
    try:
        vendor_id = int(vendor_id) if str(vendor_id or "").strip() else None
    except (TypeError, ValueError):
        vendor_id = None
    payload = {
        "key": key,
        "vendor_id": vendor_id,
        "vendor_name": str(raw_group.get("vendor_name") or "").strip(),
        "gst_no": str(raw_group.get("gst_no") or "").strip(),
        "vendor_address": str(raw_group.get("vendor_address") or "").strip(),
        "vendor_city": str(raw_group.get("vendor_city") or "").strip(),
        "vendor_state": str(raw_group.get("vendor_state") or "").strip(),
        "vendor_pincode": str(raw_group.get("vendor_pincode") or "").strip(),
        "phone_number": str(raw_group.get("phone_number") or "").strip(),
        "cheque_in_favour": str(raw_group.get("cheque_in_favour") or "").strip(),
    }
    return payload


def _build_vendor_groups_from_articles(article_rows):
    groups = []
    row_key_map = {}
    seen = {}
    for article in article_rows:
        signature = (
            str(article.vendor_id or "").strip(),
            str(article.vendor_name or "").strip(),
            str(article.gst_no or "").strip(),
            str(article.vendor_address or "").strip(),
            str(article.vendor_city or "").strip(),
            str(article.vendor_state or "").strip(),
            str(article.vendor_pincode or "").strip(),
            str(article.cheque_in_favour or "").strip(),
        )
        if not any(signature):
            continue
        key = seen.get(signature)
        if not key:
            key = f"vendor-{len(groups) + 1}"
            seen[signature] = key
            groups.append(
                {
                    "key": key,
                    "vendor_id": int(signature[0]) if signature[0].isdigit() else None,
                    "vendor_name": signature[1],
                    "gst_no": signature[2],
                    "vendor_address": signature[3],
                    "vendor_city": signature[4],
                    "vendor_state": signature[5],
                    "vendor_pincode": signature[6],
                    "phone_number": "",
                    "cheque_in_favour": signature[7],
                }
            )
        if getattr(article, "pk", None):
            row_key_map[str(article.pk)] = key
    return groups, row_key_map


FundRequestRecipientFormSet = inlineformset_factory(
    models.FundRequest,
    models.FundRequestRecipient,
    form=FundRequestRecipientForm,
    extra=0,
    can_delete=True,
)

FundRequestArticleFormSet = inlineformset_factory(
    models.FundRequest,
    models.FundRequestArticle,
    form=FundRequestArticleForm,
    extra=0,
    can_delete=True,
)


class FundRequestListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"
    model = models.FundRequest
    template_name = "order_fund_request/fund_request_list.html"
    context_object_name = "fund_requests"
    paginate_by = 20

    def get(self, request, *args, **kwargs):
        if (request.GET.get("export") or "").strip().lower() in {"xlsx", "export"}:
            return self._export_xlsx()
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        queryset = (
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles", "documents")
        )
        self._sort_key = (self.request.GET.get("sort") or "created_at").strip()
        self._sort_dir = (self.request.GET.get("dir") or "desc").strip().lower()
        if q := (self.request.GET.get("q") or "").strip():
            fund_request_number_q = Q(fund_request_number__icontains=q)
            parsed_request_number = models.parse_fund_request_sequence(q)
            if parsed_request_number is not None:
                alternate_terms = {
                    models.format_fund_request_number(f"FR-{parsed_request_number}"),
                    f"FR-{parsed_request_number}",
                    f"FR{parsed_request_number}",
                }
                for term in alternate_terms:
                    fund_request_number_q |= Q(fund_request_number__icontains=term)
            matching_ids = (
                models.FundRequest.objects.select_related("created_by")
                .annotate(
                    total_amount_text=Cast("total_amount", output_field=CharField()),
                    created_at_text=Cast("created_at", output_field=CharField()),
                    recipient_source_entry_id_text=Cast("recipients__source_entry_id", output_field=CharField()),
                    recipient_fund_requested_text=Cast("recipients__fund_requested", output_field=CharField()),
                    article_sl_no_text=Cast("articles__sl_no", output_field=CharField()),
                    article_quantity_text=Cast("articles__quantity", output_field=CharField()),
                    article_unit_price_text=Cast("articles__unit_price", output_field=CharField()),
                    article_price_including_gst_text=Cast("articles__price_including_gst", output_field=CharField()),
                    article_value_text=Cast("articles__value", output_field=CharField()),
                    article_cumulative_text=Cast("articles__cumulative", output_field=CharField()),
                    document_generated_at_text=Cast("documents__generated_at", output_field=CharField()),
                )
                .filter(
                    fund_request_number_q
                    | Q(fund_request_type__icontains=q)
                    | Q(status__icontains=q)
                    | Q(aid_type__icontains=q)
                    | Q(notes__icontains=q)
                    | Q(total_amount_text__icontains=q)
                    | Q(created_at_text__icontains=q)
                    | Q(gst_number__icontains=q)
                    | Q(supplier_name__icontains=q)
                    | Q(supplier_address__icontains=q)
                    | Q(supplier_city__icontains=q)
                    | Q(supplier_state__icontains=q)
                    | Q(supplier_pincode__icontains=q)
                    | Q(purchase_order_number__icontains=q)
                    | Q(created_by__email__icontains=q)
                    | Q(created_by__first_name__icontains=q)
                    | Q(created_by__last_name__icontains=q)
                    | Q(recipients__recipient_name__icontains=q)
                    | Q(recipients__name_of_beneficiary__icontains=q)
                    | Q(recipients__name_of_institution__icontains=q)
                    | Q(recipients__beneficiary_type__icontains=q)
                    | Q(recipients__beneficiary__icontains=q)
                    | Q(recipients__details__icontains=q)
                    | Q(recipients__address__icontains=q)
                    | Q(recipients__cheque_in_favour__icontains=q)
                    | Q(recipients__cheque_no__icontains=q)
                    | Q(recipients__notes__icontains=q)
                    | Q(recipients__district_name__icontains=q)
                    | Q(recipient_source_entry_id_text__icontains=q)
                    | Q(recipient_fund_requested_text__icontains=q)
                    | Q(recipients__aadhar_number__icontains=q)
                    | Q(articles__article_name__icontains=q)
                    | Q(articles__beneficiary__icontains=q)
                    | Q(articles__vendor_name__icontains=q)
                    | Q(articles__gst_no__icontains=q)
                    | Q(articles__vendor_address__icontains=q)
                    | Q(articles__vendor_city__icontains=q)
                    | Q(articles__vendor_state__icontains=q)
                    | Q(articles__vendor_pincode__icontains=q)
                    | Q(articles__cheque_in_favour__icontains=q)
                    | Q(articles__cheque_no__icontains=q)
                    | Q(articles__supplier_article_name__icontains=q)
                    | Q(articles__description__icontains=q)
                    | Q(article_sl_no_text__icontains=q)
                    | Q(article_quantity_text__icontains=q)
                    | Q(article_unit_price_text__icontains=q)
                    | Q(article_price_including_gst_text__icontains=q)
                    | Q(article_value_text__icontains=q)
                    | Q(article_cumulative_text__icontains=q)
                    | Q(documents__file_name__icontains=q)
                    | Q(documents__file_path__icontains=q)
                    | Q(documents__document_type__icontains=q)
                    | Q(document_generated_at_text__icontains=q)
                    | Q(documents__generated_by__email__icontains=q)
                    | Q(documents__generated_by__first_name__icontains=q)
                    | Q(documents__generated_by__last_name__icontains=q)
                )
                .values_list("pk", flat=True)
                .distinct()
            )
            queryset = queryset.filter(pk__in=matching_ids)
        if request_type := (self.request.GET.get("request_type") or "").strip():
            queryset = queryset.filter(fund_request_type=request_type)
        if status := (self.request.GET.get("status") or "").strip():
            queryset = queryset.filter(status=status)
        if supplier := (self.request.GET.get("supplier") or "").strip():
            queryset = queryset.filter(supplier_name__icontains=supplier)
        sort_fields = {
            "fund_request_number": "fund_request_number",
            "fund_request_type": "fund_request_type",
            "item_type": "aid_type",
            "total_amount": "total_amount",
            "status": "status",
            "created_at": "created_at",
            "supplier_name": "supplier_name",
        }
        sort_field = sort_fields.get(self._sort_key, "created_at")
        sort_prefix = "" if self._sort_dir == "asc" else "-"
        if sort_field == "fund_request_number":
            queryset = queryset.order_by(f"{sort_prefix}created_at", f"{sort_prefix}id")
            queryset = sorted(
                queryset,
                key=lambda fr: (
                    models.parse_fund_request_sequence(fr.fund_request_number) is None,
                    models.parse_fund_request_sequence(fr.fund_request_number) or 0,
                    fr.created_at,
                    fr.id,
                ),
                reverse=(self._sort_dir == "desc"),
            )
            return queryset
        queryset = queryset.order_by(f"{sort_prefix}{sort_field}", "-id")
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for fr in context["fund_requests"]:
            recipients = list(fr.recipients.all())
            articles = list(fr.articles.all())
            for recipient in recipients:
                recipient.display_name = _fund_request_recipient_display_name(recipient)
                recipient.beneficiary_display = _fund_request_recipient_display_name(recipient)
            fr.district_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT
            )
            fr.public_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC
            )
            fr.institutions_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.INSTITUTIONS
            )
            fr.others_recipient_count = sum(
                1 for recipient in recipients if recipient.beneficiary_type == models.RecipientTypeChoices.OTHERS
            )
            fr.article_total_quantity = sum(int(article.quantity or 0) for article in articles)
        current_sort = getattr(self, "_sort_key", "created_at")
        current_dir = getattr(self, "_sort_dir", "desc")

        def build_sort_params(column):
            params = self.request.GET.copy()
            params.pop("page", None)
            next_dir = "asc"
            if current_sort == column and current_dir == "asc":
                next_dir = "desc"
            params["sort"] = column
            params["dir"] = next_dir
            return params.urlencode()

        context["request_type_choices"] = models.FundRequestTypeChoices.choices
        context["status_choices"] = [
            (models.FundRequestStatusChoices.DRAFT, "Draft"),
            (models.FundRequestStatusChoices.SUBMITTED, "Submitted"),
        ]
        context["filters"] = {
            "q": self.request.GET.get("q", ""),
            "request_type": self.request.GET.get("request_type", ""),
            "status": self.request.GET.get("status", ""),
            "supplier": self.request.GET.get("supplier", ""),
        }
        context["current_sort"] = current_sort
        context["current_dir"] = current_dir
        context["sort_querystrings"] = {
            "fund_request_number": build_sort_params("fund_request_number"),
            "fund_request_type": build_sort_params("fund_request_type"),
            "item_type": build_sort_params("item_type"),
            "total_amount": build_sort_params("total_amount"),
            "status": build_sort_params("status"),
            "created_at": build_sort_params("created_at"),
        }
        return context

    def _event_birthday_number(self, event_year: int) -> int:
        return max(event_year - 1940, 1)

    def _beneficiary_display_for_export(self, recipient, fund_request_type):
        if fund_request_type == models.FundRequestTypeChoices.ARTICLE:
            return ""
        if not recipient:
            return ""
        return _fund_request_recipient_display_name(recipient)

    def _export_xlsx(self):
        export_status = (self.request.GET.get("export_status") or "").strip().lower()
        queryset = list(
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles")
            .order_by("fund_request_number", "created_at", "id")
        )
        if q := (self.request.GET.get("q") or "").strip():
            queryset = [
                fr
                for fr in queryset
                if q.lower() in str(fr.fund_request_number or "").lower()
                or q.lower() in str(fr.formatted_fund_request_number or "").lower()
                or q.lower() in str(fr.aid_type or "").lower()
                or q.lower() in str(fr.supplier_name or "").lower()
            ]
        if request_type := (self.request.GET.get("request_type") or "").strip():
            queryset = [fr for fr in queryset if fr.fund_request_type == request_type]
        if export_status in {
            models.FundRequestStatusChoices.DRAFT,
            models.FundRequestStatusChoices.SUBMITTED,
        }:
            queryset = [fr for fr in queryset if fr.status == export_status]

        def _fr_sort_key(fr):
            sequence = models.parse_fund_request_sequence(fr.fund_request_number)
            if sequence is not None:
                return (0, sequence)
            raw = str(fr.fund_request_number or "").strip()
            return (1, raw)

        queryset.sort(key=_fr_sort_key)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Fund Requests"

        thin = Side(style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        total_fill = PatternFill("solid", fgColor="D3D3D3")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        event_year = timezone.localdate().year
        birthday_number = self._event_birthday_number(event_year)

        current_row = 1
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = "OMSAKTHI"
        cell.font = Font(size=10, bold=True)
        cell.alignment = center
        current_row += 1

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = (
            f"MASM Makkal Nala Pani Payment Request Details for Distribution on the eve of "
            f"{birthday_number}th Birthday Celebrations of"
        )
        cell.font = Font(size=12, bold=True)
        cell.alignment = center
        current_row += 1

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        cell.value = f"His Holiness AMMA at Melmaruvathur on 03.03.{event_year}"
        cell.font = Font(size=12, bold=True)
        cell.alignment = center
        current_row += 2

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = worksheet.cell(current_row, 1)
        label = "Payment Request - MASTER LIST"
        if export_status == models.FundRequestStatusChoices.DRAFT:
            label = "Payment Request - DRAFT MASTER LIST"
        elif export_status == models.FundRequestStatusChoices.SUBMITTED:
            label = "Payment Request - SUBMITTED MASTER LIST"
        cell.value = label
        cell.font = Font(size=14, bold=True)
        cell.alignment = center
        current_row += 2

        headers = [
            "FUND REQ NO.",
            "Request Type",
            "Beneficiary",
            "Name of Beneficiary/Article",
            "Name of Institution/Article",
            "Vendor Name",
            "GST / Aadhaar Number",
            "Details",
            "Units",
            "Price incl GST",
            "Value",
            "Fund Request Value",
            "CHEQUE (OR) RTGS IN FAVOUR",
            "CHEQUE NO.",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(current_row, idx)
            cell.value = header
            cell.font = Font(size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        current_row += 1

        all_rows = []
        fr_value_map = {}
        for fr in queryset:
            if fr.fund_request_type == models.FundRequestTypeChoices.AID:
                fr_total = Decimal("0")
                for recipient in fr.recipients.all():
                    amount = Decimal(str(recipient.fund_requested or 0))
                    fr_total += amount
                    all_rows.append(
                        {
                            "fr_id": fr.id,
                            "fund_request_number": fr.formatted_fund_request_number or "",
                            "request_type": fr.aid_type or "Aid",
                            "beneficiary": self._beneficiary_display_for_export(recipient, fr.fund_request_type),
                            "name_beneficiary_article": recipient.recipient_name or recipient.name_of_beneficiary or "",
                            "name_institution_article": recipient.name_of_institution or "",
                            "vendor_name": "",
                            "gst_aadhar": recipient.aadhar_number or "",
                            "details": recipient.notes or recipient.details or "",
                            "units": 1,
                            "price_incl_gst": float(amount),
                            "value": float(amount),
                            "fund_request_value": 0,
                            "cheque_in_favour": recipient.cheque_in_favour or "",
                            "cheque_no": recipient.cheque_no or "",
                        }
                    )
                fr_value_map[fr.id] = float(fr_total)
            else:
                fr_total = Decimal("0")
                for article in fr.articles.all():
                    line_value = Decimal(str(article.value or 0))
                    fr_total += line_value
                    all_rows.append(
                        {
                            "fr_id": fr.id,
                            "fund_request_number": fr.formatted_fund_request_number or "",
                            "request_type": "Article",
                            "beneficiary": _fund_request_article_beneficiary_display(article),
                            "name_beneficiary_article": article.article_name or "",
                            "name_institution_article": article.supplier_article_name or "",
                            "vendor_name": article.vendor_name or "",
                            "gst_aadhar": article.gst_no or fr.gst_number or "",
                            "details": "",
                            "units": article.quantity or 0,
                            "price_incl_gst": float(article.price_including_gst or article.unit_price or 0),
                            "value": float(line_value),
                            "fund_request_value": 0,
                            "cheque_in_favour": article.cheque_in_favour or "",
                            "cheque_no": article.cheque_no or "",
                        }
                    )
                fr_value_map[fr.id] = float(fr_total)

        for row in all_rows:
            row["fund_request_value"] = fr_value_map.get(row["fr_id"], 0)

        fr_groups = {}
        current_fr_id = None
        group_start = current_row
        for row in all_rows:
            if row["fr_id"] != current_fr_id:
                if current_fr_id is not None:
                    fr_groups[current_fr_id] = (group_start, current_row - 1)
                current_fr_id = row["fr_id"]
                group_start = current_row

            values = [
                row["fund_request_number"],
                row["request_type"],
                row["beneficiary"],
                row["name_beneficiary_article"],
                row["name_institution_article"],
                row["vendor_name"],
                row["gst_aadhar"],
                row["details"],
                row["units"],
                row["price_incl_gst"],
                row["value"],
                row["fund_request_value"],
                row["cheque_in_favour"],
                row["cheque_no"],
            ]
            for idx, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, idx)
                cell.value = value
                cell.border = border
                cell.alignment = right if idx in {9, 10, 11, 12} else left
            current_row += 1
        if current_fr_id is not None:
            fr_groups[current_fr_id] = (group_start, current_row - 1)

        for _fr_id, (start_row, end_row) in fr_groups.items():
            if end_row > start_row:
                worksheet.merge_cells(start_row=start_row, start_column=12, end_row=end_row, end_column=12)
                worksheet.cell(start_row, 12).alignment = right

        grand_total = sum(fr_value_map.values())
        total_row = current_row
        for col in range(1, 15):
            cell = worksheet.cell(total_row, col)
            cell.fill = total_fill
            cell.border = border
        worksheet.cell(total_row, 1).value = "TOTAL"
        worksheet.cell(total_row, 1).font = Font(size=11, bold=True)
        worksheet.cell(total_row, 1).alignment = left
        worksheet.cell(total_row, 12).value = grand_total
        worksheet.cell(total_row, 12).font = Font(size=11, bold=True)
        worksheet.cell(total_row, 12).alignment = right

        widths = {
            1: 18, 2: 18, 3: 20, 4: 24, 5: 24, 6: 22, 7: 18, 8: 28,
            9: 12, 10: 15, 11: 15, 12: 18, 13: 25, 14: 15,
        }
        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        date_stamp = timezone.localtime().strftime("%Y-%m-%d")
        suffix = export_status or "all"
        if export_status == models.FundRequestStatusChoices.SUBMITTED:
            filename = f"Fund_Request_Sub_{date_stamp}.xlsx"
        elif export_status == models.FundRequestStatusChoices.DRAFT:
            filename = f"Fund_Request_dft_{date_stamp}.xlsx"
        else:
            filename = f"Fund_Request_{suffix}_{date_stamp}.xlsx"
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def _fund_request_aid_type_choices():
    names = set()
    names.update(
        filter(
            None,
            models.DistrictBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list("article__article_name", flat=True),
        )
    )
    names.update(
        filter(
            None,
            models.PublicBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list("article__article_name", flat=True),
        )
    )
    names.update(
        filter(
            None,
            models.InstitutionsBeneficiaryEntry.objects.filter(article__item_type=models.ItemTypeChoices.AID).values_list("article__article_name", flat=True),
        )
    )
    return sorted(names)


def _fund_request_article_choices(current_fund_request=None):
    all_rows = [
        row
        for row in build_order_management_rows()
        if row["item_type"] == models.ItemTypeChoices.ARTICLE
    ]
    rows_by_name = {
        str(row["article_name"]).casefold(): dict(row)
        for row in all_rows
    }
    rows = [row for row in rows_by_name.values() if row["quantity_pending"] > 0]
    article_map = {
        article.article_name.casefold(): article
        for article in models.Article.objects.filter(item_type=models.ItemTypeChoices.ARTICLE)
    }
    choices = []
    for row in rows:
        article = article_map.get(row["article_name"].casefold())
        choices.append(
            {
                "name": row["article_name"],
                "label": f"{row['article_name']} (Pending: {row['quantity_pending']})",
                "article_id": article.id if article else "",
                "default_price": str(article.cost_per_unit if article else 0),
                "pending_qty": int(row["quantity_pending"] or 0),
            }
        )
    return choices


def _aid_entry_queryset(aid_type, beneficiary_type):
    filters = {
        "article__item_type": models.ItemTypeChoices.AID,
        "article__article_name__iexact": aid_type,
    }
    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        return models.DistrictBeneficiaryEntry.objects.select_related("district", "article", "fund_request").filter(**filters)
    if beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        return models.PublicBeneficiaryEntry.objects.select_related("article", "fund_request").filter(**filters)
    return models.InstitutionsBeneficiaryEntry.objects.select_related("article", "fund_request").filter(**filters)


def _build_aid_option_payload(entry, beneficiary_type):
    amount = float(entry.total_amount or 0)
    amount_display = format(amount, ".2f").rstrip("0").rstrip(".")
    details_display = (entry.notes or "").strip() or "-"
    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        beneficiary_name = entry.district.district_name
        return {
            "source_entry_id": entry.pk,
            "application_number": entry.application_number or "",
            "display_text": f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
            "recipient_name": beneficiary_name,
            "name_of_beneficiary": entry.name_of_beneficiary or "",
            "name_of_institution": entry.name_of_institution or "",
            "details": entry.notes or "",
            "fund_requested": amount,
            "aadhar_number": entry.aadhar_number or "",
            "cheque_in_favour": entry.cheque_rtgs_in_favour or "",
            "district_name": beneficiary_name,
            "source_item": entry.article.article_name,
        }
    if beneficiary_type == models.RecipientTypeChoices.PUBLIC:
        beneficiary_name = entry.name
        return {
            "source_entry_id": entry.pk,
            "application_number": entry.application_number or "",
            "display_text": f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
            "recipient_name": beneficiary_name,
            "name_of_beneficiary": beneficiary_name,
            "name_of_institution": entry.name_of_institution or "",
            "details": entry.notes or "",
            "fund_requested": amount,
            "aadhar_number": entry.aadhar_number or "",
            "cheque_in_favour": entry.cheque_rtgs_in_favour or "",
            "district_name": "",
            "source_item": entry.article.article_name,
        }
    beneficiary_name = entry.institution_name
    return {
        "source_entry_id": entry.pk,
        "application_number": entry.application_number or "",
        "display_text": f"{entry.application_number or ''} - {beneficiary_name} - Rs.{amount_display} - {details_display}",
        "recipient_name": beneficiary_name,
        "name_of_beneficiary": entry.name_of_beneficiary or "",
        "name_of_institution": entry.name_of_institution or beneficiary_name,
        "details": entry.notes or "",
        "fund_requested": amount,
        "aadhar_number": entry.aadhar_number or "",
        "cheque_in_favour": entry.cheque_rtgs_in_favour or "",
        "district_name": "",
        "source_item": entry.article.article_name,
    }


def _get_aid_beneficiary_options(aid_type, beneficiary_type, current_fund_request=None):
    aid_type = (aid_type or "").strip()
    blocked = set()
    options = []
    if not aid_type or beneficiary_type not in {
        models.RecipientTypeChoices.DISTRICT,
        models.RecipientTypeChoices.PUBLIC,
        models.RecipientTypeChoices.INSTITUTIONS,
    }:
        return options, []

    for entry in _aid_entry_queryset(aid_type, beneficiary_type).order_by("application_number", "created_at"):
        if entry.fund_request_id and (not current_fund_request or entry.fund_request_id != current_fund_request.id):
            label = entry.fund_request.formatted_fund_request_number if entry.fund_request and entry.fund_request.fund_request_number else f"Draft #{entry.fund_request_id}"
            if entry.fund_request:
                blocked.add(f"{label} ({entry.fund_request.get_status_display()})")
            else:
                blocked.add(label)
            continue
        options.append(_build_aid_option_payload(entry, beneficiary_type))
    return options, sorted(blocked)


def _get_aid_available_beneficiary_type_choices(aid_type, current_fund_request=None):
    results = []
    for value, label in [
        (models.RecipientTypeChoices.DISTRICT, "District"),
        (models.RecipientTypeChoices.PUBLIC, "Public"),
        (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
    ]:
        options, _blocked = _get_aid_beneficiary_options(aid_type, value, current_fund_request=current_fund_request)
        if options:
            results.append({"value": value, "label": label, "count": len(options)})
    return results


class FundRequestAidOptionsView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "create_edit"

    def get(self, request, *args, **kwargs):
        aid_type = request.GET.get("aid_type") or ""
        beneficiary_type = request.GET.get("beneficiary_type") or ""
        current_id = request.GET.get("fund_request_id") or ""
        current_fund_request = None
        if current_id:
            current_fund_request = models.FundRequest.objects.filter(pk=current_id).first()
        options_by_type = {}
        blocked_by_type = {}
        for type_key in [
            models.RecipientTypeChoices.DISTRICT,
            models.RecipientTypeChoices.PUBLIC,
            models.RecipientTypeChoices.INSTITUTIONS,
        ]:
            type_options, type_blocked = _get_aid_beneficiary_options(aid_type, type_key, current_fund_request=current_fund_request)
            options_by_type[str(type_key)] = type_options
            blocked_by_type[str(type_key)] = type_blocked
        payload = {
            "available_types": _get_aid_available_beneficiary_type_choices(aid_type, current_fund_request=current_fund_request),
            "options": options_by_type.get(str(beneficiary_type), []) if beneficiary_type else [],
            "blocked": blocked_by_type.get(str(beneficiary_type), []) if beneficiary_type else [],
            "options_by_type": options_by_type,
            "blocked_by_type": blocked_by_type,
        }
        return JsonResponse(payload)


class FundRequestCreateUpdateMixin(WriteRoleMixin):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "create_edit"
    form_class = FundRequestForm
    template_name = "order_fund_request/fund_request_form.html"
    model = models.FundRequest
    success_url = reverse_lazy("ui:fund-request-list")

    def _build_formsets(self, instance: models.FundRequest | None = None):
        recipient_formset = FundRequestRecipientFormSet(self.request.POST or None, prefix="recipients", instance=instance)
        article_formset = FundRequestArticleFormSet(self.request.POST or None, prefix="articles", instance=instance)
        return recipient_formset, article_formset

    def _can_edit(self, fr: models.FundRequest):
        return _is_editable_by_user(self.request.user, fr)

    def is_purchase_order_mode(self):
        return False

    def dispatch(self, request, *args, **kwargs):
        self.object = getattr(self, "object", None)
        if self.object and not self._can_edit(self.object):
            messages.error(request, "Submitted fund requests must be reopened before editing.")
            return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[self.object.pk]))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["recipient_formset"] = kwargs.get("recipient_formset", None) or self._build_formsets(self.object)[0]
        context["article_formset"] = kwargs.get("article_formset", None) or self._build_formsets(self.object)[1]
        context["status_choices"] = models.FundRequestStatusChoices.choices
        context["aid_type_choices"] = _fund_request_aid_type_choices()
        context["article_request_choices_json"] = json.dumps(_fund_request_article_choices(self.object))
        context["current_fund_request_id"] = getattr(self.object, "pk", "") or ""
        context["purchase_order_mode"] = self.is_purchase_order_mode()
        context["back_url"] = reverse("ui:purchase-order-list") if self.is_purchase_order_mode() else reverse("ui:fund-request-list")
        vendor_groups = []
        article_vendor_group_keys = {}
        if self.request.method == "POST":
            raw_groups = self.request.POST.get("vendor_groups_json", "[]")
            try:
                parsed_groups = json.loads(raw_groups or "[]")
            except (TypeError, ValueError):
                parsed_groups = []
            vendor_groups = [group for group in (_normalize_vendor_group_payload(item) for item in parsed_groups) if group]
            for form in context["article_formset"].forms:
                article_vendor_group_keys[form.prefix] = str(self.request.POST.get(f"{form.prefix}-vendor_group_key") or "").strip()
        elif self.object:
            vendor_groups, row_key_map = _build_vendor_groups_from_articles(self.object.articles.all())
            for form in context["article_formset"].forms:
                instance_pk = getattr(getattr(form, "instance", None), "pk", None)
                article_vendor_group_keys[form.prefix] = row_key_map.get(str(instance_pk), "")
        context["vendor_groups_json"] = json.dumps(vendor_groups)
        context["article_vendor_group_keys_json"] = json.dumps(article_vendor_group_keys)
        context["vendor_options_json"] = json.dumps(
            [
                {
                    "id": vendor.id,
                    "vendor_name": vendor.vendor_name or "",
                    "gst_no": vendor.gst_number or "",
                    "vendor_address": vendor.address or "",
                    "vendor_city": vendor.city or "",
                    "vendor_state": vendor.state or "",
                    "vendor_pincode": vendor.pincode or "",
                    "phone_number": vendor.phone_number or "",
                    "cheque_in_favour": vendor.cheque_in_favour or "",
                }
                for vendor in models.Vendor.objects.filter(is_active=True).order_by("vendor_name")
            ]
        )
        return context

    def _parse_vendor_groups_from_request(self):
        raw_groups = self.request.POST.get("vendor_groups_json", "[]")
        try:
            parsed_groups = json.loads(raw_groups or "[]")
        except (TypeError, ValueError):
            parsed_groups = []
        return {
            group["key"]: group
            for group in (_normalize_vendor_group_payload(item) for item in parsed_groups)
            if group
        }

    def _apply_article_vendor_summary(self, fr, article_formset, vendor_groups):
        if fr.fund_request_type != models.FundRequestTypeChoices.ARTICLE or self.is_purchase_order_mode():
            return
        active_keys = []
        for form in article_formset.forms:
            if not getattr(form, "cleaned_data", None) or form.cleaned_data.get("DELETE", False):
                continue
            group_key = str(self.request.POST.get(f"{form.prefix}-vendor_group_key") or "").strip()
            if group_key and group_key in vendor_groups:
                active_keys.append(group_key)
        active_keys = list(dict.fromkeys(active_keys))
        if len(active_keys) == 1:
            group = vendor_groups[active_keys[0]]
            fr.supplier_name = group["vendor_name"]
            fr.gst_number = group["gst_no"]
            fr.supplier_address = group["vendor_address"]
            fr.supplier_city = group["vendor_city"]
            fr.supplier_state = group["vendor_state"]
            fr.supplier_pincode = group["vendor_pincode"]
        elif len(active_keys) > 1:
            fr.supplier_name = "Multiple Vendors"
            fr.gst_number = ""
            fr.supplier_address = ""
            fr.supplier_city = ""
            fr.supplier_state = ""
            fr.supplier_pincode = ""

    def _collect_totals(self, instance: models.FundRequest):
        for article in instance.articles.all():
            article.recompute_totals(unit_price=article.unit_price, quantity=article.quantity)
        services.sync_fund_request_totals(instance)

    def _resolve_article_record(self, article_name: str):
        article_name = (article_name or "").strip()
        if not article_name:
            return None
        article = models.Article.objects.filter(article_name__iexact=article_name).first()
        if article:
            return article
        return models.Article.objects.create(
            article_name=article_name,
            cost_per_unit=0,
            item_type=models.ItemTypeChoices.ARTICLE,
            combo=True,
            is_active=False,
        )

    def _link_aid_sources(self, fr: models.FundRequest):
        models.DistrictBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        models.PublicBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        models.InstitutionsBeneficiaryEntry.objects.filter(fund_request=fr).update(fund_request=None)
        if fr.fund_request_type != models.FundRequestTypeChoices.AID:
            return
        for recipient in fr.recipients.exclude(source_entry_id__isnull=True).exclude(source_entry_id__exact=0):
            if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT:
                models.DistrictBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)
            elif recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC:
                models.PublicBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)
            elif recipient.beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
                models.InstitutionsBeneficiaryEntry.objects.filter(pk=recipient.source_entry_id).update(fund_request=fr)

    def _is_aid_source_available(self, beneficiary_type, source_entry_id, current_fund_request=None):
        if not source_entry_id or not beneficiary_type:
            return False, "Select a valid beneficiary."
        entry = None
        if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
            entry = models.DistrictBeneficiaryEntry.objects.select_related("fund_request").filter(pk=source_entry_id).first()
        elif beneficiary_type == models.RecipientTypeChoices.PUBLIC:
            entry = models.PublicBeneficiaryEntry.objects.select_related("fund_request").filter(pk=source_entry_id).first()
        elif beneficiary_type in {models.RecipientTypeChoices.INSTITUTIONS, models.RecipientTypeChoices.OTHERS}:
            entry = models.InstitutionsBeneficiaryEntry.objects.select_related("fund_request").filter(pk=source_entry_id).first()
        if not entry:
            return False, "This beneficiary is no longer available."
        if entry.fund_request_id and (not current_fund_request or entry.fund_request_id != current_fund_request.id):
            label = entry.fund_request.formatted_fund_request_number if entry.fund_request and entry.fund_request.fund_request_number else f"Draft #{entry.fund_request_id}"
            if entry.fund_request:
                return False, f"Already present in {label} ({entry.fund_request.get_status_display()})."
            return False, f"Already present in {label}."
        return True, ""

    def _validate_fund_request_formsets(self, fr, action, recipient_formset, article_formset):
        is_valid = True
        if fr.fund_request_type == models.FundRequestTypeChoices.AID:
            if action == "submit" and not (fr.aid_type or "").strip():
                return False
            active_forms = [form for form in recipient_formset.forms if form.cleaned_data and not form.cleaned_data.get("DELETE", False)]
            if action == "submit" and not active_forms:
                recipient_formset._non_form_errors = recipient_formset.error_class(["Add at least one recipient."])
                return False
            seen_source_keys = {}
            for form in active_forms:
                beneficiary_type = form.cleaned_data.get("beneficiary_type")
                source_entry_id = form.cleaned_data.get("source_entry_id")
                if beneficiary_type and source_entry_id:
                    source_key = (str(beneficiary_type), str(source_entry_id))
                    if source_key in seen_source_keys:
                        form.add_error("beneficiary", "This recipient is already added in the same fund request.")
                        seen_source_keys[source_key].add_error("beneficiary", "This recipient is already added in the same fund request.")
                        is_valid = False
                    else:
                        seen_source_keys[source_key] = form
                if source_entry_id:
                    ok, message = self._is_aid_source_available(
                        beneficiary_type,
                        source_entry_id,
                        current_fund_request=fr if getattr(fr, "pk", None) else None,
                    )
                    if not ok:
                        form.add_error("beneficiary", message)
                        is_valid = False
                if action == "submit":
                    required_fields = ["beneficiary_type", "beneficiary", "source_entry_id", "fund_requested", "name_of_beneficiary", "name_of_institution", "details", "cheque_in_favour"]
                    if beneficiary_type != models.RecipientTypeChoices.DISTRICT:
                        required_fields.append("aadhar_number")
                    for field_name in required_fields:
                        value = form.cleaned_data.get(field_name)
                        if value in (None, "", 0, "0"):
                            form.add_error(field_name, "Required for submit.")
                            is_valid = False
        else:
            vendor_groups = self._parse_vendor_groups_from_request()
            active_forms = [form for form in article_formset.forms if form.cleaned_data and not form.cleaned_data.get("DELETE", False)]
            if action == "submit" and not active_forms:
                article_formset._non_form_errors = article_formset.error_class(["Add at least one item."])
                return False
            if action == "submit":
                for form in active_forms:
                    required_fields = ["article_name", "quantity", "unit_price"]
                    for field_name in required_fields:
                        value = form.cleaned_data.get(field_name)
                        if value in (None, "", 0, "0"):
                            form.add_error(field_name, "Required for submit.")
                            is_valid = False
                    if not self.is_purchase_order_mode():
                        group_key = str(self.request.POST.get(f"{form.prefix}-vendor_group_key") or "").strip()
                        if not group_key or group_key not in vendor_groups:
                            form.add_error("article_name", "Select a vendor.")
                            is_valid = False
                            continue
                        group = vendor_groups[group_key]
                        for value in [
                            group.get("vendor_name"),
                            group.get("gst_no"),
                            group.get("cheque_in_favour"),
                        ]:
                            if not str(value or "").strip():
                                form.add_error("article_name", "Complete the selected vendor details.")
                                is_valid = False
                                break
        return is_valid

    def _validate_article_header_fields(self, form, fr, action):
        if action != "submit" or fr.fund_request_type != models.FundRequestTypeChoices.ARTICLE or not self.is_purchase_order_mode():
            return True
        valid = True
        labels = [
            ("supplier_name", "Vendor Name" if self.is_purchase_order_mode() else "Supplier Name"),
            ("supplier_address", "Vendor Address" if self.is_purchase_order_mode() else "Address"),
            ("supplier_city", "City"),
            ("supplier_state", "State"),
            ("supplier_pincode", "Pincode"),
        ]
        for field_name, label in labels:
            value = getattr(fr, field_name, None)
            if not str(value or "").strip():
                form.add_error(field_name, f"{label} is required for submit.")
                valid = False
        return valid

    def _set_fund_request_status(self, fr, action):
        if action == "submit":
            fr.status = models.FundRequestStatusChoices.SUBMITTED
        else:
            fr.status = models.FundRequestStatusChoices.DRAFT

    def form_valid(self, form):
        action = self.request.POST.get("action", "draft")
        if action == "submit" and self.object and self.object.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(self.request, "Only draft fund requests can be submitted.")
            return HttpResponseRedirect(self.get_success_url())

        fr = form.save(commit=False)
        if self.is_purchase_order_mode():
            fr.fund_request_type = models.FundRequestTypeChoices.ARTICLE
            fr.aid_type = None
        if self.object and self.object.fund_request_number:
            fr.fund_request_number = self.object.fund_request_number
        if self.object and self.object.purchase_order_number:
            fr.purchase_order_number = self.object.purchase_order_number
        if not fr.created_by:
            fr.created_by = self.request.user
        self._set_fund_request_status(fr, action)

        recipient_formset, article_formset = self._build_formsets(fr)
        formsets_ok = recipient_formset.is_valid() and article_formset.is_valid()

        header_ok = True
        if fr.fund_request_type == models.FundRequestTypeChoices.AID and action == "submit" and not (fr.aid_type or "").strip():
            form.add_error("aid_type", "Select the aid type before submit.")
            header_ok = False
        if formsets_ok and fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE:
            self._apply_article_vendor_summary(fr, article_formset, self._parse_vendor_groups_from_request())
        header_ok = self._validate_article_header_fields(form, fr, action) and header_ok

        if not header_ok or not formsets_ok or not self._validate_fund_request_formsets(fr, action, recipient_formset, article_formset):
            messages.error(self.request, "Please fix errors in recipients/articles before saving.")
            return self.render_to_response(
                self.get_context_data(
                    form=form,
                    recipient_formset=recipient_formset,
                    article_formset=article_formset,
                )
            )

        try:
            with transaction.atomic():
                if action == "submit" and not fr.fund_request_number:
                    fr.fund_request_number = services.next_fund_request_number()
                if action == "submit" and fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE and not fr.purchase_order_number:
                    fr.purchase_order_number = purchase_order_services.next_purchase_order_number()
                fr.save()

                recipient_formset.instance = fr
                article_formset.instance = fr

                for deleted_form in getattr(recipient_formset, "deleted_forms", []):
                    deleted_instance = getattr(deleted_form, "instance", None)
                    if deleted_instance and deleted_instance.pk:
                        deleted_instance.delete()
                recipient_instances = recipient_formset.save(commit=False)
                for recipient in recipient_instances:
                    recipient.fund_request = fr
                    if recipient.beneficiary_type == models.RecipientTypeChoices.DISTRICT:
                        recipient.recipient_name = (
                            recipient.name_of_beneficiary
                            or recipient.name_of_institution
                            or recipient.district_name
                            or recipient.recipient_name
                            or recipient.beneficiary
                            or "Recipient"
                        )
                    elif recipient.beneficiary_type == models.RecipientTypeChoices.PUBLIC:
                        recipient.recipient_name = recipient.name_of_beneficiary or recipient.recipient_name or recipient.beneficiary or "Recipient"
                    elif recipient.beneficiary_type in {
                        models.RecipientTypeChoices.INSTITUTIONS,
                        models.RecipientTypeChoices.OTHERS,
                    }:
                        recipient.recipient_name = (
                            recipient.name_of_institution
                            or recipient.name_of_beneficiary
                            or recipient.recipient_name
                            or recipient.beneficiary
                            or "Recipient"
                        )
                    else:
                        recipient.recipient_name = recipient.recipient_name or recipient.name_of_beneficiary or recipient.name_of_institution or recipient.beneficiary or "Recipient"
                    recipient.save()

                for deleted_form in getattr(article_formset, "deleted_forms", []):
                    deleted_instance = getattr(deleted_form, "instance", None)
                    if deleted_instance and deleted_instance.pk:
                        deleted_instance.delete()
                vendor_groups = self._parse_vendor_groups_from_request()
                for article_form in article_formset.forms:
                    if not getattr(article_form, "cleaned_data", None) or article_form.cleaned_data.get("DELETE", False):
                        continue
                    article = article_form.save(commit=False)
                    article.fund_request = fr
                    if not article.article_id:
                        article.article = self._resolve_article_record(article.article_name)
                    if article.article and not article.article_name:
                        article.article_name = article.article.article_name
                    if self.is_purchase_order_mode():
                        article.vendor_name = fr.supplier_name
                        article.gst_no = fr.gst_number
                        article.vendor_address = fr.supplier_address
                        article.vendor_city = fr.supplier_city
                        article.vendor_state = fr.supplier_state
                        article.vendor_pincode = fr.supplier_pincode
                    else:
                        group_key = str(self.request.POST.get(f"{article_form.prefix}-vendor_group_key") or "").strip()
                        group = vendor_groups.get(group_key, {})
                        vendor_id = group.get("vendor_id")
                        article.vendor_id = int(vendor_id) if vendor_id else None
                        article.vendor_name = group.get("vendor_name", "")
                        article.gst_no = group.get("gst_no", "")
                        article.vendor_address = group.get("vendor_address", "")
                        article.vendor_city = group.get("vendor_city", "")
                        article.vendor_state = group.get("vendor_state", "")
                        article.vendor_pincode = group.get("vendor_pincode", "")
                        article.cheque_in_favour = group.get("cheque_in_favour", "")
                    article.unit_price = article.unit_price or 0
                    article.price_including_gst = article.unit_price * (article.quantity or 0)
                    article.value = article.price_including_gst
                    article.cumulative = article.value
                    article.save()

                self._link_aid_sources(fr)
                self._collect_totals(fr)
                services.sync_order_entries_from_fund_request(fr, actor=self.request.user)
                self.object = fr

                if action == "submit":
                    log_audit(
                        user=self.request.user,
                        action_type=models.ActionTypeChoices.STATUS_CHANGE,
                        entity_type="fund_request",
                        entity_id=str(fr.id),
                        details={"status": models.FundRequestStatusChoices.SUBMITTED},
                        ip_address=self.request.META.get("REMOTE_ADDR"),
                        user_agent=self.request.META.get("HTTP_USER_AGENT", ""),
                    )
                    messages.success(self.request, "Fund request submitted.")
                else:
                    messages.success(self.request, "Fund request saved as draft.")
        except IntegrityError:
            form.add_error(None, "Fund request number already exists. Please try submitting again.")
            return self.render_to_response(
                self.get_context_data(
                    form=form,
                    recipient_formset=recipient_formset,
                    article_formset=article_formset,
                )
            )
        return HttpResponseRedirect(self.get_success_url())


class FundRequestCreateView(LoginRequiredMixin, FundRequestCreateUpdateMixin, CreateView):
    pass


class FundRequestUpdateView(LoginRequiredMixin, FundRequestCreateUpdateMixin, UpdateView):
    pass


class FundRequestDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"
    model = models.FundRequest
    template_name = "order_fund_request/fund_request_detail.html"
    context_object_name = "fund_request"

    def get_queryset(self):
        return (
            models.FundRequest.objects.select_related("created_by")
            .prefetch_related("recipients", "articles")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["media_url"] = settings.MEDIA_URL
        context["can_edit"] = _is_editable_by_user(self.request.user, self.object)
        context["can_reopen"] = self.request.user.role == "admin" and self.object.status == models.FundRequestStatusChoices.SUBMITTED
        context["can_delete"] = self.request.user.role == "admin" and self.request.user.has_module_permission(
            models.ModuleKeyChoices.ORDER_FUND_REQUEST,
            "delete",
        )
        context["back_url"] = reverse("ui:fund-request-list")
        recipient_rows = list(self.object.recipients.all())
        for recipient in recipient_rows:
            recipient.beneficiary_display = _fund_request_recipient_display_name(recipient)
        context["recipient_rows"] = recipient_rows
        article_rows = list(self.object.articles.all())
        for article in article_rows:
            article.beneficiary_display = _fund_request_article_beneficiary_display(article)
        context["article_rows"] = article_rows
        return context


class FundRequestPDFView(LoginRequiredMixin, RoleRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "view"

    def get(self, request, pk):
        fund_request = get_object_or_404(
            models.FundRequest.objects.select_related("created_by").prefetch_related("recipients", "articles"),
            pk=pk,
        )
        pdf_buffer = services.generate_fund_request_pdf(fund_request)
        filename_base = fund_request.formatted_fund_request_number or f"FR-DRAFT-{fund_request.pk}"
        response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename_base}.pdf"'
        return response


class FundRequestDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "delete"
    model = models.FundRequest
    template_name = "order_fund_request/fund_request_confirm_delete.html"
    success_url = reverse_lazy("ui:fund-request-list")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        models.OrderEntry.objects.filter(fund_request=self.object).delete()
        messages.warning(self.request, "Fund request deleted.")
        return super().post(request, *args, **kwargs)


class FundRequestSubmitView(LoginRequiredMixin, WriteRoleMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "submit"
    allowed_roles = {"admin", "editor"}

    def post(self, request, pk):
        fr = models.FundRequest.objects.get(pk=pk)
        if not _is_editable_by_user(request.user, fr) or request.user.role == "viewer":
            return HttpResponse("Forbidden", status=403)
        if fr.status != models.FundRequestStatusChoices.DRAFT:
            messages.error(request, "Only draft fund requests can be submitted.")
            return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))
        fr.status = models.FundRequestStatusChoices.SUBMITTED
        if not fr.fund_request_number:
            fr.fund_request_number = services.next_fund_request_number()
        update_fields = ["status", "fund_request_number"]
        if fr.fund_request_type == models.FundRequestTypeChoices.ARTICLE and not fr.purchase_order_number:
            fr.purchase_order_number = purchase_order_services.next_purchase_order_number()
            update_fields.append("purchase_order_number")
        fr.save(update_fields=update_fields)
        log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="fund_request",
            entity_id=str(fr.id),
            details={"status": models.FundRequestStatusChoices.SUBMITTED},
            ip_address=request.META.get("REMOTE_ADDR"),
            user_agent=request.META.get("HTTP_USER_AGENT", ""),
        )
        services.sync_fund_request_totals(fr)
        services.sync_order_entries_from_fund_request(fr, actor=request.user)
        messages.success(request, "Fund request submitted.")
        return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))


class FundRequestReopenView(LoginRequiredMixin, AdminRequiredMixin, View):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "reopen"

    def post(self, request, pk):
        fr = get_object_or_404(models.FundRequest, pk=pk)
        if fr.status != models.FundRequestStatusChoices.SUBMITTED:
            messages.error(request, "Only submitted fund requests can be reopened.")
            return HttpResponseRedirect(reverse("ui:fund-request-detail", args=[fr.pk]))
        previous_status = fr.status
        fr.status = models.FundRequestStatusChoices.DRAFT
        fr.save(update_fields=["status"])
        log_audit(
            user=request.user,
            action_type=models.ActionTypeChoices.STATUS_CHANGE,
            entity_type="fund_request",
            entity_id=str(fr.id),
            details={"from": previous_status, "to": models.FundRequestStatusChoices.DRAFT},
            **get_request_audit_meta(request),
        )
        services.sync_fund_request_totals(fr)
        services.sync_order_entries_from_fund_request(fr, actor=request.user)
        messages.success(request, "Fund request reopened as draft.")
        return HttpResponseRedirect(reverse("ui:fund-request-edit", args=[fr.pk]))


class FundRequestDocumentUploadView(LoginRequiredMixin, WriteRoleMixin, FormView):
    module_key = models.ModuleKeyChoices.ORDER_FUND_REQUEST
    permission_action = "create_edit"
    template_name = "order_fund_request/fund_request_upload_document.html"
    form_class = FundRequestDocumentUploadForm

    def dispatch(self, request, *args, **kwargs):
        self.fund_request = models.FundRequest.objects.get(pk=kwargs["pk"])
        if not _is_editable_by_user(request.user, self.fund_request):
            return HttpResponse("Forbidden", status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse("ui:fund-request-detail", args=[self.fund_request.pk])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["fund_request"] = self.fund_request
        return context

    def form_valid(self, form):
        uploaded_file = form.cleaned_data["file"]
        upload_dir = os.path.join("fund-request-docs", str(self.fund_request.pk))
        relative_path = os.path.join(upload_dir, uploaded_file.name)
        from django.core.files.storage import default_storage

        stored_path = default_storage.save(relative_path, uploaded_file)
        models.FundRequestDocument.objects.create(
            fund_request=self.fund_request,
            document_type=form.cleaned_data["document_type"],
            file_path=stored_path,
            file_name=uploaded_file.name,
            generated_by=self.request.user,
        )
        messages.success(self.request, "Document uploaded.")
        return HttpResponseRedirect(self.get_success_url())
