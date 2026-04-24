from __future__ import annotations

"""Reports helper/service functions extracted from legacy web_views."""

import base64
from datetime import date
import re

from django.db.models import F
from django.utils import timezone
from pypdf import PdfReader, PdfWriter

import io
import zipfile
from pathlib import Path

from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Frame, KeepTogether, LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover - optional dependency fallback
    PILImage = None

from core.shared.format_utils import _ordinal
from core.shared.pdf_utils import _fitted_pdf_image
from core.shared.pdf_utils import _fitted_pdf_image_source
from core.shared.pdf_utils import _normalized_docx_report_logo
from core.shared.pdf_utils import _optimized_report_logo
from core.shared.pdf_utils import _pdf_guru_logo_path
from core.shared.pdf_utils import _pdf_logo_path

from core import models
from core.shared.phase2 import _phase2_normalize_text
from core.shared.phase2 import _phase2_parse_number
from core.shared.token_generation import _token_generation_saved_dataset


REPORTS_WAITING_HALL_STATE_KEY = "reports_waiting_hall_acknowledgment"
REPORTS_SHARED_LOGO_KEY = "reports_shared_logo"
REPORTS_PUBLIC_ACK_STATE_KEY = "reports_public_acknowledgment"
REPORTS_TOKEN_LOOKUP_STATE_KEY = "reports_token_lookup"
REPORTS_PUBLIC_SIGNATURE_STATE_KEY = "reports_public_signature"
REPORTS_DISTRICT_SIGNATURE_STATE_KEY = "reports_district_signature"
REPORTS_SEGREGATION_STATE_KEY = "reports_segregation"
REPORTS_DISTRIBUTION_STATE_KEY = "reports_distribution"
STAGE_DISTRIBUTION_BENEFICIARY_FILTER_CHOICES = [
    ("all", "All"),
    (models.RecipientTypeChoices.DISTRICT, "District"),
    (models.RecipientTypeChoices.PUBLIC, "Public"),
    (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
    (models.RecipientTypeChoices.OTHERS, "Others"),
]
STAGE_DISTRIBUTION_ITEM_FILTER_CHOICES = [
    ("all", "All"),
    (models.ItemTypeChoices.ARTICLE, "Article"),
    (models.ItemTypeChoices.AID, "Aid"),
]
STAGE_DISTRIBUTION_PREMISE_FILTER_CHOICES = [
    ("all", "All"),
    ("waiting_hall", "Waiting Hall Qty only"),
    ("masm_hall", "Masm Hall Qty only"),
]


def _reports_active_session():
    session = models.EventSession.objects.filter(is_active=True).order_by("-event_year", "session_name", "-id").first()
    if session:
        return session
    return models.EventSession.objects.order_by("-event_year", "session_name", "-id").first()


def _reports_parse_date(raw_value: str | None) -> date:
    raw = str(raw_value or "").strip()
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return timezone.localdate()


def _reports_waiting_hall_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "ignored_keys": [],
        "rows": [],
        "headers": [],
        "beneficiary_type_filter": "",
        "item_type_filter": models.ItemTypeChoices.AID,
    }


def _reports_waiting_hall_session_state(request):
    state = request.session.get(REPORTS_WAITING_HALL_STATE_KEY) or {}
    merged = _reports_waiting_hall_default_state()
    merged.update(state)
    return merged


def _reports_shared_logo_state(request):
    state = request.session.get(REPORTS_SHARED_LOGO_KEY) or {}
    return {
        "logo_name": str(state.get("logo_name") or ""),
        "logo_content_type": str(state.get("logo_content_type") or ""),
        "logo_base64": str(state.get("logo_base64") or ""),
    }


def _reports_set_shared_logo_state(request, *, uploaded_logo):
    request.session[REPORTS_SHARED_LOGO_KEY] = {
        "logo_name": str(getattr(uploaded_logo, "name", "") or ""),
        "logo_content_type": str(getattr(uploaded_logo, "content_type", "") or "image/png"),
        "logo_base64": base64.b64encode(uploaded_logo.read()).decode("ascii"),
    }
    request.session.modified = True


def _reports_public_ack_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "template_name": "",
        "template_base64": "",
        "template_content_type": "application/pdf",
        "template_fields": [],
        "field_map": {},
    }


def _reports_public_ack_session_state(request):
    state = request.session.get(REPORTS_PUBLIC_ACK_STATE_KEY) or {}
    merged = _reports_public_ack_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    merged["template_fields"] = [dict(field) for field in list(merged.get("template_fields") or []) if isinstance(field, dict)]
    merged["field_map"] = {
        str(key or "").strip(): str(value or "").strip()
        for key, value in dict(merged.get("field_map") or {}).items()
        if str(key or "").strip()
    }
    return merged


def _reports_set_public_ack_state(request, state):
    request.session[REPORTS_PUBLIC_ACK_STATE_KEY] = state
    request.session.modified = True


def _reports_public_signature_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "selected_items": [],
        "sort_modes": [],
    }


def _reports_public_signature_session_state(request):
    state = request.session.get(REPORTS_PUBLIC_SIGNATURE_STATE_KEY) or {}
    merged = _reports_public_signature_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    merged["selected_items"] = [
        str(item or "").strip()
        for item in list(merged.get("selected_items") or [])
        if str(item or "").strip()
    ]
    sort_modes = [
        str(mode or "").strip()
        for mode in list(merged.get("sort_modes") or [])
        if str(mode or "").strip()
    ]
    valid_modes = []
    for mode in sort_modes:
        if mode in {"application_number", "item_name", "token_number"} and mode not in valid_modes:
            valid_modes.append(mode)
    merged["sort_modes"] = valid_modes
    return merged


def _reports_set_public_signature_state(request, state):
    request.session[REPORTS_PUBLIC_SIGNATURE_STATE_KEY] = state
    request.session.modified = True


def _reports_district_signature_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
    }


def _reports_district_signature_session_state(request):
    state = request.session.get(REPORTS_DISTRICT_SIGNATURE_STATE_KEY) or {}
    merged = _reports_district_signature_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    return merged


def _reports_set_district_signature_state(request, state):
    request.session[REPORTS_DISTRICT_SIGNATURE_STATE_KEY] = state
    request.session.modified = True


def _reports_simple_report_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
    }


def _reports_simple_report_session_state(request, state_key: str):
    state = request.session.get(state_key) or {}
    merged = _reports_simple_report_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    return merged


def _reports_set_simple_report_state(request, state_key: str, state):
    request.session[state_key] = state
    request.session.modified = True


SEGREGATION_BENEFICIARY_FILTER_CHOICES = [
    ("all", "All"),
    (models.RecipientTypeChoices.DISTRICT, "District"),
    (models.RecipientTypeChoices.PUBLIC, "Public"),
    (models.RecipientTypeChoices.INSTITUTIONS, "Institutions"),
    (models.RecipientTypeChoices.OTHERS, "Others"),
]

SEGREGATION_ITEM_FILTER_CHOICES = [
    ("all", "All"),
    (models.ItemTypeChoices.ARTICLE, "Article"),
    (models.ItemTypeChoices.AID, "Aid"),
]


def _segregation_pick_value(row: dict, aliases: list[str], default=""):
    item = dict(row or {})
    normalized = {
        _phase2_normalize_text(key): value
        for key, value in item.items()
        if str(key or "").strip()
    }
    for alias in aliases:
        normalized_alias = _phase2_normalize_text(alias)
        if normalized_alias not in normalized:
            continue
        value = normalized.get(normalized_alias)
        if value == 0:
            return value
        if str(value or "").strip():
            return value
    return default


def _segregation_display_text(*values):
    for value in values:
        if value == 0:
            return "0"
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _segregation_resolved_item_type(raw_value: str | None, *, default=models.ItemTypeChoices.ARTICLE) -> str:
    value = str(raw_value or "").strip()
    if not value and raw_value == "":
        return ""
    allowed_item_types = {choice[0] for choice in SEGREGATION_ITEM_FILTER_CHOICES}
    if value in allowed_item_types:
        return value
    normalized_value = _phase2_normalize_text(value)
    if normalized_value == _phase2_normalize_text(models.ItemTypeChoices.ARTICLE):
        return models.ItemTypeChoices.ARTICLE
    if normalized_value == _phase2_normalize_text(models.ItemTypeChoices.AID):
        return models.ItemTypeChoices.AID
    if normalized_value == "all":
        return ""
    return default


def _segregation_type_order(value: str) -> int:
    order = {
        models.RecipientTypeChoices.DISTRICT: 0,
        models.RecipientTypeChoices.PUBLIC: 1,
        models.RecipientTypeChoices.INSTITUTIONS: 2,
        models.RecipientTypeChoices.OTHERS: 3,
    }
    return order.get(str(value or "").strip(), 99)


def _segregation_normalize_row(row: dict) -> dict:
    item = dict(row or {})
    application_number = _segregation_display_text(
        _segregation_pick_value(item, ["Application Number", "App No", "application_number"]),
    )
    beneficiary_type = _segregation_display_text(
        _segregation_pick_value(item, ["Beneficiary Type", "beneficiary_type"]),
    )
    item_type = _segregation_display_text(
        _segregation_pick_value(item, ["Item Type", "item_type"]),
    )
    district_name = _segregation_display_text(
        _segregation_pick_value(item, ["District", "district"]),
    )
    beneficiary_name = _segregation_display_text(
        _segregation_pick_value(item, ["Beneficiary Name", "beneficiary_name", "Name"]),
    )
    names_value = _segregation_display_text(
        _segregation_pick_value(item, ["Names", "Beneficiary Name", "Name"]),
    )
    item_name = _segregation_display_text(
        _segregation_pick_value(
            item,
            ["Token Name", "Requested Item", "Article Name", "Article", "Item"],
        ),
    )
    waiting_hall_quantity = _phase2_parse_number(
        _segregation_pick_value(item, ["Waiting Hall Quantity", "waiting_hall_quantity"], 0)
    )
    token_quantity = _phase2_parse_number(
        _segregation_pick_value(item, ["Token Quantity", "Token Qty", "token_quantity"], 0)
    )
    sequence_no = _phase2_parse_number(
        _segregation_pick_value(item, ["Sequence No", "Sequence List", "sequence_no"], 0)
    )
    start_token_no = _phase2_parse_number(
        _segregation_pick_value(item, ["Start Token No", "Start Token No.", "token_start"], 0)
    )
    end_token_no = _phase2_parse_number(
        _segregation_pick_value(item, ["End Token No", "End Token No.", "token_end"], 0)
    )
    if token_quantity <= 0 and start_token_no > 0 and end_token_no >= start_token_no:
        token_quantity = end_token_no - start_token_no + 1

    if beneficiary_type == models.RecipientTypeChoices.DISTRICT:
        beneficiary_label = _segregation_display_text(district_name, beneficiary_name, names_value, application_number)
    else:
        if application_number and beneficiary_name:
            beneficiary_label = f"{application_number} - {beneficiary_name}"
        else:
            beneficiary_label = _segregation_display_text(names_value, beneficiary_name, application_number)

    return {
        "application_number": application_number,
        "beneficiary_type": beneficiary_type,
        "item_type": item_type,
        "district_name": district_name,
        "beneficiary_name": beneficiary_name,
        "beneficiary_label": beneficiary_label,
        "item_name": item_name,
        "waiting_hall_quantity": waiting_hall_quantity,
        "token_quantity": token_quantity,
        "sequence_no": sequence_no,
        "start_token_no": start_token_no,
        "end_token_no": end_token_no,
    }


def _segregation_normalize_dataset(dataset: dict) -> dict:
    rows = []
    for row in list(dataset.get("rows") or []):
        normalized = _segregation_normalize_row(row)
        if not (
            normalized["beneficiary_label"]
            or normalized["item_name"]
            or normalized["waiting_hall_quantity"]
            or normalized["token_quantity"]
        ):
            continue
        rows.append(normalized)
    return {
        "rows": rows,
        "headers": list(dataset.get("headers") or []),
    }


def _segregation_filter_rows(
    rows: list[dict],
    *,
    beneficiary_types: list[str] | tuple[str, ...] | set[str] | None = None,
    item_types: list[str] | tuple[str, ...] | set[str] | None = None,
) -> list[dict]:
    filtered_rows = []
    beneficiary_type_set = {str(value or "").strip() for value in list(beneficiary_types or []) if str(value or "").strip()}
    item_type_set = {str(value or "").strip() for value in list(item_types or []) if str(value or "").strip()}
    for row in list(rows or []):
        row_beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        row_item_type = str(row.get("item_type") or "").strip()
        if beneficiary_type_set and row_beneficiary_type not in beneficiary_type_set:
            continue
        if item_type_set and row_item_type not in item_type_set:
            continue
        filtered_rows.append(dict(row))
    return filtered_rows


def _segregation_build_file1(rows: list[dict]) -> dict:
    grouped_map: dict[tuple[str, str], dict] = {}
    for row in list(rows or []):
        waiting_hall_quantity = int(row.get("waiting_hall_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        beneficiary_label = str(row.get("beneficiary_label") or "").strip()
        if waiting_hall_quantity <= 0 or not item_name or not beneficiary_label:
            continue
        group_key = (
            str(row.get("beneficiary_type") or "").strip(),
            beneficiary_label,
        )
        group = grouped_map.setdefault(
            group_key,
            {
                "beneficiary_type": group_key[0],
                "beneficiary_label": beneficiary_label,
                "sort_sequence": _segregation_type_order(group_key[0]),
                "items": {},
                "total_quantity": 0,
            },
        )
        group["items"][item_name] = int(group["items"].get(item_name) or 0) + waiting_hall_quantity
        group["total_quantity"] += waiting_hall_quantity

    groups = []
    row_count = 0
    total_quantity = 0
    for _, group in sorted(
        grouped_map.items(),
        key=lambda item: (
            item[1]["sort_sequence"],
            str(item[1]["beneficiary_label"]).casefold(),
        ),
    ):
        items = [
            {"article_name": item_name, "quantity": quantity}
            for item_name, quantity in sorted(group["items"].items(), key=lambda entry: entry[0].casefold())
        ]
        row_count += len(items)
        total_quantity += int(group["total_quantity"] or 0)
        groups.append(
            {
                "beneficiary_type": group["beneficiary_type"],
                "beneficiary_label": group["beneficiary_label"],
                "items": items,
                "total_quantity": int(group["total_quantity"] or 0),
            }
        )
    return {
        "groups": groups,
        "beneficiary_count": len(groups),
        "row_count": row_count,
        "total_quantity": total_quantity,
    }


def _segregation_build_file2(rows: list[dict]) -> dict:
    grouped_map: dict[str, dict] = {}
    for row in list(rows or []):
        waiting_hall_quantity = int(row.get("waiting_hall_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        beneficiary_label = str(row.get("beneficiary_label") or "").strip()
        if waiting_hall_quantity <= 0 or not item_name or not beneficiary_label:
            continue
        article_group = grouped_map.setdefault(
            item_name,
            {
                "article_name": item_name,
                "beneficiaries": {},
                "total_quantity": 0,
            },
        )
        article_group["beneficiaries"][beneficiary_label] = int(article_group["beneficiaries"].get(beneficiary_label) or 0) + waiting_hall_quantity
        article_group["total_quantity"] += waiting_hall_quantity

    groups = []
    beneficiary_row_count = 0
    total_quantity = 0
    for article_name, group in sorted(grouped_map.items(), key=lambda item: item[0].casefold()):
        beneficiaries = [
            {"beneficiary_label": beneficiary_label, "quantity": quantity}
            for beneficiary_label, quantity in sorted(group["beneficiaries"].items(), key=lambda entry: entry[0].casefold())
        ]
        beneficiary_row_count += len(beneficiaries)
        total_quantity += int(group["total_quantity"] or 0)
        groups.append(
            {
                "article_name": article_name,
                "beneficiaries": beneficiaries,
                "total_quantity": int(group["total_quantity"] or 0),
            }
        )
    return {
        "groups": groups,
        "article_count": len(groups),
        "row_count": beneficiary_row_count,
        "total_quantity": total_quantity,
    }


def _segregation_build_file3(rows: list[dict]) -> dict:
    stage_map: dict[tuple[int, str], dict] = {}
    for row in list(rows or []):
        token_quantity = int(row.get("token_quantity") or 0)
        item_name = str(row.get("item_name") or "").strip()
        if token_quantity <= 0 or not item_name:
            continue
        sequence_no = int(row.get("sequence_no") or 0)
        key = (sequence_no, item_name)
        stage_row = stage_map.setdefault(
            key,
            {
                "sequence_no": sequence_no,
                "item_name": item_name,
                "token_quantity": 0,
                "start_token_no": 0,
                "end_token_no": 0,
            },
        )
        stage_row["token_quantity"] += token_quantity
        start_token_no = int(row.get("start_token_no") or 0)
        end_token_no = int(row.get("end_token_no") or 0)
        if start_token_no > 0:
            if stage_row["start_token_no"] <= 0:
                stage_row["start_token_no"] = start_token_no
            else:
                stage_row["start_token_no"] = min(stage_row["start_token_no"], start_token_no)
        if end_token_no > 0:
            stage_row["end_token_no"] = max(stage_row["end_token_no"], end_token_no)

    rows_list = [
        dict(value)
        for _, value in sorted(
            stage_map.items(),
            key=lambda item: (
                item[0][0] <= 0,
                item[0][0] if item[0][0] > 0 else 10**9,
                item[0][1].casefold(),
            ),
        )
    ]
    return {
        "rows": rows_list,
        "row_count": len(rows_list),
        "total_token_quantity": sum(int(row.get("token_quantity") or 0) for row in rows_list),
    }


def _segregation_master_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Seq No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
            "Beneficiary Type": str(row.get("beneficiary_type") or ""),
            "Beneficiary": str(row.get("beneficiary_label") or ""),
            "Application No": str(row.get("application_number") or ""),
            "Item Type": str(row.get("item_type") or ""),
            "Item": str(row.get("item_name") or ""),
            "Waiting Hall Qty": int(row.get("waiting_hall_quantity") or 0),
            "Token Qty": int(row.get("token_quantity") or 0),
            "Start Token": int(row.get("start_token_no") or 0) if int(row.get("start_token_no") or 0) > 0 else "",
            "End Token": int(row.get("end_token_no") or 0) if int(row.get("end_token_no") or 0) > 0 else "",
            "District": str(row.get("district_name") or ""),
        }
        for row in list(rows or [])
    ]


def _segregation_file1_sheet_rows(groups: list[dict]) -> list[dict]:
    rows = []
    for group in list(groups or []):
        for item in list(group.get("items") or []):
            rows.append(
                {
                    "Beneficiary": str(group.get("beneficiary_label") or ""),
                    "Article": str(item.get("article_name") or ""),
                    "Quantity": int(item.get("quantity") or 0),
                    "Signature": "",
                }
            )
    return rows


def _segregation_file2_sheet_rows(groups: list[dict]) -> list[dict]:
    rows = []
    for group in list(groups or []):
        for beneficiary in list(group.get("beneficiaries") or []):
            rows.append(
                {
                    "Article": str(group.get("article_name") or ""),
                    "Beneficiary": str(beneficiary.get("beneficiary_label") or ""),
                    "Waiting Hall Quantity": int(beneficiary.get("quantity") or 0),
                }
            )
    return rows


def _segregation_file3_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Seq No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
            "Item": str(row.get("item_name") or ""),
            "Token Qty": int(row.get("token_quantity") or 0),
            "Start Token": int(row.get("start_token_no") or 0) if int(row.get("start_token_no") or 0) > 0 else "",
            "End Token": int(row.get("end_token_no") or 0) if int(row.get("end_token_no") or 0) > 0 else "",
        }
        for row in list(rows or [])
    ]


def _stage_distribution_pick_value(row: dict, aliases: list[str], default=""):
    item = dict(row or {})
    for alias in list(aliases or []):
        if alias in item and item.get(alias) not in {None, ""}:
            return item.get(alias)
    return default


def _stage_distribution_display_text(*values):
    parts = []
    for value in values:
        text = str(value or "").strip()
        if text:
            parts.append(text)
    return " - ".join(parts)


def _stage_distribution_normalize_row(row: dict) -> dict:
    item = dict(row or {})
    sequence_no = _phase2_parse_number(_stage_distribution_pick_value(item, ["Sequence No", "Seq No", "sequence_no", "Sequence"], 0))
    item_name = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Article", "Requested Item", "Token Name", "Item", "article_name", "item_name"]),
    )
    beneficiary_name = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Beneficiary Name", "Beneficiary", "Name", "beneficiary_name"]),
    )
    application_number = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Application Number", "App No", "application_number"]),
    )
    names = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Names", "Name", "names"]),
    )
    beneficiary_type = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Beneficiary Type", "beneficiary_type"]),
    )
    item_type = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Item Type", "item_type"]),
    )
    waiting_hall_quantity = _phase2_parse_number(
        _stage_distribution_pick_value(item, ["Waiting Hall Quantity", "waiting_hall_quantity"], 0)
    )
    token_quantity = _phase2_parse_number(
        _stage_distribution_pick_value(item, ["Token Quantity", "Token Qty", "token_quantity"], 0)
    )
    start_token_no = _phase2_parse_number(
        _stage_distribution_pick_value(item, ["Start Token No", "Start Token", "token_start"], 0)
    )
    end_token_no = _phase2_parse_number(
        _stage_distribution_pick_value(item, ["End Token No", "End Token", "token_end"], 0)
    )
    if token_quantity <= 0 and start_token_no > 0 and end_token_no >= start_token_no:
        token_quantity = end_token_no - start_token_no + 1
    premise = _stage_distribution_display_text(
        _stage_distribution_pick_value(item, ["Premise", "premise"]),
    )
    if not premise:
        if waiting_hall_quantity > 0:
            premise = "waiting_hall"
        elif token_quantity > 0:
            premise = "masm_hall"
        else:
            premise = "all"
    token_display = str(int(start_token_no or 0))
    if end_token_no and end_token_no > start_token_no:
        token_display = f"{int(start_token_no)} - {int(end_token_no)}"

    return {
        "sequence_no": int(sequence_no or 0),
        "item_name": item_name,
        "beneficiary_name": beneficiary_name,
        "application_number": application_number,
        "names": names,
        "beneficiary_type": beneficiary_type,
        "item_type": item_type,
        "waiting_hall_quantity": int(waiting_hall_quantity or 0),
        "token_quantity": int(token_quantity or 0),
        "start_token_no": int(start_token_no or 0),
        "end_token_no": int(end_token_no or 0),
        "premise": premise,
        "token_display": token_display,
    }


def _stage_distribution_normalize_dataset(dataset: dict) -> dict:
    rows = []
    for row in list(dataset.get("rows") or []):
        normalized = _stage_distribution_normalize_row(row)
        if not (
            normalized["item_name"]
            or normalized["beneficiary_name"]
            or normalized["token_quantity"]
            or normalized["waiting_hall_quantity"]
        ):
            continue
        rows.append(normalized)
    return {
        "rows": rows,
        "headers": list(dataset.get("headers") or []),
    }


def _stage_distribution_filter_rows(
    rows: list[dict],
    *,
    beneficiary_types: list[str] | tuple[str, ...] | set[str] | None = None,
    item_types: list[str] | tuple[str, ...] | set[str] | None = None,
    premise: str = "all",
    seq_start: int | None = None,
    seq_end: int | None = None,
) -> list[dict]:
    beneficiary_type_set = {str(value or "").strip() for value in list(beneficiary_types or []) if str(value or "").strip()}
    item_type_set = {str(value or "").strip() for value in list(item_types or []) if str(value or "").strip()}
    premise_value = str(premise or "all").strip().lower() or "all"
    filtered_rows = []
    for row in list(rows or []):
        row_beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        row_item_type = str(row.get("item_type") or "").strip()
        row_premise = str(row.get("premise") or "all").strip().lower() or "all"
        row_sequence = int(row.get("sequence_no") or 0)
        if beneficiary_type_set and row_beneficiary_type not in beneficiary_type_set:
            continue
        if item_type_set and row_item_type not in item_type_set:
            continue
        if premise_value != "all" and row_premise != premise_value:
            continue
        if seq_start is not None and row_sequence and row_sequence < seq_start:
            continue
        if seq_end is not None and row_sequence and row_sequence > seq_end:
            continue
        if seq_start is not None and seq_end is not None and row_sequence <= 0:
            continue
        filtered_rows.append(dict(row))
    filtered_rows.sort(
        key=lambda row: (
            int(row.get("sequence_no") or 0) <= 0,
            int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else 10**9,
            str(row.get("beneficiary_name") or "").casefold(),
            str(row.get("item_name") or "").casefold(),
        )
    )
    return filtered_rows


def _stage_distribution_build_file1(rows: list[dict]) -> dict:
    filtered_rows = [
        row
        for row in list(rows or [])
        if str(row.get("item_name") or "").strip()
        and (str(row.get("beneficiary_name") or "").strip() or str(row.get("application_number") or "").strip())
    ]
    filtered_rows.sort(
        key=lambda row: (
            int(row.get("sequence_no") or 0) <= 0,
            int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else 10**9,
            str(row.get("item_name") or "").casefold(),
            str(row.get("beneficiary_name") or "").casefold(),
        )
    )
    grouped_rows: list[dict] = []
    grouped_map: dict[tuple[str, str], dict] = {}
    for row in filtered_rows:
        names_value = str(row.get("names") or "").strip()
        if not names_value:
            beneficiary_name = str(row.get("beneficiary_name") or "").strip()
            application_number = str(row.get("application_number") or "").strip()
            if beneficiary_name and application_number and str(row.get("beneficiary_type") or "").strip().lower() != "district":
                names_value = f"{application_number} - {beneficiary_name}"
            else:
                names_value = beneficiary_name or application_number
        group_key = (str(row.get("item_name") or "").strip().casefold(), names_value.strip().casefold())
        group = grouped_map.get(group_key)
        if group is None:
            group = {
                "article_name": str(row.get("item_name") or ""),
                "beneficiary_name": names_value,
                "sequence_no": int(row.get("sequence_no") or 0),
                "token_start": 0,
                "token_end": 0,
                "token_quantity": 0,
            }
            grouped_map[group_key] = group
            grouped_rows.append(group)
        sequence_no = int(row.get("sequence_no") or 0)
        if sequence_no > 0 and (group["sequence_no"] <= 0 or sequence_no < group["sequence_no"]):
            group["sequence_no"] = sequence_no
        token_qty = int(row.get("token_quantity") or 0)
        if token_qty <= 0:
            token_qty = int(row.get("waiting_hall_quantity") or 0)
        group["token_quantity"] += token_qty
        start_token = int(row.get("start_token_no") or 0)
        end_token = int(row.get("end_token_no") or 0)
        if start_token <= 0 and end_token <= 0:
            continue
        if start_token <= 0:
            start_token = end_token
        if end_token <= 0:
            end_token = start_token
        if group["token_start"] <= 0 or start_token < group["token_start"]:
            group["token_start"] = start_token
        if end_token > group["token_end"]:
            group["token_end"] = end_token
    rendered_rows = []
    token_total = 0
    for group in grouped_rows:
        token_total += int(group.get("token_quantity") or 0)
        token_start = int(group.get("token_start") or 0)
        token_end = int(group.get("token_end") or 0)
        token_number = ""
        if token_start > 0 and token_end > 0:
            token_number = str(token_start) if token_start == token_end else f"{token_start} - {token_end}"
        elif token_start > 0:
            token_number = str(token_start)
        rendered_rows.append(
            {
                "article_name": str(group.get("article_name") or ""),
                "beneficiary_name": str(group.get("beneficiary_name") or ""),
                "token_number": token_number,
                "sequence_no": int(group.get("sequence_no") or 0),
            }
        )
    return {
        "rows": rendered_rows,
        "row_count": len(rendered_rows),
        "total_token_quantity": token_total,
    }


def _stage_distribution_file1_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Seq No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
            "Article": str(row.get("article_name") or ""),
            "Beneficiary": str(row.get("beneficiary_name") or ""),
            "Token Number": str(row.get("token_number") or ""),
        }
        for row in list(rows or [])
    ]


def _stage_distribution_names_value(row: dict) -> str:
    names_value = str(row.get("names") or "").strip()
    if names_value:
        return names_value
    beneficiary_name = str(row.get("beneficiary_name") or "").strip()
    application_number = str(row.get("application_number") or "").strip()
    if beneficiary_name and application_number and str(row.get("beneficiary_type") or "").strip().lower() != "district":
        return f"{application_number} - {beneficiary_name}"
    return beneficiary_name or application_number


def _stage_distribution_selected_quantity(row: dict, premise: str = "all") -> int:
    waiting_qty = int(row.get("waiting_hall_quantity") or 0)
    token_qty = int(row.get("token_quantity") or 0)
    premise_value = str(premise or "all").strip().lower() or "all"
    if premise_value == "waiting_hall":
        return waiting_qty
    if premise_value == "masm_hall":
        return token_qty
    return waiting_qty + token_qty


def _stage_distribution_build_beneficiary_article_file(
    rows: list[dict],
    *,
    beneficiary_types: set[str],
    premise: str = "all",
) -> dict:
    grouped: dict[str, dict[str, int]] = {}
    for row in list(rows or []):
        beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        if beneficiary_type not in beneficiary_types:
            continue
        article_name = str(row.get("item_name") or "").strip()
        beneficiary_label = _stage_distribution_names_value(row).strip()
        if not article_name or not beneficiary_label:
            continue
        quantity = _stage_distribution_selected_quantity(row, premise)
        if quantity <= 0:
            continue
        grouped.setdefault(beneficiary_label, {})
        grouped[beneficiary_label][article_name] = grouped[beneficiary_label].get(article_name, 0) + quantity

    groups = []
    grand_total = 0
    row_count = 0
    for beneficiary_label in sorted(grouped.keys(), key=lambda value: value.casefold()):
        article_map = grouped.get(beneficiary_label) or {}
        items = []
        total_quantity = 0
        for article_name in sorted(article_map.keys(), key=lambda value: value.casefold()):
            quantity = int(article_map.get(article_name) or 0)
            if quantity <= 0:
                continue
            items.append({"article_name": article_name, "quantity": quantity})
            total_quantity += quantity
        if not items:
            continue
        groups.append(
            {
                "group_label": beneficiary_label,
                "items": items,
                "total_quantity": total_quantity,
            }
        )
        row_count += len(items)
        grand_total += total_quantity
    return {
        "groups": groups,
        "row_count": row_count,
        "grand_total": grand_total,
    }


def _stage_distribution_build_article_beneficiary_file(rows: list[dict], *, premise: str = "all") -> dict:
    grouped: dict[str, dict[tuple[str, str], int]] = {}
    beneficiary_type_order = {
        models.RecipientTypeChoices.DISTRICT: 0,
        models.RecipientTypeChoices.PUBLIC: 1,
        models.RecipientTypeChoices.INSTITUTIONS: 2,
        models.RecipientTypeChoices.OTHERS: 3,
    }
    for row in list(rows or []):
        article_name = str(row.get("item_name") or "").strip()
        beneficiary_label = _stage_distribution_names_value(row).strip()
        beneficiary_type = str(row.get("beneficiary_type") or "").strip()
        if not article_name or not beneficiary_label:
            continue
        quantity = _stage_distribution_selected_quantity(row, premise)
        if quantity <= 0:
            continue
        grouped.setdefault(article_name, {})
        beneficiary_key = (beneficiary_type, beneficiary_label)
        grouped[article_name][beneficiary_key] = grouped[article_name].get(beneficiary_key, 0) + quantity

    groups = []
    grand_total = 0
    row_count = 0
    for article_name in sorted(grouped.keys(), key=lambda value: value.casefold()):
        beneficiary_map = grouped.get(article_name) or {}
        items = []
        total_quantity = 0
        sorted_beneficiary_keys = sorted(
            beneficiary_map.keys(),
            key=lambda value: (
                beneficiary_type_order.get(str(value[0] or "").strip(), 99),
                str(value[1] or "").casefold(),
            ),
        )
        for beneficiary_type, beneficiary_label in sorted_beneficiary_keys:
            quantity = int(beneficiary_map.get((beneficiary_type, beneficiary_label)) or 0)
            if quantity <= 0:
                continue
            items.append({"beneficiary_name": beneficiary_label, "quantity": quantity})
            total_quantity += quantity
        if not items:
            continue
        groups.append(
            {
                "group_label": article_name,
                "items": items,
                "total_quantity": total_quantity,
            }
        )
        row_count += len(items)
        grand_total += total_quantity
    return {
        "groups": groups,
        "row_count": row_count,
        "grand_total": grand_total,
    }


def _stage_distribution_header_table(styles, *, custom_logo=None, title_text: str = "Beneficiaries List"):
    header_title = ParagraphStyle(
        "stage_distribution_title_grouped",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        textColor=colors.black,
    )
    header_sub = ParagraphStyle(
        "stage_distribution_sub_grouped",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica",
        fontSize=9.3,
        leading=9.2,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.black,
    )
    left_logo = _fitted_pdf_image_source(custom_logo or _pdf_guru_logo_path(), max_width_mm=18, max_height_mm=18)
    right_logo = _fitted_pdf_image(_pdf_logo_path(), max_width_mm=18, max_height_mm=18)
    district_prog_line_one, district_prog_line_two = _district_signature_programme_lines()
    return Table(
        [[
            left_logo,
            [
                Paragraph(
                    "OM SAKTHI",
                    ParagraphStyle(
                        "stage-distribution-om-grouped",
                        parent=header_title,
                        fontSize=8.8,
                        leading=9.0,
                        spaceAfter=0,
                        textColor=colors.red,
                    ),
                ),
                Paragraph(district_prog_line_one, header_sub),
                Paragraph(district_prog_line_two, header_sub),
                Paragraph(str(title_text or "Beneficiaries List"), header_title),
            ],
            right_logo,
        ]],
        colWidths=[22 * mm, 154 * mm, 22 * mm],
        style=TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        ),
    )


def generate_stage_distribution_grouped_pdf(
    groups: list[dict],
    *,
    section_title: str,
    item_value_key: str,
    name_column_label: str = "Name",
    header_title_text: str | None = None,
    custom_logo=None,
) -> io.BytesIO:
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "stage_distribution_grouped_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.2,
        leading=10.2,
        alignment=1,
        textColor=colors.black,
    )
    body_style = ParagraphStyle(
        "stage_distribution_grouped_body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.6,
        leading=9.6,
        textColor=colors.black,
    )
    body_bold_style = ParagraphStyle(
        "stage_distribution_grouped_body_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )
    body_center_bold_style = ParagraphStyle(
        "stage_distribution_grouped_center_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
        alignment=1,
    )
    subtotal_label_style = ParagraphStyle(
        "stage_distribution_grouped_subtotal_label",
        parent=body_style,
        fontName="Helvetica-Bold",
        alignment=2,
        textColor=colors.HexColor("#1d4ed8"),
    )
    subtotal_value_style = ParagraphStyle(
        "stage_distribution_grouped_subtotal_value",
        parent=body_style,
        fontName="Helvetica-Bold",
        alignment=1,
        textColor=colors.HexColor("#1d4ed8"),
    )
    section_style = ParagraphStyle(
        "stage_distribution_grouped_section",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.1,
        leading=10.2,
        textColor=colors.HexColor("#334155"),
    )
    empty_style = ParagraphStyle(
        "stage_distribution_grouped_empty",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.black,
    )
    stage_border = colors.HexColor("#5b6572")
    stage_light_border = colors.HexColor("#7c8794")

    story = [
        _stage_distribution_header_table(
            styles,
            custom_logo=custom_logo,
            title_text=header_title_text or "Beneficiaries List",
        ),
        Spacer(1, 2 * mm),
    ]
    grand_total = 0
    if not groups:
        story.append(Paragraph("No rows available for this report.", empty_style))
    else:
        table_rows = [[
            Paragraph(str(name_column_label or "Name"), header_style),
            Paragraph("Qty", header_style),
        ]]
        group_header_row_indexes: list[int] = []
        group_subtotal_row_indexes: list[int] = []
        group_end_row_indexes: list[int] = []
        detail_row_indexes: list[int] = []

        for group in list(groups or []):
            group_label = str(group.get("group_label") or "").strip()
            items = list(group.get("items") or [])
            group_total = int(group.get("total_quantity") or 0)
            grand_total += group_total

            if group_label:
                group_header_row_indexes.append(len(table_rows))
                table_rows.append(
                    [
                        Paragraph(escape(group_label), body_bold_style),
                        Paragraph("", body_style),
                    ]
                )

            for item in items:
                item_value = str(item.get(item_value_key) or "")
                if not item_value:
                    item_value = str(item.get("article_name") or item.get("beneficiary_name") or "")
                detail_row_indexes.append(len(table_rows))
                table_rows.append(
                    [
                        Paragraph(f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{escape(item_value)}", body_style),
                        Paragraph(str(int(item.get("quantity") or 0)), body_center_bold_style),
                    ]
                )
            if items:
                group_subtotal_row_indexes.append(len(table_rows))
                table_rows.append(
                    [
                        Paragraph("Total", subtotal_label_style),
                        Paragraph(str(group_total), subtotal_value_style),
                    ]
                )
            if items:
                group_end_row_indexes.append(len(table_rows) - 1)
            elif group_label:
                group_end_row_indexes.append(len(table_rows) - 1)

        table_rows.append(
            [
                Paragraph("Grand Total", body_bold_style),
                Paragraph(str(grand_total), body_center_bold_style),
            ]
        )
        grand_total_row_index = len(table_rows) - 1

        table = _segregation_pdf_table(table_rows, col_widths=[160 * mm, 36 * mm])
        table_style_commands = [
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
            ("BOX", (0, 0), (-1, -1), 0.65, stage_border),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, stage_border),
            ("GRID", (0, 1), (-1, -1), 0, colors.white),
            ("LINEAFTER", (0, 0), (0, -1), 0.45, stage_light_border),
            ("BACKGROUND", (0, grand_total_row_index), (-1, grand_total_row_index), colors.HexColor("#e2e8f0")),
            ("LINEABOVE", (0, grand_total_row_index), (-1, grand_total_row_index), 0.6, stage_border),
            ("TOPPADDING", (0, 0), (-1, -1), 3.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.2),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_index in group_header_row_indexes:
            table_style_commands.extend(
                [
                    ("SPAN", (0, row_index), (1, row_index)),
                    ("BACKGROUND", (0, row_index), (1, row_index), colors.HexColor("#eff6ff")),
                    ("LINEABOVE", (0, row_index), (1, row_index), 0.45, stage_light_border),
                ]
            )
        for row_index in group_subtotal_row_indexes:
            table_style_commands.extend(
                [
                    ("BACKGROUND", (0, row_index), (1, row_index), colors.HexColor("#f8fafc")),
                    ("LINEABOVE", (0, row_index), (1, row_index), 0.45, stage_light_border),
                ]
            )
        for row_index in group_end_row_indexes:
            if row_index != grand_total_row_index:
                table_style_commands.append(("LINEBELOW", (0, row_index), (1, row_index), 0.45, stage_light_border))
        for row_index in detail_row_indexes:
            table_style_commands.extend(
                [
                    ("LINEABOVE", (0, row_index), (1, row_index), 0, colors.white),
                    ("LINEBELOW", (0, row_index), (1, row_index), 0, colors.white),
                ]
            )
        table.setStyle(TableStyle(table_style_commands))
        story.append(table)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    doc.build(story, canvasmaker=_NumberedPdfCanvas)
    buffer.seek(0)
    return buffer


class _NumberedPdfCanvas(canvas.Canvas):
    def __init__(self, *args, footer_text: str | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
        self._footer_text = footer_text or ""

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.black)
            if self._footer_text:
                self.drawString(9 * mm, 6 * mm, self._footer_text)
            self.drawCentredString(A4[0] / 2, 6 * mm, f"Page {self._pageNumber}/{total_pages}")
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)


def generate_stage_distribution_file1_pdf(
    rows: list[dict],
    *,
    seq_start: int | None = None,
    seq_end: int | None = None,
    custom_logo=None,
) -> io.BytesIO:
    styles = getSampleStyleSheet()
    stage_border = colors.HexColor("#5b6572")
    stage_light_border = colors.HexColor("#7c8794")
    left_logo = _fitted_pdf_image_source(custom_logo or _pdf_guru_logo_path(), max_width_mm=18, max_height_mm=18)
    right_logo = _fitted_pdf_image(_pdf_logo_path(), max_width_mm=18, max_height_mm=18)
    header_title = ParagraphStyle(
        "stage_distribution_title",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        textColor=colors.black,
    )
    header_sub = ParagraphStyle(
        "stage_distribution_sub",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica",
        fontSize=9.3,
        leading=9.2,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.black,
    )
    header_style = ParagraphStyle(
        "stage_distribution_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.4,
        leading=10.4,
        alignment=1,
        textColor=colors.black,
    )
    body_style = ParagraphStyle(
        "stage_distribution_body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=9.8,
        textColor=colors.black,
    )
    body_bold_style = ParagraphStyle(
        "stage_distribution_body_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )
    body_center_bold_style = ParagraphStyle(
        "stage_distribution_body_center_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
        alignment=1,
    )
    district_prog_line_one, district_prog_line_two = _district_signature_programme_lines()
    story = [
        Table(
            [[
                left_logo,
                [
                    Paragraph(
                        "OM SAKTHI",
                        ParagraphStyle(
                            "stage-distribution-om",
                            parent=header_title,
                            fontSize=8.8,
                            leading=9.0,
                            spaceAfter=0,
                            textColor=colors.red,
                        ),
                    ),
                    Paragraph(district_prog_line_one, header_sub),
                    Paragraph(district_prog_line_two, header_sub),
                    Paragraph("Beneficiaries List", header_title),
                ],
                right_logo,
            ]],
            colWidths=[22 * mm, 154 * mm, 22 * mm],
            style=TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (0, 0), "LEFT"),
                    ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            ),
        ),
    ]

    if not rows:
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph("No rows available for this report.", ParagraphStyle("stage-distribution-empty", parent=styles["BodyText"], alignment=1, fontName="Helvetica", fontSize=10, textColor=colors.black)))
    else:
        story.append(Spacer(1, 3 * mm))
        table_rows = [[
            Paragraph("Article Name", header_style),
            Paragraph("Beneficiary Names", header_style),
            Paragraph("Token Number", header_style),
        ]]
        for row in list(rows or []):
            table_rows.append(
                [
                    Paragraph(escape(str(row.get("article_name") or "")), body_style),
                    Paragraph(escape(str(row.get("beneficiary_name") or "")), body_style),
                    Paragraph(escape(str(row.get("token_number") or "")), body_center_bold_style),
                ]
            )
        table = _segregation_pdf_table(table_rows, col_widths=[80 * mm, 80 * mm, 36 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (2, 0), (2, -1), "CENTER"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                    ("BOX", (0, 0), (-1, -1), 0.65, stage_border),
                    ("INNERGRID", (0, 0), (-1, -1), 0.45, stage_light_border),
                    ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
                ]
            )
        )
        story.append(table)

    sequence_numbers = [int(row.get("sequence_no") or 0) for row in list(rows or []) if int(row.get("sequence_no") or 0) > 0]
    footer_start = int(seq_start or 0) if int(seq_start or 0) > 0 else (min(sequence_numbers) if sequence_numbers else 0)
    footer_end = int(seq_end or 0) if int(seq_end or 0) > 0 else (max(sequence_numbers) if sequence_numbers else 0)
    seq_range_text = ""
    if footer_start > 0 and footer_end > 0:
        seq_range_text = f"Seq No : {footer_start} - {footer_end}"
    elif footer_start > 0:
        seq_range_text = f"Seq No : {footer_start}"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    doc.build(story, canvasmaker=lambda *args, **kwargs: _NumberedPdfCanvas(*args, footer_text=seq_range_text, **kwargs))
    buffer.seek(0)
    return buffer


def _reports_token_lookup_default_state():
    return {
        "loaded": False,
        "synced_at": "",
        "source": "",
        "session_id": None,
        "rows": [],
        "headers": [],
        "filters": {
            "token_number": "",
            "application_number": "",
            "beneficiary_name": "",
            "item_name": "",
            "item_type": "",
        },
    }


def _reports_token_lookup_session_state(request):
    state = request.session.get(REPORTS_TOKEN_LOOKUP_STATE_KEY) or {}
    merged = _reports_token_lookup_default_state()
    merged.update(state)
    merged["rows"] = [dict(row) for row in list(merged.get("rows") or [])]
    merged["headers"] = [str(header or "") for header in list(merged.get("headers") or []) if str(header or "").strip()]
    filters = dict(merged.get("filters") or {})
    merged["filters"] = {
        "token_number": str(filters.get("token_number") or "").strip(),
        "application_number": str(filters.get("application_number") or "").strip(),
        "beneficiary_name": str(filters.get("beneficiary_name") or "").strip(),
        "item_name": str(filters.get("item_name") or "").strip(),
        "item_type": str(filters.get("item_type") or "").strip(),
    }
    return merged


def _reports_set_token_lookup_state(request, state):
    request.session[REPORTS_TOKEN_LOOKUP_STATE_KEY] = state
    request.session.modified = True


def _reports_token_lookup_display_value(*values):
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text or value == 0:
            return text
    return ""


def _reports_token_lookup_normalize_row(row):
    item = dict(row or {})
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    beneficiary_name = _reports_token_lookup_display_value(
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    return {
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
        "application_number": _reports_token_lookup_display_value(
            item.get("Application Number"),
            item.get("application_number"),
            item.get("App No"),
        ),
        "beneficiary_name": beneficiary_name,
        "beneficiary_type": beneficiary_type,
        "district": _reports_token_lookup_display_value(
            item.get("District"),
            item.get("district"),
            beneficiary_name if beneficiary_type == models.RecipientTypeChoices.DISTRICT else "",
        ),
        "item_name": _reports_token_lookup_display_value(
            item.get("Requested Item"),
            item.get("Token Name"),
            item.get("Article"),
        ),
        "item_type": _reports_token_lookup_display_value(
            item.get("Item Type"),
            item.get("item_type"),
        ),
        "total_value": _reports_token_lookup_display_value(
            item.get("Total Value"),
            item.get("total_value"),
            item.get("Total Amount"),
            item.get("Amount"),
        ),
        "token_quantity": int(token_quantity),
        "sequence_no": int(_phase2_parse_number(item.get("Sequence No")) or 0),
    }


def _reports_token_lookup_rows_from_session(session):
    rows = list(
        models.TokenGenerationRow.objects.filter(session=session).order_by(
            "sort_order",
            F("sequence_no").asc(nulls_last=True),
            "requested_item",
            "application_number",
            "id",
        )
    )
    prepared_rows = []
    for row in rows:
        row_data = dict(row.row_data or {})
        row_data["Application Number"] = row.application_number or row_data.get("Application Number") or ""
        row_data["Beneficiary Name"] = row.beneficiary_name or row_data.get("Beneficiary Name") or ""
        row_data["Requested Item"] = row.requested_item or row_data.get("Requested Item") or ""
        row_data["Beneficiary Type"] = row.beneficiary_type or row_data.get("Beneficiary Type") or ""
        row_data["Sequence No"] = row.sequence_no if row.sequence_no is not None else row_data.get("Sequence No") or ""
        row_data["Start Token No"] = row.start_token_no if row.start_token_no is not None else row_data.get("Start Token No") or 0
        row_data["End Token No"] = row.end_token_no if row.end_token_no is not None else row_data.get("End Token No") or 0
        existing_token_quantity = row_data.get("Token Quantity")
        if existing_token_quantity in {None, ""}:
            existing_token_quantity = max(
                (row.end_token_no or 0) - (row.start_token_no or 0) + 1,
                0,
            )
        row_data["Token Quantity"] = existing_token_quantity
        row_data["Total Value"] = (
            row_data.get("Total Value")
            or row_data.get("total_value")
            or row_data.get("Total Amount")
            or row_data.get("Amount")
            or ""
        )
        normalized = _reports_token_lookup_normalize_row(row_data)
        if int(normalized.get("token_quantity") or 0) > 0:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_token_lookup_data_rows(rows):
    return [
        normalized
        for normalized in (_reports_token_lookup_normalize_row(row) for row in list(rows or []))
        if int(normalized.get("token_quantity") or 0) > 0
    ]


def _reports_token_lookup_filters_from_post(post_data):
    return {
        "token_number": str(post_data.get("token_number") or "").strip(),
        "application_number": str(post_data.get("application_number") or "").strip(),
        "beneficiary_name": str(post_data.get("beneficiary_name") or "").strip(),
        "item_name": str(post_data.get("item_name") or "").strip(),
        "item_type": str(post_data.get("item_type") or "").strip(),
    }


def _reports_token_lookup_filter_rows(rows, filters):
    filtered = list(rows or [])
    token_number = _phase2_parse_number((filters or {}).get("token_number"))
    application_number = str((filters or {}).get("application_number") or "").strip().casefold()
    beneficiary_name = str((filters or {}).get("beneficiary_name") or "").strip().casefold()
    item_name = str((filters or {}).get("item_name") or "").strip().casefold()
    item_type = str((filters or {}).get("item_type") or "").strip().casefold()

    if token_number:
        filtered = [
            row for row in filtered
            if int(row.get("token_start") or 0) <= int(token_number) <= int(row.get("token_end") or row.get("token_start") or 0)
        ]
    if application_number:
        filtered = [row for row in filtered if application_number in str(row.get("application_number") or "").casefold()]
    if beneficiary_name:
        filtered = [row for row in filtered if beneficiary_name in str(row.get("beneficiary_name") or "").casefold()]
    if item_name:
        filtered = [row for row in filtered if item_name in str(row.get("item_name") or "").casefold()]
    if item_type:
        filtered = [row for row in filtered if item_type in str(row.get("item_type") or "").casefold()]
    return filtered


def _reports_token_lookup_choice_values(rows, key):
    return sorted(
        {
            str(row.get(key) or "").strip()
            for row in list(rows or [])
            if str(row.get(key) or "").strip()
        },
        key=lambda value: value.casefold(),
    )


def _reports_public_signature_normalize_row(row):
    item = dict(row or {})
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    if beneficiary_type != models.RecipientTypeChoices.PUBLIC:
        return None
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    if token_quantity <= 0:
        return None
    application_number = _reports_token_lookup_display_value(
        item.get("Application Number"),
        item.get("application_number"),
        item.get("App No"),
    )
    beneficiary_name = _reports_token_lookup_display_value(
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    item_name = _reports_token_lookup_display_value(
        item.get("Requested Item"),
        item.get("requested_item"),
        item.get("Token Name"),
        item.get("Article"),
    )
    return {
        "application_number": application_number,
        "beneficiary_name": beneficiary_name,
        "item_name": item_name,
        "item_type": _reports_token_lookup_display_value(
            item.get("Item Type"),
            item.get("item_type"),
        ),
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
        "token_quantity": int(token_quantity),
    }


def _reports_public_signature_rows_from_dataset(rows):
    prepared_rows = []
    for row in list(rows or []):
        normalized = _reports_public_signature_normalize_row(row)
        if normalized:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_public_signature_rows_from_session(session):
    dataset = _token_generation_saved_dataset(session) if session else {"rows": []}
    return _reports_public_signature_rows_from_dataset(dataset.get("rows") or [])


def _reports_public_signature_item_options(rows):
    counts = {}
    for row in list(rows or []):
        item_name = str(row.get("item_name") or "").strip()
        if not item_name:
            continue
        counts[item_name] = counts.get(item_name, 0) + 1
    return [
        {
            "item_name": item_name,
            "row_count": counts[item_name],
        }
        for item_name in sorted(counts.keys(), key=lambda value: value.casefold())
    ]


def _reports_public_signature_sort_rows(rows, sort_modes):
    selected_modes = [
        str(mode or "").strip().casefold()
        for mode in list(sort_modes or [])
        if str(mode or "").strip()
    ]
    normalized_modes = []
    for mode in selected_modes:
        if mode in {"application_number", "item_name", "token_number"} and mode not in normalized_modes:
            normalized_modes.append(mode)
    selected_modes = normalized_modes
    if not selected_modes:
        return list(rows or [])

    def _row_key(row):
        application_number = _public_signature_app_sort_key(str(row.get("application_number") or ""))
        item_name = str(row.get("item_name") or "").strip().casefold()
        token_number = int(row.get("token_start") or 0)
        values = {
            "application_number": application_number,
            "item_name": item_name,
            "token_number": token_number,
        }
        return tuple(values[mode] for mode in selected_modes)

    return sorted(list(rows or []), key=_row_key)


def _reports_district_signature_normalize_row(row):
    item = dict(row or {})
    beneficiary_type = _reports_token_lookup_display_value(
        item.get("Beneficiary Type"),
        item.get("beneficiary_type"),
    )
    if beneficiary_type != models.RecipientTypeChoices.DISTRICT:
        return None
    district_name = _reports_token_lookup_display_value(
        item.get("District"),
        item.get("district"),
        item.get("Beneficiary Name"),
        item.get("beneficiary_name"),
        item.get("Names"),
        item.get("Name"),
    )
    item_name = _reports_token_lookup_display_value(
        item.get("Requested Item"),
        item.get("requested_item"),
        item.get("Token Name"),
        item.get("Article"),
    )
    total_quantity = max(_phase2_parse_number(item.get("Quantity")) or 0, 0)
    token_start = _phase2_parse_number(item.get("Start Token No"))
    token_end = _phase2_parse_number(item.get("End Token No"))
    if token_start is None:
        token_start = _phase2_parse_number(item.get("Token No"))
    if token_end is None:
        token_end = token_start
    token_quantity = max(_phase2_parse_number(item.get("Token Quantity")) or 0, 0)
    if token_quantity <= 0 and token_start and token_end and token_end >= token_start:
        token_quantity = int(token_end - token_start + 1)
    if not district_name or not item_name or (total_quantity <= 0 and token_quantity <= 0):
        return None
    return {
        "district_name": district_name,
        "item_name": item_name,
        "total_quantity": int(total_quantity),
        "token_quantity": int(token_quantity),
        "token_start": int(token_start or 0),
        "token_end": int(token_end or token_start or 0),
    }


def _reports_district_signature_rows_from_dataset(rows):
    prepared_rows = []
    for row in list(rows or []):
        normalized = _reports_district_signature_normalize_row(row)
        if normalized:
            prepared_rows.append(normalized)
    return prepared_rows


def _reports_district_signature_rows_from_session(session):
    dataset = _token_generation_saved_dataset(session) if session else {"rows": []}
    return _reports_district_signature_rows_from_dataset(dataset.get("rows") or [])


def _reports_district_signature_grouped(rows):
    grouped = {}
    for row in list(rows or []):
        district_name = str(row.get("district_name") or "").strip()
        item_name = str(row.get("item_name") or "").strip()
        if not district_name or not item_name:
            continue
        district_bucket = grouped.setdefault(district_name, {})
        item_bucket = district_bucket.setdefault(
            item_name,
            {
                "item_name": item_name,
                "total_quantity": 0,
                "token_quantity": 0,
                "start_token": None,
                "end_token": None,
            },
        )
        item_bucket["total_quantity"] += int(row.get("total_quantity") or 0)
        current_token_qty = int(row.get("token_quantity") or 0)
        item_bucket["token_quantity"] += current_token_qty
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or 0)
        if current_token_qty > 0 and token_start > 0:
            item_bucket["start_token"] = token_start if item_bucket["start_token"] is None else min(item_bucket["start_token"], token_start)
            item_bucket["end_token"] = token_end if item_bucket["end_token"] is None else max(item_bucket["end_token"], token_end)

    districts = []
    total_quantity = 0
    total_token_quantity = 0
    total_item_count = 0
    for district_name in sorted(grouped.keys(), key=lambda value: value.casefold()):
        items = [grouped[district_name][name] for name in sorted(grouped[district_name].keys(), key=lambda value: value.casefold())]
        district_total_qty = sum(int(item.get("total_quantity") or 0) for item in items)
        district_token_qty = sum(int(item.get("token_quantity") or 0) for item in items)
        positive_starts = [int(item["start_token"]) for item in items if item.get("start_token")]
        positive_ends = [int(item["end_token"]) for item in items if item.get("end_token")]
        districts.append(
            {
                "district_name": district_name,
                "items": items,
                "total_quantity": district_total_qty,
                "token_quantity": district_token_qty,
                "start_token": min(positive_starts) if positive_starts else None,
                "end_token": max(positive_ends) if positive_ends else None,
            }
        )
        total_item_count += len(items)
        total_quantity += district_total_qty
        total_token_quantity += district_token_qty
    return {
        "districts": districts,
        "district_count": len(districts),
        "item_count": total_item_count,
        "total_quantity": total_quantity,
        "total_token_quantity": total_token_quantity,
    }


def _reports_public_ack_data_rows(rows):
    prepared_rows = []
    has_beneficiary_type = any(
        str(row.get("Beneficiary Type") or row.get("beneficiary_type") or "").strip()
        for row in rows
    )
    for row in rows:
        item = dict(row or {})
        beneficiary_type = str(item.get("Beneficiary Type") or item.get("beneficiary_type") or "").strip()
        quantity = _phase2_parse_number(item.get("Quantity"))
        if quantity is None:
            quantity = _phase2_parse_number(item.get("Token Quantity"))
        if quantity is None:
            quantity = 0
        if has_beneficiary_type and beneficiary_type and _phase2_normalize_text(beneficiary_type) != "public":
            continue
        if quantity <= 0 and has_beneficiary_type:
            continue
        if not has_beneficiary_type and quantity <= 0:
            continue
        item["Quantity"] = quantity
        prepared_rows.append(item)
    return prepared_rows


def _reports_public_ack_template_fields(template_bytes):
    fields = _public_acknowledgment_pdf_fields(template_bytes)
    return [
        {
            "field_name": field["field_name"],
            "field_key": field["field_key"],
        }
        for field in fields
    ]


def _reports_public_ack_column_options(headers):
    fallback = [
        "District",
        "Address",
        "Application Number",
        "Start Token No",
        "Beneficiary Name",
        "Mobile",
        "Aadhar Number",
        "Requested Item",
        "Token Name",
        "Total Value",
        "Cost Per Unit",
        "Cheque / RTGS in Favour",
        "Name of Institution",
        "Quantity",
    ]
    options = []
    seen = set()
    for header in list(headers or []) + fallback:
        text = str(header or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        options.append(text)
    return options


def _reports_public_ack_default_field_map(headers, template_fields):
    normalized_headers = {_phase2_normalize_text(header): header for header in (headers or []) if str(header or "").strip()}
    candidates = {
        "district": ["District"],
        "address": ["Address"],
        "app_no": ["Application Number"],
        "token": ["Start Token No", "Token No", "Sequence No"],
        "bf_name": ["Beneficiary Name", "Name of Beneficiary", "Name"],
        "mobile": ["Mobile"],
        "aadhar": ["Aadhar Number", "Aadhaar Number"],
        "article": ["Requested Item", "Token Name", "Article"],
        "value_aid": ["Total Value", "Cost Per Unit"],
        "cheque_no": ["Cheque / RTGS in Favour", "Cheque No"],
    }
    mapping = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        selected = ""
        for candidate in candidates.get(field_key, []):
            match = normalized_headers.get(_phase2_normalize_text(candidate))
            if match:
                selected = match
                break
        mapping[field_key] = selected
    return mapping


def _reports_public_ack_field_map_from_post(post_data, template_fields):
    mapping = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        mapping[field_key] = str(post_data.get(f"public_ack_map__{field_key}") or "").strip()
    return mapping


def _reports_public_ack_field_map_with_defaults(headers, template_fields, existing_map=None):
    defaults = _reports_public_ack_default_field_map(headers, template_fields)
    existing_map = {
        str(key or "").strip(): str(value or "").strip()
        for key, value in dict(existing_map or {}).items()
        if str(key or "").strip()
    }
    merged = {}
    for field in template_fields or []:
        field_key = str(field.get("field_key") or "").strip()
        if field_key in existing_map:
            merged[field_key] = existing_map[field_key]
        else:
            merged[field_key] = defaults.get(field_key, "")
    return merged


def _reports_public_ack_normalize_dataset(rows):
    return _reports_public_ack_data_rows(rows)


def _reports_waiting_hall_grouped_data(rows, ignored_keys=None, beneficiary_type_filter="", item_type_filter=""):
    ignored = set(ignored_keys or [])
    selected_type = str(beneficiary_type_filter or "").strip().lower()
    selected_item_type = str(item_type_filter or "").strip().lower()
    if selected_type == "all":
        selected_type = ""
    if selected_item_type == "all":
        selected_item_type = ""
    grouped = {}
    for row in rows:
        beneficiary_type = str(row.get("Beneficiary Type") or row.get("beneficiary_type") or "").strip()
        if selected_type and beneficiary_type.lower() != selected_type:
            continue
        item_type = str(row.get("Item Type") or row.get("item_type") or "").strip()
        if selected_item_type and item_type.lower() != selected_item_type:
            continue
        waiting_quantity = _phase2_parse_number(row.get("Waiting Hall Quantity"))
        if not waiting_quantity or waiting_quantity <= 0:
            continue
        raw_name = str(
            row.get("District")
            or row.get("Names")
            or row.get("Beneficiary Name")
            or row.get("Name of Beneficiary")
            or row.get("Name of Institution")
            or row.get("district")
            or row.get("beneficiary_name")
            or row.get("application_number")
            or row.get("Application Number")
            or ""
        ).strip()
        item_name = str(row.get("Requested Item") or row.get("requested_item") or "").strip()
        if not raw_name or not item_name:
            continue
        beneficiary_type_label = beneficiary_type or "Unknown"
        key = f"{beneficiary_type_label}||{raw_name}||{item_name}"
        group_key = f"{beneficiary_type_label}||{raw_name}"
        group_bucket = grouped.setdefault(
            group_key,
            {
                "entity_name": raw_name,
                "entity_kind": beneficiary_type_label,
                "label_prefix": "District" if beneficiary_type_label == models.RecipientTypeChoices.DISTRICT else "Beneficiary",
                "items_map": {},
            },
        )
        entry = group_bucket["items_map"].setdefault(
            item_name,
            {
                "key": key,
                "entity_name": raw_name,
                "entity_kind": beneficiary_type_label,
                "requested_item": item_name,
                "item_type": item_type,
                "quantity": 0,
            },
        )
        entry["quantity"] += int(waiting_quantity or 0)
    districts = []
    total_items = 0
    total_quantity = 0
    available_keys = []
    for group_key in sorted(grouped.keys(), key=lambda value: value.lower()):
        group_meta = grouped[group_key]
        raw_items = [group_meta["items_map"][name] for name in sorted(group_meta["items_map"].keys(), key=lambda value: value.lower())]
        available_keys.extend(item["key"] for item in raw_items)
        filtered_items = [item for item in raw_items if item["key"] not in ignored]
        district_total = sum(int(item["quantity"] or 0) for item in filtered_items)
        total_items += len(filtered_items)
        total_quantity += district_total
        districts.append(
            {
                "entity_name": group_meta["entity_name"],
                "entity_kind": group_meta["entity_kind"],
                "label_prefix": group_meta["label_prefix"],
                "items": filtered_items,
                "raw_items": raw_items,
                "total_quantity": district_total,
                "item_count": len(filtered_items),
            }
        )
    return {
        "districts": districts,
        "district_count": len(districts),
        "item_count": total_items,
        "total_quantity": total_quantity,
        "available_keys": available_keys,
    }

def _waiting_hall_programme_lines(event_age_label: str, event_date) -> tuple[str, str]:
    age_label = str(event_age_label or "").strip() or "Event"
    if hasattr(event_date, "strftime"):
        line_one_date = event_date.strftime("%d-%m-%Y")
        line_two_date = event_date.strftime("%d.%m.%Y")
    else:
        line_one_date = str(event_date or "").strip()
        line_two_date = line_one_date
    line_one = f"MASM Social Welfare Programme on the eve of {age_label} Birthday ({line_one_date})"
    line_two = f"His Holiness AMMA at Melmaruvathur on {line_two_date}"
    return line_one, line_two


def _district_signature_programme_lines() -> tuple[str, str]:
    current_year = timezone.localdate().year
    birthday_number = max(current_year - 1940, 1)
    line_one = (
        "MASM Social Welfare Programme on the eve of "
        f"{_ordinal(birthday_number)} Birthday (03-03-{current_year})"
    )
    line_two = f"His Holiness AMMA at Melmaruvathur on 03.03.{current_year}"
    return line_one, line_two


def _waiting_hall_pack_pages(groups: list[dict]) -> list[list[dict]]:
    pages: list[list[dict]] = []
    current_page: list[dict] = []
    current_units = 0
    page_capacity = 18
    for group in groups:
        row_count = len(group.get("items") or [])
        block_units = 6 + row_count
        if not current_page:
            current_page = [group]
            current_units = block_units
            continue
        if len(current_page) < 2 and (current_units + block_units) <= page_capacity:
            current_page.append(group)
            current_units += block_units
        else:
            pages.append(current_page)
            current_page = [group]
            current_units = block_units
    if current_page:
        pages.append(current_page)
    return pages


def _waiting_hall_logo_data_uri(source, mime_type: str | None = None) -> str:
    if not source:
        return ""
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.exists():
            return ""
        raw = path.read_bytes()
        mime = mime_type or "image/png"
    else:
        raw = bytes(source)
        mime = mime_type or "image/png"
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _waiting_hall_layout_profile(item_count: int, *, paired: bool = False) -> dict:
    if item_count >= 14:
        return {
            "logo_width_mm": 24,
            "logo_height_mm": 28,
            "logo_col_width_mm": 26,
            "header_col_width_mm": 132,
            "title_font": 10,
            "title_leading": 11,
            "sub_font": 8,
            "sub_leading": 10,
            "section_font": 10,
            "section_leading": 11,
            "district_font": 13,
            "district_leading": 15,
            "header_font": 14,
            "header_leading": 16,
            "body_font": 11,
            "body_leading": 13,
            "note_font": 12,
            "note_leading": 14,
            "table_top_bottom_padding": 4,
            "space_after_header_mm": 3,
            "space_after_table_mm": 4,
            "space_before_signature_mm": 8,
        }
    if item_count >= 11:
        return {
            "logo_width_mm": 26,
            "logo_height_mm": 30,
            "logo_col_width_mm": 29,
            "header_col_width_mm": 126,
            "title_font": 10,
            "title_leading": 11,
            "sub_font": 9,
            "sub_leading": 10,
            "section_font": 11,
            "section_leading": 12,
            "district_font": 14,
            "district_leading": 16,
            "header_font": 16,
            "header_leading": 18,
            "body_font": 12,
            "body_leading": 14,
            "note_font": 13,
            "note_leading": 15,
            "table_top_bottom_padding": 5,
            "space_after_header_mm": 3,
            "space_after_table_mm": 5,
            "space_before_signature_mm": 10,
        }
    return {
        "logo_width_mm": 28,
        "logo_height_mm": 34,
        "logo_col_width_mm": 32,
        "header_col_width_mm": 120,
        "title_font": 11,
        "title_leading": 12,
        "sub_font": 10,
        "sub_leading": 12,
        "section_font": 12,
        "section_leading": 13,
        "district_font": 15,
        "district_leading": 18,
        "header_font": 18,
        "header_leading": 20,
        "body_font": 14,
        "body_leading": 16,
        "note_font": 14,
        "note_leading": 16,
        "table_top_bottom_padding": 8,
        "space_after_header_mm": 4,
        "space_after_table_mm": 6,
        "space_before_signature_mm": 12,
    }


def _waiting_hall_section_flowables(
    group: dict,
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    right_logo_source=None,
    paired=False,
):
    styles = getSampleStyleSheet()
    item_count = len(group.get("items") or [])
    profile = _waiting_hall_layout_profile(item_count)
    title_style = ParagraphStyle(
        "waiting-hall-om",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["title_font"],
        leading=profile["title_leading"],
        textColor=colors.red,
        spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "waiting-hall-sub",
        parent=styles["BodyText"],
        alignment=1,
        fontSize=profile["sub_font"],
        leading=profile["sub_leading"],
        textColor=colors.black,
        spaceAfter=1,
    )
    section_title_style = ParagraphStyle(
        "waiting-hall-section-title",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["section_font"],
        leading=profile["section_leading"],
        textColor=colors.black,
        spaceAfter=6,
    )
    district_style = ParagraphStyle(
        "waiting-hall-district",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["district_font"],
        leading=profile["district_leading"],
        spaceAfter=8,
    )
    table_header_text = ParagraphStyle(
        "waiting-hall-table-header-text",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=profile["header_font"],
        leading=profile["header_leading"],
    )
    table_header_qty_text = ParagraphStyle(
        "waiting-hall-table-header-qty-text",
        parent=table_header_text,
        alignment=1,
    )
    table_body_text = ParagraphStyle(
        "waiting-hall-table-body-text",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=profile["body_font"],
        leading=profile["body_leading"],
    )
    table_body_qty_text = ParagraphStyle(
        "waiting-hall-table-body-qty-text",
        parent=table_body_text,
        alignment=1,
    )
    note_style = ParagraphStyle(
        "waiting-hall-note",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=profile["note_font"],
        leading=profile["note_leading"],
    )
    signature_style = ParagraphStyle(
        "waiting-hall-signature",
        parent=styles["BodyText"],
        alignment=2,
        fontName="Helvetica-Bold",
        fontSize=profile["note_font"],
        leading=profile["note_leading"],
    )
    header_line_one, header_line_two = _waiting_hall_programme_lines(event_age_label, event_date)
    left_logo = _fitted_pdf_image_source(custom_logo, max_width_mm=profile["logo_width_mm"], max_height_mm=profile["logo_height_mm"])
    right_logo = _fitted_pdf_image_source(right_logo_source, max_width_mm=profile["logo_width_mm"], max_height_mm=profile["logo_height_mm"])
    header = Table(
        [[left_logo, [
            Paragraph("OM SAKTHI", title_style),
            Paragraph(header_line_one, sub_style),
            Paragraph(header_line_two, sub_style),
            Paragraph("District Waiting Hall Articles Collection List", section_title_style),
        ], right_logo]],
        colWidths=[profile["logo_col_width_mm"] * mm, profile["header_col_width_mm"] * mm, profile["logo_col_width_mm"] * mm],
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    table_rows = [
        [Paragraph("<b>Article list</b>", table_header_text), Paragraph("<b>Qty</b>", table_header_qty_text)],
    ]
    for item in group.get("items") or []:
        table_rows.append([
            Paragraph(escape(str(item.get("requested_item") or "")), table_body_text),
            Paragraph(str(item.get("quantity") or 0), table_body_qty_text),
        ])
    table_rows.append([
        Paragraph("<b>Total</b>", table_body_text),
        Paragraph(f"<b>{group.get('total_quantity') or 0}</b>", table_body_qty_text),
    ])
    items_table = Table(table_rows, colWidths=[144 * mm, 28 * mm], repeatRows=1)
    items_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), profile["table_top_bottom_padding"]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), profile["table_top_bottom_padding"]),
    ]))
    return [
        header,
        Spacer(1, profile["space_after_header_mm"] * mm),
        Paragraph(
            f"{escape(str(group.get('label_prefix') or 'Beneficiary'))}: {escape(str(group.get('entity_name') or '-').upper())}",
            district_style,
        ),
        items_table,
        Spacer(1, profile["space_after_table_mm"] * mm),
        Paragraph(
            "I Acknowledge that I have received the above articles",
            ParagraphStyle(
                "waiting-hall-note-smaller",
                parent=note_style,
                fontSize=max(note_style.fontSize - 2, 8),
                leading=max(note_style.leading - 2, 10),
            ),
        ),
        Spacer(1, (profile["space_before_signature_mm"] + 6) * mm),
        Paragraph("Signature", signature_style),
    ]


def _waiting_hall_section_story(
    group: dict,
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    right_logo_source=None,
    paired=False,
):
    return KeepTogether(
        _waiting_hall_section_flowables(
            group,
            event_age_label=event_age_label,
            event_date=event_date,
            custom_logo=custom_logo,
            right_logo_source=right_logo_source,
            paired=paired,
        )
    )


def generate_waiting_hall_acknowledgment_pdf(
    groups: list[dict],
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=12 * mm,
    )
    story = []
    optimized_custom_logo, _ = _optimized_report_logo(custom_logo)
    optimized_right_logo, _ = _optimized_report_logo(_pdf_logo_path(), "image/png")
    packed_pages = _waiting_hall_pack_pages(groups)
    for page_index, page_groups in enumerate(packed_pages):
        for group_index, group in enumerate(page_groups):
            story.append(
                _waiting_hall_section_story(
                    group,
                    event_age_label=event_age_label,
                    event_date=event_date,
                    custom_logo=optimized_custom_logo,
                    right_logo_source=optimized_right_logo,
                    paired=False,
                )
            )
            if group_index < len(page_groups) - 1:
                story.append(Spacer(1, 10 * mm))
        if page_index < len(packed_pages) - 1:
            story.append(PageBreak())
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_waiting_hall_acknowledgment_doc(
    groups: list[dict],
    *,
    event_age_label: str,
    event_date,
    custom_logo=None,
    custom_logo_mime_type: str | None = None,
) -> bytes:
    def _twips_from_mm(value_mm: float) -> int:
        return int(round(float(value_mm) * 56.6929))

    def _twips_from_pt(value_pt: float) -> int:
        return int(round(float(value_pt) * 20))

    def _half_points(value_pt: float) -> int:
        return int(round(float(value_pt) * 2))

    def _emu_from_mm(value_mm: float) -> int:
        return int(round(float(value_mm) * 36000))

    def _xml_text(value: str) -> str:
        return escape(str(value or ""))

    def _w_paragraph(text: str = "", *, align: str | None = None, bold: bool = False, size_pt: float | None = None,
                     color: str | None = None,
                     spacing_before_pt: float | None = None, spacing_after_pt: float | None = None) -> str:
        paragraph_props = []
        if align:
            paragraph_props.append(f'<w:jc w:val="{align}"/>')
        spacing_attrs = []
        if spacing_before_pt is not None:
            spacing_attrs.append(f'w:before="{_twips_from_pt(spacing_before_pt)}"')
        if spacing_after_pt is not None:
            spacing_attrs.append(f'w:after="{_twips_from_pt(spacing_after_pt)}"')
        if spacing_attrs:
            paragraph_props.append(f"<w:spacing {' '.join(spacing_attrs)}/>")
        ppr = f"<w:pPr>{''.join(paragraph_props)}</w:pPr>" if paragraph_props else ""
        run_props = []
        if bold:
            run_props.append("<w:b/>")
        if size_pt is not None:
            run_props.append(f'<w:sz w:val="{_half_points(size_pt)}"/>')
        if color:
            run_props.append(f'<w:color w:val="{color}"/>')
        rpr = f"<w:rPr>{''.join(run_props)}</w:rPr>" if run_props else ""
        return f"<w:p>{ppr}<w:r>{rpr}<w:t xml:space=\"preserve\">{_xml_text(text)}</w:t></w:r></w:p>"

    def _w_empty_paragraph(*, spacing_after_pt: float | None = None) -> str:
        return _w_paragraph("", spacing_after_pt=spacing_after_pt)

    def _w_table_cell(inner_xml: str, *, width_twips: int, align: str | None = None, shading: str | None = None,
                      borders: bool = True) -> str:
        tc_pr = [f'<w:tcW w:w="{width_twips}" w:type="dxa"/>', '<w:vAlign w:val="center"/>']
        if not borders:
            tc_pr.append(
                "<w:tcBorders><w:top w:val=\"nil\"/><w:left w:val=\"nil\"/><w:bottom w:val=\"nil\"/><w:right w:val=\"nil\"/></w:tcBorders>"
            )
        if shading:
            tc_pr.append(f'<w:shd w:val="clear" w:color="auto" w:fill="{shading}"/>')
        if align:
            inner_xml = inner_xml.replace("<w:p>", f"<w:p><w:pPr><w:jc w:val=\"{align}\"/></w:pPr>", 1)
        return f"<w:tc><w:tcPr>{''.join(tc_pr)}</w:tcPr>{inner_xml}</w:tc>"

    def _w_image_paragraph(rel_id: str, *, width_mm: float, height_mm: float, docpr_id: int) -> str:
        cx = _emu_from_mm(width_mm)
        cy = _emu_from_mm(height_mm)
        return f"""
        <w:p>
          <w:pPr><w:jc w:val="center"/></w:pPr>
          <w:r>
            <w:drawing>
              <wp:inline distT="0" distB="0" distL="0" distR="0">
                <wp:extent cx="{cx}" cy="{cy}"/>
                <wp:docPr id="{docpr_id}" name="Picture {docpr_id}"/>
                <a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
                  <a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">
                    <pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
                      <pic:nvPicPr>
                        <pic:cNvPr id="{docpr_id}" name="Picture {docpr_id}"/>
                        <pic:cNvPicPr/>
                      </pic:nvPicPr>
                      <pic:blipFill>
                        <a:blip r:embed="{rel_id}"/>
                        <a:stretch><a:fillRect/></a:stretch>
                      </pic:blipFill>
                      <pic:spPr>
                        <a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>
                        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
                      </pic:spPr>
                    </pic:pic>
                  </a:graphicData>
                </a:graphic>
              </wp:inline>
            </w:drawing>
          </w:r>
        </w:p>
        """

    def _docx_supported_image(image_bytes: bytes | None, mime_type: str | None):
        if not image_bytes:
            return None, None
        normalized_mime = (mime_type or "image/png").strip().lower()
        if normalized_mime in {"image/png", "image/jpeg", "image/jpg"}:
            return image_bytes, "image/jpeg" if normalized_mime == "image/jpg" else normalized_mime
        if PILImage is not None:
            try:
                source = io.BytesIO(image_bytes)
                with PILImage.open(source) as image:
                    converted = io.BytesIO()
                    image.convert("RGBA").save(converted, format="PNG")
                    return converted.getvalue(), "image/png"
            except Exception:
                return None, None
        return None, None

    header_line_one, header_line_two = _waiting_hall_programme_lines(event_age_label, event_date)
    optimized_custom_logo, optimized_custom_logo_mime = _normalized_docx_report_logo(custom_logo, custom_logo_mime_type)
    optimized_right_logo, optimized_right_logo_mime = _normalized_docx_report_logo(_pdf_logo_path(), "image/png")
    left_logo_bytes, left_logo_mime = _docx_supported_image(optimized_custom_logo, optimized_custom_logo_mime)
    right_logo_bytes, right_logo_mime = _docx_supported_image(optimized_right_logo, optimized_right_logo_mime)

    rel_counter = 2
    image_counter = 1
    docpr_counter = 1
    media_files: list[tuple[str, bytes, str, str]] = []
    image_rel_ids: dict[str, str] = {}

    def _register_image(key: str, image_bytes: bytes | None, mime_type: str | None) -> str | None:
        nonlocal rel_counter, image_counter
        if not image_bytes or not mime_type:
            return None
        if key in image_rel_ids:
            return image_rel_ids[key]
        extension = "png" if mime_type == "image/png" else "jpeg"
        target = f"media/image{image_counter}.{extension}"
        rel_id = f"rId{rel_counter}"
        rel_counter += 1
        image_counter += 1
        media_files.append((target, image_bytes, mime_type, rel_id))
        image_rel_ids[key] = rel_id
        return rel_id

    left_logo_rel = _register_image("left_logo", left_logo_bytes, left_logo_mime)
    right_logo_rel = _register_image("right_logo", right_logo_bytes, right_logo_mime)

    body_parts = []
    content_width_twips = _twips_from_mm(186)
    logo_width_twips = _twips_from_mm(30)
    center_width_twips = content_width_twips - (2 * logo_width_twips)
    item_col_twips = _twips_from_mm(145)
    qty_col_twips = _twips_from_mm(27)

    for page_index, page_groups in enumerate(_waiting_hall_pack_pages(groups)):
        for group_index, group in enumerate(page_groups):
            item_count = len(group.get("items") or [])
            profile = _waiting_hall_layout_profile(item_count)

            left_cell_xml = _w_empty_paragraph()
            if left_logo_rel:
                logo_side_mm = min(profile["logo_width_mm"], profile["logo_height_mm"])
                left_cell_xml = _w_image_paragraph(
                    left_logo_rel,
                    width_mm=logo_side_mm,
                    height_mm=logo_side_mm,
                    docpr_id=docpr_counter,
                )
                docpr_counter += 1

            right_cell_xml = _w_empty_paragraph()
            if right_logo_rel:
                logo_side_mm = min(profile["logo_width_mm"], profile["logo_height_mm"])
                right_cell_xml = _w_image_paragraph(
                    right_logo_rel,
                    width_mm=logo_side_mm,
                    height_mm=logo_side_mm,
                    docpr_id=docpr_counter,
                )
                docpr_counter += 1

            title_cell_xml = "".join([
                _w_paragraph("OM SAKTHI", align="center", bold=True, size_pt=profile["title_font"], color="C1121F", spacing_after_pt=1),
                _w_paragraph(header_line_one, align="center", size_pt=profile["sub_font"], spacing_after_pt=0),
                _w_paragraph(header_line_two, align="center", size_pt=profile["sub_font"], spacing_after_pt=1),
                _w_paragraph(
                    "District Waiting Hall Articles Collection List",
                    align="center",
                    bold=True,
                    size_pt=profile["section_font"],
                    spacing_after_pt=2,
                ),
            ])

            header_table = f"""
            <w:tbl>
              <w:tblPr>
                <w:tblW w:w="{content_width_twips}" w:type="dxa"/>
                <w:tblBorders>
                  <w:top w:val="nil"/><w:left w:val="nil"/><w:bottom w:val="nil"/><w:right w:val="nil"/>
                  <w:insideH w:val="nil"/><w:insideV w:val="nil"/>
                </w:tblBorders>
              </w:tblPr>
              <w:tblGrid>
                <w:gridCol w:w="{logo_width_twips}"/>
                <w:gridCol w:w="{center_width_twips}"/>
                <w:gridCol w:w="{logo_width_twips}"/>
              </w:tblGrid>
              <w:tr>
                {_w_table_cell(left_cell_xml, width_twips=logo_width_twips, borders=False)}
                {_w_table_cell(title_cell_xml, width_twips=center_width_twips, borders=False)}
                {_w_table_cell(right_cell_xml, width_twips=logo_width_twips, borders=False)}
              </w:tr>
            </w:tbl>
            """
            body_parts.append(header_table)
            body_parts.append(
                _w_paragraph(
                    f"{str(group.get('label_prefix') or 'Beneficiary')}: {str(group.get('entity_name') or '-').upper()}",
                    align="center",
                    bold=True,
                    size_pt=profile["district_font"],
                    spacing_after_pt=4,
                )
            )

            row_xml = [
                "<w:tr>"
                + _w_table_cell(_w_paragraph("Article list", bold=True, size_pt=profile["header_font"]), width_twips=item_col_twips, shading="F3F4F6")
                + _w_table_cell(_w_paragraph("Qty", align="center", bold=True, size_pt=profile["header_font"]), width_twips=qty_col_twips, shading="F3F4F6")
                + "</w:tr>"
            ]
            for item in (group.get("items") or []):
                row_xml.append(
                    "<w:tr>"
                    + _w_table_cell(_w_paragraph(str(item.get("requested_item") or ""), bold=True, size_pt=profile["body_font"]), width_twips=item_col_twips)
                    + _w_table_cell(_w_paragraph(str(int(item.get("quantity") or 0)), align="center", bold=True, size_pt=profile["body_font"]), width_twips=qty_col_twips)
                    + "</w:tr>"
                )
            row_xml.append(
                "<w:tr>"
                + _w_table_cell(_w_paragraph("Total", bold=True, size_pt=profile["body_font"]), width_twips=item_col_twips)
                + _w_table_cell(_w_paragraph(str(int(group.get("total_quantity") or 0)), align="center", bold=True, size_pt=profile["body_font"]), width_twips=qty_col_twips)
                + "</w:tr>"
            )
            body_parts.append(
                f"""
                <w:tbl>
                  <w:tblPr>
                    <w:tblW w:w="{item_col_twips + qty_col_twips}" w:type="dxa"/>
                    <w:jc w:val="center"/>
                    <w:tblBorders>
                      <w:top w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:left w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:bottom w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:right w:val="single" w:sz="8" w:space="0" w:color="000000"/>
                      <w:insideH w:val="single" w:sz="6" w:space="0" w:color="000000"/>
                      <w:insideV w:val="single" w:sz="6" w:space="0" w:color="000000"/>
                    </w:tblBorders>
                  </w:tblPr>
                  <w:tblGrid>
                    <w:gridCol w:w="{item_col_twips}"/>
                    <w:gridCol w:w="{qty_col_twips}"/>
                  </w:tblGrid>
                  {''.join(row_xml)}
                </w:tbl>
                """
            )
            body_parts.append(
                _w_paragraph(
                    "I Acknowledge that I have received the above articles",
                    align="center",
                    bold=True,
                    size_pt=profile["note_font"],
                    spacing_before_pt=8,
                    spacing_after_pt=12,
                )
            )
            body_parts.append(
                _w_paragraph(
                    "Signature",
                    align="right",
                    bold=True,
                    size_pt=profile["note_font"],
                    spacing_before_pt=profile["note_leading"] * 2,
                    spacing_after_pt=0,
                )
            )

            if group_index < len(page_groups) - 1:
                body_parts.append(_w_empty_paragraph(spacing_after_pt=18))
        if page_index < len(_waiting_hall_pack_pages(groups)) - 1:
            body_parts.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')

    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
      <w:style w:type="paragraph" w:default="1" w:styleId="Normal">
        <w:name w:val="Normal"/>
        <w:qFormat/>
        <w:rPr>
          <w:rFonts w:ascii="Arial" w:hAnsi="Arial"/>
          <w:sz w:val="22"/>
        </w:rPr>
      </w:style>
    </w:styles>
    """

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:document
      xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
      xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
      xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
      xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
      xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">
      <w:body>
        {''.join(body_parts)}
        <w:sectPr>
          <w:pgSz w:w="11906" w:h="16838"/>
          <w:pgMar w:top="680" w:right="680" w:bottom="680" w:left="680" w:header="708" w:footer="708" w:gutter="0"/>
        </w:sectPr>
      </w:body>
    </w:document>
    """

    image_defaults: dict[str, str] = {}
    document_rels = [
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
    ]
    for target, _image_bytes, mime_type, rel_id in media_files:
        extension = target.rsplit(".", 1)[-1]
        content_type = "image/png" if mime_type == "image/png" else "image/jpeg"
        image_defaults[extension] = content_type
        document_rels.append(
            f'<Relationship Id="{rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{target}"/>'
        )

    content_types_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
      <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
      <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
      <Default Extension="xml" ContentType="application/xml"/>
      {''.join(f'<Default Extension="{extension}" ContentType="{content_type}"/>' for extension, content_type in image_defaults.items())}
      <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
      <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
    </Types>
    """

    package_rels_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
      <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
    </Relationships>
    """

    document_rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
      {''.join(document_rels)}
    </Relationships>
    """

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as docx_zip:
        docx_zip.writestr("[Content_Types].xml", content_types_xml)
        docx_zip.writestr("_rels/.rels", package_rels_xml)
        docx_zip.writestr("word/document.xml", document_xml)
        docx_zip.writestr("word/styles.xml", styles_xml)
        docx_zip.writestr("word/_rels/document.xml.rels", document_rels_xml)
        for target, image_bytes, _mime_type, _rel_id in media_files:
            docx_zip.writestr(f"word/{target}", image_bytes)
    return buffer.getvalue()


def _public_acknowledgment_pdf_fields(template_bytes: bytes | None) -> list[dict]:
    if not template_bytes:
        return []
    reader = PdfReader(io.BytesIO(template_bytes))
    seen = set()
    fields: list[dict] = []
    for page in reader.pages:
        for annotation_ref in page.get("/Annots") or []:
            annotation = annotation_ref.get_object()
            if str(annotation.get("/Subtype") or "") != "/Widget":
                continue
            if str(annotation.get("/FT") or "") != "/Tx":
                continue
            field_name = str(annotation.get("/T") or "").strip()
            if not field_name or field_name in seen:
                continue
            seen.add(field_name)
            field_key = re.sub(r"[^a-z0-9]+", "_", field_name.casefold()).strip("_") or f"field_{len(fields) + 1}"
            fields.append(
                {
                    "field_name": field_name,
                    "field_key": field_key,
                }
            )
    return fields


def _public_acknowledgment_resolve_value(row: dict, column_name: str | None) -> str:
    if not column_name:
        return ""
    column = str(column_name or "").strip()
    if not column:
        return ""
    blank_tokens = {"n/a", "na", "none", "null", "-"}

    def _clean(value: object) -> str:
        text = str(value or "").strip()
        if text.casefold() in blank_tokens:
            return ""
        return text

    lookup_keys = {
        column,
        column.replace("_", " "),
        column.replace(" ", "_"),
        column.casefold(),
        column.replace("_", " ").casefold(),
        column.replace(" ", "_").casefold(),
    }
    for key, value in row.items():
        key_text = str(key or "").strip()
        if key_text in lookup_keys or key_text.casefold() in {item.casefold() for item in lookup_keys}:
            return _clean(value)
    return _clean(row.get(column) or row.get(column.replace("_", " ")) or row.get(column.replace(" ", "_")) or "")


def _public_acknowledgment_is_aid_row(row: dict) -> bool:
    item_type = str(row.get("Item Type") or row.get("item_type") or "").strip().casefold()
    return item_type == str(models.ItemTypeChoices.AID).strip().casefold()


def _public_acknowledgment_field_rects(template_bytes: bytes) -> dict[str, list[float]]:
    reader = PdfReader(io.BytesIO(template_bytes))
    rects: dict[str, list[float]] = {}
    if not reader.pages:
        return rects
    for annotation_ref in reader.pages[0].get("/Annots") or []:
        annotation = annotation_ref.get_object()
        if str(annotation.get("/Subtype") or "") != "/Widget":
            continue
        field_name = str(annotation.get("/T") or "").strip()
        rect = annotation.get("/Rect")
        if not field_name or not rect:
            continue
        rects[field_name] = [float(value) for value in rect]
    return rects


def _public_acknowledgment_draw_field(
    overlay_canvas,
    field_name: str,
    value: str,
    rect: list[float],
) -> None:
    x1, y1, x2, y2 = rect
    width = max(x2 - x1, 0)
    height = max(y2 - y1, 0)
    text = str(value or "").strip()
    if not text:
        return

    if field_name == "address":
        style = ParagraphStyle(
            "public_ack_address",
            fontName="Helvetica",
            fontSize=12,
            leading=13,
            textColor=colors.black,
            alignment=0,
        )
        paragraph = Paragraph(text.replace("\n", "<br/>"), style)
        frame = Frame(x1 + 6, y1 + 4, max(width - 10, 1), max(height - 8, 1), showBoundary=0)
        frame.addFromList([paragraph], overlay_canvas)
        return

    font_name = "Helvetica"
    font_size = 16
    if field_name in {"district", "bf name", "article", "cheque_no"}:
        font_name = "Helvetica-Bold"
        font_size = 17 if field_name != "district" else 19
    elif field_name in {"App no", "token"}:
        font_name = "Helvetica-Bold"
        font_size = 16
    elif field_name in {"mobile", "Aadhar", "value_aid"}:
        font_name = "Helvetica-Bold"
        font_size = 16

    overlay_canvas.setFillColor(colors.black)
    overlay_canvas.setFont(font_name, font_size)
    text_width = overlay_canvas.stringWidth(text, font_name, font_size)
    if text_width > width - 8:
        shrink = max((width - 8) / max(text_width, 1), 0.6)
        font_size = max(round(font_size * shrink), 8)
        overlay_canvas.setFont(font_name, font_size)
        text_width = overlay_canvas.stringWidth(text, font_name, font_size)
    x = x1 + max((width - text_width) / 2, 2)
    y = y1 + max((height - font_size) / 2, 2)
    overlay_canvas.drawString(x, y, text)


def generate_public_acknowledgment_pdf(
    template_bytes: bytes,
    rows: list[dict],
    field_map: dict[str, str],
) -> io.BytesIO:
    normalized_map = {str(key or "").strip(): str(value or "").strip() for key, value in (field_map or {}).items()}
    if not rows:
        rows = [{}]
    writer = PdfWriter()
    rect_map = _public_acknowledgment_field_rects(template_bytes)
    for row in rows:
        template_reader = PdfReader(io.BytesIO(template_bytes))
        if not template_reader.pages:
            raise ValueError("Uploaded PDF template has no pages.")
        page = template_reader.pages[0]
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)
        overlay_buffer = io.BytesIO()
        overlay_canvas = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
        for field_name, column_name in normalized_map.items():
            if not field_name:
                continue
            if field_name == "value_aid" and not _public_acknowledgment_is_aid_row(row):
                continue
            value = _public_acknowledgment_resolve_value(row, column_name)
            rect = rect_map.get(field_name)
            if not rect:
                continue
            _public_acknowledgment_draw_field(overlay_canvas, field_name, value, rect)
        overlay_canvas.save()
        overlay_buffer.seek(0)
        overlay_reader = PdfReader(overlay_buffer)
        if overlay_reader.pages:
            page.merge_page(overlay_reader.pages[0])
        if "/Annots" in page:
            del page["/Annots"]
        writer.add_page(page)
    buffer = io.BytesIO()
    writer.write(buffer)
    buffer.seek(0)
    return buffer


def _public_signature_app_sort_key(value: str) -> tuple[str, int, str]:
    text = str(value or "").strip()
    match = re.match(r"^([A-Za-z]+)\s*0*([0-9]+)$", text)
    if match:
        return (match.group(1).casefold(), int(match.group(2)), text.casefold())
    return ("", 0, text.casefold())


def generate_public_signature_pdf(rows: list[dict]) -> io.BytesIO:
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "public_signature_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=12,
        alignment=1,
        textColor=colors.black,
    )
    body_style = ParagraphStyle(
        "public_signature_body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=11,
        textColor=colors.black,
    )
    body_center_style = ParagraphStyle(
        "public_signature_body_center",
        parent=body_style,
        alignment=1,
    )
    body_bold_style = ParagraphStyle(
        "public_signature_body_bold",
        parent=body_style,
        fontName="Helvetica-Bold",
    )
    body_center_bold_style = ParagraphStyle(
        "public_signature_body_center_bold",
        parent=body_center_style,
        fontName="Helvetica-Bold",
    )

    sorted_rows = list(rows or [])

    table_rows = [[
        Paragraph("Sl No", header_style),
        Paragraph("App No", header_style),
        Paragraph("Name", header_style),
        Paragraph("Requested Item", header_style),
        Paragraph("Token No", header_style),
        Paragraph("Signature", header_style),
    ]]

    for index, row in enumerate(sorted_rows, start=1):
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or token_start or 0)
        token_text = str(token_start) if token_end <= token_start else f"{token_start} - {token_end}"
        table_rows.append(
            [
                Paragraph(str(index), body_center_bold_style),
                Paragraph(str(row.get("application_number") or ""), body_style),
                Paragraph(str(row.get("beneficiary_name") or ""), body_style),
                Paragraph(str(row.get("item_name") or ""), body_style),
                Paragraph(token_text, body_center_bold_style),
                Paragraph("", body_style),
            ]
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    table = LongTable(
        table_rows,
        colWidths=[12 * mm, 20 * mm, 46 * mm, 54 * mm, 18 * mm, 48 * mm],
        repeatRows=1,
        splitByRow=1,
    )
    table.hAlign = "CENTER"
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9fbbe7")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (4, 0), (4, -1), "CENTER"),
                ("ALIGN", (5, 0), (5, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4.5),
                ("TOPPADDING", (0, 0), (-1, 0), 7),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ]
        )
    )
    doc.build([table], canvasmaker=_NumberedPdfCanvas)
    buffer.seek(0)
    return buffer



def generate_public_signature_xlsx(rows: list[dict]) -> io.BytesIO:
    sorted_rows = list(rows or [])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Public Signature"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="9FBBE7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Sl No", "App No", "Name", "Requested Item", "Token No", "Signature"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for index, row in enumerate(sorted_rows, start=1):
        token_start = int(row.get("token_start") or 0)
        token_end = int(row.get("token_end") or token_start or 0)
        token_text = str(token_start) if token_end <= token_start else f"{token_start} - {token_end}"
        worksheet.append([
            index,
            str(row.get("application_number") or ""),
            str(row.get("beneficiary_name") or ""),
            str(row.get("item_name") or ""),
            token_text,
            "",
        ])

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=6):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = center if idx in {1, 2, 5} else left
            cell.font = Font(size=10.5, bold=idx in {1, 5})

    widths = {1: 10, 2: 16, 3: 32, 4: 34, 5: 16, 6: 40}
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


class _DistrictSignatureNumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.setFont("Helvetica", 8)
            self.setFillColor(colors.black)
            self.drawCentredString(A4[0] / 2, 8 * mm, f"Page {self._pageNumber} of {total_pages}")
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)


def generate_district_signature_xlsx(
    districts: list[dict],
    *,
    custom_logo: bytes | None = None,
    custom_logo_mime_type: str | None = None,
) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "District Signature"

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    header_fill = PatternFill("solid", fgColor="9FBBE7")
    district_fill = PatternFill("solid", fgColor="F4F8FD")
    title_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.3
    worksheet.page_margins.bottom = 0.3
    worksheet.page_margins.header = 0.1
    worksheet.page_margins.footer = 0.1

    worksheet.row_dimensions[1].height = 22

    left_logo_bytes = None
    right_logo_bytes = None
    try:
        left_logo_source = custom_logo or _pdf_guru_logo_path()
        left_logo_mime = custom_logo_mime_type or ("image/png" if custom_logo else "image/jpeg")
        left_logo_bytes, _ = _normalized_docx_report_logo(left_logo_source, left_logo_mime, canvas_width_px=120, canvas_height_px=120)
        right_logo_bytes, _ = _normalized_docx_report_logo(_pdf_logo_path(), "image/png", canvas_width_px=120, canvas_height_px=120)
    except Exception:
        left_logo_bytes = None
        right_logo_bytes = None

    if left_logo_bytes:
        left_logo = XLImage(io.BytesIO(left_logo_bytes))
        left_logo.width = 32
        left_logo.height = 32
        worksheet.add_image(left_logo, "A1")
    if right_logo_bytes:
        right_logo = XLImage(io.BytesIO(right_logo_bytes))
        right_logo.width = 32
        right_logo.height = 32
        worksheet.add_image(right_logo, "F1")

    current_year = timezone.localdate().year
    birthday_number = max(current_year - 1940, 1)

    # Keep row 1 as the table header row for downstream tooling/tests.

    headers = ["District", "Total Qty", "Token Qty", "Start Token", "End Token", "Signature"]
    header_row_index = 1
    for col_index, header_value in enumerate(headers, start=1):
        cell = worksheet.cell(header_row_index, col_index, header_value)
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = outer_border
        cell.alignment = center

    row_index = 2
    for district in list(districts or []):
        district_name = str(district.get("district_name") or "").strip()
        if not district_name:
            continue

        block_start = row_index
        worksheet.append([district_name, "", "", "", "", ""])
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        for col in range(1, 7):
            worksheet.cell(row_index, col).border = outer_border
        district_cell = worksheet.cell(row_index, 1)
        district_cell.font = Font(size=10.5, bold=True)
        district_cell.fill = district_fill
        district_cell.alignment = center
        row_index += 1

        for item in district.get("items") or []:
            worksheet.append([
                str(item.get("item_name") or ""),
                int(item.get("total_quantity") or 0),
                int(item.get("token_quantity") or 0) if item.get("token_quantity") not in {None, "", 0} else "",
                _district_signature_display_token(item.get("start_token")),
                _district_signature_display_token(item.get("end_token")),
                "",
            ])
            for col in range(1, 7):
                worksheet.cell(row_index, col).border = border
                worksheet.cell(row_index, col).alignment = center if col in {2, 3, 4, 5} else left
                worksheet.cell(row_index, col).font = Font(size=10)
            row_index += 1

        worksheet.append([
            f"{district_name} Total",
            int(district.get("total_quantity") or 0),
            int(district.get("token_quantity") or 0) if district.get("token_quantity") not in {None, "", 0} else "",
            "",
            "",
            "",
        ])
        for col in range(1, 7):
            worksheet.cell(row_index, col).border = outer_border
            worksheet.cell(row_index, col).alignment = center if col in {2, 3} else left
            worksheet.cell(row_index, col).font = Font(size=10.5, bold=True)
            worksheet.cell(row_index, col).fill = district_fill

        # Merge the signature column across the district block so it reads as a
        # single signing area instead of repeated per-row cells.
        worksheet.merge_cells(start_row=block_start, start_column=6, end_row=row_index, end_column=6)
        signature_cell = worksheet.cell(block_start, 6)
        signature_cell.border = outer_border
        signature_cell.alignment = center

        row_index += 1

        # Insert a clean visual separator between districts without disturbing
        # the district box borders.
        worksheet.append(["", "", "", "", "", ""])
        separator_row = row_index
        worksheet.row_dimensions[separator_row].height = 8
        row_index += 1

    widths = {1: 34, 2: 12, 3: 12, 4: 14, 5: 14, 6: 28}
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _district_signature_display_token(value):
    if value in {None, "", 0}:
        return ""
    return str(int(value))


def generate_district_signature_pdf(
    districts: list[dict],
    *,
    custom_logo: bytes | None = None,
) -> io.BytesIO:
    styles = getSampleStyleSheet()
    district_border = colors.HexColor("#5b6572")
    district_light_border = colors.HexColor("#7c8794")
    left_logo_source = custom_logo or _pdf_guru_logo_path()
    left_logo = _fitted_pdf_image_source(left_logo_source, max_width_mm=18, max_height_mm=18)
    right_logo = _fitted_pdf_image(_pdf_logo_path(), max_width_mm=18, max_height_mm=18)
    header_title = ParagraphStyle(
        "district_signature_title",
        parent=styles["Heading2"],
        alignment=1,
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        textColor=colors.black,
    )
    header_sub = ParagraphStyle(
        "district_signature_sub",
        parent=styles["BodyText"],
        alignment=1,
        fontName="Helvetica",
        fontSize=9.3,
        leading=9.2,
        spaceBefore=0,
        spaceAfter=0,
        textColor=colors.black,
    )
    header_style = ParagraphStyle(
        "district_signature_header",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.3,
        leading=10.3,
        alignment=1,
        textColor=colors.black,
    )
    district_style = ParagraphStyle(
        "district_signature_district",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.4,
        leading=10.4,
        alignment=1,
        textColor=colors.black,
    )
    item_style = ParagraphStyle(
        "district_signature_item",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=9.8,
        textColor=colors.black,
    )
    total_style = ParagraphStyle(
        "district_signature_total",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.9,
        leading=9.9,
        textColor=colors.black,
    )
    numeric_style = ParagraphStyle(
        "district_signature_numeric",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=9.8,
        alignment=1,
        textColor=colors.black,
    )
    numeric_bold_style = ParagraphStyle(
        "district_signature_numeric_bold",
        parent=numeric_style,
        fontName="Helvetica-Bold",
    )

    table_rows = [[
        Paragraph("District", header_style),
        Paragraph("Total Qty", header_style),
        Paragraph("Token Qty", header_style),
        Paragraph("Start Token", header_style),
        Paragraph("End Token", header_style),
        Paragraph("Signature", header_style),
    ]]
    style_commands = [
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9fbbe7")),
        ("BOX", (0, 0), (-1, 0), 0.7, district_border),
        ("INNERGRID", (0, 0), (-1, 0), 0.55, district_light_border),
        ("ALIGN", (1, 0), (4, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ]

    district_prog_line_one, district_prog_line_two = _district_signature_programme_lines()
    header_rows = [[
        left_logo,
        [
            Paragraph("OM SAKTHI", ParagraphStyle("district-signature-om", parent=header_title, fontSize=8.8, leading=9.0, spaceAfter=0, textColor=colors.red)),
            Paragraph(district_prog_line_one, header_sub),
            Paragraph(district_prog_line_two, header_sub),
            Paragraph("District Token Distribution Register", header_title),
        ],
        right_logo,
    ]]

    row_index = 1
    for district in list(districts or []):
        district_name = str(district.get("district_name") or "").strip()
        if not district_name:
            continue
        block_start = row_index
        table_rows.append([
            Paragraph(district_name, district_style),
            "",
            "",
            "",
            "",
            "",
        ])
        style_commands.extend([
            ("SPAN", (0, row_index), (4, row_index)),
            ("BACKGROUND", (0, row_index), (4, row_index), colors.HexColor("#f4f8fd")),
            ("ALIGN", (0, row_index), (4, row_index), "CENTER"),
            ("LINEBELOW", (0, row_index), (4, row_index), 0.75, district_light_border),
            ("TOPPADDING", (0, row_index), (5, row_index), 8),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 8),
        ])
        row_index += 1

        for item in district.get("items") or []:
            table_rows.append([
                Paragraph(str(item.get("item_name") or ""), item_style),
                Paragraph(str(int(item.get("total_quantity") or 0)), numeric_style),
                Paragraph(_district_signature_display_token(item.get("token_quantity")), numeric_style),
                Paragraph(_district_signature_display_token(item.get("start_token")), numeric_style),
                Paragraph(_district_signature_display_token(item.get("end_token")), numeric_style),
                "",
            ])
            style_commands.extend([
                ("LINEBELOW", (0, row_index), (4, row_index), 0.45, district_light_border),
                ("TOPPADDING", (0, row_index), (5, row_index), 4),
                ("BOTTOMPADDING", (0, row_index), (5, row_index), 4),
            ])
            row_index += 1

        table_rows.append([
            Paragraph(f"{district_name} Total", total_style),
            Paragraph(str(int(district.get("total_quantity") or 0)), numeric_bold_style),
            Paragraph(_district_signature_display_token(district.get("token_quantity")), numeric_bold_style),
            "",
            "",
            "",
        ])
        style_commands.extend([
            ("LINEABOVE", (0, row_index), (4, row_index), 1.0, district_border),
            ("LINEBELOW", (0, row_index), (5, row_index), 1.0, district_border),
            ("TOPPADDING", (0, row_index), (5, row_index), 5),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 6),
            ("BOX", (0, block_start), (5, row_index), 1.2, district_border),
            ("INNERGRID", (0, block_start), (4, row_index), 0.45, district_light_border),
            ("LINEBEFORE", (5, block_start), (5, row_index), 1.2, district_border),
        ])
        row_index += 1
        table_rows.append(["", "", "", "", "", ""])
        style_commands.extend([
            ("SPAN", (0, row_index), (5, row_index)),
            ("TOPPADDING", (0, row_index), (5, row_index), 4),
            ("BOTTOMPADDING", (0, row_index), (5, row_index), 4),
        ])
        row_index += 1

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=5 * mm,
        bottomMargin=8 * mm,
    )
    table = LongTable(
        table_rows,
        colWidths=[82 * mm, 14 * mm, 14 * mm, 16 * mm, 16 * mm, 45 * mm],
        repeatRows=1,
        splitByRow=1,
    )
    table.hAlign = "CENTER"
    table.setStyle(TableStyle(style_commands))
    header_table = Table(header_rows, colWidths=[18 * mm, 144 * mm, 18 * mm])
    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story = [
        header_table,
        Spacer(1, 0.8 * mm),
        table,
    ]
    doc.build(story, canvasmaker=_DistrictSignatureNumberedCanvas)
    buffer.seek(0)
    return buffer


def _segregation_pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "segregation_title",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11.8,
            leading=14,
            alignment=1,
            textColor=colors.black,
            spaceAfter=4,
        ),
        "section": ParagraphStyle(
            "segregation_section",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9.5,
            leading=11.5,
            textColor=colors.black,
            spaceBefore=3,
            spaceAfter=3,
        ),
        "header": ParagraphStyle(
            "segregation_header",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.8,
            leading=10,
            alignment=1,
            textColor=colors.black,
        ),
        "body": ParagraphStyle(
            "segregation_body",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.4,
            leading=9.8,
            textColor=colors.black,
        ),
        "body_bold": ParagraphStyle(
            "segregation_body_bold",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.4,
            leading=9.8,
            textColor=colors.black,
        ),
        "body_center": ParagraphStyle(
            "segregation_body_center",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.4,
            leading=9.8,
            alignment=1,
            textColor=colors.black,
        ),
        "body_center_bold": ParagraphStyle(
            "segregation_body_center_bold",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.4,
            leading=9.8,
            alignment=1,
            textColor=colors.black,
        ),
        "empty": ParagraphStyle(
            "segregation_empty",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=8.8,
            leading=10.2,
            alignment=1,
            textColor=colors.HexColor("#475569"),
        ),
    }


def _segregation_pdf_table(table_rows, *, col_widths):
    table = LongTable(table_rows, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.hAlign = "CENTER"
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.55, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, 0), 4),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 1), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ]
        )
    )
    return table


def generate_segregation_file1_pdf(groups: list[dict]) -> io.BytesIO:
    styles = _segregation_pdf_styles()
    body_style = styles["body"]
    body_center_style = styles["body_center"]
    body_bold_style = styles["body_bold"]
    body_center_bold_style = styles["body_center_bold"]
    header_style = styles["header"]
    story = [Paragraph("File 1 : Beneficiary-wise Article List (Waiting Hall)", styles["title"])]

    if not groups:
        story.append(Paragraph("No rows available for this report.", styles["empty"]))
    else:
        for index, group in enumerate(list(groups or []), start=1):
            story.append(Paragraph(f"{index}. {escape(str(group.get('beneficiary_label') or ''))}", styles["section"]))
            table_rows = [[
                Paragraph("Sl No", header_style),
                Paragraph("Article", header_style),
                Paragraph("Qty", header_style),
                Paragraph("Signature", header_style),
            ]]
            for item_index, item in enumerate(list(group.get("items") or []), start=1):
                table_rows.append(
                    [
                        Paragraph(str(item_index), body_center_bold_style),
                        Paragraph(escape(str(item.get("article_name") or "")), body_style),
                        Paragraph(str(int(item.get("quantity") or 0)), body_center_bold_style),
                        Paragraph("", body_style),
                    ]
                )
            table_rows.append(
                [
                    Paragraph("", body_center_style),
                    Paragraph("Total", body_bold_style),
                    Paragraph(str(int(group.get("total_quantity") or 0)), body_center_bold_style),
                    Paragraph("", body_style),
                ]
            )
            table = _segregation_pdf_table(table_rows, col_widths=[14 * mm, 112 * mm, 18 * mm, 44 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (2, 0), (2, -1), "CENTER"),
                        ("BACKGROUND", (0, len(table_rows) - 1), (-1, len(table_rows) - 1), colors.HexColor("#eff6ff")),
                    ]
                )
            )
            story.extend([table, Spacer(1, 5 * mm)])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    doc.build(story, canvasmaker=_NumberedPdfCanvas)
    buffer.seek(0)
    return buffer


def generate_segregation_file2_pdf(groups: list[dict]) -> io.BytesIO:
    styles = _segregation_pdf_styles()
    body_style = styles["body"]
    body_center_bold_style = styles["body_center_bold"]
    header_style = styles["header"]
    story = [Paragraph("File 2: Article-wise Beneficiaries", styles["title"])]

    if not groups:
        story.append(Paragraph("No rows available for this report.", styles["empty"]))
    else:
        for index, group in enumerate(list(groups or []), start=1):
            story.append(Paragraph(f"{index}. {escape(str(group.get('article_name') or ''))}", styles["section"]))
            table_rows = [[
                Paragraph("Sl No", header_style),
                Paragraph("Beneficiary", header_style),
                Paragraph("Waiting Hall Qty", header_style),
            ]]
            for item_index, item in enumerate(list(group.get("beneficiaries") or []), start=1):
                table_rows.append(
                    [
                        Paragraph(str(item_index), body_center_bold_style),
                        Paragraph(escape(str(item.get("beneficiary_label") or "")), body_style),
                        Paragraph(str(int(item.get("quantity") or 0)), body_center_bold_style),
                    ]
                )
            table_rows.append(
                [
                    Paragraph("", styles["body_center"]),
                    Paragraph("Total", styles["body_bold"]),
                    Paragraph(str(int(group.get("total_quantity") or 0)), body_center_bold_style),
                ]
            )
            table = _segregation_pdf_table(table_rows, col_widths=[14 * mm, 128 * mm, 34 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (2, 0), (2, -1), "CENTER"),
                        ("BACKGROUND", (0, len(table_rows) - 1), (-1, len(table_rows) - 1), colors.HexColor("#eff6ff")),
                    ]
                )
            )
            story.extend([table, Spacer(1, 5 * mm)])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    doc.build(story, canvasmaker=_NumberedPdfCanvas)
    buffer.seek(0)
    return buffer


def generate_segregation_file3_pdf(rows: list[dict]) -> io.BytesIO:
    styles = _segregation_pdf_styles()
    body_style = styles["body"]
    body_center_bold_style = styles["body_center_bold"]
    header_style = styles["header"]
    story = [Paragraph("File 3: Stage Moving List", styles["title"])]

    if not rows:
        story.append(Paragraph("No rows available for this report.", styles["empty"]))
    else:
        table_rows = [[
            Paragraph("Seq No", header_style),
            Paragraph("Item", header_style),
            Paragraph("Token Qty", header_style),
            Paragraph("Start Token", header_style),
            Paragraph("End Token", header_style),
        ]]
        for row in list(rows or []):
            table_rows.append(
                [
                    Paragraph(str(row.get("sequence_no") or ""), body_center_bold_style),
                    Paragraph(escape(str(row.get("item_name") or "")), body_style),
                    Paragraph(str(int(row.get("token_quantity") or 0)), body_center_bold_style),
                    Paragraph(str(row.get("start_token_no") or ""), body_center_bold_style),
                    Paragraph(str(row.get("end_token_no") or ""), body_center_bold_style),
                ]
            )
        table = _segregation_pdf_table(table_rows, col_widths=[18 * mm, 92 * mm, 22 * mm, 24 * mm, 24 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (2, 0), (4, -1), "CENTER"),
                ]
            )
        )
        story.append(table)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=portrait(A4),
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    doc.build(story, canvasmaker=_NumberedPdfCanvas)
    buffer.seek(0)
    return buffer


def _segregation_write_sheet(worksheet, title: str, rows: list[dict]):
    worksheet.title = title
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = list(rows[0].keys()) if rows else []
    if not headers:
        worksheet.append(["No data"])
        worksheet["A1"].font = Font(size=11, bold=True)
        worksheet["A1"].alignment = center
        worksheet.column_dimensions["A"].width = 24
        return

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = center if isinstance(cell.value, (int, float)) else left
            cell.font = Font(size=10.5)

    for column_index, header in enumerate(headers, start=1):
        max_length = max(len(str(header or "")), *(len(str(worksheet.cell(row_idx, column_index).value or "")) for row_idx in range(2, worksheet.max_row + 1)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 36)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def generate_segregation_xlsx(*, master_rows: list[dict], file1_rows: list[dict], file2_rows: list[dict], file3_rows: list[dict]) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    _segregation_write_sheet(worksheet, "Master Data", master_rows)

    _segregation_write_sheet(workbook.create_sheet(), "File 1", file1_rows)
    _segregation_write_sheet(workbook.create_sheet(), "File 2", file2_rows)
    _segregation_write_sheet(workbook.create_sheet(), "File 3", file3_rows)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _stage_distribution_master_sheet_rows(rows: list[dict]) -> list[dict]:
    return [
        {
            "Application Number": str(row.get("application_number") or ""),
            "Beneficiary Type": str(row.get("beneficiary_type") or ""),
            "District": str(row.get("district") or ""),
            "Beneficiary Name": str(row.get("beneficiary_name") or ""),
            "Names": str(row.get("names") or ""),
            "Item Name": str(row.get("item_name") or ""),
            "Item Type": str(row.get("item_type") or ""),
            "Premise": str(row.get("premise") or ""),
            "Waiting Hall Qty": int(row.get("waiting_hall_quantity") or 0),
            "Token Qty": int(row.get("token_quantity") or 0),
            "Start Token No": int(row.get("start_token_no") or 0) if int(row.get("start_token_no") or 0) > 0 else "",
            "End Token No": int(row.get("end_token_no") or 0) if int(row.get("end_token_no") or 0) > 0 else "",
            "Sequence No": int(row.get("sequence_no") or 0) if int(row.get("sequence_no") or 0) > 0 else "",
        }
        for row in list(rows or [])
    ]


def _stage_distribution_write_file1_sheet(worksheet, rows: list[dict]):
    worksheet.title = "File 1"
    thin = Side(style="thin", color="7C8794")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Seq No", "Article Name", "Beneficiary Names", "Token Number"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row in list(rows or []):
        worksheet.append(
            [
                row.get("Seq No", ""),
                row.get("Article", ""),
                row.get("Beneficiary", ""),
                row.get("Token Number", ""),
            ]
        )

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=4):
        for index, cell in enumerate(row, start=1):
            cell.border = border
            cell.font = Font(size=10)
            cell.alignment = center if index in {1, 4} else left

    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 44
    worksheet.column_dimensions["D"].width = 18
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def _stage_distribution_write_grouped_sheet(
    worksheet,
    *,
    title: str,
    name_column_label: str,
    groups: list[dict],
):
    worksheet.title = title
    thin = Side(style="thin", color="7C8794")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    group_fill = PatternFill("solid", fgColor="EFF6FF")
    subtotal_fill = PatternFill("solid", fgColor="F8FAFC")
    grand_total_fill = PatternFill("solid", fgColor="E2E8F0")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    worksheet.append([name_column_label, "Qty"])
    for cell in worksheet[1]:
        cell.font = Font(size=11, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    grand_total = 0
    for group in list(groups or []):
        group_label = str(group.get("group_label") or "")
        items = list(group.get("items") or [])
        group_total = int(group.get("total_quantity") or 0)
        grand_total += group_total

        worksheet.append([group_label, ""])
        row_index = worksheet.max_row
        worksheet.cell(row_index, 1).font = Font(size=10.5, bold=True)
        worksheet.cell(row_index, 1).fill = group_fill
        worksheet.cell(row_index, 1).border = border
        worksheet.cell(row_index, 1).alignment = left
        worksheet.cell(row_index, 2).fill = group_fill
        worksheet.cell(row_index, 2).border = border
        worksheet.cell(row_index, 2).alignment = center

        for item in items:
            item_value = str(item.get("article_name") or item.get("beneficiary_name") or "")
            quantity = int(item.get("quantity") or 0)
            worksheet.append([f"        {item_value}", quantity])
            item_row = worksheet.max_row
            worksheet.cell(item_row, 1).font = Font(size=10)
            worksheet.cell(item_row, 1).border = border
            worksheet.cell(item_row, 1).alignment = left
            worksheet.cell(item_row, 2).font = Font(size=10, bold=True)
            worksheet.cell(item_row, 2).border = border
            worksheet.cell(item_row, 2).alignment = center

        worksheet.append(["Total", group_total])
        subtotal_row = worksheet.max_row
        worksheet.cell(subtotal_row, 1).font = Font(size=10, bold=True, color="1D4ED8")
        worksheet.cell(subtotal_row, 1).alignment = right
        worksheet.cell(subtotal_row, 1).fill = subtotal_fill
        worksheet.cell(subtotal_row, 1).border = border
        worksheet.cell(subtotal_row, 2).font = Font(size=10, bold=True, color="1D4ED8")
        worksheet.cell(subtotal_row, 2).alignment = center
        worksheet.cell(subtotal_row, 2).fill = subtotal_fill
        worksheet.cell(subtotal_row, 2).border = border

    worksheet.append(["Grand Total", grand_total])
    total_row = worksheet.max_row
    worksheet.cell(total_row, 1).font = Font(size=11, bold=True)
    worksheet.cell(total_row, 1).fill = grand_total_fill
    worksheet.cell(total_row, 1).alignment = left
    worksheet.cell(total_row, 1).border = border
    worksheet.cell(total_row, 2).font = Font(size=11, bold=True)
    worksheet.cell(total_row, 2).fill = grand_total_fill
    worksheet.cell(total_row, 2).alignment = center
    worksheet.cell(total_row, 2).border = border

    worksheet.column_dimensions["A"].width = 58
    worksheet.column_dimensions["B"].width = 12
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def generate_stage_distribution_xlsx(
    *,
    master_rows: list[dict],
    file1_rows: list[dict],
    file2_groups: list[dict],
    file3_groups: list[dict],
    file4_groups: list[dict],
    file5_groups: list[dict],
    file5_title: str,
) -> io.BytesIO:
    workbook = Workbook()
    master_sheet = workbook.active
    _segregation_write_sheet(master_sheet, "Master Data", master_rows)

    _stage_distribution_write_file1_sheet(workbook.create_sheet(), file1_rows)
    _stage_distribution_write_grouped_sheet(
        workbook.create_sheet(),
        title="File 2",
        name_column_label="District Name",
        groups=file2_groups,
    )
    _stage_distribution_write_grouped_sheet(
        workbook.create_sheet(),
        title="File 3",
        name_column_label="Public Name",
        groups=file3_groups,
    )
    _stage_distribution_write_grouped_sheet(
        workbook.create_sheet(),
        title="File 4",
        name_column_label="Institution Name",
        groups=file4_groups,
    )
    _stage_distribution_write_grouped_sheet(
        workbook.create_sheet(),
        title="File 5",
        name_column_label="Article Name",
        groups=file5_groups,
    )

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
