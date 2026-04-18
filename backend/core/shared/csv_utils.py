"""Shared CSV/file upload helper functions reused across modules."""

import csv
import io

from openpyxl import load_workbook

def _csv_reader_from_upload(uploaded_file):
    uploaded_file.seek(0)
    return csv.DictReader(io.StringIO(uploaded_file.read().decode("utf-8-sig")))

def _tabular_rows_from_upload(uploaded_file):
    name = str(getattr(uploaded_file, "name", "") or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        raw_headers = list(values[0] or [])
        headers = [str(header or "").strip() for header in raw_headers]
        rows = []
        for row_values in values[1:]:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell_value = row_values[index] if index < len(row_values) else ""
                row[header] = "" if cell_value is None else str(cell_value)
            if any(str(value or "").strip() for value in row.values()):
                rows.append(row)
        return headers, rows
    reader = _csv_reader_from_upload(uploaded_file)
    return list(reader.fieldnames or []), list(reader)

